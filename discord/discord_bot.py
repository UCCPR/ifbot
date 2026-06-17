"""
ZMDBot Discord Version - 完整功能实现
支持抽卡、战斗、队伍、签到、排行榜等功能
"""

import os
import random
import json
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

import discord
from discord import app_commands
from discord.ext import commands
from PIL import Image, ImageDraw, ImageFont
import openpyxl

# 导入战斗系统
try:
    from battle_system import BattleSystem, format_battle_result, format_boss_result, get_battle_help
    BATTLE_SYSTEM_LOADED = True
except ImportError as e:
    BATTLE_SYSTEM_LOADED = False
    class BattleSystem:
        def __init__(self, data): pass
        def start_battle(self, p_team, e_team, challenger="player", initial_player_sp=0): 
            return {"winner": "player", "rounds": 1, "log": [], "player_units": [], "enemy_units": []}
        def start_boss_battle(self, player_team, boss_card_id, initial_sp=300): 
            return {"boss_name": "???", "boss_starting_hp": 15000000, "boss_ending_hp": 15000000, 
                    "damage_dealt": 0, "damage_percent": 0, "rounds": 0, "player_survived": 0, 
                    "player_total": 0, "boss_killed": False, "log": [], "player_units": [], "enemy_units": []}
        def get_character(self, card_id): return None
        def _get_fallback_character(self, card_id): return None
    def format_battle_result(result): return "战斗系统未加载"
    def format_boss_result(result, include_log=False): return "战斗系统未加载"
    def get_battle_help(): return "战斗系统未加载"

# 导入GIF渲染器
try:
    from gif_renderer import battle_to_gif_new
    GIF_RENDERER_LOADED = True
except ImportError as e:
    GIF_RENDERER_LOADED = False
    def battle_to_gif_new(result): return None

# ========== 配置 ==========
CONFIG = {
    "TOKEN": "YOUR_DISCORD_BOT_TOKEN",
    "OWNER_ID": 0,
    "GUILD_ID": 0,
    "PREFIX": "!",
    "GACHA_COOLDOWN": 60,
    "MAX_DAILY_GACHA": 10,
    "DAILY_REWARD": 30000,
    "GACHA_COST": 5000,
    "TEN_GACHA_COST": 45000
}

# 目录配置
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
INFO_DIR = BASE_DIR.parent / "info"
OUTPUT_DIR = BASE_DIR.parent / "output"
ICON_DIR = BASE_DIR.parent / "iconimage"
BACKUP_DIR = BASE_DIR.parent / "backup"
XLSX_FILE = BASE_DIR.parent / "卡牌信息.xlsx"
BATTLE_XLSX = BASE_DIR.parent / "cards_completed.xlsx"

# 确保目录存在
for dir_path in [DATA_DIR, INFO_DIR, OUTPUT_DIR, BACKUP_DIR]:
    dir_path.mkdir(parents=True, exist_ok=True)

# 全局变量
CHARACTERS = {}  # 抽卡角色数据
BATTLE_CHARACTERS = {}  # 战斗角色数据
BATTLE_INSTANCE = None
USER_DATA = {}
PITY_DATA = {}
RANKING_DATA = {}

# 抽卡权重配置
GACHA_WEIGHTS = {
    "3star": 6,
    "2star": 3,
    "1star": 1
}

# ========== 日志模块 ==========
def log_info(message: str):
    """记录普通信息"""
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "discord_bot.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")
    print(f"[INFO] {message}", flush=True)

def log_error(message: str):
    """记录错误信息"""
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "discord_error.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] ERR {message}\n")
    print(f"[ERROR] {message}", flush=True)

# ========== 数据加载模块 ==========
def load_characters():
    """加载抽卡角色数据"""
    global CHARACTERS
    CHARACTERS = {}
    
    if not XLSX_FILE.exists():
        log_error(f"抽卡数据文件不存在: {XLSX_FILE}")
        return
    
    try:
        wb = openpyxl.load_workbook(XLSX_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            char = {
                "id": str(row[0]) if row[0] else "",
                "name": row[1] if row[1] else "未知",
                "rarity": row[2] if row[2] else 1,
                "element": row[3] if row[3] else "",
                "role": row[4] if row[4] else "",
                "type": row[5] if row[5] else "",
                "atk": row[6] if row[6] else 0,
                "def": row[7] if row[7] else 0,
                "hp": row[8] if row[8] else 0,
                "sp_cost": row[9] if row[9] else 0,
                "special_skill": row[10] if row[10] else "",
                "image_url": row[11] if row[11] else "",
                "description": row[12] if row[12] else ""
            }
            CHARACTERS[char["id"]] = char
        
        log_info(f"[抽卡] 加载了 {len(CHARACTERS)} 个角色")
    except Exception as e:
        log_error(f"加载抽卡数据失败: {e}")

def load_battle_characters():
    """加载战斗角色数据"""
    global BATTLE_CHARACTERS, BATTLE_INSTANCE
    BATTLE_CHARACTERS = {}
    
    if not BATTLE_XLSX.exists():
        log_error(f"战斗数据文件不存在: {BATTLE_XLSX}")
        return
    
    try:
        wb = openpyxl.load_workbook(BATTLE_XLSX)
        ws = wb.active
        characters_data = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            char = {
                "id": str(row[0]) if row[0] else "",
                "name": row[1] if row[1] else "未知",
                "rarity": row[2] if row[2] else 1,
                "element": row[3] if row[3] else "",
                "role": row[4] if row[4] else "",
                "atk": row[5] if row[5] else 0,
                "def": row[6] if row[6] else 0,
                "hp": row[7] if row[7] else 0,
                "sp_cost": row[8] if row[8] else 0,
                "special_skill": row[9] if row[9] else "",
                "image_url": row[10] if row[10] else "",
                "description": row[11] if row[11] else ""
            }
            BATTLE_CHARACTERS[char["id"]] = char
            characters_data.append(char)
        
        BATTLE_INSTANCE = BattleSystem(characters_data)
        log_info(f"[战斗] 加载了 {len(BATTLE_CHARACTERS)} 个角色")
    except Exception as e:
        log_error(f"加载战斗数据失败: {e}")

# ========== 用户数据模块 ==========
def get_user_data_path(user_id):
    """获取用户数据文件路径"""
    return DATA_DIR / f"user_{user_id}.json"

def load_user_data(user_id):
    """加载用户数据"""
    user_id = str(user_id)
    if user_id in USER_DATA:
        return USER_DATA[user_id]
    
    file_path = get_user_data_path(user_id)
    if file_path.exists():
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                USER_DATA[user_id] = json.load(f)
        except Exception as e:
            log_error(f"用户 {user_id} 的数据文件格式错误: {e}")
            USER_DATA[user_id] = get_default_user_data()
    else:
        USER_DATA[user_id] = get_default_user_data()
    
    return USER_DATA[user_id]

def get_default_user_data():
    """获取默认用户数据"""
    return {
        "gacha": 0,
        "cards": [],
        "team": {"battle_cards": [], "assist_cards": []},
        "presets": [],
        "active_preset": 0,
        "battle_history": [],
        "signin": {
            "last_date": "",
            "streak": 0,
            "total_days": 0
        }
    }

def save_user_data(user_id):
    """保存用户数据"""
    user_id = str(user_id)
    if user_id in USER_DATA:
        file_path = get_user_data_path(user_id)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(USER_DATA[user_id], f, ensure_ascii=False, indent=2)

# ========== 保底数据模块 ==========
def get_pity_data_path(user_id):
    """获取保底数据文件路径"""
    return DATA_DIR / f"pity_{user_id}.json"

def load_pity_data(user_id):
    """加载保底数据"""
    user_id = str(user_id)
    if user_id in PITY_DATA:
        return PITY_DATA[user_id]
    
    file_path = get_pity_data_path(user_id)
    if file_path.exists():
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                PITY_DATA[user_id] = json.load(f)
        except Exception as e:
            log_error(f"用户 {user_id} 的保底数据文件格式错误: {e}")
            PITY_DATA[user_id] = get_default_pity_data()
    else:
        PITY_DATA[user_id] = get_default_pity_data()
    
    return PITY_DATA[user_id]

def get_default_pity_data():
    """获取默认保底数据"""
    return {
        "normal_pity": 0,
        "ten_pity": 0,
        "last_3star": 0,
        "total_gacha": 0
    }

def save_pity_data(user_id):
    """保存保底数据"""
    user_id = str(user_id)
    if user_id in PITY_DATA:
        file_path = get_pity_data_path(user_id)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(PITY_DATA[user_id], f, ensure_ascii=False, indent=2)

# ========== 排行榜模块 ==========
def load_ranking_data():
    """加载排行榜数据"""
    global RANKING_DATA
    file_path = DATA_DIR / "ranking.json"
    
    if file_path.exists():
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                RANKING_DATA = json.load(f)
        except Exception as e:
            log_error(f"加载排行榜数据失败: {e}")
            RANKING_DATA = {"power": [], "gacha": []}
    else:
        RANKING_DATA = {"power": [], "gacha": []}

def save_ranking_data():
    """保存排行榜数据"""
    file_path = DATA_DIR / "ranking.json"
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(RANKING_DATA, f, ensure_ascii=False, indent=2)

# ========== 抽卡逻辑 ==========
def gacha_draw(count=1):
    """执行抽卡"""
    if not CHARACTERS:
        return {"results": [], "remaining_pity": 0, "got_3star": False}
    
    results = []
    got_3star = False
    
    for _ in range(count):
        # 根据权重随机抽卡
        total_weight = sum(GACHA_WEIGHTS.values())
        rand = random.randint(1, total_weight)
        current = 0
        
        rarity = "1star"
        for r, w in GACHA_WEIGHTS.items():
            current += w
            if rand <= current:
                rarity = r
                break
        
        # 获取对应稀有度的角色列表
        chars_of_rarity = [c for c in CHARACTERS.values() if c["rarity"] == int(rarity[0])]
        if not chars_of_rarity:
            chars_of_rarity = list(CHARACTERS.values())
        
        card = random.choice(chars_of_rarity)
        results.append(card.copy())
        
        if card["rarity"] == 3:
            got_3star = True
    
    return {"results": results, "remaining_pity": 0, "got_3star": got_3star}

# ========== 签到系统 ==========
def signin(user_id):
    """执行签到"""
    user_data = load_user_data(user_id)
    today = datetime.now().strftime("%Y-%m-%d")
    
    if user_data["signin"]["last_date"] == today:
        return {"success": False, "message": "今天已经签到过了", "streak": user_data["signin"]["streak"]}
    
    # 更新签到记录
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    if user_data["signin"]["last_date"] == yesterday:
        user_data["signin"]["streak"] += 1
    else:
        user_data["signin"]["streak"] = 1
    
    user_data["signin"]["last_date"] = today
    user_data["signin"]["total_days"] += 1
    user_data["gacha"] += CONFIG["DAILY_REWARD"]
    
    save_user_data(user_id)
    
    return {
        "success": True,
        "message": f"签到成功！获得 {CONFIG['DAILY_REWARD']} 呱太",
        "streak": user_data["signin"]["streak"],
        "total_days": user_data["signin"]["total_days"],
        "gacha": user_data["gacha"]
    }

# ========== 卡牌图片生成 ==========
def generate_card_image(card):
    """生成卡牌图片"""
    try:
        icon_path = ICON_DIR / f"card_cutin_{card['id']}.png"
        if icon_path.exists():
            img = Image.open(icon_path)
            img = img.resize((150, 150), Image.Resampling.LANCZOS)
        else:
            img = Image.new('RGBA', (150, 150), (30, 30, 30))
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("C:/Windows/Fonts/msgothic.ttc", 12)
            except:
                font = ImageFont.load_default()
            draw.text((10, 60), card['name'][:8], fill=(255, 255, 255), font=font)
        
        draw = ImageDraw.Draw(img)
        border_color = {3: (255, 215, 0), 2: (147, 112, 219), 1: (34, 139, 34)}.get(card['rarity'], (128, 128, 128))
        draw.rectangle([(0, 0), (img.width-1, img.height-1)], outline=border_color, width=3)
        
        return img
    except Exception as e:
        log_error(f"生成卡牌图片失败: {e}")
        return None

# ========== Discord Bot Cogs ==========
class GachaCog(commands.Cog):
    """抽卡相关命令"""
    
    def __init__(self, bot):
        self.bot = bot
    
    @app_commands.command(name="gacha", description="进行单抽")
    async def gacha(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        pity_data = load_pity_data(interaction.user.id)
        
        if user_data["gacha"] < CONFIG["GACHA_COST"]:
            await interaction.response.send_message(f"呱太不足！需要 {CONFIG['GACHA_COST']} 呱太，当前 {user_data['gacha']} 呱太", ephemeral=True)
            return
        
        # 扣除呱太
        user_data["gacha"] -= CONFIG["GACHA_COST"]
        
        # 执行抽卡
        result = gacha_draw(1)
        card = result["results"][0]
        
        # 更新保底
        pity_data["normal_pity"] += 1
        pity_data["total_gacha"] += 1
        
        if card["rarity"] == 3:
            pity_data["normal_pity"] = 0
            pity_data["last_3star"] = pity_data["total_gacha"]
        
        # 保存卡牌
        user_data["cards"].append(card)
        
        # 保存数据
        save_user_data(interaction.user.id)
        save_pity_data(interaction.user.id)
        
        # 发送结果
        rarity_text = {3: "⭐⭐⭐", 2: "⭐⭐", 1: "⭐"}.get(card["rarity"], "⭐")
        
        embed = discord.Embed(title="抽卡结果", description=f"{rarity_text} {card['name']}")
        img = generate_card_image(card)
        
        if img:
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            file = discord.File(buffer, filename="card.png")
            embed.set_image(url="attachment://card.png")
            embed.set_footer(text=f"剩余呱太: {user_data['gacha']} | 保底进度: {pity_data['normal_pity']}/10")
            await interaction.response.send_message(embed=embed, file=file)
        else:
            await interaction.response.send_message(f"抽卡成功！{rarity_text} {card['name']}\n剩余呱太: {user_data['gacha']}")
    
    @app_commands.command(name="tenpull", description="进行十连抽")
    async def tenpull(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        pity_data = load_pity_data(interaction.user.id)
        
        if user_data["gacha"] < CONFIG["TEN_GACHA_COST"]:
            await interaction.response.send_message(f"呱太不足！需要 {CONFIG['TEN_GACHA_COST']} 呱太，当前 {user_data['gacha']} 呱太", ephemeral=True)
            return
        
        # 扣除呱太
        user_data["gacha"] -= CONFIG["TEN_GACHA_COST"]
        
        # 执行抽卡
        result = gacha_draw(10)
        cards = result["results"]
        
        # 更新保底
        pity_data["ten_pity"] += 1
        pity_data["total_gacha"] += 10
        
        got_3star = False
        for card in cards:
            user_data["cards"].append(card)
            if card["rarity"] == 3:
                pity_data["ten_pity"] = 0
                pity_data["last_3star"] = pity_data["total_gacha"]
                got_3star = True
        
        # 保存数据
        save_user_data(interaction.user.id)
        save_pity_data(interaction.user.id)
        
        # 发送结果
        three_stars = [c for c in cards if c["rarity"] == 3]
        two_stars = [c for c in cards if c["rarity"] == 2]
        one_stars = [c for c in cards if c["rarity"] == 1]
        
        result_text = "十连抽结果：\n"
        if three_stars:
            result_text += "⭐⭐⭐: " + ", ".join([c["name"] for c in three_stars]) + "\n"
        if two_stars:
            result_text += "⭐⭐: " + ", ".join([c["name"] for c in two_stars]) + "\n"
        if one_stars:
            result_text += "⭐: " + ", ".join([c["name"] for c in one_stars])
        
        embed = discord.Embed(title="十连抽结果", description=result_text)
        embed.set_footer(text=f"剩余呱太: {user_data['gacha']} | 保底进度: {pity_data['ten_pity']}/10")
        
        await interaction.response.send_message(embed=embed)
    
    @app_commands.command(name="collection", description="查看卡牌收藏")
    async def collection(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        cards = user_data["cards"]
        
        if not cards:
            await interaction.response.send_message("你还没有抽过卡哦！", ephemeral=True)
            return
        
        three_stars = [c for c in cards if c["rarity"] == 3]
        two_stars = [c for c in cards if c["rarity"] == 2]
        one_stars = [c for c in cards if c["rarity"] == 1]
        
        embed = discord.Embed(
            title=f"我的卡牌收藏",
            description=f"总计: {len(cards)} 张\n⭐⭐⭐: {len(three_stars)} | ⭐⭐: {len(two_stars)} | ⭐: {len(one_stars)}"
        )
        
        if three_stars:
            embed.add_field(name="⭐⭐⭐", value=", ".join([c["name"] for c in three_stars][:10]))
        if two_stars:
            embed.add_field(name="⭐⭐", value=", ".join([c["name"] for c in two_stars][:10]))
        if one_stars:
            embed.add_field(name="⭐", value=", ".join([c["name"] for c in one_stars][:10]))
        
        await interaction.response.send_message(embed=embed, ephemeral=True)
    
    @app_commands.command(name="pity", description="查看保底进度")
    async def pity(self, interaction: discord.Interaction):
        pity_data = load_pity_data(interaction.user.id)
        
        embed = discord.Embed(title="保底进度")
        embed.add_field(name="单抽保底", value=f"{pity_data['normal_pity']}/10", inline=False)
        embed.add_field(name="十连保底", value=f"{pity_data['ten_pity']}/10", inline=False)
        embed.add_field(name="累计抽卡", value=str(pity_data['total_gacha']), inline=False)
        
        await interaction.response.send_message(embed=embed, ephemeral=True)

class BattleCog(commands.Cog):
    """战斗相关命令"""
    
    def __init__(self, bot):
        self.bot = bot
    
    @app_commands.command(name="battle", description="进行战斗")
    async def battle(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        team = user_data["team"]
        
        if not team["battle_cards"]:
            await interaction.response.send_message("请先设置队伍！使用 /setteam 命令", ephemeral=True)
            return
        
        # 随机选择敌方队伍
        all_cards = list(BATTLE_CHARACTERS.values())
        if not all_cards:
            await interaction.response.send_message("战斗数据未加载！", ephemeral=True)
            return
        
        enemy_cards = random.sample(all_cards, min(3, len(all_cards)))
        enemy_team = {"battle_cards": enemy_cards}
        
        # 进行战斗
        try:
            result = BATTLE_INSTANCE.start_battle(team, enemy_team)
            winner = result.get("winner", "unknown")
            rounds = result.get("round", 0)
            
            # 保存战斗记录
            user_data["battle_history"].append({
                "time": datetime.now().isoformat(),
                "result": winner,
                "rounds": rounds,
                "log": result.get("log", []),
                "player_units": result.get("player_units", []),
                "enemy_units": result.get("enemy_units", []),
                "parsable_log": result.get("parsable_log", [])
            })
            user_data["battle_history"] = user_data["battle_history"][-10:]
            save_user_data(interaction.user.id)
            
            # 生成GIF
            gif_path = None
            if GIF_RENDERER_LOADED:
                gif_path = battle_to_gif_new(result)
            
            # 发送结果
            embed = discord.Embed(
                title="战斗结果",
                description=f"{'🎉 胜利！' if winner == 'player' else '💀 失败...'} 共 {rounds} 回合"
            )
            
            if gif_path and os.path.exists(gif_path):
                file = discord.File(gif_path, filename="battle.gif")
                embed.set_image(url="attachment://battle.gif")
                await interaction.response.send_message(embed=embed, file=file)
            else:
                await interaction.response.send_message(embed=embed)
                
        except Exception as e:
            log_error(f"战斗失败: {e}")
            await interaction.response.send_message(f"战斗发生错误: {str(e)}", ephemeral=True)
    
    @app_commands.command(name="battlegif", description="生成最近战斗的GIF")
    async def battlegif(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        
        if not user_data["battle_history"]:
            await interaction.response.send_message("没有战斗记录！", ephemeral=True)
            return
        
        last_battle = user_data["battle_history"][-1]
        
        if GIF_RENDERER_LOADED:
            gif_path = battle_to_gif_new(last_battle)
            
            if gif_path and os.path.exists(gif_path):
                file = discord.File(gif_path, filename="battle.gif")
                embed = discord.Embed(title="战斗回放")
                embed.set_image(url="attachment://battle.gif")
                await interaction.response.send_message(embed=embed, file=file)
            else:
                await interaction.response.send_message("生成GIF失败！", ephemeral=True)
        else:
            await interaction.response.send_message("GIF渲染器未加载！", ephemeral=True)

class TeamCog(commands.Cog):
    """队伍相关命令"""
    
    def __init__(self, bot):
        self.bot = bot
    
    @app_commands.command(name="team", description="查看当前队伍")
    async def team(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        team = user_data["team"]
        
        if not team["battle_cards"]:
            await interaction.response.send_message("你还没有设置队伍！", ephemeral=True)
            return
        
        embed = discord.Embed(title="我的队伍")
        for i, card in enumerate(team["battle_cards"]):
            embed.add_field(
                name=f"位置{i+1}",
                value=f"{card['name']}\n⭐{card['rarity']} | HP: {card['hp']} | ATK: {card['atk']}",
                inline=False
            )
        
        await interaction.response.send_message(embed=embed, ephemeral=True)
    
    @app_commands.command(name="setteam", description="设置战斗队伍")
    @app_commands.describe(card1="第一张卡牌ID", card2="第二张卡牌ID", card3="第三张卡牌ID")
    async def setteam(self, interaction: discord.Interaction, card1: str, card2: str, card3: str):
        user_data = load_user_data(interaction.user.id)
        user_cards = {c["id"]: c for c in user_data["cards"]}
        
        team_cards = []
        for card_id in [card1, card2, card3]:
            if card_id in user_cards:
                team_cards.append(user_cards[card_id])
            elif card_id in BATTLE_CHARACTERS:
                team_cards.append(BATTLE_CHARACTERS[card_id])
            else:
                await interaction.response.send_message(f"卡牌 {card_id} 不存在！", ephemeral=True)
                return
        
        user_data["team"] = {"battle_cards": team_cards, "assist_cards": []}
        save_user_data(interaction.user.id)
        
        card_names = ", ".join([c["name"] for c in team_cards])
        await interaction.response.send_message(f"队伍设置成功！\n成员: {card_names}", ephemeral=True)

class SocialCog(commands.Cog):
    """社交相关命令"""
    
    def __init__(self, bot):
        self.bot = bot
    
    @app_commands.command(name="signin", description="每日签到")
    async def signin(self, interaction: discord.Interaction):
        result = signin(interaction.user.id)
        
        if result["success"]:
            embed = discord.Embed(
                title="签到成功！",
                description=f"获得 {CONFIG['DAILY_REWARD']} 呱太\n连续签到: {result['streak']} 天\n累计签到: {result['total_days']} 天"
            )
            embed.set_footer(text=f"当前呱太: {result['gacha']}")
            await interaction.response.send_message(embed=embed)
        else:
            await interaction.response.send_message(result["message"], ephemeral=True)
    
    @app_commands.command(name="balance", description="查看呱太余额")
    async def balance(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        
        embed = discord.Embed(title="呱太余额")
        embed.add_field(name="当前呱太", value=str(user_data["gacha"]), inline=False)
        embed.add_field(name="单抽", value=f"{CONFIG['GACHA_COST']} 呱太", inline=True)
        embed.add_field(name="十连", value=f"{CONFIG['TEN_GACHA_COST']} 呱太", inline=True)
        
        await interaction.response.send_message(embed=embed, ephemeral=True)
    
    @app_commands.command(name="profile", description="查看个人资料")
    async def profile(self, interaction: discord.Interaction):
        user_data = load_user_data(interaction.user.id)
        pity_data = load_pity_data(interaction.user.id)
        
        cards = user_data["cards"]
        three_stars = [c for c in cards if c["rarity"] == 3]
        
        embed = discord.Embed(title=f"{interaction.user.name} 的资料")
        embed.add_field(name="呱太", value=str(user_data["gacha"]), inline=True)
        embed.add_field(name="卡牌总数", value=str(len(cards)), inline=True)
        embed.add_field(name="三星卡牌", value=str(len(three_stars)), inline=True)
        embed.add_field(name="连续签到", value=f"{user_data['signin']['streak']} 天", inline=True)
        embed.add_field(name="累计签到", value=f"{user_data['signin']['total_days']} 天", inline=True)
        embed.add_field(name="累计抽卡", value=str(pity_data["total_gacha"]), inline=True)
        
        await interaction.response.send_message(embed=embed, ephemeral=True)

class AdminCog(commands.Cog):
    """管理员命令"""
    
    def __init__(self, bot):
        self.bot = bot
    
    @app_commands.command(name="reload", description="重新加载数据")
    async def reload(self, interaction: discord.Interaction):
        if interaction.user.id != CONFIG["OWNER_ID"]:
            await interaction.response.send_message("你没有权限执行此命令！", ephemeral=True)
            return
        
        load_characters()
        load_battle_characters()
        load_ranking_data()
        
        await interaction.response.send_message("数据已重新加载！", ephemeral=True)
    
    @app_commands.command(name="backup", description="备份数据")
    async def backup(self, interaction: discord.Interaction):
        if interaction.user.id != CONFIG["OWNER_ID"]:
            await interaction.response.send_message("你没有权限执行此命令！", ephemeral=True)
            return
        
        today = datetime.now().strftime("%Y-%m-%d")
        backup_dir = BACKUP_DIR / today
        backup_dir.mkdir(exist_ok=True)
        
        # 备份用户数据
        user_files = list(DATA_DIR.glob("*.json"))
        for file in user_files:
            dest_file = backup_dir / file.name
            with open(file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        await interaction.response.send_message(f"备份完成！共备份 {len(user_files)} 个文件", ephemeral=True)

class ZMDiscordBot(commands.Bot):
    """主机器人类"""
    
    def __init__(self):
        intents = discord.Intents.default()
        intents.message_content = True
        super().__init__(command_prefix=CONFIG["PREFIX"], intents=intents)
    
    async def setup_hook(self):
        """设置钩子"""
        await self.add_cog(GachaCog(self))
        await self.add_cog(BattleCog(self))
        await self.add_cog(TeamCog(self))
        await self.add_cog(SocialCog(self))
        await self.add_cog(AdminCog(self))
        
        # 同步命令
        try:
            if CONFIG["GUILD_ID"]:
                guild = self.get_guild(CONFIG["GUILD_ID"])
                if guild:
                    self.tree.copy_global_to(guild=guild)
                    synced = await self.tree.sync(guild=guild)
                    print(f"已同步 {len(synced)} 个命令到服务器")
                else:
                    synced = await self.tree.sync()
                    print(f"已全局同步 {len(synced)} 个命令")
            else:
                synced = await self.tree.sync()
                print(f"已全局同步 {len(synced)} 个命令")
        except Exception as e:
            print(f"同步命令失败: {e}")
    
    async def on_ready(self):
        """机器人就绪"""
        print(f"Logged in as {self.user} (ID: {self.user.id})")
        print("------")

def main():
    """主函数"""
    # 加载配置
    config_path = BASE_DIR / "config.json"
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            CONFIG.update(json.load(f))
    
    if not CONFIG["TOKEN"] or CONFIG["TOKEN"] == "YOUR_DISCORD_BOT_TOKEN":
        print("请先在 config.json 中设置 Discord Bot Token！")
        return
    
    # 加载数据
    load_characters()
    load_battle_characters()
    load_ranking_data()
    
    # 启动机器人
    bot = ZMDiscordBot()
    bot.run(CONFIG["TOKEN"])

if __name__ == "__main__":
    main()