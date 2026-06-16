#!/usr/bin/env python3
"""
ZMDBot KOOK Version - 使用Webhook方式与KOOK通信
支持抽卡、战斗、队伍、签到、排行榜等功能
运行在Ubuntu服务器上

根据KOOK官方文档实现：
- https://developer.kookapp.cn/doc/webhook
- https://developer.kookapp.cn/doc/event/event-introduction
"""

import os
import sys
import random
import json
import uuid
import hashlib
import zlib
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

try:
    from Crypto.Cipher import AES
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False

sys.path.insert(0, str(Path(__file__).parent))

from flask import Flask, request, jsonify
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import requests

try:
    from battle_system import BattleSystem, format_battle_result, format_boss_result, get_battle_help
    BATTLE_SYSTEM_LOADED = True
    print("[INFO] 战斗系统加载成功")
except ImportError as e:
    BATTLE_SYSTEM_LOADED = False
    print(f"[WARN] 战斗系统加载失败: {e}")
    class BattleSystem:
        def __init__(self, data): pass
        def start_battle(self, p_team, e_team, challenger="player", initial_player_sp=0): 
            return {"winner": "player", "rounds": 1, "log": [], "player_units": [], "enemy_units": []}
        def start_boss_battle(self, player_team, boss_card_id, initial_sp=300): 
            return {"boss_name": "???", "boss_starting_hp": 15000000, "boss_ending_hp": 15000000, 
                    "damage_dealt": 0, "damage_percent": 0, "rounds": 0, "player_survived": 0, 
                    "player_total": 0, "boss_killed": False, "log": [], "player_units": [], "enemy_units": []}
        def get_character(self, card_id): return None
        def _get_fallback_character(self, card_id): return None
    def format_battle_result(result): return "战斗系统未加载"
    def format_boss_result(result, include_log=False): return "战斗系统未加载"
    def get_battle_help(): return "战斗系统未加载"

try:
    from gif_renderer import battle_to_gif_new, battle_to_gif_bytes
    GIF_RENDERER_LOADED = True
    print("[INFO] GIF渲染器加载成功")
except ImportError as e:
    GIF_RENDERER_LOADED = False
    print(f"[WARN] GIF渲染器加载失败: {e}")
    def battle_to_gif_new(result): return None
    def battle_to_gif_bytes(result): return None

# 导入配队系统
try:
    from team_system import (
        load_team_data,
        save_team_data,
        get_user_3star_cards,
        build_team_image,
        build_vs_team_image,
        build_3star_cards_image,
        set_team_card,
        clear_team_card,
        clear_all_team,
        get_team_info,
        auto_build_team,
        auto_save_preset,
        load_preset,
        load_presets,
        save_presets,
        list_presets_info,
        get_defense_slot,
        set_defense_slot,
        get_defense_team,
        get_defense_team_info
    )
    TEAM_SYSTEM_LOADED = True
    print("[INFO] 配队系统加载成功")
except ImportError as e:
    TEAM_SYSTEM_LOADED = False
    print(f"[WARN] 配队系统加载失败: {e}")
    # 如果加载失败，定义空函数作为备用
    def load_team_data(user_id): return {"battle_cards": [], "assist_cards": []}
    def save_team_data(user_id, data): pass
    def get_user_3star_cards(user_id, characters=None): return []
    def build_team_image(team_data, characters): return None
    def build_vs_team_image(player_team, enemy_team, characters): return None
    def build_3star_cards_image(user_id, characters, page=1, page_size=50): return None, [], 0
    def auto_build_team(user_id, characters): return {"success": False, "message": "配队系统未加载", "team": None}
    def auto_save_preset(user_id): return -1
    def load_preset(user_id, slot): return False
    def load_presets(user_id): return {"presets": [], "active_slot": 0}
    def save_presets(user_id, data): pass
    def list_presets_info(user_id, characters): return "配队系统未加载"
    def set_team_card(user_id, position, card_id, card_type="battle"): return False
    def clear_team_card(user_id, position, card_type="battle"): return False
    def clear_all_team(user_id): pass
    def get_team_info(user_id, characters): return "配队系统未加载"
    def get_defense_slot(user_id): return 1
    def set_defense_slot(user_id, slot): return False
    def get_defense_team(user_id): return {"battle_cards": [], "assist_cards": []}
    def get_defense_team_info(user_id, characters): return "配队系统未加载"

# 基础目录定义（必须在load_config之前）
BASE_DIR = Path(__file__).parent

def load_config():
    """从config.json加载配置（与QQ版一致）"""
    config_path = BASE_DIR / "config.json"
    default_config = {
        "WEBHOOK_SECRET": "",
        "BOT_TOKEN": "",
        "VERIFY_TOKEN": "",
        "ENCRYPT_KEY": "",
        "OWNER_ID": "0",
        "SERVER_ID": "0",
        "HOST": "0.0.0.0",
        "PORT": 5000,
        
        "PITY_LIMIT": 150,
        "GACHA_COST": 300,
        "GACHA10_COST": 3000,
        "GACHA10_COOLDOWN_SECONDS": 60,
        "GET_GACHA_COOLDOWN_SECONDS": 60,
        "LIMITED_GACHA_COST": 15000,
        "LIMITED_GACHA_COOLDOWN_SECONDS": 480,
        "GET_GACHA_REWARD": 10000,
        "DAILY_REWARD": 30000,
        
        "MYSTERY_BOX_CHANCE": 0.02,
        "MUTATION_NO_CHANGE": 0.88,
        "MUTATION_1_TO_2": 0.08,
        "MUTATION_1_TO_3": 0.02,
        "MUTATION_2_TO_3": 0.05,
        "BOX_OPEN_TIMEOUT": 300,
        
        "THREE_STAR_POOL_RED_COST": 1500,
        "THREE_STAR_POOL_BLUE_COST": 350,
        
        "FES_LIMIT_PROB": 0.25,
        "PERIOD_LIMIT_PROB": 0.35,
        "OTHER_3STAR_PROB": 0.40,
        
        "MYSTERY_BOX_2STAR_PROB": 65,
        "MYSTERY_BOX_3STAR_PROB": 35,
        "NORMAL_BOX_1STAR_PROB": 72,
        "NORMAL_BOX_2STAR_PROB": 23,
        "NORMAL_BOX_3STAR_PROB": 3,
        
        "GACHA_1STAR_PROB": 72,
        "GACHA_2STAR_PROB": 23,
        "GACHA_3STAR_PROB": 3,
        
        "BOSS_CARD_ID": "100430006",
        "BOSS_BATTLE_COOLDOWN_SECONDS": 60,
        
        "DEBUG_MODE": False,
        "LOG_LEVEL": "INFO"
    }
    
    if config_path.exists():
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                loaded_config = json.load(f)
                # 合并配置，以加载的配置为准
                default_config.update(loaded_config)
                print("[INFO] 配置文件加载成功")
        except json.JSONDecodeError as e:
            print(f"[ERROR] 配置文件格式错误: {e}")
        except IOError as e:
            print(f"[ERROR] 读取配置文件失败: {e}")
    else:
        print("[WARNING] 配置文件不存在，使用默认配置")
    
    return default_config

CONFIG = load_config()

# 验证BOT_TOKEN是否已配置
if not CONFIG["BOT_TOKEN"] or CONFIG["BOT_TOKEN"] == "请在这里输入你的KOOK Bot Token":
    print("[ERROR] 请先在 config.json 中设置 KOOK Bot Token！")
    exit(1)

INFO_DIR = BASE_DIR / "info"
OUTPUT_DIR = BASE_DIR / "output"
ICON_DIR = BASE_DIR / "iconimage"
LEVEL_DIR = BASE_DIR / "level"
BACKUP_DIR = BASE_DIR / "backup"
XLSX_FILE = BASE_DIR / "卡牌信息.xlsx"
BATTLE_XLSX = BASE_DIR / "cards_completed.xlsx"

# 备份记录文件（与QQ版一致）
BACKUP_RECORD_FILE = INFO_DIR / "last_backup.json"

# 裁剪比例（与QQ版一致）
CROP_LEFT_RATIO = 0.25
CROP_RIGHT_RATIO = 0.75
CROP_TOP_RATIO = 0.15
CROP_BOTTOM_RATIO = 0.65

# 确保目录存在（与QQ版一致）
INFO_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

CHARACTERS = {}
BATTLE_CHARACTERS = {}
BATTLE_INSTANCE = None
USER_DATA = {}
PITY_DATA = {}
RANKING_DATA = {}
BOX_SESSIONS = {}  # 盲盒会话状态 {user_id: {"boxes": [], "opened": [], "created_at": timestamp, "characters": {}}}

# 抽卡概率配置（从CONFIG加载，与QQ版一致）
GACHA_1STAR_PROB = CONFIG["GACHA_1STAR_PROB"]
GACHA_2STAR_PROB = CONFIG["GACHA_2STAR_PROB"]
GACHA_3STAR_PROB = CONFIG["GACHA_3STAR_PROB"]

GACHA_WEIGHTS = [GACHA_1STAR_PROB, GACHA_2STAR_PROB, GACHA_3STAR_PROB]

# 盲盒配置（从CONFIG加载）
MYSTERY_BOX_CHANCE = CONFIG["MYSTERY_BOX_CHANCE"]
MUTATION_NO_CHANGE = CONFIG["MUTATION_NO_CHANGE"]
MUTATION_1_TO_2 = CONFIG["MUTATION_1_TO_2"]
MUTATION_1_TO_3 = CONFIG["MUTATION_1_TO_3"]
MUTATION_2_TO_3 = CONFIG["MUTATION_2_TO_3"]

NORMAL_BOX_1STAR_PROB = CONFIG["NORMAL_BOX_1STAR_PROB"]
NORMAL_BOX_2STAR_PROB = CONFIG["NORMAL_BOX_2STAR_PROB"]
NORMAL_BOX_3STAR_PROB = CONFIG["NORMAL_BOX_3STAR_PROB"]
MYSTERY_BOX_2STAR_PROB = CONFIG["MYSTERY_BOX_2STAR_PROB"]
MYSTERY_BOX_3STAR_PROB = CONFIG["MYSTERY_BOX_3STAR_PROB"]

NORMAL_BOX_WEIGHTS = [NORMAL_BOX_1STAR_PROB, NORMAL_BOX_2STAR_PROB, NORMAL_BOX_3STAR_PROB]
MYSTERY_BOX_WEIGHTS = [MYSTERY_BOX_2STAR_PROB, MYSTERY_BOX_3STAR_PROB]

# 三星内部分配概率（从CONFIG加载）
FES_LIMIT_PROB = CONFIG["FES_LIMIT_PROB"]
PERIOD_LIMIT_PROB = CONFIG["PERIOD_LIMIT_PROB"]
OTHER_3STAR_PROB = CONFIG["OTHER_3STAR_PROB"]

KOOK_API_URL = "https://www.kookapp.cn/api/v3"

# 三星池子配置（从CONFIG加载，与QQ版一致）
THREE_STAR_POOL_RED_COST = CONFIG["THREE_STAR_POOL_RED_COST"]
THREE_STAR_POOL_BLUE_COST = CONFIG["THREE_STAR_POOL_BLUE_COST"]

# BOSS战配置（从CONFIG加载，与QQ版一致）
BOSS_BATTLE_COOLDOWN_SECONDS = CONFIG["BOSS_BATTLE_COOLDOWN_SECONDS"]
BOSS_CARD_ID = CONFIG["BOSS_CARD_ID"]

app = Flask(__name__)

# 用于存储已处理的事件sn，防止重复处理
PROCESSED_EVENTS = set()
# 最大存储的事件sn数量（防止内存溢出）
MAX_EVENTS = 10000

# BOSS战冷却时间 {user_id: last_timestamp}
BOSS_BATTLE_COOLDOWN = {}

def log_info(message: str):
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "kook_bot.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")
    print(f"[INFO] {message}", flush=True)

def log_error(message: str):
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "kook_error.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] ERR {message}\n")
    print(f"[ERROR] {message}", flush=True)

# ========== 抽卡记录备份模块（与QQ版一致）==========
def get_last_backup_date() -> str:
    """获取最后一次备份的日期"""
    if BACKUP_RECORD_FILE.exists():
        try:
            with open(BACKUP_RECORD_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("last_backup_date", "")
        except:
            return ""
    return ""

def set_last_backup_date(date_str: str):
    """设置最后一次备份的日期"""
    with open(BACKUP_RECORD_FILE, "w", encoding="utf-8") as f:
        json.dump({"last_backup_date": date_str}, f)

def backup_pity_records():
    """备份所有抽卡记录（每天第一次启动时执行）"""
    today = datetime.now().strftime("%Y-%m-%d")
    last_backup = get_last_backup_date()
    
    if last_backup == today:
        print(f"[INFO] 今日({today})已备份过抽卡记录，跳过")
        return False
    
    try:
        today_backup_dir = BACKUP_DIR / today
        today_backup_dir.mkdir(exist_ok=True)
        
        # 备份 pity_*.json 文件（抽卡记录）
        pity_files = list(INFO_DIR.glob("pity_*.json"))
        for pity_file in pity_files:
            dest_file = today_backup_dir / pity_file.name
            with open(pity_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 gacha_*.json 文件（呱太数据）
        gacha_files = list(INFO_DIR.glob("gacha_*.json"))
        for gacha_file in gacha_files:
            dest_file = today_backup_dir / gacha_file.name
            with open(gacha_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 signin_*.json 文件（签到数据）
        signin_files = list(INFO_DIR.glob("signin_*.json"))
        for signin_file in signin_files:
            dest_file = today_backup_dir / signin_file.name
            with open(signin_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 team_*.json 文件（队伍配置）
        team_files = list(INFO_DIR.glob("team_*.json"))
        for team_file in team_files:
            dest_file = today_backup_dir / team_file.name
            with open(team_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 ranking.json（排行榜）
        ranking_file = INFO_DIR / "ranking.json"
        if ranking_file.exists():
            dest_file = today_backup_dir / ranking_file.name
            with open(ranking_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 更新备份记录
        set_last_backup_date(today)
        
        total_files = len(pity_files) + len(gacha_files) + len(signin_files) + len(team_files)
        if ranking_file.exists(): total_files += 1
        
        print(f"[INFO] 抽卡记录备份完成！日期: {today}, 文件数: {total_files}")
        return True
        
    except Exception as e:
        print(f"[ERROR] 抽卡记录备份失败: {e}")
        return False

def load_config():
    config_path = BASE_DIR / "config.json"
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            CONFIG.update(json.load(f))

def decrypt_message(encrypted_data):
    """根据KOOK官方文档解密消息"""
    if not CONFIG.get("ENCRYPT_KEY") or not CRYPTO_AVAILABLE:
        return encrypted_data
    
    try:
        encrypt_key = CONFIG["ENCRYPT_KEY"]
        encrypt_key = encrypt_key.ljust(32, '\0')
        encrypt_key = encrypt_key.encode('utf-8')
        
        str_data = base64.b64decode(encrypted_data)
        iv = str_data[0:16]
        cipher_text = str_data[16:]
        
        cipher = AES.new(encrypt_key, AES.MODE_CBC, iv)
        decrypted = cipher.decrypt(base64.b64decode(cipher_text))
        return decrypted.decode('utf-8').rstrip('\0')
    except Exception as e:
        log_error(f"解密消息失败: {e}")
        return encrypted_data

def decompress_data(data):
    """解压zlib压缩的数据（支持多种解压方式）"""
    try:
        # 尝试标准zlib解压
        return zlib.decompress(data, zlib.MAX_WBITS | 16).decode('utf-8')
    except Exception:
        try:
            # 尝试不带gzip头的解压
            return zlib.decompress(data).decode('utf-8')
        except Exception:
            try:
                # 尝试其他解压方式
                return zlib.decompress(data, -zlib.MAX_WBITS).decode('utf-8')
            except Exception:
                # 如果解压失败，尝试直接解码为UTF-8（可能未压缩）
                try:
                    return data.decode('utf-8')
                except Exception as e:
                    log_error(f"解压失败，数据可能不是UTF-8: {e}")
                    # 返回原始字节数据的十六进制表示用于调试
                    return data.hex()

def find_character_icon(chara_id, stars: int) -> str:
    """根据角色ID和星级查找图标文件（支持多种文件名格式）"""
    try:
        if isinstance(chara_id, float):
            chara_id_int = int(chara_id)
        elif isinstance(chara_id, str):
            chara_id_int = int(float(chara_id))
        else:
            chara_id_int = int(chara_id)
    except (ValueError, TypeError):
        log_error(f"无效的chara_id: {chara_id}, 类型: {type(chara_id)}")
        return None
    
    patterns = [
        f"card_cutin_{chara_id_int:09d}.png",
        f"card_cutin_{chara_id_int}.png",
        f"{chara_id_int:03d}.png",
        f"{chara_id_int}.png",
        f"{chara_id_int:09d}.png",
        f"chara_{chara_id_int:03d}.png",
        f"chara_{chara_id_int}.png",
    ]
    
    for pattern in patterns:
        icon_path = ICON_DIR / pattern
        if icon_path.exists():
            return str(icon_path)
    
    return None

def load_characters():
    """从卡牌信息.xlsx 加载角色数据（用于抽卡，含1/2/3星）"""
    global CHARACTERS
    CHARACTERS = {}
    
    if not XLSX_FILE.exists():
        log_error(f"抽卡数据文件不存在: {XLSX_FILE}")
        return
    
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
        characters = []
        
        for sheet_name, type_name in [('BattleCard资源', 'battle'), ('AssistCard资源', 'assist')]:
            if sheet_name not in wb.sheetnames:
                log_info(f"[抽卡] 未找到sheet: {sheet_name}")
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    card_id = str(row[5]) if row[5] is not None else None
                    chara_id = row[6]
                    init_stars = row[1]
                    card_name = row[4]
                    limit_type = row[0] or ""
                    if not (card_id and chara_id and init_stars and card_name):
                        continue
                    stars_val = int(float(str(init_stars)))
                    icon_path = find_character_icon(chara_id, stars_val)
                    characters.append({
                        "id": card_id, "chara_id": chara_id,
                        "name": str(card_name),
                        "rarity": stars_val,
                        "element": str(row[3]) if row[3] else "红",
                        "limit_type": str(limit_type),
                        "icon_path": icon_path, "type": type_name,
                        "hp": 5000, "attack": 3000, "defense": 2000,
                        "dexterity": 1000, "speed": 500,
                    })
                except Exception as e:
                    continue
            log_info(f"[抽卡] 加载了 {sum(1 for c in characters if c['type']==type_name)} 个{type_name}角色")
        
        # 转换为字典
        for char in characters:
            CHARACTERS[char["id"]] = char
        
        log_info(f"[抽卡] 共加载 {len(CHARACTERS)} 个角色")
        wb.close()
    except Exception as e:
        log_error(f"加载抽卡数据失败: {e}")

def load_battle_characters():
    """从 cards_completed.xlsx 加载战斗数据（3星卡详细数值）"""
    global BATTLE_CHARACTERS, BATTLE_INSTANCE
    BATTLE_CHARACTERS = {}
    
    if not BATTLE_XLSX.exists():
        log_error(f"战斗数据文件不存在: {BATTLE_XLSX}")
        return
    
    try:
        wb = openpyxl.load_workbook(BATTLE_XLSX, data_only=True)
        characters_data = []
        
        for sheet_name, card_type in [('b卡', 'battle'), ('a卡', 'assist')]:
            if sheet_name not in wb.sheetnames:
                log_info(f"[战斗] 未找到sheet: {sheet_name}")
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    card_id = str(int(row[0])) if row[0] is not None else None
                    if not card_id:
                        continue
                    name_raw = str(row[1]) if row[1] else ""
                    name = name_raw.split('|')[0].strip() if '|' in name_raw else name_raw.strip()
                    side = str(row[3]) if row[3] else "科学"
                    direction_raw = str(row[4]) if row[4] else ""
                    attribute = str(row[5]) if row[5] else "红"
                    attack_type_str = str(row[6]) if row[6] else "物理"
                    sk1_desc = str(row[8]) if row[8] else ""
                    sk1_cd = int(float(row[9])) if row[9] is not None else 0
                    sk2_desc = str(row[10]) if row[10] else ""
                    sk2_cd = int(float(row[11])) if row[11] is not None else 0
                    sk3_desc = str(row[12]) if row[12] else ""
                    sk3_cd = int(float(row[13])) if row[13] is not None else 0
                    pas1 = str(row[14]) if row[14] else ""
                    pas2 = str(row[15]) if row[15] else ""
                    phys_atk = int(float(row[16])) if row[16] is not None else 5000
                    magic_atk = int(float(row[17])) if row[17] is not None else 5000
                    phys_def = int(float(row[18])) if row[18] is not None else 3000
                    magic_def = int(float(row[19])) if row[19] is not None else 3000
                    hp = int(float(row[20])) if row[20] is not None else 10000
                    dexterity = int(float(row[21])) if row[21] is not None else 1000
                    is_physical = '物理' in attack_type_str
                    attack = phys_atk if is_physical else magic_atk
                    defense = (phys_def + magic_def) // 2
                    color_map = {'赤': '红', '緑': '绿', '青': '蓝'}
                    attr = attribute.strip()
                    is_super = attr.startswith('超')
                    base = attr[1:] if is_super else attr
                    base = color_map.get(base, base)
                    attr_normalized = ('超' + base) if is_super else base
                    offsets = set()
                    for ch, off in [('↖',-1),('↙',-1),('←',-1),('↑',0),('↓',0),('↗',1),('↘',1),('→',1)]:
                        if ch in str(direction_raw): offsets.add(off)
                    dire_list = sorted(offsets) if offsets else [0]
                    acquire = str(row[2]) if row[2] else ""
                    limit_type = ""
                    if 'fes' in acquire.lower(): limit_type = "フェス限定"
                    elif '限定' in acquire or '期間' in acquire: limit_type = "期間限定"
                    elif '活動' in acquire or '活动' in acquire: limit_type = "期間限定"
                    icon_path = find_character_icon(card_id, 3)
                    
                    char = {
                        "id": card_id, "chara_id": card_id,
                        "name": name,
                        "rarity": 3,
                        "element": attr_normalized,
                        "limit_type": limit_type,
                        "icon_path": icon_path,
                        "type": card_type,
                        "hp": hp,
                        "attack": attack,
                        "defense": defense,
                        "dexterity": dexterity,
                        "phys_atk": phys_atk,
                        "magic_atk": magic_atk,
                        "attack_directions": dire_list,
                        "attack_type": attack_type_str,
                        "side": side,
                        "skill1": {"cd": sk1_cd, "description": sk1_desc, "keywords": "", "condition": ""},
                        "skill2": {"cd": sk2_cd, "description": sk2_desc, "keywords": "", "condition": ""},
                        "skill3": {"cd": sk3_cd, "description": sk3_desc, "keywords": "", "condition": ""},
                        "passive1_text": pas1,
                        "passive2_text": pas2,
                    }
                    BATTLE_CHARACTERS[card_id] = char
                    characters_data.append(char)
                except Exception as e:
                    continue
            log_info(f"[战斗] 加载了 {sum(1 for c in characters_data if c['type']==card_type)} 个{card_type}角色")
        
        if characters_data:
            BATTLE_INSTANCE = BattleSystem(characters_data)
        
        log_info(f"[战斗] 共加载 {len(BATTLE_CHARACTERS)} 个角色")
        wb.close()
    except Exception as e:
        log_error(f"加载战斗数据失败: {e}")

def get_user_data_path(user_id):
    return INFO_DIR / f"gacha_{user_id}.json"

def load_user_data(user_id):
    user_id = str(user_id)
    if user_id in USER_DATA:
        return USER_DATA[user_id]
    
    file_path = get_user_data_path(user_id)
    if file_path.exists():
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                USER_DATA[user_id] = json.load(f)
        except Exception as e:
            log_error(f"用户 {user_id} 的数据文件格式错误: {e}")
            USER_DATA[user_id] = get_default_user_data()
    else:
        USER_DATA[user_id] = get_default_user_data()
    
    return USER_DATA[user_id]

def get_default_user_data():
    return {
        "gacha": 0,
        "cards": [],
        "team": {"battle_cards": [], "assist_cards": []},
        "presets": [],
        "active_preset": 0,
        "defense_slot": 0,
        "battle_history": [],
        "signin": {
            "last_date": "",
            "streak": 0,
            "total_days": 0
        },
        "red_crystal": 0,
        "blue_crystal": 0,
        "fes_count": 0,
        "period_count": 0,
        "other_3star_count": 0,
        "total_2stars": 0,
        "card_collection": {},
        "recent_3stars": []
    }

def save_user_data(user_id):
    user_id = str(user_id)
    if user_id in USER_DATA:
        file_path = get_user_data_path(user_id)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(USER_DATA[user_id], f, ensure_ascii=False, indent=2)

def get_pity_data_path(user_id):
    return INFO_DIR / f"pity_{user_id}.json"

def load_pity_data(user_id):
    user_id = str(user_id)
    if user_id in PITY_DATA:
        return PITY_DATA[user_id]
    
    file_path = get_pity_data_path(user_id)
    if file_path.exists():
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                PITY_DATA[user_id] = json.load(f)
        except Exception as e:
            log_error(f"用户 {user_id} 的保底数据文件格式错误: {e}")
            PITY_DATA[user_id] = get_default_pity_data()
    else:
        PITY_DATA[user_id] = get_default_pity_data()
    
    return PITY_DATA[user_id]

def get_default_pity_data():
    """获取默认的抽卡记录数据（与QQ版一致）"""
    return {
        "total_draws": 0,
        "pity_count": 0,
        "fes_pity_count": 0,
        "total_3stars": 0,
        "red_crystal": 0,
        "blue_crystal": 0,
        "recent_3stars": [],  # 最近获得的三星卡记录，按时间排序，存储card_id
        "card_collection": {},  # 卡片收藏 {card_id: {"name": "", "stars": 3, "limit_type": "", "count": 1, "last_time": timestamp}}
        "fes_count": 0,        # フェス限定数量
        "period_count": 0,      # 期間限定数量
        "other_3star_count": 0, # 其他三星数量
        "total_2stars": 0       # 二星数量
    }

# ========== 碎片系统（与QQ版一致）==========
def add_red_crystal(user_id: str, amount: int):
    """增加用户的红色碎片"""
    pity_data = load_pity_data(user_id)
    pity_data["red_crystal"] = pity_data.get("red_crystal", 0) + amount
    save_pity_data(user_id)
    return pity_data["red_crystal"]

def add_blue_crystal(user_id: str, amount: int):
    """增加用户的蓝色碎片"""
    pity_data = load_pity_data(user_id)
    pity_data["blue_crystal"] = pity_data.get("blue_crystal", 0) + amount
    save_pity_data(user_id)
    return pity_data["blue_crystal"]

def add_recent_3star(user_id: str, card_id: str, chara_id: str, name: str, limit_type: str = None):
    """添加最近获得的三星卡记录（最多保存10个）"""
    pity_data = load_pity_data(user_id)
    recent = pity_data.get("recent_3stars", [])
    
    record = {
        "card_id": card_id,
        "chara_id": chara_id,
        "name": name,
        "limit_type": limit_type,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    recent.insert(0, record)
    
    if len(recent) > 10:
        recent = recent[:10]
    
    pity_data["recent_3stars"] = recent
    save_pity_data(user_id)
    return recent

def add_card_collection(user_id: str, card_id: str, name: str, stars: int, limit_type: str = None):
    """添加卡片到收藏并更新数量统计（支持重复计数）"""
    pity_data = load_pity_data(user_id)
    collection = pity_data.get("card_collection", {})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    now_timestamp = datetime.now().timestamp()
    
    if card_id not in collection:
        collection[card_id] = {
            "name": name,
            "stars": stars,
            "limit_type": limit_type,
            "count": 1,
            "first_time": now,
            "last_time": now_timestamp
        }
        
        if stars == 3:
            if limit_type == "フェス限定":
                pity_data["fes_count"] = pity_data.get("fes_count", 0) + 1
            elif limit_type == "期間限定":
                pity_data["period_count"] = pity_data.get("period_count", 0) + 1
            else:
                pity_data["other_3star_count"] = pity_data.get("other_3star_count", 0) + 1
        elif stars == 2:
            pity_data["total_2stars"] = pity_data.get("total_2stars", 0) + 1
    else:
        collection[card_id]["count"] = collection[card_id].get("count", 1) + 1
        collection[card_id]["last_time"] = now_timestamp
    
    pity_data["card_collection"] = collection
    save_pity_data(user_id)
    return collection

# ========== FES统计模块（与QQ版一致）==========
FES_STATS_FILE = INFO_DIR / "fes_stats.json"

def load_fes_stats() -> dict:
    """加载全服FES角色获取统计"""
    if FES_STATS_FILE.exists():
        try:
            with open(FES_STATS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_fes_stats(stats: dict):
    """保存全服FES角色获取统计"""
    with open(FES_STATS_FILE, "w", encoding="utf-8") as f:
        json.dump(stats, f, indent=2, ensure_ascii=False)

def increment_fes_count(card_id: str, name: str) -> int:
    """增加FES角色获取计数并返回新的计数"""
    stats = load_fes_stats()
    if card_id not in stats:
        stats[card_id] = {"count": 0, "name": name}
    stats[card_id]["count"] += 1
    save_fes_stats(stats)
    return stats[card_id]["count"]

# ========== 呱太数据模块（与QQ版一致）==========
def get_gacha_file(user_id: str) -> Path:
    """获取用户呱太记录文件路径"""
    return INFO_DIR / f"gacha_{user_id}.json"

def load_gacha_data(user_id: str) -> dict:
    """加载用户的呱太数据"""
    gacha_file = get_gacha_file(user_id)
    if gacha_file.exists():
        try:
            with open(gacha_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {"gacha": 0}
    return {"gacha": 0}

def save_gacha_data(user_id: str, data: dict):
    """保存用户的呱太数据"""
    gacha_file = get_gacha_file(user_id)
    with open(gacha_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

def get_gacha_count(user_id: str) -> int:
    """获取用户的呱太数量"""
    gacha_data = load_gacha_data(user_id)
    return gacha_data.get("gacha", 0)

def add_gacha(user_id: str, amount: int) -> int:
    """增加用户的呱太数量"""
    gacha_data = load_gacha_data(user_id)
    gacha_data["gacha"] = gacha_data.get("gacha", 0) + amount
    save_gacha_data(user_id, gacha_data)
    return gacha_data["gacha"]

def spend_gacha(user_id: str, amount: int) -> bool:
    """消耗用户的呱太数量，返回是否成功"""
    gacha_data = load_gacha_data(user_id)
    current = gacha_data.get("gacha", 0)
    if current >= amount:
        gacha_data["gacha"] = current - amount
        save_gacha_data(user_id, gacha_data)
        return True
    return False

# ========== 签到系统（与QQ版一致）==========
def get_signin_file(user_id: str) -> Path:
    """获取用户签到记录文件路径"""
    return INFO_DIR / f"signin_{user_id}.json"

def load_signin_data(user_id: str) -> dict:
    """加载用户的签到数据"""
    signin_file = get_signin_file(user_id)
    if signin_file.exists():
        try:
            with open(signin_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {"last_signin": "", "streak": 0, "total_days": 0}
    return {"last_signin": "", "streak": 0, "total_days": 0}

def save_signin_data(user_id: str, data: dict):
    """保存用户的签到数据"""
    signin_file = get_signin_file(user_id)
    with open(signin_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

def can_signin(user_id: str) -> bool:
    """检查用户今天是否可以签到"""
    signin_data = load_signin_data(user_id)
    last_signin = signin_data.get("last_signin", "")
    today = datetime.now().strftime("%Y-%m-%d")
    return last_signin != today

def signin(user_id: str) -> dict:
    """执行签到（与QQ版一致），返回结果"""
    signin_data = load_signin_data(user_id)
    today = datetime.now().strftime("%Y-%m-%d")
    
    if signin_data.get("last_signin") == today:
        return {"success": False, "message": "今日已签到"}
    
    streak = signin_data.get("streak", 0) + 1
    total_days = signin_data.get("total_days", 0) + 1
    
    signin_data["last_signin"] = today
    signin_data["streak"] = streak
    signin_data["total_days"] = total_days
    
    save_signin_data(user_id, signin_data)
    
    # 添加签到奖励（与QQ版一致）
    add_gacha(user_id, CONFIG["DAILY_REWARD"])
    current_gacha = get_gacha_count(user_id)
    
    return {"success": True, "streak": streak, "total_days": total_days, "gacha": current_gacha}

def get_defense_slot(user_id):
    """获取用户的防守队预设槽位"""
    user_data = load_user_data(user_id)
    return user_data.get("defense_slot", 0)

def set_defense_slot(user_id, slot):
    """设置用户的防守队预设槽位"""
    user_data = load_user_data(user_id)
    if 0 <= slot <= len(user_data.get("presets", [])):
        user_data["defense_slot"] = slot
        save_user_data(user_id)
        return True
    return False

def get_defense_team(user_id):
    """获取用户的防守队配置"""
    user_data = load_user_data(user_id)
    defense_slot = user_data.get("defense_slot", 0)
    presets = user_data.get("presets", [])
    
    if defense_slot > 0 and defense_slot <= len(presets):
        return presets[defense_slot - 1]
    
    # 如果没有设置防守队预设，返回主队伍
    return user_data.get("team", {"battle_cards": [], "assist_cards": []})

def save_pity_data(user_id):
    user_id = str(user_id)
    if user_id in PITY_DATA:
        file_path = get_pity_data_path(user_id)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(PITY_DATA[user_id], f, ensure_ascii=False, indent=2)

def save_rolling_battle_log(user_id, result):
    """保存战斗日志，只保留最近3次（与QQ版一致）"""
    log_path = INFO_DIR / f"battle_{user_id}.json"
    result["saved_at"] = datetime.now().strftime("%m-%d %H:%M:%S")
    
    # 读取现有日志
    logs = []
    if log_path.exists():
        try:
            with open(log_path, "r", encoding="utf-8") as f:
                logs = json.load(f)
        except:
            logs = []
    
    # 添加新日志并保留最近3次
    logs.insert(0, result)
    logs = logs[:3]
    
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)
    
    return f"battle_{user_id} (共{len(logs)}次)"

def get_characters():
    """获取预加载的角色数据（如果未预加载则自动加载）"""
    global CHARACTERS
    if not CHARACTERS:
        load_characters()
    return list(CHARACTERS.values())

def generate_ai_team(difficulty=2):
    """
    生成AI队伍（根据难度等级）
      1=简单: 随机选卡，不刻意匹配
      2=普通: 随机但偏向同色同攻 (默认)
      3=困难: 倾向最优组合
      4=极难: 最优组合+属性克制
      5=地狱: 最优组合+属性克制+等级压制
    """
    characters = get_characters()
    if not characters:
        return {"battle_cards": [], "assist_cards": []}
    
    import random
    battle_cards = []
    available = [c for c in characters if c.get("type") == "battle"]
    
    for i in range(6):
        if available:
            card = random.choice(available)
            battle_cards.append(str(card["id"]))
            if difficulty < 5:
                available = [c for c in available if c["id"] != card["id"]]
    
    assist_cards = []
    available_assist = [c for c in characters if c.get("type") == "assist"]
    for i in range(6):
        if available_assist:
            card = random.choice(available_assist)
            assist_cards.append(str(card["id"]))
            if difficulty < 5:
                available_assist = [c for c in available_assist if c["id"] != card["id"]]
        else:
            assist_cards.append(None)
    
    return {"battle_cards": battle_cards, "assist_cards": assist_cards}

def load_ranking_data():
    """加载排行榜数据（与QQ版一致）"""
    global RANKING_DATA
    file_path = INFO_DIR / "ranking.json"
    
    if file_path.exists():
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                RANKING_DATA = json.load(f)
        except Exception as e:
            log_error(f"加载排行榜数据失败: {e}")
            RANKING_DATA = {"power": [], "gacha": []}
    else:
        RANKING_DATA = {"power": [], "gacha": []}

def save_ranking_data():
    """保存排行榜数据（与QQ版一致）"""
    file_path = INFO_DIR / "ranking.json"
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(RANKING_DATA, f, ensure_ascii=False, indent=2)

def send_kook_message(channel_id, content, files=None, quote_msg_id=None):
    """
    发送消息到KOOK频道
    
    Args:
        channel_id: 频道ID
        content: 消息内容
        files: 要发送的文件路径（可选）
        quote_msg_id: 要回复的消息ID（可选）
    
    Returns:
        发送是否成功
    """
    url = f"{KOOK_API_URL}/message/create"
    headers = {"Authorization": f"Bot {CONFIG['BOT_TOKEN']}"}
    
    data = {
        "channel_id": channel_id,
        "content": content
    }
    
    # 添加回复消息ID
    if quote_msg_id:
        data["quote"] = quote_msg_id
    
    try:
        if files:
            with open(files, 'rb') as f:
                files_data = {"file": f}
                response = requests.post(url, headers=headers, data=data, files=files_data)
        else:
            response = requests.post(url, headers=headers, json=data)
        
        if response.status_code == 200:
            result = response.json()
            if result.get("code") == 0:
                return True
            else:
                log_error(f"发送消息失败: {result.get('message', 'unknown error')}")
                return False
        else:
            log_error(f"发送消息失败: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        log_error(f"发送消息异常: {e}")
        return False

def upload_kook_image(image) -> str:
    """
    上传图片到KOOK并返回URL（根据KOOK官方文档，asset/create接口返回url字段）
    
    Args:
        image: PIL Image对象
    
    Returns:
        URL字符串，失败返回None
    """
    url = f"{KOOK_API_URL}/asset/create"
    headers = {"Authorization": f"Bot {CONFIG['BOT_TOKEN']}"}
    
    try:
        buffer = BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)
        
        files = {
            "file": ("card.png", buffer, "image/png")
        }
        
        response = requests.post(url, headers=headers, files=files)
        
        if response.status_code == 200:
            result = response.json()
            if result.get("code") == 0:
                # 根据KOOK官方文档，返回的是url字段，不是file_id
                return result.get("data", {}).get("url")
            else:
                log_error(f"上传图片失败: {result.get('message', 'unknown error')}")
                return None
        else:
            log_error(f"上传图片失败: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        log_error(f"上传图片异常: {e}")
        return None

def send_kook_message_with_image(channel_id, content, image, quote_msg_id=None):
    """
    发送带有图片的消息到KOOK频道（使用官方推荐的卡片消息格式）
    
    Args:
        channel_id: 频道ID
        content: 消息内容
        image: PIL Image对象
        quote_msg_id: 要回复的消息ID（可选）
    
    Returns:
        发送是否成功
    """
    # 先上传图片获取URL（根据KOOK官方文档，asset/create返回url字段）
    image_url = upload_kook_image(image)
    if not image_url:
        log_error("上传图片失败，无法发送图片消息")
        return False
    
    log_info(f"图片上传成功: {image_url}")
    
    # 使用卡片消息格式发送（KOOK官方推荐）
    url = f"{KOOK_API_URL}/message/create"
    headers = {"Authorization": f"Bot {CONFIG['BOT_TOKEN']}"}
    
    # 构建卡片消息JSON（官方推荐格式）
    card = {
        "type": 10,
        "channel_id": channel_id,
        "content": json.dumps([{
            "type": "card",
            "modules": [
                {
                    "type": "section",
                    "text": {
                        "type": "plain-text",
                        "content": content if content else ""
                    }
                },
                {
                    "type": "container",
                    "elements": [
                        {
                            "type": "image",
                            "src": image_url  # 直接使用上传返回的URL
                        }
                    ]
                }
            ]
        }])
    }
    
    # 添加回复消息ID
    if quote_msg_id:
        card["quote"] = quote_msg_id
    
    try:
        response = requests.post(url, headers=headers, json=card)
        
        if response.status_code == 200:
            result = response.json()
            if result.get("code") == 0:
                return True
            else:
                log_error(f"发送消息失败: {result.get('message', 'unknown error')}")
                return False
        else:
            log_error(f"发送消息失败: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        log_error(f"发送消息异常: {e}")
        return False

def upload_kook_gif(gif_buffer) -> str:
    """
    上传GIF到KOOK并返回URL（根据KOOK官方文档，asset/create返回url字段）
    
    Args:
        gif_buffer: BytesIO对象
    
    Returns:
        URL字符串，失败返回None
    """
    url = f"{KOOK_API_URL}/asset/create"
    headers = {"Authorization": f"Bot {CONFIG['BOT_TOKEN']}"}
    
    try:
        gif_buffer.seek(0)
        
        filename = f"battle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.gif"
        
        files = {
            'file': (filename, gif_buffer, 'image/gif')
        }
        
        response = requests.post(url, headers=headers, files=files)
        
        if response.status_code == 200:
            result = response.json()
            if result.get("code") == 0:
                # 根据KOOK官方文档，返回的是url字段，不是file_id
                return result.get("data", {}).get("url")
            else:
                log_error(f"上传GIF失败: {result.get('message', 'unknown error')}")
                return None
        else:
            log_error(f"上传GIF失败: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        log_error(f"上传GIF异常: {e}")
        return None

def send_kook_gif_bytes(channel_id, content, gif_buffer, quote_msg_id=None):
    """
    发送BytesIO格式的GIF文件到KOOK（不保存到本地）
    
    Args:
        channel_id: 频道ID
        content: 附带的消息内容
        gif_buffer: BytesIO对象
        quote_msg_id: 要回复的消息ID（可选）
    
    Returns:
        发送是否成功
    """
    # 先上传GIF获取URL（根据KOOK官方文档，asset/create返回url字段）
    gif_url = upload_kook_gif(gif_buffer)
    if not gif_url:
        log_error("上传GIF失败，无法发送")
        return False
    
    log_info(f"GIF上传成功: {gif_url}")
    
    # 使用卡片消息格式发送（KOOK官方推荐）
    url = f"{KOOK_API_URL}/message/create"
    headers = {"Authorization": f"Bot {CONFIG['BOT_TOKEN']}"}
    
    # 构建卡片消息JSON（官方推荐格式）
    card = {
        "type": 10,
        "channel_id": channel_id,
        "content": json.dumps([{
            "type": "card",
            "modules": [
                {
                    "type": "section",
                    "text": {
                        "type": "plain-text",
                        "content": content if content else ""
                    }
                },
                {
                    "type": "container",
                    "elements": [
                        {
                            "type": "image",
                            "src": gif_url  # 直接使用上传返回的URL
                        }
                    ]
                }
            ]
        }])
    }
    
    # 添加回复消息ID
    if quote_msg_id:
        card["quote"] = quote_msg_id
    
    try:
        response = requests.post(url, headers=headers, json=card)
        
        if response.status_code == 200:
            result = response.json()
            if result.get("code") == 0:
                log_info(f"发送GIF成功到频道 {channel_id}")
                return True
            else:
                log_error(f"发送GIF消息失败: {result.get('message', 'unknown error')}")
                return False
        else:
            log_error(f"发送GIF消息失败: {response.status_code} - {response.text}")
            return False
            
    except Exception as e:
        log_error(f"发送GIF异常: {e}")
        return False

def send_kook_card_message(channel_id, card, quote_msg_id=None):
    url = f"{KOOK_API_URL}/message/create"
    headers = {"Authorization": f"Bot {CONFIG['BOT_TOKEN']}"}
    
    data = {
        "channel_id": channel_id,
        "type": 10,
        "content": json.dumps(card, ensure_ascii=False)
    }
    
    # 添加回复消息ID
    if quote_msg_id:
        data["quote"] = quote_msg_id
    
    try:
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            return True
        else:
            log_error(f"发送卡片消息失败: {response.text}")
            return False
    except Exception as e:
        log_error(f"发送卡片消息异常: {e}")
        return False

def create_kook_card(title, description, fields=None, color="#0099ff"):
    card = {
        "type": "card",
        "theme": "secondary",
        "modules": [
            {
                "type": "header",
                "text": {
                    "type": "plain-text",
                    "content": title
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "plain-text",
                    "content": description
                }
            }
        ]
    }
    
    if fields:
        for field in fields:
            card["modules"].append({
                "type": "section",
                "text": {
                    "type": "plain-text",
                    "content": f"{field['name']}: {field['value']}"
                }
            })
    
    return card

# ========== 盲盒系统（黑呱太）==========
def has_pending_boxes(user_id: str) -> bool:
    """检查用户是否有未开完的盲盒"""
    if user_id not in BOX_SESSIONS:
        return False
    
    session = BOX_SESSIONS[user_id]
    boxes = session.get("boxes", [])
    opened = session.get("opened", [])
    
    if len(boxes) == len(opened):
        return False
    
    return True

def get_box_session(user_id: str) -> dict:
    """获取用户的盲盒会话"""
    return BOX_SESSIONS.get(user_id)

def create_box_session(user_id: str, boxes: list, characters: dict):
    """创建盲盒会话"""
    BOX_SESSIONS[user_id] = {
        "boxes": boxes,
        "opened": [],
        "characters": characters,
        "created_at": datetime.now().timestamp()
    }

def clear_box_session(user_id: str):
    """清除盲盒会话"""
    if user_id in BOX_SESSIONS:
        del BOX_SESSIONS[user_id]

def is_valid_box_index(boxes: list, index_str: str, opened_indices: list = None) -> tuple:
    """
    验证盲盒索引是否合法
    返回: (是否有效, 索引列表或错误信息)
    """
    if opened_indices is None:
        opened_indices = []
    
    index_str = index_str.strip()
    
    if index_str == "全部开" or index_str == "全开":
        valid_indices = [i for i in range(len(boxes)) if i not in opened_indices]
        return (True, valid_indices)
    
    if index_str == "剩下的全部开" or index_str == "剩余全开":
        valid_indices = [i for i in range(len(boxes)) if i not in opened_indices]
        return (True, valid_indices)
    
    try:
        index = int(index_str)
        if index < 1 or index > len(boxes):
            return (False, f"超出范围（1-{len(boxes)}）")
        if index - 1 in opened_indices:
            return (False, f"{index}号已开过")
        return (True, [index - 1])
    except ValueError:
        return (False, "请输入数字或「全部开」")

def open_mystery_box(box: dict, characters: dict) -> dict:
    """打开黑色盲盒，获取角色"""
    stars = box["stars"]
    available = [c for c in characters.values() if c["rarity"] == stars]
    if available:
        character = random.choice(available)
    else:
        character = random.choice(list(characters.values()))
    
    return {
        "stars": stars,
        "is_mystery": box["is_mystery"],
        "character": character,
        "opened": True
    }

def select_3star_character(characters: dict, is_fes_pity: bool = False) -> dict:
    """
    根据限定种类概率选择三星角色（与QQ版一致）
    期間限定35%，フェス限定25%，其余三星40%
    如果是フェス保底，则必定返回フェス限定角色
    """
    all_3stars = [c for c in characters.values() if c["rarity"] == 3]
    if not all_3stars:
        return random.choice(list(characters.values()))
    
    if is_fes_pity:
        fes_chars = [c for c in all_3stars if c.get("limit_type") == "フェス限定"]
        if fes_chars:
            return random.choice(fes_chars)
        return random.choice(all_3stars)
    
    period_limited = [c for c in all_3stars if c.get("limit_type") == "期間限定"]
    fes_limited = [c for c in all_3stars if c.get("limit_type") == "フェス限定"]
    other_3stars = [c for c in all_3stars if c.get("limit_type") not in ["期間限定", "フェス限定"]]
    
    categories = []
    weights = []
    
    if period_limited:
        categories.append(period_limited)
        weights.append(PERIOD_LIMIT_PROB)
    if fes_limited:
        categories.append(fes_limited)
        weights.append(FES_LIMIT_PROB)
    if other_3stars:
        categories.append(other_3stars)
        weights.append(OTHER_3STAR_PROB)
    
    if not categories:
        return random.choice(all_3stars)
    
    selected_category = random.choices(categories, weights=weights, k=1)[0]
    return random.choice(selected_category)

def draw_mystery_box(characters: dict, user_id: str = None, is_pity: bool = False) -> dict:
    """
    抽取一个盲盒（未开的盒子）（与QQ版一致）
    返回: {"stars": 星级, "is_mystery": 是否黑色盲盒, "character": 角色对象}
    
    黑色盲盒概率: 2%
    黑色盲盒必定是2星或3星（65%/35%）
    正常盲盒遵循标准抽卡概率（72%/23%/5%）
    """
    if not characters:
        return {"stars": 1, "is_mystery": False, "character": None}
    
    if is_pity:
        character = select_3star_character(characters, is_fes_pity=True)
        return {
            "stars": 3,
            "is_mystery": False,
            "character": character
        }
    
    is_mystery = random.random() < MYSTERY_BOX_CHANCE
    
    if is_mystery:
        stars = random.choices([2, 3], weights=[MYSTERY_BOX_2STAR_PROB, MYSTERY_BOX_3STAR_PROB], k=1)[0]
        
        if stars == 3:
            character = select_3star_character(characters, is_fes_pity=False)
        else:
            available = [c for c in characters.values() if c["rarity"] == 2]
            character = random.choice(available) if available else random.choice(list(characters.values()))
        
        return {
            "stars": stars,
            "is_mystery": True,
            "character": character
        }
    else:
        stars = random.choices([1, 2, 3], weights=[NORMAL_BOX_1STAR_PROB, NORMAL_BOX_2STAR_PROB, NORMAL_BOX_3STAR_PROB], k=1)[0]
        
        if stars == 3:
            character = select_3star_character(characters, is_fes_pity=False)
        else:
            available = [c for c in characters.values() if c["rarity"] == stars]
            character = random.choice(available) if available else random.choice(list(characters.values()))
        
        return {
            "stars": stars,
            "is_mystery": False,
            "character": character
        }

def apply_mutation(box_info: dict, characters: dict) -> tuple:
    """
    对盲盒结果应用突变（与QQ版一致）
    返回: (突变后的box_info, 是否发生突变, 突变描述)
    
    突变概率:
    - 1星→2星: 8%
    - 1星→3星: 2%
    - 2星→3星: 5%
    - 不突变: 88%
    """
    if box_info["is_mystery"]:
        return box_info, False, None
    
    original_stars = box_info["stars"]
    roll = random.random()
    new_stars = original_stars
    mutation_occurred = False
    mutation_text = None
    
    if original_stars == 1:
        if roll < MUTATION_1_TO_2:
            new_stars = 2
            mutation_occurred = True
            mutation_text = "1星→2星"
        elif roll < MUTATION_1_TO_2 + MUTATION_1_TO_3:
            new_stars = 3
            mutation_occurred = True
            mutation_text = "1星→3星"
    elif original_stars == 2:
        if roll < MUTATION_2_TO_3:
            new_stars = 3
            mutation_occurred = True
            mutation_text = "2星→3星"
    
    if mutation_occurred:
        if new_stars == 3:
            character = select_3star_character(characters, is_fes_pity=False)
        else:
            available = [c for c in characters.values() if c["rarity"] == new_stars]
            character = random.choice(available) if available else random.choice(list(characters.values()))
        
        return {
            "stars": new_stars,
            "is_mystery": box_info["is_mystery"],
            "character": character
        }, True, mutation_text
    
    return box_info, False, None

def gacha_draw(user_id=None, count=1):
    """
    执行抽卡（与QQ版完全一致，包含黑呱太系统）
    保底机制：每150抽必出フェス限定3星
    盲盒机制：2%概率出现黑色盲盒（必定2星或3星）
    突变机制：1星→2星(8%)、1星→3星(2%)、2星→3星(5%)
    """
    if not CHARACTERS:
        return {"results": [], "remaining_pity": 0, "got_3star": False, "got_fes_3star": False}
    
    if count not in [1, 10]:
        log_error(f"抽卡数量错误: {count}")
        return {"results": [], "remaining_pity": 0, "got_3star": False, "got_fes_3star": False}
    
    results = []
    got_3star = False
    got_fes_3star = False
    remaining_pity = 0
    
    for _ in range(count):
        is_pity = False
        
        if user_id:
            pity_data = load_pity_data(user_id)
            fes_pity_count = pity_data.get("fes_pity_count", 0)
            remaining_before = max(0, CONFIG["PITY_LIMIT"] - fes_pity_count)
            
            if remaining_before == 1:
                is_pity = True
        
        box_info = draw_mystery_box(CHARACTERS, user_id, is_pity)
        
        if not box_info["is_mystery"]:
            box_info, mutated, mutation_text = apply_mutation(box_info, CHARACTERS)
        
        character = box_info["character"]
        if character:
            results.append(character.copy())
            if box_info["stars"] == 3:
                got_3star = True
                if character.get("limit_type") == "フェス限定":
                    got_fes_3star = True
    
    if user_id:
        for _ in range(count):
            update_pity(user_id, got_3star, got_fes_3star)
        remaining_pity = get_remaining_pity(user_id)
    
    return {"results": results, "remaining_pity": remaining_pity, "got_3star": got_3star, "got_fes_3star": got_fes_3star}

def get_remaining_pity(user_id: str) -> int:
    """获取用户距离フェス限定三星保底还剩多少抽（与QQ版一致）"""
    pity_data = load_pity_data(user_id)
    fes_pity_count = pity_data.get("fes_pity_count", 0)
    return max(0, CONFIG["PITY_LIMIT"] - fes_pity_count)

def handle_box_open(user_id: str, channel_id: str, open_input: str, msg_id=None):
    """处理盲盒开启请求（与QQ版一致）"""
    session = get_box_session(user_id)
    if not session:
        send_kook_message(channel_id, "没有找到盲盒会话！", quote_msg_id=msg_id)
        return

    boxes = session["boxes"]
    characters = session.get("characters") or CHARACTERS
    opened = session["opened"].copy()

    valid, result = is_valid_box_index(boxes, open_input, session["opened"])
    if not valid:
        send_kook_message(channel_id, f"输入无效：{result}\n请输入数字（如1、2、3）或「全部开」", quote_msg_id=msg_id)
        return

    indices = result
    indices = [i for i in indices if i not in opened]
    
    if not indices:
        send_kook_message(channel_id, "这些都已经开过了！", quote_msg_id=msg_id)
        return

    new_opened = []
    opened_results = []
    mutation_messages = []

    red_crystal_gained = 0
    blue_crystal_gained = 0

    for idx in indices:
        box = boxes[idx]
        
        if box["is_mystery"] and not box.get("opened"):
            box = open_mystery_box(box, characters)
            boxes[idx] = box
        
        box, mutated, mutation_text = apply_mutation(box, characters)
        boxes[idx] = box
        
        if mutated and mutation_text:
            mutation_messages.append(f"#{idx+1}发生了{mutation_text}突变！")
        
        character = box.get("character")
        if not character:
            stars = box["stars"]
            available = [c for c in characters.values() if c["rarity"] == stars]
            if available:
                character = random.choice(available)
            else:
                character = random.choice(list(characters.values()))
            box["character"] = character
            boxes[idx] = box
        
        stars = box["stars"]
        card_id = str(character.get("id", ""))
        limit_type = character.get("limit_type", "")
        chara_name = character.get("name", "")

        fes_message = ""
        if stars == 3 and limit_type == "フェス限定":
            fes_count = increment_fes_count(card_id, chara_name)
            fes_message = f"✨ 恭喜！这是全服第{fes_count}个「{chara_name}」！"

        if stars == 1:
            add_red_crystal(user_id, 1)
            red_crystal_gained += 1
        elif stars == 2:
            add_blue_crystal(user_id, 1)
            blue_crystal_gained += 1
            add_card_collection(user_id, card_id, chara_name, stars, limit_type)
        elif stars == 3:
            add_recent_3star(user_id, card_id, str(character.get("chara_id", "")), chara_name, limit_type)
            add_card_collection(user_id, card_id, chara_name, stars, limit_type)

        got_3star = (stars == 3)
        is_fes_3star = got_3star and (limit_type == "フェス限定")
        update_pity(user_id, got_3star, is_fes_3star)

        if fes_message:
            opened_results.append({
                "index": idx + 1,
                "stars": stars,
                "name": chara_name,
                "fes_message": fes_message
            })
        else:
            opened_results.append({
                "index": idx + 1,
                "stars": stars,
                "name": chara_name
            })

        new_opened.append(idx)

    session["boxes"] = boxes
    session["opened"].extend(new_opened)

    stars_display = {1: "⭐", 2: "⭐⭐", 3: "⭐⭐⭐"}
    result_lines = []
    for r in opened_results:
        result_lines.append(f"{r['index']}. {stars_display.get(r['stars'], '⭐')} {r['name']}")

    result_text = "\n".join(result_lines)

    crystal_summary = ""
    if red_crystal_gained > 0 or blue_crystal_gained > 0:
        parts = []
        if red_crystal_gained > 0:
            parts.append(f"🔴红色碎片 x{red_crystal_gained}")
        if blue_crystal_gained > 0:
            parts.append(f"🔵蓝色碎片 x{blue_crystal_gained}")
        crystal_summary = f"\n本次获得: {' + '.join(parts)}"

    all_opened = session["opened"]
    summary_img_bytes = create_box_summary_image(boxes, all_opened, characters)

    if summary_img_bytes:
        img = Image.open(BytesIO(summary_img_bytes))
        result_text += crystal_summary
        
        remaining = len(boxes) - len(all_opened)
        if remaining > 0:
            remaining_hint = f"\n还有{remaining}个未开，输入「/开 剩下的全部开」可以一键开启"
            box_hints = [str(i+1) for i in range(len(boxes)) if i not in all_opened]
            hint = " | ".join(box_hints[:5])
            if len(box_hints) > 5:
                hint += "\n" + " | ".join(box_hints[5:])
            remaining_hint += f"\n可选: {hint}"
            result_text += remaining_hint
        else:
            result_text += "\n所有盲盒已开完！"
            clear_box_session(user_id)

        fes_messages = [r.get("fes_message") for r in opened_results if r.get("fes_message")]
        if fes_messages:
            result_text = "\n".join(fes_messages) + "\n" + result_text

        if mutation_messages:
            result_text = "\n".join(mutation_messages) + "\n" + result_text

        send_kook_message_with_image(channel_id, result_text, img, quote_msg_id=msg_id)
    else:
        send_kook_message(channel_id, result_text + crystal_summary, quote_msg_id=msg_id)

def signin(user_id):
    user_data = load_user_data(user_id)
    today = datetime.now().strftime("%Y-%m-%d")
    
    if user_data["signin"]["last_date"] == today:
        return {"success": False, "message": "今天已经签到过了", "streak": user_data["signin"]["streak"]}
    
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    if user_data["signin"]["last_date"] == yesterday:
        user_data["signin"]["streak"] += 1
    else:
        user_data["signin"]["streak"] = 1
    
    user_data["signin"]["last_date"] = today
    user_data["signin"]["total_days"] += 1
    user_data["gacha"] += CONFIG["DAILY_REWARD"]
    
    save_user_data(user_id)
    
    return {
        "success": True,
        "message": f"签到成功！获得 {CONFIG['DAILY_REWARD']} 呱太",
        "streak": user_data["signin"]["streak"],
        "total_days": user_data["signin"]["total_days"],
        "gacha": user_data["gacha"]
    }

def get_level_image(stars: int, layer_type: str) -> str:
    """
    获取星级框或背景图片（与QQ版一致）
    stars: 1, 2, 3
    layer_type: "bg" (背景) 或 "frame" (框)
    """
    star_idx = stars - 1
    layer_idx = 0 if layer_type == "bg" else 1
    filename = f"gacha_tmb_{star_idx:02d}_{layer_idx:02d}"
    
    if stars == 3 and layer_type == "bg":
        filename += "_b"
    
    path = LEVEL_DIR / f"{filename}.png"
    if path.exists():
        return str(path)
    else:
        log_error(f"找不到星级图片: {path}")
        return None

def find_attribute_icon(attribute: str) -> str:
    """根据属性名称查找属性图标文件（与QQ版一致）"""
    if not attribute:
        return None
    
    attr_name = str(attribute).strip()
    patterns = [
        f"common_tmb_label_element_{attr_name}.png",
        f"attr_{attr_name}.png",
        f"attribute_{attr_name}.png",
        f"type_{attr_name}.png",
        f"{attr_name}_attr.png",
        f"{attr_name}.png"
    ]
    
    for pattern in patterns:
        path = LEVEL_DIR / pattern
        if path.exists():
            return str(path)
    
    log_error(f"找不到属性图标: attribute={attribute}")
    return None

def find_type_icon(card_type: str) -> str:
    """根据卡牌类型查找Battle/Assist图标文件（与QQ版一致）"""
    if not card_type:
        return None
    
    type_name = str(card_type).strip().lower()
    label_name = "battle" if type_name == "battle" else "assist"
    patterns = [
        f"gacha_tmb_label_{label_name}.png",
        f"battle_{type_name}.png",
        f"{type_name}_icon.png",
    ]
    
    for pattern in patterns:
        path = LEVEL_DIR / pattern
        if path.exists():
            return str(path)
    
    log_error(f"找不到类型图标: type={card_type}")
    return None

def composite_card(card: dict) -> bytes:
    """
    合成卡牌图片（与QQ版一致）：外框 + 内卡（背景+角色+星级框） + 属性图标 + BA图标
    返回PNG格式的字节数据
    """
    stars = card.get("rarity", 1)
    card_id = card.get("id", card.get("chara_id", ""))
    
    # 查找角色图标
    icon_path = None
    patterns = [
        ICON_DIR / f"card_cutin_{card_id}.png",
        ICON_DIR / f"card_cutin_{card['name']}.png",
        ICON_DIR / f"{card_id}.png",
        ICON_DIR / f"{card['name']}.png"
    ]
    
    for pattern in patterns:
        if pattern.exists():
            icon_path = str(pattern)
            break
    
    if not icon_path:
        log_error(f"找不到角色图标: card_id={card_id}, name={card['name']}")
        img = Image.new('RGBA', (150, 150), (100, 100, 100, 255))
        bio = BytesIO()
        img.save(bio, format='PNG')
        return bio.getvalue()
    
    try:
        bg_path = get_level_image(stars, "bg")
        frame_path = get_level_image(stars, "frame")
        
        if not bg_path or not frame_path:
            log_error(f"背景或框图片不存在: bg={bg_path}, frame={frame_path}")
            return None
        
        bg_img = Image.open(bg_path).convert('RGBA')
        inner_frame_img = Image.open(frame_path).convert('RGBA')
        char_img = Image.open(icon_path).convert('RGBA')
        
        # 加载属性图标
        attribute = card.get("attribute")
        attr_icon_path = find_attribute_icon(attribute) if attribute else None
        attr_img = None
        if attr_icon_path and os.path.exists(attr_icon_path):
            attr_img = Image.open(attr_icon_path).convert('RGBA')
        
        # 加载Battle/Assist图标
        card_type = card.get("type", "battle")
        type_icon_path = find_type_icon(card_type)
        type_img = None
        if type_icon_path and os.path.exists(type_icon_path):
            type_img = Image.open(type_icon_path).convert('RGBA')
        
        CARD_SIZE = 122
        
        # 裁剪角色图
        char_width, char_height = char_img.size
        crop_left = int(char_width * CROP_LEFT_RATIO)
        crop_right = int(char_width * CROP_RIGHT_RATIO)
        crop_top = int(char_height * CROP_TOP_RATIO)
        crop_bottom = int(char_height * CROP_BOTTOM_RATIO)
        
        char_img_cropped = char_img.crop((crop_left, crop_top, crop_right, crop_bottom))
        
        # 缩放裁剪后的图片
        cropped_width, cropped_height = char_img_cropped.size
        cropped_ratio = cropped_width / cropped_height
        
        if cropped_ratio > 1:
            char_target_width = int(CARD_SIZE * 0.85)
            char_target_height = int(char_target_width / cropped_ratio)
        else:
            char_target_height = int(CARD_SIZE * 0.85)
            char_target_width = int(char_target_height * cropped_ratio)
        
        char_img_resized = char_img_cropped.resize(
            (char_target_width, char_target_height),
            Image.Resampling.LANCZOS
        )
        
        # 合成卡牌
        output = Image.new('RGBA', (CARD_SIZE, CARD_SIZE), (0, 0, 0, 0))
        char_x = (CARD_SIZE - char_target_width) // 2
        char_y = (CARD_SIZE - char_target_height) // 2
        
        output.paste(bg_img, (0, 0))
        output.paste(char_img_resized, (char_x, char_y), char_img_resized)
        output.paste(inner_frame_img, (0, 0), inner_frame_img)
        
        # 属性图标放在左下角
        if attr_img:
            attr_w, attr_h = attr_img.size
            output.paste(attr_img, (5, CARD_SIZE - attr_h - 5), attr_img)
        
        # BA图标放在底部居中
        if type_img:
            type_w, type_h = type_img.size
            type_x = (CARD_SIZE - type_w) // 2
            type_y = CARD_SIZE - type_h
            output.paste(type_img, (type_x, type_y), type_img)
        
        bio = BytesIO()
        output.save(bio, format='PNG')
        return bio.getvalue()
    
    except Exception as e:
        log_error(f"合成卡牌图片失败: {e}")
        return None

def generate_card_image(card, size=(150, 150)):
    """
    生成卡牌图片（与QQ版一致）
    
    Args:
        card: 卡牌数据
        size: 输出尺寸，默认(150, 150)
    
    Returns:
        PIL Image对象
    """
    try:
        img_bytes = composite_card(card)
        if img_bytes:
            img = Image.open(BytesIO(img_bytes))
            # 压缩图片到指定大小
            img = img.resize(size, Image.Resampling.LANCZOS)
            return img
        
        # 如果合成失败，使用简单的边框绘制
        icon_path = ICON_DIR / f"card_cutin_{card['id']}.png"
        if icon_path.exists():
            img = Image.open(icon_path)
            img = img.resize(size, Image.Resampling.LANCZOS)
        else:
            img = Image.new('RGBA', size, (30, 30, 30))
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
            except:
                font = ImageFont.load_default()
            draw.text((10, 60), card['name'][:8], fill=(255, 255, 255), font=font)
        
        draw = ImageDraw.Draw(img)
        border_color = {3: (255, 215, 0), 2: (147, 112, 219), 1: (34, 139, 34)}.get(card['rarity'], (128, 128, 128))
        draw.rectangle([(0, 0), (img.width-1, img.height-1)], outline=border_color, width=3)
        
        return img
    except Exception as e:
        log_error(f"生成卡牌图片失败: {e}")
        return None

def get_mystery_box_image(stars: int) -> str:
    """获取盲盒图片路径（使用02图层作为盲盒封面）"""
    star_idx = stars - 1
    filename = f"gacha_tmb_{star_idx:02d}_02.png"
    path = LEVEL_DIR / filename
    if path.exists():
        return str(path)
    
    # 如果02不存在，尝试用01（框）作为后备
    filename = f"gacha_tmb_{star_idx:02d}_01.png"
    path = LEVEL_DIR / filename
    if path.exists():
        return str(path)
    
    log_error(f"找不到盲盒图片: gacha_tmb_{star_idx:02d}_02")
    return None

def get_black_box_image() -> str:
    """获取黑色盲盒图片路径"""
    filename = "gacha_tmb_04_02.png"
    path = LEVEL_DIR / filename
    if path.exists():
        return str(path)
    log_error(f"找不到黑色盲盒图片: {filename}")
    return None

def create_opened_box_card(box_info: dict) -> bytes:
    """创建已开盲盒的卡牌图片（显示角色）"""
    if box_info["is_mystery"] and not box_info.get("opened"):
        return create_box_card(box_info, {})
    
    stars = box_info["stars"]
    character = box_info.get("character")
    
    if not character:
        log_error("已开盲盒没有角色信息")
        return create_box_card(box_info, {})
    
    return composite_card(character)

def create_box_summary_image(boxes: list, opened_indices: list, characters: dict) -> bytes:
    """
    创建盲盒汇总图片（带背景）
    boxes: 所有盲盒列表
    opened_indices: 已开的盲盒索引列表
    """
    if not boxes:
        return None
    
    single_card_img = None
    for i, box in enumerate(boxes):
        img_bytes = create_box_card(box, characters)
        if img_bytes:
            single_card_img = Image.open(BytesIO(img_bytes))
            break
    
    if not single_card_img:
        return None
    
    card_width, card_height = single_card_img.size
    count = len(boxes)
    
    if count <= 5:
        cols = count
        rows = 1
    else:
        cols = 5
        rows = (count + 4) // 5

    gap = 18
    cards_total_width = card_width * cols + gap * (cols - 1)
    cards_total_height = card_height * rows + gap * (rows - 1)
    
    bg_path = None
    for bg_name in ["gacha_tmb_bg_11.png", "gacha_tmb_11_bg.png", "gacha_bg_11.png"]:
        test_path = LEVEL_DIR / bg_name
        if test_path.exists():
            bg_path = str(test_path)
            break
    
    if not bg_path:
        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
            test_path = LEVEL_DIR / bg_name
            if test_path.exists():
                bg_path = str(test_path)
                break
    
    if bg_path:
        log_info(f"使用汇总背景图片: {bg_path}")
        bg_img = Image.open(bg_path).convert('RGB')
        
        bg_w, bg_h = bg_img.size
        final_w = int(bg_w * 2 * 0.264)
        final_h = int(bg_h * 2 * 0.264)
        bg_img_resized = bg_img.resize((final_w, final_h), Image.Resampling.LANCZOS)
        
        output = Image.new('RGB', (final_w, final_h), (50, 50, 50))
        output.paste(bg_img_resized, (0, 0))
        
        cards_x = (final_w - cards_total_width) // 2
        cards_y = (final_h - cards_total_height) // 2
    else:
        output = Image.new('RGB', (cards_total_width, cards_total_height), (50, 50, 50))
        cards_x = 0
        cards_y = 0
    
    for i, box in enumerate(boxes):
        row = i // cols
        col = i % cols
        
        if i in opened_indices:
            img_bytes = create_opened_box_card(box)
        else:
            img_bytes = create_box_card(box, characters)
        
        if img_bytes:
            card_img = Image.open(BytesIO(img_bytes))
            x = cards_x + col * (card_width + gap)
            y = cards_y + row * (card_height + gap)
            output.paste(card_img, (x, y))
    
    bio = BytesIO()
    output.save(bio, format='JPEG', optimize=True, quality=60)
    return bio.getvalue()

def create_box_card(box_info: dict, characters: dict) -> bytes:
    """
    创建盲盒卡牌图片（未开的样子）
    对于黑色盲盒，使用gacha_tmb_04_02作为中心图
    对于3星卡，如果是フェス限定，使用gacha_tmb_03_02_b.png
    """
    stars = box_info["stars"]
    is_mystery = box_info.get("is_mystery", False)
    character = box_info.get("character", None)
    
    # 获取背景和框
    if is_mystery:
        # 黑色盲盒使用04_00和04_01
        bg_path = LEVEL_DIR / "gacha_tmb_04_00.png"
        frame_path = LEVEL_DIR / "gacha_tmb_04_01.png"
        center_path = LEVEL_DIR / "gacha_tmb_04_02.png"
    else:
        bg_path = get_level_image(stars, "bg")
        frame_path = get_level_image(stars, "frame")
        # 盲盒使用02图层作为中心
        star_idx = stars - 1
        
        # 判断是否是フェス限定3星
        is_fes = False
        if character and stars == 3:
            limit_type = character.get("limit_type", "")
            if limit_type == "フェス限定":
                is_fes = True
        
        # 根据是否是フェス限定选择不同的盲盒图
        if is_fes:
            # フェス限定3星使用特殊盲盒图
            center_path = LEVEL_DIR / "gacha_tmb_03_02_b.png"
            # 如果不存在则使用普通3星盲盒图
            if not center_path.exists():
                center_path = LEVEL_DIR / f"gacha_tmb_{star_idx:02d}_02.png"
        else:
            center_path = LEVEL_DIR / f"gacha_tmb_{star_idx:02d}_02.png"
    
    if not bg_path or not frame_path or not center_path:
        log_error(f"盲盒图片不全: bg={bg_path}, frame={frame_path}, center={center_path}")
        return None
    
    bg_img = Image.open(bg_path).convert('RGBA')
    frame_img = Image.open(frame_path).convert('RGBA')
    center_img = Image.open(center_path).convert('RGBA')
    
    bg_width, bg_height = bg_img.size
    
    # 创建输出画布
    output = Image.new('RGBA', (bg_width, bg_height), (0, 0, 0, 0))
    output.paste(bg_img, (0, 0))
    
    # 调整中心图大小以适应框
    center_width, center_height = center_img.size
    center_ratio = center_width / center_height
    frame_ratio = bg_width / bg_height
    
    if center_ratio > frame_ratio:
        target_width = int(bg_width * 0.85)
        target_height = int(target_width / center_ratio)
    else:
        target_height = int(bg_height * 0.85)
        target_width = int(target_height * center_ratio)
    
    center_resized = center_img.resize((target_width, target_height), Image.Resampling.LANCZOS)
    center_x = (bg_width - target_width) // 2
    center_y = (bg_height - target_height) // 2
    
    output.paste(center_resized, (center_x, center_y), center_resized)
    output.paste(frame_img, (0, 0), frame_img)
    
    # 转换为RGB
    output_rgb = Image.new('RGB', (bg_width, bg_height), (255, 255, 255))
    output_rgb.paste(output, (0, 0), output)
    
    bio = BytesIO()
    output_rgb.save(bio, format='JPEG', optimize=True, quality=60)
    return bio.getvalue()

def generate_gacha10_image(cards):
    """
    生成十连抽卡的合成图片（与QQ版一致，2行5列显示所有10张卡）
    
    Args:
        cards: 10张卡牌数据列表
    
    Returns:
        PIL Image对象（合成后的图片）
    """
    try:
        # 生成每张卡牌的图片（不压缩，保持原始尺寸）
        card_images = []
        for card in cards:
            img_bytes = composite_card(card)
            if img_bytes:
                card_images.append(Image.open(BytesIO(img_bytes)))
        
        if not card_images:
            return None
        
        # 每张卡的尺寸（原始大小）
        card_width, card_height = card_images[0].size
        
        # 创建2行5列的大图
        gap = 18
        total_width = card_width * 5 + gap * 4
        total_height = card_height * 2 + gap
        
        # ========== 尝试加载十连背景图片（与QQ版一致）==========
        bg_path = None
        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
            test_path = LEVEL_DIR / bg_name
            if test_path.exists():
                bg_path = str(test_path)
                break
        
        if bg_path:
            # 使用十连背景，缩放到原来的2倍（与QQ版一致）
            log_info(f"使用十连背景图片: {bg_path}")
            bg_img = Image.open(bg_path).convert('RGB')
            bg_w, bg_h = bg_img.size
            
            # 与QQ版一致：final_w = int(bg_w * 2.0 * 0.264)
            final_w = int(bg_w * 2.0 * 0.264)
            final_h = int(bg_h * 2.0 * 0.264)
            bg_img_scaled = bg_img.resize((final_w, final_h), Image.Resampling.LANCZOS)
            
            # RGB画布，卡牌区域居中于背景
            output = Image.new('RGB', (final_w, final_h), (50, 50, 50))
            output.paste(bg_img_scaled, (0, 0))
            
            cards_x = (final_w - total_width) // 2
            cards_y = (final_h - total_height) // 2
        else:
            # 无背景图时使用纯色画布
            output = Image.new('RGB', (total_width, total_height), (50, 50, 50))
            cards_x = 0
            cards_y = 0
        
        # 将卡牌粘贴到背景上（居中放置）
        for idx, img in enumerate(card_images):
            row = idx // 5
            col = idx % 5
            x = cards_x + col * (card_width + gap)
            y = cards_y + row * (card_height + gap)
            # 转换为RGB模式以避免paste问题
            if img.mode == 'RGBA':
                output.paste(img, (x, y), img)
            else:
                output.paste(img, (x, y))
        
        return output
    except Exception as e:
        log_error(f"生成十连抽卡图片失败: {e}")
        return None

def handle_command(raw_message, user_id, channel_id, nickname, msg_id=None):
    # 检查是否以斜杠开头（避免误触）
    if not raw_message.startswith('/'):
        return
    
    # 移除斜杠前缀
    cmd = raw_message[1:]
    
    if '十连' in cmd or 'gacha10' in cmd.lower():
        handle_gacha(10, user_id, channel_id, msg_id)
    elif '单抽' in cmd or 'gacha' in cmd.lower():
        handle_gacha(1, user_id, channel_id, msg_id)
    elif '开 ' in cmd or '开箱' in cmd or has_pending_boxes(user_id):
        # 检查是否有未开完的盲盒
        if has_pending_boxes(user_id):
            # 移除@符号和空格
            cleaned_cmd = cmd.replace('@', '').replace(' ', '').replace('\u3000', '').strip()
            cleaned_cmd = cleaned_cmd.replace('，', ',').replace(',', '')
            
            # 检查是否是开箱命令
            open_commands = ['全部开', '剩下的全部开']
            is_open_command = any(cmd in cleaned_cmd for cmd in open_commands)
            
            # 检查是否是数字输入
            has_valid_input = False
            if cleaned_cmd.isdigit():
                has_valid_input = True
            else:
                import re
                match = re.search(r'选择[0-9]+', cleaned_cmd)
                if match:
                    has_valid_input = True
                else:
                    num_match = re.search(r'[0-9]+', cleaned_cmd)
                    if num_match:
                        has_valid_input = True
            
            if is_open_command or has_valid_input:
                handle_box_open(user_id, channel_id, cleaned_cmd, msg_id)
            else:
                session = get_box_session(user_id)
                remaining = max(0, len(session["boxes"]) - len(session["opened"]))
                reply = f"你还有{remaining}个未开！请输入要开的编号（如1、选择1）或「全部开」"
                send_kook_message(channel_id, reply, quote_msg_id=msg_id)
        else:
            send_kook_message(channel_id, "没有找到未开的盲盒！", quote_msg_id=msg_id)
    elif '限定十连' in cmd or '限定池' in cmd:
        handle_limited_gacha(user_id, channel_id, msg_id)
    elif '帮助' in cmd or 'help' in cmd.lower():
        handle_help(user_id, channel_id, raw_message, msg_id)
    elif '获取呱太' in cmd or 'getgacha' in cmd.lower():
        handle_get_gacha(user_id, channel_id, msg_id)
    elif '签到' in cmd or 'signin' in cmd.lower():
        handle_signin(user_id, channel_id, msg_id)
    elif '队伍' in cmd or '配队' in cmd.lower():
        handle_team(user_id, channel_id, cmd, msg_id)
    elif '防守队' in cmd:
        handle_defense_team(user_id, channel_id, cmd, msg_id)
    elif '个人记录' in cmd or '记录' in cmd.lower():
        handle_personal_info(user_id, channel_id, msg_id)
    elif '兑换呱太' in cmd or '兑换' in cmd.lower():
        handle_exchange_crystal(user_id, channel_id, msg_id)
    elif '抽卡榜单' in cmd or '抽卡排行' in cmd:
        handle_gacha_leaderboard(user_id, channel_id, msg_id)
    elif '排行榜' in cmd or '排行' in cmd.lower():
        handle_ranking(user_id, channel_id, msg_id)
    elif '详细信息' in cmd:
        handle_show_details(user_id, channel_id, msg_id)
    elif '战斗日志' in cmd or '战斗GIF' in cmd or '战斗gif' in cmd:
        handle_battle_log(user_id, channel_id, gen_gif=('GIF' in cmd or 'gif' in cmd), msg_id=msg_id)
    elif '三星池子' in cmd or '红抽' in cmd or '蓝抽' in cmd:
        handle_3star_pool(user_id, channel_id, cmd, msg_id)
    elif '挑战' in cmd:
        import re
        rank_match = re.search(r'挑战\s*(\d+)', cmd)
        if rank_match:
            handle_challenge(user_id, channel_id, cmd, msg_id)
        else:
            send_kook_message(channel_id, "请指定挑战的排名！格式：/挑战 1~10", quote_msg_id=msg_id)
    elif 'BOSS战' in cmd or 'boss战' in cmd.lower():
        handle_boss_battle(user_id, channel_id, cmd, msg_id)
    elif '战斗' in cmd or '对战' in cmd or '决斗' in cmd:
        handle_battle(user_id, channel_id, cmd, msg_id)
    elif '收藏' in cmd or '卡牌' in cmd:
        handle_collection(user_id, channel_id, msg_id)
    elif '保底' in cmd:
        handle_pity(user_id, channel_id, msg_id)
    elif '余额' in cmd or '呱太' in cmd:
        handle_balance(user_id, channel_id, msg_id)
    elif '资料' in cmd:
        handle_profile(user_id, channel_id, nickname, msg_id)
    elif '重新加载' in cmd and user_id == CONFIG["OWNER_ID"]:
        handle_reload(user_id, channel_id, msg_id)
    elif '备份' in cmd and user_id == CONFIG["OWNER_ID"]:
        handle_backup(user_id, channel_id, msg_id)
    elif '备份战斗录' in cmd and user_id == CONFIG["OWNER_ID"]:
        handle_backup_battle_logs(user_id, channel_id, msg_id)

def update_pity(user_id: str, got_3star: bool = False, is_fes_3star: bool = False):
    """更新用户的抽卡记录（与QQ版完全一致）"""
    pity_data = load_pity_data(user_id)
    
    # 如果抽到3星，重置保底计数并增加三星个数
    if got_3star:
        pity_data["pity_count"] = 0
        pity_data["total_3stars"] = pity_data.get("total_3stars", 0) + 1
        # 如果抽到フェス限定三星，重置フェス保底计数
        if is_fes_3star:
            pity_data["fes_pity_count"] = 0
    else:
        pity_data["pity_count"] = pity_data.get("pity_count", 0) + 1
        pity_data["fes_pity_count"] = pity_data.get("fes_pity_count", 0) + 1
    
    pity_data["total_draws"] = pity_data.get("total_draws", 0) + 1
    
    save_pity_data(user_id)

def handle_gacha(count, user_id, channel_id, msg_id=None):
    """处理抽卡请求（盲盒模式，与QQ版一致）"""
    user_data = load_user_data(user_id)
    pity_data = load_pity_data(user_id)
    
    cost = CONFIG["GACHA10_COST"] if count == 10 else CONFIG["GACHA_COST"]
    
    # 检查呱太是否足够
    if user_data["gacha"] < cost:
        send_kook_message(channel_id, f"呱太不足！需要 {cost} 呱太，当前 {user_data['gacha']} 呱太", quote_msg_id=msg_id)
        return
    
    # 十连冷却检查
    if count == 10:
        now_ts = datetime.now().timestamp()
        last_gacha = pity_data.get("last_ten_gacha", 0)
        remaining = CONFIG["GACHA10_COOLDOWN_SECONDS"] - (now_ts - last_gacha)
        if remaining > 0:
            send_kook_message(channel_id, f"十连冷却中！请等待 {int(remaining)} 秒后再试~", quote_msg_id=msg_id)
            return
    
    # 扣除呱太
    user_data["gacha"] -= cost
    save_user_data(user_id)
    
    # 检查是否触发FES保底
    fes_pity_count = pity_data.get("fes_pity_count", 0)
    is_fes_pity = fes_pity_count >= CONFIG["PITY_LIMIT"]
    
    if is_fes_pity:
        log_info(f"用户 {user_id} 触发FES限定三星保底！")
    
    # 生成盲盒
    boxes = []
    has_2star = False
    
    for i in range(count):
        is_pity_draw = is_fes_pity and (i == count - 1)
        box = draw_mystery_box(CHARACTERS, user_id, is_pity_draw)
        
        if box["stars"] == 2:
            has_2star = True
        
        boxes.append(box)
    
    # 十连保底：如果没有二星，将最后一个非保底盲盒改为二星
    if count == 10 and not has_2star and not is_fes_pity:
        for i in range(len(boxes)-1, -1, -1):
            if not boxes[i].get("is_mystery", False):
                boxes[i]["stars"] = 2
                available = [c for c in CHARACTERS.values() if c["rarity"] == 2]
                if available:
                    boxes[i]["character"] = random.choice(available)
                else:
                    boxes[i]["character"] = random.choice(list(CHARACTERS.values()))
                log_info(f"十连保底触发，将第{i+1}个盲盒改为二星")
                break
    
    # 创建盲盒会话
    create_box_session(user_id, boxes, CHARACTERS)
    
    # 生成盲盒图片
    box_images = []
    for box in boxes:
        img_bytes = create_box_card(box, CHARACTERS)
        if img_bytes:
            box_images.append(Image.open(BytesIO(img_bytes)))
    
    if not box_images:
        log_error("生成盲盒图片失败")
        clear_box_session(user_id)
        send_kook_message(channel_id, "生成盲盒图片失败！", quote_msg_id=msg_id)
        return
    
    # 合成盲盒大图
    card_width, card_height = box_images[0].size
    if count == 1:
        output = box_images[0]
    else:
        gap = 18
        cols = 5
        rows = (count + 4) // 5
        cards_total_width = card_width * cols + gap * (cols - 1)
        cards_total_height = card_height * rows + gap * (rows - 1)
        
        bg_path = None
        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
            test_path = LEVEL_DIR / bg_name
            if test_path.exists():
                bg_path = str(test_path)
                break
        
        if bg_path:
            log_info(f"使用十连背景图片: {bg_path}")
            bg_img = Image.open(bg_path).convert('RGB')
            
            bg_w, bg_h = bg_img.size
            final_w = int(bg_w * 2.0 * 0.264)
            final_h = int(bg_h * 2.0 * 0.264)
            bg_img_resized = bg_img.resize((final_w, final_h), Image.Resampling.LANCZOS)
            
            output = Image.new('RGB', (final_w, final_h), (50, 50, 50))
            output.paste(bg_img_resized, (0, 0))
            
            cards_x = (final_w - cards_total_width) // 2
            cards_y = (final_h - cards_total_height) // 2
        else:
            output = Image.new('RGB', (cards_total_width, cards_total_height), (50, 50, 50))
            cards_x = 0
            cards_y = 0
        
        for i, img in enumerate(box_images):
            row = i // cols
            col = i % cols
            x = cards_x + col * (card_width + gap)
            y = cards_y + row * (card_height + gap)
            output.paste(img, (x, y))
    
    # 构建提示消息
    box_hints = [f"选择{i+1}" for i in range(count)]
    hint_text = " | ".join(box_hints[:5])
    if count > 5:
        hint_text += "\n" + " | ".join(box_hints[5:10])
    
    fes_pity_text = ""
    if is_fes_pity:
        fes_pity_text = "🎉 触发FES保底，今日时运为man！\n"
    
    current_gacha = user_data["gacha"]
    total_draws = pity_data.get("total_gacha", 0)
    total_3stars = user_data.get("fes_count", 0) + user_data.get("period_count", 0) + user_data.get("other_3star_count", 0)
    remaining_pity = get_remaining_pity(user_id)
    
    info_text = (
        f"当前呱太: {current_gacha}\n"
        f"累计次数: {total_draws}\n"
        f"累计3星: {total_3stars}\n"
        f"距离FES保底: {remaining_pity}次\n"
    )
    prompt_text = f"{fes_pity_text}{info_text}\n请输入「/开 编号」开启盲盒，如「/开 1」或「/开 全部开」"
    
    send_kook_message_with_image(channel_id, prompt_text, output, quote_msg_id=msg_id)

def handle_limited_gacha(user_id, channel_id, msg_id=None):
    """处理限定十连抽卡请求（必出FES/期间限定三星）"""
    user_data = load_user_data(user_id)
    pity_data = load_pity_data(user_id)
    
    cost = CONFIG["LIMITED_GACHA_COST"]
    
    if user_data["gacha"] < cost:
        send_kook_message(channel_id, f"呱太不足！需要 {cost} 呱太，当前 {user_data['gacha']} 呱太", quote_msg_id=msg_id)
        return
    
    # 限定池冷却检查
    now_ts = datetime.now().timestamp()
    last_limited = pity_data.get("last_limited_gacha", 0)
    remaining = CONFIG["LIMITED_GACHA_COOLDOWN_SECONDS"] - (now_ts - last_limited)
    if remaining > 0:
        send_kook_message(channel_id, f"限定池冷却中！请等待 {int(remaining // 60)} 分钟 {int(remaining % 60)} 秒后再试~", quote_msg_id=msg_id)
        return
    
    user_data["gacha"] -= cost
    pity_data["last_limited_gacha"] = now_ts
    
    # 限定池必出FES或期间限定三星
    limited_chars = [c for c in CHARACTERS.values() if c["rarity"] == 3 and c.get("limit_type") in ["フェス限定", "期間限定"]]
    if not limited_chars:
        send_kook_message(channel_id, "没有可用的限定角色！", quote_msg_id=msg_id)
        return
    
    results = []
    # 1个限定三星 + 9个随机卡
    limited_char = random.choice(limited_chars)
    results.append(limited_char.copy())
    
    for _ in range(9):
        stars = random.choices(population=[1, 2, 3], weights=GACHA_WEIGHTS, k=1)[0]
        available = [c for c in CHARACTERS.values() if c["rarity"] == stars]
        char = random.choice(available) if available else random.choice(list(CHARACTERS.values()))
        results.append(char.copy())
    
    # 更新数据
    pity_data["total_gacha"] = pity_data.get("total_gacha", 0) + 10
    
    for card in results:
        user_data["cards"].append(card)
        if card["rarity"] == 3:
            pity_data["pity_count"] = 0
            pity_data["ten_pity"] = 0
            pity_data["normal_pity"] = 0
            pity_data["last_3star"] = pity_data["total_gacha"]
            
            limit_type = card.get("limit_type", "")
            if limit_type == "フェス限定":
                user_data["fes_count"] = user_data.get("fes_count", 0) + 1
                pity_data["fes_pity_count"] = 0
            elif limit_type == "期間限定":
                user_data["period_count"] = user_data.get("period_count", 0) + 1
            else:
                user_data["other_3star_count"] = user_data.get("other_3star_count", 0) + 1
        elif card["rarity"] == 2:
            user_data["total_2stars"] = user_data.get("total_2stars", 0) + 1
    
    save_user_data(user_id)
    save_pity_data(user_id)
    
    # 发送结果
    three_stars = [c for c in results if c["rarity"] == 3]
    two_stars = [c for c in results if c["rarity"] == 2]
    one_stars = [c for c in results if c["rarity"] == 1]
    
    result_text = "🎉 限定十连抽结果：\n"
    if three_stars:
        result_text += "⭐⭐⭐: " + ", ".join([c["name"] for c in three_stars]) + "\n"
    if two_stars:
        result_text += "⭐⭐: " + ", ".join([c["name"] for c in two_stars]) + "\n"
    if one_stars:
        result_text += "⭐: " + ", ".join([c["name"] for c in one_stars]) + "\n"
    
    result_text += f"\n剩余呱太: {user_data['gacha']}"
    send_kook_message(channel_id, result_text, quote_msg_id=msg_id)

def handle_help(user_id, channel_id, raw_message=None, msg_id=None):
    """分章节帮助系统（与QQ版一致，支持二级内容查询）"""
    # 解析命令参数
    cmd = raw_message[1:] if raw_message and raw_message.startswith('/') else ''
    
    # 二级帮助内容
    help_sections = {
        "抽卡": f"""
╔══════════════════════════════╗
║        🎴 抽卡帮助           ║
╠══════════════════════════════╣
║                              ║
║  /单抽        - 消耗{CONFIG['GACHA_COST']}呱太抽卡      ║
║  /十连        - 消耗{CONFIG['GACHA10_COST']}呱太十连抽 ║
║  /限定十连    - 消耗{CONFIG['LIMITED_GACHA_COST']}呱太  ║
║                （必出限定三星）                         ║
║                              ║
║  抽卡概率：                   ║
║  ⭐ 1星: 72%                 ║
║  ⭐⭐ 2星: 23%                ║
║  ⭐⭐⭐ 3星: 5%                ║
║                              ║
║  保底机制：                   ║
║  每{CONFIG['PITY_LIMIT']}抽必出フェス限定三星          ║
╚══════════════════════════════╝
""",
        "战斗": """
╔══════════════════════════════╗
║        ⚔️ 战斗帮助           ║
╠══════════════════════════════╣
║                              ║
║  /战斗        - AI对战        ║
║  /BOSS战      - 挑战BOSS      ║
║  /挑战 排名   - 挑战排行榜玩家 ║
║  /战斗日志    - 查看战斗记录   ║
║  /战斗GIF     - 生成战斗回放   ║
║                              ║
║  /排行榜      - 战力排行榜     ║
║  /抽卡榜单    - 抽卡排行榜     ║
╚══════════════════════════════╝
""",
        "队伍": """
╔══════════════════════════════╗
║        📋 队伍帮助           ║
╠══════════════════════════════╣
║                              ║
║  /队伍        - 查看当前队伍   ║
║  /队伍 自动配队 - AI配队      ║
║  /队伍 我的卡  - 查看三星卡    ║
║  /队伍 设置 X Y              ║
║                - 设置位置X为卡Y║
║  /队伍 切换 N  - 切换预设槽N  ║
║  /队伍 清空    - 清空队伍      ║
║                              ║
║  /防守队      - 查看防守队     ║
║  /防守队 设置N - 设置防守队    ║
╚══════════════════════════════╝
""",
        "经济": f"""
╔══════════════════════════════╗
║        💰 经济帮助           ║
╠══════════════════════════════╣
║                              ║
║  /签到        - 每日签到      ║
║                (+{CONFIG['DAILY_REWARD']}呱太)         ║
║  /获取呱太    - 获取呱太      ║
║                (+{CONFIG['GET_GACHA_REWARD']}呱太)     ║
║  /兑换呱太    - 碎片换呱太    ║
║                红碎片1:5      ║
║                蓝碎片1:20     ║
║                              ║
║  /三星池子    - 三星抽卡      ║
║  /红抽        - 红碎片抽三星  ║
║  /蓝抽        - 蓝碎片抽三星  ║
╚══════════════════════════════╝
""",
        "其他": """
╔══════════════════════════════╗
║        📦 其他帮助           ║
╠══════════════════════════════╣
║                              ║
║  /个人记录    - 查看个人统计   ║
║  /详细信息    - 详细数据面板   ║
║  /收藏        - 卡牌收藏      ║
║  /保底        - 保底进度      ║
║  /余额        - 呱太余额      ║
║                              ║
║  管理员命令：                 ║
║  /重新加载    - 重载配置      ║
║  /备份        - 备份数据      ║
╚══════════════════════════════╝
"""
    }
    
    # 检查是否有二级参数
    for section in help_sections:
        if section in cmd:
            send_kook_message(channel_id, help_sections[section], quote_msg_id=msg_id)
            return
    
    # 默认显示帮助总览
    help_text = f"""
╔══════════════════════════════╗
║     🃏 小千 帮助总览         ║
╠══════════════════════════════╣
║                              ║
║  📖 请选择章节：              ║
║                              ║
║  🎴 /帮助 抽卡 — 抽卡相关     ║
║  ⚔️ /帮助 战斗 — 对战/BOSS/排行║
║  📋 /帮助 队伍 — 配队/防守队  ║
║  💰 /帮助 经济 — 货币/签到    ║
║  📦 /帮助 其他 — 个人记录等   ║
║                              ║
║  例：发送「/帮助 抽卡」       ║
║                              ║
║  常用命令速查：               ║
║  /单抽 ({CONFIG['GACHA_COST']}呱太)   /十连 ({CONFIG['GACHA10_COST']}呱太) ║
║  /限定十连 ({CONFIG['LIMITED_GACHA_COST']}呱太)  /签到 (+{CONFIG['DAILY_REWARD']}) ║
║  /获取呱太 (+{CONFIG['GET_GACHA_REWARD']})  /队伍  /战斗  ║
╚══════════════════════════════╝
"""
    send_kook_message(channel_id, help_text, quote_msg_id=msg_id)

def handle_get_gacha(user_id, channel_id, msg_id=None):
    user_data = load_user_data(user_id)
    pity_data = load_pity_data(user_id)
    
    # 冷却检查
    now_ts = datetime.now().timestamp()
    last_get_gacha = pity_data.get("last_get_gacha", 0)
    remaining = CONFIG["GET_GACHA_COOLDOWN_SECONDS"] - (now_ts - last_get_gacha)
    if remaining > 0:
        send_kook_message(channel_id, f"获取呱太冷却中！请等待 {int(remaining)} 秒后再试~", quote_msg_id=msg_id)
        return
    
    user_data["gacha"] += CONFIG["GET_GACHA_REWARD"]
    pity_data["last_get_gacha"] = now_ts
    
    save_user_data(user_id)
    save_pity_data(user_id)
    
    send_kook_message(channel_id, f"获得{CONFIG['GET_GACHA_REWARD']}呱太！当前呱太: {user_data['gacha']}", quote_msg_id=msg_id)

def handle_signin(user_id, channel_id, msg_id=None):
    result = signin(user_id)
    
    if result["success"]:
        card = create_kook_card(
            "签到成功！",
            f"获得 {CONFIG['DAILY_REWARD']} 呱太",
            [
                {"name": "连续签到", "value": f"{result['streak']} 天"},
                {"name": "累计签到", "value": f"{result['total_days']} 天"},
                {"name": "当前呱太", "value": str(result['gacha'])}
            ]
        )
        send_kook_card_message(channel_id, card, quote_msg_id=msg_id)
    else:
        send_kook_message(channel_id, result["message"], quote_msg_id=msg_id)

def handle_team(user_id, channel_id, raw_message, msg_id=None):
    """
    处理队伍配置请求（与QQ版一致）
    命令格式:
    - /队伍 - 查看当前队伍
    - /队伍 我的卡 - 查看拥有的3星卡
    - /队伍 设置 位置 序号 - 设置指定位置的卡
    - /队伍 清除 位置 - 清除指定位置的卡
    - /队伍 清空 - 清空整个队伍
    - /队伍 自动配队 - AI自动配队
    - /队伍 切换 N - 切换到预设槽位N
    - /队伍 预设 - 查看所有预设槽位
    """
    if not TEAM_SYSTEM_LOADED:
        send_kook_message(channel_id, "配队系统未加载！", quote_msg_id=msg_id)
        return
    
    # 自动配队命令
    if '自动配队' in raw_message or '自动' in raw_message:
        result = auto_build_team(user_id, CHARACTERS)
        
        if not result["success"]:
            send_kook_message(channel_id, result["message"], quote_msg_id=msg_id)
            return
        
        # 保存配队结果
        if result["team"]:
            save_team_data(user_id, result["team"])
            save_user_data(user_id)
        
        # 生成队伍图片（与QQ版一致）
        team_data = load_team_data(user_id)
        img_path = build_team_image(team_data, CHARACTERS)
        
        if img_path and os.path.exists(img_path):
            from PIL import Image
            img = Image.open(img_path)
            send_kook_message_with_image(channel_id, "🤖 自动配队完成！", img, quote_msg_id=msg_id)
            # 删除临时文件
            os.remove(img_path)
        else:
            send_kook_message(channel_id, "🤖 自动配队完成！", quote_msg_id=msg_id)
        return
    
    # 查看我的卡（支持翻页，与QQ版一致）
    if '我的卡' in raw_message:
        # 获取用户当前查看的页码（从session或默认第1页）
        team_session_file = INFO_DIR / f"team_session_{user_id}.json"
        current_page = 1
        if team_session_file.exists():
            try:
                with open(team_session_file, "r", encoding="utf-8") as f:
                    session_data = json.load(f)
                    current_page = session_data.get("cards_page", 1)
            except:
                current_page = 1
        
        # 获取总页数
        user_cards = get_user_3star_cards(user_id, CHARACTERS)
        total_pages = max(1, (len(user_cards) + 50 - 1) // 50)
        
        # 处理翻页
        import re
        page_match = re.search(r'我的卡\s+(第)?(\d+)(页)?', raw_message)
        if page_match:
            target_page = int(page_match.group(2))
            current_page = max(1, min(target_page, total_pages))
        elif '下一页' in raw_message:
            current_page += 1
            if current_page > total_pages:
                current_page = total_pages
        elif '上一页' in raw_message:
            current_page -= 1
            if current_page < 1:
                current_page = 1
        
        # 确保页码在有效范围内
        current_page = max(1, min(current_page, total_pages))
        
        # 保存当前页码
        with open(team_session_file, "w", encoding="utf-8") as f:
            json.dump({"cards_page": current_page}, f)
        
        # 显示用户拥有的三星卡（50张一页）
        img_path, current_cards, total_pages = build_3star_cards_image(user_id, CHARACTERS, current_page, 50)
        
        if not current_cards:
            send_kook_message(channel_id, "你还没有三星卡~", quote_msg_id=msg_id)
            return
        
        # 构建消息（只有图片和页码提示）
        page_info = f"第{current_page}/{total_pages}页"
        if total_pages > 1:
            if current_page < total_pages:
                page_info += " | 输入「/队伍 我的卡 下一页」查看下一页"
            if current_page > 1:
                page_info += " | 输入「/队伍 我的卡 上一页」查看上一页"
            page_info += " | 输入「/队伍 我的卡 页码」跳转到指定页"
        
        # 使用提示
        current_page_size = len(current_cards)
        usage_hint = f"设置: /队伍 设置 位置 序号(1-{current_page_size}) | 切换预设: /队伍 切换 1~6"
        
        if img_path and os.path.exists(img_path):
            from PIL import Image
            img = Image.open(img_path)
            send_kook_message_with_image(channel_id, f"{page_info}\n{usage_hint}", img, quote_msg_id=msg_id)
            os.remove(img_path)
        else:
            send_kook_message(channel_id, f"{page_info}\n{usage_hint}", quote_msg_id=msg_id)
        return
    
    # 切换预设
    import re
    switch_match = re.search(r'切换\s*(\d+)', raw_message)
    if switch_match:
        slot = int(switch_match.group(1))
        if load_preset(user_id, slot):
            send_kook_message(channel_id, f"已切换到预设槽位{slot}！", quote_msg_id=msg_id)
        else:
            send_kook_message(channel_id, f"切换失败！预设槽位{slot}不存在或为空", quote_msg_id=msg_id)
        return
    
    # 查看预设
    if '预设' in raw_message:
        presets_info = list_presets_info(user_id, CHARACTERS)
        send_kook_message(channel_id, presets_info, quote_msg_id=msg_id)
        return
    
    # 设置队伍位置
    set_match = re.search(r'设置\s*(\d+)\s+(\d+)', raw_message)
    if set_match:
        position = int(set_match.group(1))
        card_idx = int(set_match.group(2))
        
        user_cards = get_user_3star_cards(user_id, CHARACTERS)
        if card_idx <= 0 or card_idx > len(user_cards):
            send_kook_message(channel_id, f"卡牌序号超出范围！你有{len(user_cards)}张三星卡", quote_msg_id=msg_id)
            return
        
        if position < 1 or position > 6:
            send_kook_message(channel_id, "位置必须在1~6之间！", quote_msg_id=msg_id)
            return
        
        card = user_cards[card_idx - 1]
        if set_team_card(user_id, position, str(card["id"]), "battle"):
            send_kook_message(channel_id, f"位置{position}已设置为: {card['name']}", quote_msg_id=msg_id)
        else:
            send_kook_message(channel_id, "设置失败！", quote_msg_id=msg_id)
        return
    
    # 清除队伍位置
    clear_match = re.search(r'清除\s*(\d+)', raw_message)
    if clear_match:
        position = int(clear_match.group(1))
        if clear_team_card(user_id, position, "battle"):
            send_kook_message(channel_id, f"已清除位置{position}的卡牌", quote_msg_id=msg_id)
        else:
            send_kook_message(channel_id, "清除失败！", quote_msg_id=msg_id)
        return
    
    # 清空队伍
    if '清空' in raw_message:
        clear_all_team(user_id)
        send_kook_message(channel_id, "队伍已清空！", quote_msg_id=msg_id)
        return
    
    # 默认：查看当前队伍（显示队伍图片）
    team_data = load_team_data(user_id)
    img_path = build_team_image(team_data, CHARACTERS)
    
    if img_path and os.path.exists(img_path):
        from PIL import Image
        img = Image.open(img_path)
        send_kook_message_with_image(channel_id, "当前队伍：", img, quote_msg_id=msg_id)
        os.remove(img_path)
    else:
        team_info = get_team_info(user_id, CHARACTERS)
        send_kook_message(channel_id, team_info, quote_msg_id=msg_id)

def handle_defense_team(user_id, channel_id, raw_message=None, msg_id=None):
    """处理防守队请求（与QQ版一致）"""
    if not TEAM_SYSTEM_LOADED:
        send_kook_message(channel_id, "配队系统未加载！", quote_msg_id=msg_id)
        return
    
    # 设置防守队
    if raw_message and "设置" in raw_message:
        import re
        slot_match = re.search(r'设置\s*(\d+)', raw_message)
        if slot_match:
            slot = int(slot_match.group(1))
            if set_defense_slot(user_id, slot):
                send_kook_message(channel_id, f"防守队已设置为预设槽{slot}！", quote_msg_id=msg_id)
            else:
                send_kook_message(channel_id, f"无效的预设槽位！", quote_msg_id=msg_id)
            return
    
    # 查看防守队信息
    defense_info = get_defense_team_info(user_id, CHARACTERS)
    send_kook_message(channel_id, defense_info, quote_msg_id=msg_id)

def handle_personal_info(user_id, channel_id, msg_id=None):
    """显示个人记录（与QQ版一致）"""
    user_data = load_user_data(user_id)
    pity_data = load_pity_data(user_id)
    
    cards = user_data["cards"]
    three_stars = [c for c in cards if c["rarity"] == 3]
    
    # 计算保底进度
    fes_pity_remaining = max(0, CONFIG["PITY_LIMIT"] - pity_data.get("fes_pity_count", 0))
    
    # 统计限定卡数量
    fes_count = user_data.get("fes_count", 0)
    period_count = user_data.get("period_count", 0)
    other_3star_count = user_data.get("other_3star_count", 0)
    
    # 战力
    power = calculate_power(user_id)
    
    card = create_kook_card("📊 个人记录", "")
    
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": "🐸 资源"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   呱太: {user_data['gacha']}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   红色碎片: {user_data.get('red_crystal', 0)}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   蓝色碎片: {user_data.get('blue_crystal', 0)}"}
    })
    
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": "\n🎫 抽卡统计"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   累计抽卡: {pity_data.get('total_gacha', 0)}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   三星总数: {pity_data.get('total_3stars', 0)}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   保底进度: {fes_pity_remaining}/{CONFIG['PITY_LIMIT']}"}
    })
    
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": "\n⭐ 卡牌收藏"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   卡牌总数: {len(cards)}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   三星卡牌: {len(three_stars)}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   フェス限定: {fes_count}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   期間限定: {period_count}"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   其他三星: {other_3star_count}"}
    })
    
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": "\n⚔️ 战力"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   总战力: {power}"}
    })
    
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": "\n📅 签到"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   连续签到: {user_data['signin']['streak']} 天"}
    })
    card["modules"].append({
        "type": "section",
        "text": {"type": "plain-text", "content": f"   累计签到: {user_data['signin']['total_days']} 天"}
    })
    
    send_kook_card_message(channel_id, card, quote_msg_id=msg_id)

def handle_exchange_crystal(user_id, channel_id, msg_id=None):
    """处理碎片兑换呱太请求（与QQ版一致）"""
    try:
        pity_data = load_pity_data(user_id)
        
        # 获取用户碎片数量
        red_crystal = pity_data.get("red_crystal", 0)
        blue_crystal = pity_data.get("blue_crystal", 0)
        
        if red_crystal == 0 and blue_crystal == 0:
            send_kook_message(channel_id, "你没有可以兑换的碎片！", quote_msg_id=msg_id)
            return
        
        # 计算兑换数量（与QQ版一致：红碎片1:5，蓝碎片1:20）
        red_amount = red_crystal * 5
        blue_amount = blue_crystal * 20
        total_amount = red_amount + blue_amount
        
        # 清空碎片
        pity_data["red_crystal"] = 0
        pity_data["blue_crystal"] = 0
        
        # 添加呱太（使用独立的呱太数据模块）
        add_gacha(user_id, total_amount)
        
        save_pity_data(user_id, pity_data)
        
        # 构建回复消息
        parts = []
        if red_crystal > 0:
            parts.append(f"🔴红色碎片 x{red_crystal} → {red_amount} 呱太")
        if blue_crystal > 0:
            parts.append(f"🔵蓝色碎片 x{blue_crystal} → {blue_amount} 呱太")
        
        parts_str = '\n'.join(parts)
        current_gacha = get_gacha_count(user_id)
        card = create_kook_card("💎 兑换成功！", parts_str)
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"总共获得: {total_amount} 呱太"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"当前呱太: {current_gacha}"}
        })
        
        send_kook_card_message(channel_id, card, quote_msg_id=msg_id)
        log_info(f"碎片兑换 [{user_id}]: red={red_crystal}, blue={blue_crystal}, total={total_amount}")
        
    except Exception as e:
        log_error(f"碎片兑换失败: {e}")
        send_kook_message(channel_id, f"碎片兑换失败: {str(e)}", quote_msg_id=msg_id)

def get_gacha_leaderboard() -> list:
    """获取抽卡榜单（三星个数前10名）"""
    leaderboard = []

    # 使用INFO_DIR，与QQ版一致
    for user_file in INFO_DIR.glob("user_*.json"):
        try:
            user_id = user_file.stem.replace("user_", "")
            user_data = load_user_data(user_id)
            pity_data = load_pity_data(user_id)

            total_draws = pity_data.get("total_gacha", 0)
            total_3stars = pity_data.get("total_3stars", 0)

            # 计算三星率
            rate = round(total_3stars / total_draws * 100, 2) if total_draws > 0 else 0.0

            leaderboard.append({
                "user_id": user_id,
                "total_draws": total_draws,
                "total_3stars": total_3stars,
                "rate": rate,
                "fes_count": user_data.get("fes_count", 0),
                "period_count": user_data.get("period_count", 0),
                "other_3star_count": user_data.get("other_3star_count", 0),
                "total_2stars": user_data.get("total_2stars", 0)
            })
        except Exception as e:
            log_error(f"读取用户数据失败 {user_file}: {e}")

    # 按三星个数降序排序，相同则按总抽卡数升序
    leaderboard.sort(key=lambda x: (-x["total_3stars"], x["total_draws"]))

    return leaderboard[:10]

def handle_gacha_leaderboard(user_id, channel_id, msg_id=None):
    """处理抽卡榜单查询请求（按三星个数排行）"""
    try:
        leaderboard = get_gacha_leaderboard()

        if not leaderboard:
            send_kook_message(channel_id, "暂无抽卡数据！", quote_msg_id=msg_id)
            return

        rank_emojis = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]

        card = create_kook_card("🎰 抽卡榜单 TOP 10", "（按三星个数排行）")

        for i, player in enumerate(leaderboard):
            rank = i + 1
            emoji = rank_emojis[i] if i < len(rank_emojis) else f"{rank}."

            # 标记查询者
            is_self = player["user_id"] == user_id
            self_mark = " 👈(你)" if is_self else ""

            # 格式化三星率
            rate_str = f"{player['rate']:.1f}%" if player['total_draws'] > 0 else "0.0%"

            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"{emoji} {player['user_id']}{self_mark}\n   ⭐三星: {player['total_3stars']} | 🎫总抽: {player['total_draws']} | 📊三星率: {rate_str}"}
            })

        send_kook_card_message(channel_id, card, quote_msg_id=msg_id)
        log_info(f"查询抽卡榜单 [{user_id}]")

    except Exception as e:
        log_error(f"查询抽卡榜单失败: {e}")
        send_kook_message(channel_id, f"查询抽卡榜单失败: {str(e)}", quote_msg_id=msg_id)

def calculate_power(user_id) -> int:
    """计算用户战力：フェス限定数*10 + 期間限定数*8 + 其他三星数*7 + 二星数*3"""
    user_data = load_user_data(user_id)
    fes = user_data.get("fes_count", 0)
    period = user_data.get("period_count", 0)
    other_3star = user_data.get("other_3star_count", 0)
    two_star = user_data.get("total_2stars", 0)
    return fes * 10 + period * 8 + other_3star * 7 + two_star * 3

def get_leaderboard() -> list:
    """获取排行榜（战力前10名）"""
    leaderboard = []
    
    # 遍历所有用户数据文件（使用INFO_DIR，与QQ版一致）
    for user_file in INFO_DIR.glob("user_*.json"):
        try:
            user_id = user_file.stem.replace("user_", "")
            user_data = load_user_data(user_id)
            
            # 计算战力
            power = calculate_power(user_id)
            
            # 获取基本信息
            total_draws = load_pity_data(user_id).get("total_gacha", 0)
            
            leaderboard.append({
                "user_id": user_id,
                "power": power,
                "total_draws": total_draws,
                "fes_count": user_data.get("fes_count", 0),
                "period_count": user_data.get("period_count", 0),
                "other_3star_count": user_data.get("other_3star_count", 0),
                "total_2stars": user_data.get("total_2stars", 0)
            })
        except Exception as e:
            log_error(f"读取用户数据失败 {user_file}: {e}")
    
    # 按战力降序排序
    leaderboard.sort(key=lambda x: x["power"], reverse=True)
    
    # 返回前10名
    return leaderboard[:10]

def handle_ranking(user_id, channel_id, msg_id=None):
    """处理排行榜查询请求"""
    try:
        # 获取排行榜数据
        leaderboard = get_leaderboard()
        
        if not leaderboard:
            send_kook_message(channel_id, "暂无排行榜数据！", quote_msg_id=msg_id)
            return
        
        # 构建排行榜消息
        rank_emojis = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]
        
        card = create_kook_card("🏆 战力排行榜", "")
        
        for i, player in enumerate(leaderboard):
            rank = i + 1
            emoji = rank_emojis[i] if i < len(rank_emojis) else f"{rank}."
            
            # 获取用户自己的战力
            if player["user_id"] == user_id:
                power_info = f"⚔️{player['power']} (你)"
            else:
                power_info = f"⚔️{player['power']}"
            
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"{emoji} {player['user_id']} {power_info}\n   累计抽卡: {player['total_draws']}"}
            })
        
        send_kook_card_message(channel_id, card, quote_msg_id=msg_id)
        log_info(f"查询排行榜 [{user_id}]")
        
    except Exception as e:
        log_error(f"查询排行榜失败: {e}")
        send_kook_message(channel_id, f"查询排行榜失败: {str(e)}", quote_msg_id=msg_id)

def handle_show_details(user_id, channel_id, msg_id=None):
    """处理详细信息查询请求"""
    try:
        user_data = load_user_data(user_id)
        pity_data = load_pity_data(user_id)
        
        # 获取统计信息
        total_draws = pity_data.get("total_gacha", 0)
        total_3stars = pity_data.get("total_3stars", 0)
        total_2stars = user_data.get("total_2stars", 0)
        fes_count = user_data.get("fes_count", 0)
        period_count = user_data.get("period_count", 0)
        other_3star_count = user_data.get("other_3star_count", 0)
        
        # 计算三星率
        rate = round(total_3stars / total_draws * 100, 2) if total_draws > 0 else 0.0
        
        # 计算战力
        power = calculate_power(user_id)
        
        # 获取碎片数量
        red_crystal = user_data.get("red_crystal", 0)
        blue_crystal = user_data.get("blue_crystal", 0)
        
        # 获取呱太数量
        gacha = user_data.get("gacha", 0)
        
        # 获取签到信息
        signin = user_data.get("signin", {})
        streak = signin.get("streak", 0)
        total_days = signin.get("total_days", 0)
        
        # 获取最近获得的三星
        recent_3stars = user_data.get("recent_3stars", [])
        
        # 获取卡牌收藏数量
        card_collection = user_data.get("card_collection", {})
        collection_count = len(card_collection)
        
        # 构建卡片消息
        card = create_kook_card("📊 详细信息", "")
        
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": "🎫 抽卡统计"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   总抽卡: {total_draws}"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   ⭐三星: {total_3stars}"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   ⭐⭐二星: {total_2stars}"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   📊三星率: {rate:.1f}%"}
        })
        
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": "\n🎖️ 战力构成"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   ⚔️总战力: {power}"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   🔥フェス限定: {fes_count} (+{fes_count * 10})"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   ⏰期間限定: {period_count} (+{period_count * 8})"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   ⭐其他三星: {other_3star_count} (+{other_3star_count * 7})"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   ⭐⭐二星: {total_2stars} (+{total_2stars * 3})"}
        })
        
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": "\n💎 资源"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   🐸呱太: {gacha}"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   🔴红色碎片: {red_crystal}"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   🔵蓝色碎片: {blue_crystal}"}
        })
        
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": "\n📅 签到"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   连续签到: {streak} 天"}
        })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"   累计签到: {total_days} 天"}
        })
        
        # 最近获得的三星
        if recent_3stars:
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": "\n✨ 最近获得"}
            })
            for item in recent_3stars[:5]:
                limit_badge = f" [{item.get('limit_type')}]" if item.get('limit_type') else ""
                card["modules"].append({
                    "type": "section",
                    "text": {"type": "plain-text", "content": f"   ⭐ {item.get('name', '')}{limit_badge}"}
                })
        
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"\n📚 卡牌收藏: {collection_count} 张"}
        })
        
        send_kook_card_message(channel_id, card, quote_msg_id=msg_id)
        log_info(f"查询详细信息 [{user_id}]")
        
    except Exception as e:
        log_error(f"查询详细信息失败: {e}")
        send_kook_message(channel_id, f"查询详细信息失败: {str(e)}", quote_msg_id=msg_id)

def handle_battle_log(user_id, channel_id, gen_gif=False, msg_id=None):
    user_data = load_user_data(user_id)
    
    if not user_data["battle_history"]:
        send_kook_message(channel_id, "没有战斗记录！", quote_msg_id=msg_id)
        return
    
    last_battle = user_data["battle_history"][-1]
    
    if gen_gif and GIF_RENDERER_LOADED:
        # 使用BytesIO直接发送GIF，不保存到本地
        gif_buffer = battle_to_gif_bytes(last_battle)
        
        if gif_buffer:
            send_kook_gif_bytes(channel_id, "战斗回放", gif_buffer, quote_msg_id=msg_id)
        else:
            send_kook_message(channel_id, "生成GIF失败！", quote_msg_id=msg_id)
    else:
        log_text = "\n".join(last_battle.get("log", [])[:20])
        send_kook_message(channel_id, f"最近战斗日志：\n{log_text}", quote_msg_id=msg_id)

def select_3star_from_pool() -> dict:
    """从三星池子中随机抽取一个角色（只返回三星角色）"""
    three_star_chars = [c for c in CHARACTERS.values() if c.get("rarity") == 3]
    
    if not three_star_chars:
        return None
    
    # 随机选择
    selected = random.choice(three_star_chars)
    return selected

def handle_3star_pool(user_id, channel_id, raw_message, msg_id=None):
    """
    处理三星池子抽卡请求
    命令格式:
    - 三星池子 - 显示三星池子介绍
    - 红抽 - 使用红色碎片抽卡
    - 蓝抽 - 使用蓝色碎片抽卡
    """
    try:
        user_data = load_user_data(user_id)
        
        # 检测抽卡类型
        if '红抽' in raw_message:
            crystal_type = "red"
            crystal_name = "红色碎片"
            cost = THREE_STAR_POOL_RED_COST
            current = user_data.get("red_crystal", 0)
        elif '蓝抽' in raw_message:
            crystal_type = "blue"
            crystal_name = "蓝色碎片"
            cost = THREE_STAR_POOL_BLUE_COST
            current = user_data.get("blue_crystal", 0)
        else:
            # 显示三星池子介绍
            red_crystal = user_data.get("red_crystal", 0)
            blue_crystal = user_data.get("blue_crystal", 0)
            
            card = create_kook_card("★★★ 三星池子 ★★★", "")
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": "消耗说明:"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"🔴 红色碎片: {THREE_STAR_POOL_RED_COST}个/抽"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"🔵 蓝色碎片: {THREE_STAR_POOL_BLUE_COST}个/抽"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": "\n你当前拥有:"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"🔴 红色碎片: {red_crystal}个"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"🔵 蓝色碎片: {blue_crystal}个"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": "\n可抽取次数:"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"🔴 红色碎片: {red_crystal // THREE_STAR_POOL_RED_COST}次"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"🔵 蓝色碎片: {blue_crystal // THREE_STAR_POOL_BLUE_COST}次"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": "\n输入「三星池子红抽」使用红色碎片抽"}
            })
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": "输入「三星池子蓝抽」使用蓝色碎片抽"}
            })
            
            send_kook_card_message(channel_id, card, quote_msg_id=msg_id)
            return
        
        # 检查碎片是否足够
        if current < cost:
            send_kook_message(channel_id, f"{crystal_name}不足！需要{cost}个，当前拥有{current}个", quote_msg_id=msg_id)
            return
        
        # 消耗碎片
        if crystal_type == "red":
            user_data["red_crystal"] = current - cost
        else:
            user_data["blue_crystal"] = current - cost
        
        # 抽取角色
        selected = select_3star_from_pool()
        
        if not selected:
            send_kook_message(channel_id, "三星池子为空，请联系管理员！", quote_msg_id=msg_id)
            return
        
        # 更新用户数据
        card_id = str(selected.get("id", ""))
        limit_type = selected.get("limit_type", "")
        chara_name = selected.get("name", "")
        
        # 添加到卡牌收藏
        if card_id not in user_data.get("card_collection", {}):
            user_data["card_collection"][card_id] = {
                "name": chara_name,
                "stars": 3,
                "limit_type": limit_type,
                "count": 1,
                "first_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "last_time": datetime.now().timestamp()
            }
            # 更新计数
            if limit_type == "フェス限定":
                user_data["fes_count"] = user_data.get("fes_count", 0) + 1
            elif limit_type == "期間限定":
                user_data["period_count"] = user_data.get("period_count", 0) + 1
            else:
                user_data["other_3star_count"] = user_data.get("other_3star_count", 0) + 1
        else:
            user_data["card_collection"][card_id]["count"] += 1
            user_data["card_collection"][card_id]["last_time"] = datetime.now().timestamp()
        
        # 添加到最近获得记录
        recent_3stars = user_data.get("recent_3stars", [])
        recent_3stars.insert(0, {
            "card_id": card_id,
            "name": chara_name,
            "limit_type": limit_type,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        if len(recent_3stars) > 10:
            recent_3stars = recent_3stars[:10]
        user_data["recent_3stars"] = recent_3stars
        
        save_user_data(user_id)
        
        # 生成卡牌图片
        img = generate_card_image(selected)
        
        # 发送结果
        remaining = current - cost
        card = create_kook_card("💎 三星池子抽卡", f"使用{cost}个{crystal_name}！")
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"⭐⭐⭐ {chara_name}"}
        })
        if limit_type:
            card["modules"].append({
                "type": "section",
                "text": {"type": "plain-text", "content": f"限定类型: {limit_type}"}
            })
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": f"剩余{crystal_name}: {remaining}个"}
        })
        
        send_kook_card_message(channel_id, card, quote_msg_id=msg_id)
        
        if img:
            send_kook_message_with_image(channel_id, "", img)
        
        log_info(f"三星池子 [{user_id}]: type={crystal_type}, card={chara_name}")
        
    except Exception as e:
        log_error(f"三星池失败: {e}")
        send_kook_message(channel_id, f"三星池失败: {str(e)}", quote_msg_id=msg_id)

def handle_challenge(user_id, channel_id, raw_message, msg_id=None):
    """
    处理挑战排名请求
    命令格式: 挑战排名 <排名>
    挑战排行榜上指定排名的玩家
    """
    try:
        import re
        
        # 解析目标排名
        match = re.search(r'挑战排名\s*(\d+)', raw_message)
        if not match:
            send_kook_message(channel_id, "请输入要挑战的排名！格式：挑战排名 1", quote_msg_id=msg_id)
            return
        
        target_rank = int(match.group(1))
        if target_rank < 1:
            send_kook_message(channel_id, "请输入有效的排名！", quote_msg_id=msg_id)
            return
        
        # 获取排行榜
        leaderboard = get_leaderboard()
        
        if target_rank > len(leaderboard):
            send_kook_message(channel_id, f"排行榜只有 {len(leaderboard)} 位玩家！", quote_msg_id=msg_id)
            return
        
        # 获取目标玩家
        target_player = leaderboard[target_rank - 1]
        target_user_id = target_player["user_id"]
        
        # 不能挑战自己
        if target_user_id == user_id:
            send_kook_message(channel_id, "不能挑战自己！", quote_msg_id=msg_id)
            return
        
        # 加载双方队伍
        attacker_data = load_user_data(user_id)
        defender_data = load_user_data(target_user_id)
        
        attacker_team = attacker_data.get("team", {"battle_cards": [], "assist_cards": []})
        defender_team = get_defense_team(target_user_id)
        
        # 检查攻击者队伍
        attacker_battle_cards = attacker_team.get("battle_cards", [])
        attacker_has_cards = any(card for card in attacker_battle_cards)
        
        if not attacker_has_cards:
            send_kook_message(channel_id, "你的队伍还没有配置战斗卡！请先使用「队伍 我的卡」命令配置队伍。", quote_msg_id=msg_id)
            return
        
        # 检查防守者队伍
        defender_battle_cards = defender_team.get("battle_cards", [])
        defender_has_cards = any(card for card in defender_battle_cards)
        
        if not defender_has_cards:
            send_kook_message(channel_id, f"玩家 {target_user_id} 还没有设置防守队！", quote_msg_id=msg_id)
            return
        
        # 检查战斗系统
        if not BATTLE_SYSTEM_LOADED or BATTLE_INSTANCE is None:
            send_kook_message(channel_id, "战斗系统未加载，请稍后重试", quote_msg_id=msg_id)
            return
        
        # 发送战斗开始消息
        send_kook_message(channel_id, f"⚔️ 挑战排名第{target_rank}位的 {target_user_id}！", quote_msg_id=msg_id)
        
        # 执行战斗
        log_info(f"挑战排名开始: {user_id} vs {target_user_id}")
        result = BATTLE_INSTANCE.start_battle(
            attacker_team,
            defender_team
        )
        
        # 保存战斗日志
        attacker_data["battle_history"].append({
            "time": datetime.now().isoformat(),
            "result": "win" if result.get("player_win") else "lose",
            "opponent": target_user_id,
            "opponent_rank": target_rank,
            "rounds": result.get("rounds", 0),
            "log": result.get("log", []),
            "player_units": result.get("player_units", []),
            "enemy_units": result.get("enemy_units", []),
            "parsable_log": result.get("parsable_log", [])
        })
        attacker_data["battle_history"] = attacker_data["battle_history"][-10:]
        save_user_data(user_id)
        
        # 格式化结果
        rounds = result.get("rounds", 0)
        player_win = result.get("player_win", False)
        
        if player_win:
            result_text = f"🎉 挑战成功！击败了排名第{target_rank}位的 {target_user_id}\n战斗回合: {rounds}"
        else:
            result_text = f"💔 挑战失败！输给了排名第{target_rank}位的 {target_user_id}\n战斗回合: {rounds}"
        
        # 如果开启了GIF渲染器，生成战斗GIF
        if GIF_RENDERER_LOADED:
            gif_buffer = battle_to_gif_bytes(result)
            if gif_buffer:
                send_kook_message(channel_id, result_text)
                send_kook_gif_bytes(channel_id, "", gif_buffer)
            else:
                send_kook_message(channel_id, result_text)
        else:
            send_kook_message(channel_id, result_text)
        
        log_info(f"挑战排名结束: {user_id} vs {target_user_id}, result={player_win}")
        
    except Exception as e:
        log_error(f"挑战排名失败: {e}")
        send_kook_message(channel_id, f"挑战排名失败: {str(e)}", quote_msg_id=msg_id)

def handle_boss_battle(user_id, channel_id, raw_message, msg_id=None):
    """
    处理BOSS战请求
    命令格式: BOSS战 / boss战
    BOSS由CONFIG中的BOSS_CARD_ID指定，初始SP=300
    """
    try:
        if not BATTLE_SYSTEM_LOADED or BATTLE_INSTANCE is None:
            send_kook_message(channel_id, "战斗系统未加载，请稍后重试", quote_msg_id=msg_id)
            return

        # BOSS战冷却检查
        now_ts = datetime.now().timestamp()
        last_boss = BOSS_BATTLE_COOLDOWN.get(user_id, 0)
        remaining = BOSS_BATTLE_COOLDOWN_SECONDS - (now_ts - last_boss)
        if remaining > 0:
            send_kook_message(channel_id, f"BOSS战冷却中！请等待 {int(remaining)} 秒后再试~", quote_msg_id=msg_id)
            log_info(f"BOSS战冷却 [{user_id}]: 还需等待 {int(remaining)} 秒")
            return

        # 加载玩家队伍
        user_data = load_user_data(user_id)
        player_team = user_data.get("team", {"battle_cards": [], "assist_cards": []})
        player_battle_cards = player_team.get("battle_cards", [])
        player_has_cards = any(card for card in player_battle_cards)

        if not player_has_cards:
            send_kook_message(channel_id, "你的队伍还没有配置战斗卡！请先使用「队伍 我的卡」命令配置队伍。", quote_msg_id=msg_id)
            return

        # 查找BOSS角色
        boss_char = BATTLE_INSTANCE.get_character(str(BOSS_CARD_ID))
        if not boss_char:
            boss_char = BATTLE_INSTANCE._get_fallback_character(str(BOSS_CARD_ID))
        boss_name = boss_char.name if boss_char else "未知BOSS"

        # 发送战斗开始消息
        send_kook_message(channel_id, f"⚔️ BOSS战 VS 【{boss_name}】 (HP: 15,000,000)", quote_msg_id=msg_id)

        # 执行BOSS战
        log_info(f"BOSS战开始: {user_id} vs BOSS({BOSS_CARD_ID} {boss_name})")
        BOSS_BATTLE_COOLDOWN[user_id] = datetime.now().timestamp()
        result = BATTLE_INSTANCE.start_boss_battle(
            player_team, str(BOSS_CARD_ID), initial_sp=300
        )

        # 保存战斗日志
        user_data["battle_history"].append({
            "time": datetime.now().isoformat(),
            "result": "player" if result.get("boss_killed") else "boss",
            "rounds": result.get("rounds", 0),
            "log": result.get("log", []),
            "player_units": result.get("player_units", []),
            "enemy_units": result.get("enemy_units", []),
            "parsable_log": result.get("parsable_log", []),
            "boss_damage": result.get("damage_dealt", 0),
            "boss_damage_percent": result.get("damage_percent", 0),
            "boss_killed": result.get("boss_killed", False)
        })
        user_data["battle_history"] = user_data["battle_history"][-10:]
        save_user_data(user_id)

        # 格式化并发送结果
        damage_dealt = result.get("damage_dealt", 0)
        damage_percent = result.get("damage_percent", 0)
        rounds = result.get("rounds", 0)
        boss_killed = result.get("boss_killed", False)

        if boss_killed:
            result_text = f"🎉 BOSS击杀成功！\n对 {boss_name} 造成了 {damage_dealt:,} 点伤害！\n战斗回合: {rounds}"
        else:
            result_text = f"⚔️ BOSS战结束\n对 {boss_name} 造成了 {damage_dealt:,} 点伤害 ({damage_percent:.1f}%)\n战斗回合: {rounds}"

        # 如果开启了GIF渲染器，生成战斗GIF
        if GIF_RENDERER_LOADED:
            gif_buffer = battle_to_gif_bytes(result)
            if gif_buffer:
                send_kook_message(channel_id, result_text)
                send_kook_gif_bytes(channel_id, "", gif_buffer)
            else:
                send_kook_message(channel_id, result_text)
        else:
            send_kook_message(channel_id, result_text)

        log_info(f"BOSS战结束: {user_id}, damage={damage_dealt}, pct={damage_percent}%")

    except Exception as e:
        log_error(f"BOSS战失败: {e}")
        send_kook_message(channel_id, f"BOSS战失败: {str(e)}", quote_msg_id=msg_id)

def handle_battle(user_id, channel_id, raw_message, msg_id=None):
    """
    处理对战请求（与QQ版一致）
    命令格式:
    - /战斗[1-5] - AI对战，可选难度1~5（默认2）
    - /战斗 对战说明 - 显示战斗帮助
    """
    try:
        if not BATTLE_SYSTEM_LOADED or BATTLE_INSTANCE is None:
            send_kook_message(channel_id, "战斗系统未加载，请稍后重试", quote_msg_id=msg_id)
            return
        
        # 战斗帮助说明
        if '说明' in raw_message:
            help_text = """
⚔️ 战斗系统说明 ⚔️

命令格式：
  /战斗          - AI对战（难度2）
  /战斗1~5       - AI对战（指定难度）
  /对战说明      - 显示此帮助

难度等级：
  1 - 简单
  2 - 普通（默认）
  3 - 困难
  4 - 极难
  5 - 地狱

战斗规则：
  • 双方各6张战斗卡对战
  • 先击败对方全部战斗卡获胜
  • 支援卡提供被动效果
  • 技能消耗SP，回合恢复SP
            """.strip()
            send_kook_message(channel_id, help_text, quote_msg_id=msg_id)
            return
        
        # 加载玩家队伍
        player_team = load_team_data(user_id)
        player_battle_cards = player_team.get("battle_cards", [])
        player_has_cards = any(card for card in player_battle_cards)
        
        if not player_has_cards:
            send_kook_message(channel_id, "你的队伍还没有配置战斗卡！请先使用「/队伍」命令配置队伍。", quote_msg_id=msg_id)
            return
        
        # 读取活跃预设信息
        active_slot = 0
        try:
            pdata = load_presets(user_id)
            active_slot = pdata.get("active_slot", 0)
        except Exception:
            pass
        
        # 解析难度: 战斗5, 战斗 3, 对战2 等
        import re
        diff_match = re.search(r'(?:战斗|对战|决斗)\s*([1-5])', raw_message)
        ai_difficulty = int(diff_match.group(1)) if diff_match else 2
        
        # AI对手（带难度）
        difficulty_names = {1: "简单", 2: "普通", 3: "困难", 4: "极难", 5: "地狱"}
        enemy_team = generate_ai_team(ai_difficulty)
        enemy_name = f"AI({difficulty_names[ai_difficulty]})"
        
        # 发送双方VS配队图
        characters = get_characters()
        vs_img_path = build_vs_team_image(player_team, enemy_team, characters)
        if vs_img_path and os.path.exists(vs_img_path):
            from PIL import Image
            vs_img = Image.open(vs_img_path)
            send_kook_message_with_image(channel_id, f"⚔️ VS {enemy_name}", vs_img, quote_msg_id=msg_id)
            os.remove(vs_img_path)
        else:
            send_kook_message(channel_id, f"⚔️ VS {enemy_name}", quote_msg_id=msg_id)
        
        # 执行战斗
        log_info(f"战斗开始: {user_id} vs AI难度{ai_difficulty}")
        result = BATTLE_INSTANCE.start_battle(player_team, enemy_team, challenger="player")
        
        # 保存战斗日志（与QQ版一致，滚动保留最近3次）
        battle_log_key = save_rolling_battle_log(user_id, result)
        
        # 生成简短结果
        winner = result.get("winner", "unknown")
        rounds = result.get("rounds", 0)
        
        if winner == "player":
            result_text = f"🏆 胜利！经过 {rounds} 回合的激战，你击败了 {enemy_name}！"
        else:
            result_text = f"💀 失败... 经过 {rounds} 回合的激战，你被 {enemy_name} 击败了..."
        
        # 统计双方存活情况
        player_alive = sum(1 for u in result.get("player_units", []) if u.get("alive") and not u.get("is_assist"))
        enemy_alive = sum(1 for u in result.get("enemy_units", []) if u.get("alive") and not u.get("is_assist"))
        
        result_text += f"\n📊 最终状态: 我方存活 {player_alive}/6, 敌方存活 {enemy_alive}/6"
        if active_slot > 0:
            result_text += f"\n📋 使用预设: 槽{active_slot}"
        
        # 发送结果
        send_kook_message(channel_id, result_text, quote_msg_id=msg_id)
        
        # 发送战斗GIF（如果可用）
        if GIF_RENDERER_LOADED:
            gif_buffer = battle_to_gif_bytes(result)
            if gif_buffer:
                send_kook_gif_bytes(channel_id, "", gif_buffer)
        
        log_info(f"战斗结束: {user_id} vs AI难度{ai_difficulty}, winner={winner}, rounds={rounds}")
        
    except Exception as e:
        log_error(f"战斗失败: {e}")
        send_kook_message(channel_id, f"战斗失败: {str(e)}", quote_msg_id=msg_id)

def handle_collection(user_id, channel_id, msg_id=None):
    user_data = load_user_data(user_id)
    cards = user_data["cards"]
    
    if not cards:
        send_kook_message(channel_id, "你还没有抽过卡哦！", quote_msg_id=msg_id)
        return
    
    three_stars = [c for c in cards if c["rarity"] == 3]
    two_stars = [c for c in cards if c["rarity"] == 2]
    one_stars = [c for c in cards if c["rarity"] == 1]
    
    card = create_kook_card(
        "我的卡牌收藏",
        f"总计: {len(cards)} 张\n⭐⭐⭐: {len(three_stars)} | ⭐⭐: {len(two_stars)} | ⭐: {len(one_stars)}",
        []
    )
    
    if three_stars:
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": "⭐⭐⭐: " + ", ".join([c["name"] for c in three_stars][:10])}
        })
    if two_stars:
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": "⭐⭐: " + ", ".join([c["name"] for c in two_stars][:10])}
        })
    if one_stars:
        card["modules"].append({
            "type": "section",
            "text": {"type": "plain-text", "content": "⭐: " + ", ".join([c["name"] for c in one_stars][:10])}
        })
    
    send_kook_card_message(channel_id, card, quote_msg_id=msg_id)

def handle_pity(user_id, channel_id, msg_id=None):
    """显示保底进度（与QQ版一致）"""
    pity_data = load_pity_data(user_id)
    
    fes_pity_count = pity_data.get("fes_pity_count", 0)
    fes_pity_remaining = max(0, CONFIG["PITY_LIMIT"] - fes_pity_count)
    
    card = create_kook_card(
        "保底进度",
        "",
        [
            {"name": "フェス保底", "value": f"{fes_pity_remaining}/{CONFIG['PITY_LIMIT']}"},
            {"name": "累计抽卡", "value": str(pity_data.get("total_draws", 0))},
            {"name": "累计三星", "value": str(pity_data.get("total_3stars", 0))}
        ]
    )
    send_kook_card_message(channel_id, card, quote_msg_id=msg_id)

def handle_balance(user_id, channel_id, msg_id=None):
    user_data = load_user_data(user_id)
    
    card = create_kook_card(
        "呱太余额",
        "",
        [
            {"name": "当前呱太", "value": str(user_data["gacha"])},
            {"name": "单抽", "value": f"{CONFIG['GACHA_COST']} 呱太"},
            {"name": "十连", "value": f"{CONFIG['GACHA10_COST']} 呱太"}
        ]
    )
    send_kook_card_message(channel_id, card, quote_msg_id=msg_id)

def handle_profile(user_id, channel_id, nickname, msg_id=None):
    handle_personal_info(user_id, channel_id, msg_id)

def handle_reload(user_id, channel_id, msg_id=None):
    load_characters()
    load_battle_characters()
    load_ranking_data()
    send_kook_message(channel_id, "数据已重新加载！", quote_msg_id=msg_id)

def handle_backup(user_id, channel_id, msg_id=None):
    today = datetime.now().strftime("%Y-%m-%d")
    backup_dir = BACKUP_DIR / today
    backup_dir.mkdir(exist_ok=True)
    
    user_files = list(DATA_DIR.glob("*.json"))
    for file in user_files:
        dest_file = backup_dir / file.name
        with open(file, "rb") as src:
            with open(dest_file, "wb") as dst:
                dst.write(src.read())
    
    send_kook_message(channel_id, f"备份完成！共备份 {len(user_files)} 个文件", quote_msg_id=msg_id)

def handle_backup_battle_logs(user_id, channel_id, msg_id=None):
    import shutil
    backup_dir = BACKUP_DIR / "battle_logs" / datetime.now().strftime("%Y%m%d")
    backup_dir.mkdir(parents=True, exist_ok=True)
    
    info_dir = BASE_DIR.parent / "info"
    battle_files = list(info_dir.glob("battle_*.json"))
    
    for battle_file in battle_files:
        dest = backup_dir / battle_file.name
        shutil.copy(str(battle_file), str(dest))
    
    send_kook_message(channel_id, f"✅ 已备份 {len(battle_files)} 个战斗日志文件")

@app.route('/webhook', methods=['GET', 'POST'])
def webhook():
    # 处理GET请求 - URL验证
    if request.method == 'GET':
        try:
            challenge = request.args.get('challenge', '')
            verify_token = request.args.get('verify_token', '')
            
            # 检查verify_token是否匹配
            if CONFIG.get("VERIFY_TOKEN"):
                if verify_token != CONFIG["VERIFY_TOKEN"]:
                    log_error("GET验证token不匹配")
                    return jsonify({"status": "error", "message": "验证token不匹配"}), 401
            
            log_info(f"收到GET Challenge验证请求")
            return jsonify({"challenge": challenge}), 200
        except Exception as e:
            log_error(f"处理GET请求失败: {e}")
            return jsonify({"status": "error", "message": str(e)}), 500
    
    # 处理POST请求 - 事件推送
    try:
        body = request.get_data()
        
        try:
            try:
                decompressed_body = decompress_data(body)
                data = json.loads(decompressed_body)
            except:
                data = json.loads(body.decode('utf-8'))
        except Exception as e:
            log_error(f"解析数据失败: {e}")
            return jsonify({"status": "error", "message": "解析数据失败"}), 400
        
        if "encrypt" in data:
            data_str = decrypt_message(data["encrypt"])
            data = json.loads(data_str)
        
        if "d" not in data or "s" not in data:
            return jsonify({"status": "error", "message": "缺少数据字段"}), 400
        
        sn = data.get("s", "")
        event_data = data["d"]
        
        # 检查事件唯一性，避免重复处理
        if sn:
            if sn in PROCESSED_EVENTS:
                log_info(f"重复事件，已跳过: {sn}")
                return jsonify({"status": "success"})
            
            # 添加到已处理集合
            PROCESSED_EVENTS.add(sn)
            
            # 清理旧事件，防止内存溢出
            if len(PROCESSED_EVENTS) > MAX_EVENTS:
                PROCESSED_EVENTS.pop()
        
        channel_type = event_data.get("channel_type", "")
        
        if channel_type == "WEBHOOK_CHALLENGE":
            challenge = event_data.get("challenge", "")
            verify_token = event_data.get("verify_token", "")
            
            # 检查verify_token是否匹配
            if CONFIG.get("VERIFY_TOKEN"):
                if verify_token != CONFIG["VERIFY_TOKEN"]:
                    log_error("POST验证token不匹配")
                    return jsonify({"status": "error", "message": "验证token不匹配"}), 401
            
            log_info(f"收到POST Challenge验证请求")
            return jsonify({"challenge": challenge}), 200
        
        event_type = event_data.get("type", 1)
        target_id = event_data.get("target_id", "")
        author_id = event_data.get("author_id", "")
        content = event_data.get("content", "").strip()
        msg_id = event_data.get("msg_id", "")
        
        # 调试日志：打印接收到的事件数据
        log_info(f"事件类型: {event_type}, 目标ID: {target_id}, 作者ID: {author_id}, 消息ID: {msg_id}")
        log_info(f"消息内容: '{content}'")
        
        # 忽略机器人自己发送的消息
        if author_id == "1":
            log_info("忽略机器人自己发送的消息")
            return jsonify({"status": "success"}), 200
        
        # 处理文本类消息（文字消息、KMarkdown消息）
        # type=1: 文字消息, type=9: KMarkdown消息
        if event_type in [1, 9] and content:
            author_info = event_data.get("extra", {}).get("author", {})
            nickname = author_info.get("nickname", "未知用户")
            
            log_info(f"收到消息 [用户:{author_id}] [{nickname}]: {content}")
            
            # 清理消息内容中的特殊格式
            import re
            
            # 移除@用户格式 (met)xxx(met)
            content = re.sub(r'\(met\)[^\(]+\(met\)', '', content).strip()
            
            # 移除@角色格式 (rol)xxx(rol)
            content = re.sub(r'\(rol\)[^\(]+\(rol\)', '', content).strip()
            
            # 移除频道链接格式 (chn)xxx(chn)
            content = re.sub(r'\(chn\)[^\(]+\(chn\)', '', content).strip()
            
            # 移除表情格式 (emj)xxx(emj)
            content = re.sub(r'\(emj\)[^\(]+\(emj\)', '', content).strip()
            
            # 移除markdown格式（保留纯文本）
            content = re.sub(r'\*\*([^*]+)\*\*', r'\1', content)  # 粗体
            content = re.sub(r'\*([^*]+)\*', r'\1', content)      # 斜体
            content = re.sub(r'`([^`]+)`', r'\1', content)        # 行内代码
            content = re.sub(r'~~([^~]+)~~', r'\1', content)      # 删除线
            
            # 移除首尾空格
            content = content.strip()
            
            # 使用线程处理命令，确保快速响应
            import threading
            threading.Thread(target=handle_command, args=(content, author_id, target_id, nickname, msg_id), daemon=True).start()
        elif event_type == 255:
            # 系统消息，记录但不处理
            log_info(f"系统消息 - extra: {event_data.get('extra', {})}")
        else:
            log_info(f"未处理的消息类型 - 类型: {event_type}, 内容长度: {len(content) if content else 0}")
        
        # 1秒内返回200响应
        return jsonify({"status": "success"}), 200
    
    except Exception as e:
        log_error(f"处理Webhook失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

def main():
    load_config()
    
    if not CONFIG["BOT_TOKEN"] or CONFIG["BOT_TOKEN"] == "YOUR_KOOK_BOT_TOKEN":
        print("请先在 config.json 中设置 KOOK Bot Token！")
        return
    
    backup_pity_records()
    
    load_characters()
    load_battle_characters()
    load_ranking_data()
    
    log_info(f"KOOK Bot 启动中...")
    log_info(f"监听地址: http://{CONFIG['HOST']}:{CONFIG['PORT']}/webhook")
    log_info(f"战斗系统: {'已加载' if BATTLE_SYSTEM_LOADED else '未加载'}")
    log_info(f"GIF渲染器: {'已加载' if GIF_RENDERER_LOADED else '未加载'}")
    
    app.run(host=CONFIG["HOST"], port=CONFIG["PORT"], debug=False)

if __name__ == "__main__":
    main()