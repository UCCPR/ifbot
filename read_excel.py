"""读取Excel文件结构"""
import openpyxl
from pathlib import Path

XLSX_FILE = Path(__file__).parent / "卡牌信息.xlsx"

wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)

print("=" * 60)
print("Excel 文件中的所有 Sheet 名称:")
print("=" * 60)
for name in wb.sheetnames:
    print(f"  - {name}")

print()

# 检查 BattleCard资源 sheet
if 'BattleCard资源' in wb.sheetnames:
    print("=" * 60)
    print("BattleCard资源 Sheet 结构:")
    print("=" * 60)
    sheet = wb['BattleCard资源']
    
    # 打印表头（前2行）
    print("\n表头（前2行）:")
    for row_idx in range(1, 3):
        row_data = []
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data.append(f"[{col_idx}]{cell.value}")
        print(f"  行{row_idx}: {' | '.join(row_data)}")
    
    # 打印前3行数据
    print("\n前3行数据示例:")
    for row_idx in range(2, 5):
        row_data = []
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data.append(f"[{col_idx}]{cell.value}")
        print(f"  行{row_idx}: {' | '.join(row_data)}")

print()

# 检查 AssistCard资源 sheet
if 'AssistCard资源' in wb.sheetnames:
    print("=" * 60)
    print("AssistCard资源 Sheet 结构:")
    print("=" * 60)
    sheet = wb['AssistCard资源']
    
    # 打印表头（前2行）
    print("\n表头（前2行）:")
    for row_idx in range(1, 3):
        row_data = []
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data.append(f"[{col_idx}]{cell.value}")
        print(f"  行{row_idx}: {' | '.join(row_data)}")
    
    # 打印前3行数据
    print("\n前3行数据示例:")
    for row_idx in range(2, 5):
        row_data = []
        for col_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data.append(f"[{col_idx}]{cell.value}")
        print(f"  行{row_idx}: {' | '.join(row_data)}")

wb.close()
