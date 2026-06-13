"""
自动抽卡QQ Bot - 基于Napcat
支持一抽和十连，生成角色卡牌图片
"""

import os
import random
import json
import base64
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

import requests
from flask import Flask, request, jsonify
from PIL import Image
import openpyxl

# 导入配队系统（延迟到log_info定义后再记录日志）
try:
    from team_system import (
        load_team_data,
        save_team_data,
        get_user_3star_cards,
        build_team_image,
        build_vs_team_image,
        build_3star_cards_image,
        set_team_card,
        clear_team_card,
        clear_all_team,
        get_team_info,
        auto_build_team,
        auto_save_preset,
        load_preset,
        load_presets,
        save_presets,
        list_presets_info
    )
    TEAM_SYSTEM_LOADED = True
except ImportError as e:
    TEAM_SYSTEM_LOADED = False
    # 如果加载失败，定义空函数作为备用
    def load_team_data(user_id): return {"battle_cards": [], "assist_cards": []}
    def save_team_data(user_id, data): pass
    def get_user_3star_cards(user_id, characters=None): return []
    def build_team_image(team_data, characters): return None
    def build_vs_team_image(player_team, enemy_team, characters): return None
    def build_3star_cards_image(user_id, characters, page=1, page_size=50): return None, [], 0
    def auto_build_team(user_id, characters): return {"success": False, "message": "配队系统未加载", "team": None}
    def auto_save_preset(user_id): return -1
    def load_preset(user_id, slot): return False
    def load_presets(user_id): return {"presets": [], "active_slot": 0}
    def save_presets(user_id, data): pass
    def list_presets_info(user_id, characters): return "配队系统未加载"
    def set_team_card(user_id, position, card_id, card_type="battle"): return False
    def clear_team_card(user_id, position, card_type="battle"): return False
    def clear_all_team(user_id): pass
    def get_team_info(user_id, characters): return "配队系统未加载"

# 导入战斗系统
try:
    from battle_system import BattleSystem, format_battle_result, format_boss_result, get_battle_help
    BATTLE_SYSTEM_LOADED = True
    BATTLE_INSTANCE = None  # 战斗系统实例
except ImportError as e:
    BATTLE_SYSTEM_LOADED = False
    class BattleSystem:
        def __init__(self, data): pass
        def start_battle(self, p_team, e_team, challenger="player", initial_player_sp=0): return {"winner": "player", "rounds": 1, "log": [], "player_units": [], "enemy_units": []}
        def start_boss_battle(self, player_team, boss_card_id, initial_sp=90): return {"boss_name": "???", "boss_starting_hp": 15000000, "boss_ending_hp": 15000000, "damage_dealt": 0, "damage_percent": 0, "rounds": 0, "player_survived": 0, "player_total": 0, "boss_killed": False, "log": [], "player_units": [], "enemy_units": []}
        def get_character(self, card_id): return None
        def _get_fallback_character(self, card_id): return None
    def format_battle_result(result): return "战斗系统未加载"
    def format_boss_result(result): return "战斗系统未加载"
    def get_battle_help(): return "战斗系统未加载"


# ========== 配置 ==========
BASE_DIR = Path(__file__).parent
ICON_DIR = BASE_DIR / "iconimage"
LEVEL_DIR = BASE_DIR / "level"
XLSX_FILE = BASE_DIR / "卡牌信息.xlsx"         # 抽卡卡池（含1/2/3星）
BATTLE_XLSX = BASE_DIR / "cards_completed.xlsx"  # 战斗数值（3星详细数据）
INFO_DIR = BASE_DIR / "info"
OUTPUT_DIR = BASE_DIR / "output"

# 确保目录存在
INFO_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# ========== 日志模块（必须在其他模块之前定义）==========
def log_info(message: str):
    """记录普通信息"""
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "gacha_info.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")
    print(f"[INFO] {message}")


def log_error(message: str):
    """记录错误信息"""
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "gacha_error.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] ERR {message}\n")
    print(f"[ERROR] {message}")


# 从配置文件导入配置信息
# 请在 config.py 文件中填写实际的配置值
try:
    from config import (
        NAPCAT_HOST,
        NAPCAT_PORT,
        NAPCAT_TOKEN,
        FLASK_HOST,
        FLASK_PORT,
        # 保底配置
        PITY_LIMIT,
        # 呱太配置
        GACHA_COST,
        GACHA10_COST,
        GACHA10_COOLDOWN_SECONDS,
        GET_GACHA_REWARD,
        DAILY_REWARD,
        # 开箱配置
        MYSTERY_BOX_CHANCE,
        MUTATION_NO_CHANGE,
        MUTATION_1_TO_2,
        MUTATION_1_TO_3,
        MUTATION_2_TO_3,
        BOX_OPEN_TIMEOUT,
        # 三星池子配置
        THREE_STAR_POOL_RED_COST,
        THREE_STAR_POOL_BLUE_COST,
        # 抽卡概率（三星内部分配）
        FES_LIMIT_PROB,
        PERIOD_LIMIT_PROB,
        OTHER_3STAR_PROB,
        # 盲盒星级概率
        MYSTERY_BOX_2STAR_PROB,
        MYSTERY_BOX_3STAR_PROB,
        NORMAL_BOX_1STAR_PROB,
        NORMAL_BOX_2STAR_PROB,
        NORMAL_BOX_3STAR_PROB,
        # 抽卡星级概率（呱太抽卡）
        GACHA_1STAR_PROB,
        GACHA_2STAR_PROB,
        GACHA_3STAR_PROB,
        # 管理员
        ADMIN_QQ,
        # BOSS战
        BOSS_BATTLE_COOLDOWN_SECONDS
    )
except ImportError:
    # 如果配置文件不存在，使用默认值
    NAPCAT_HOST = "127.0.0.1"
    NAPCAT_PORT = 3000
    NAPCAT_TOKEN = ""
    FLASK_HOST = "0.0.0.0"
    FLASK_PORT = 5000
    # 保底配置
    PITY_LIMIT = 150  # 150抽必出フェス限定三星
    # 呱太配置
    GACHA_COST = 300        # 单抽价格
    GACHA10_COST = 3000     # 十连价格
    GACHA10_COOLDOWN_SECONDS = 60  # 十连冷却时间（秒）
    GET_GACHA_REWARD = 10000  # 获取呱太奖励
    DAILY_REWARD = 30000     # 每日签到奖励
    # 开箱配置
    MYSTERY_BOX_CHANCE = 0.02  # 黑色盲盒概率
    MUTATION_NO_CHANGE = 0.88  # 不突变概率
    MUTATION_1_TO_2 = 0.08    # 1星→2星概率
    MUTATION_1_TO_3 = 0.02    # 1星→3星概率
    MUTATION_2_TO_3 = 0.05    # 2星→3星概率
    BOX_OPEN_TIMEOUT = 300     # 盲盒开启超时时间（秒）
    # 三星池子配置
    THREE_STAR_POOL_RED_COST = 1500   # 红色碎片消耗
    THREE_STAR_POOL_BLUE_COST = 350   # 蓝色碎片消耗
    # 抽卡概率（三星内部分配）
    FES_LIMIT_PROB = 0.25     # フェス限定概率
    PERIOD_LIMIT_PROB = 0.35  # 期間限定概率
    OTHER_3STAR_PROB = 0.40   # 其他三星概率
    # 不跳过的十连和单抽星级概率
    MYSTERY_BOX_2STAR_PROB = 65  # 黑色盲盒2星概率（权重）
    MYSTERY_BOX_3STAR_PROB = 35  # 黑色盲盒3星概率（权重）
    NORMAL_BOX_1STAR_PROB = 72   # 正常盲盒1星概率（权重）
    NORMAL_BOX_2STAR_PROB = 23   # 正常盲盒2星概率（权重）
    NORMAL_BOX_3STAR_PROB = 3    # 正常盲盒3星概率（权重）
    # 十连跳过概率
    GACHA_1STAR_PROB = 72   # 1星概率（权重）
    GACHA_2STAR_PROB = 23   # 2星概率（权重）
    GACHA_3STAR_PROB = 3    # 3星概率（权重）
    # BOSS战
    BOSS_BATTLE_COOLDOWN_SECONDS = 60
    ADMIN_QQ = ""  # 管理员QQ

# 定义概率权重常量（避免代码重复）
GACHA_WEIGHTS = [GACHA_1STAR_PROB, GACHA_2STAR_PROB, GACHA_3STAR_PROB]
NORMAL_BOX_WEIGHTS = [NORMAL_BOX_1STAR_PROB, NORMAL_BOX_2STAR_PROB, NORMAL_BOX_3STAR_PROB]
MYSTERY_BOX_WEIGHTS = [MYSTERY_BOX_2STAR_PROB, MYSTERY_BOX_3STAR_PROB]

# 定义角色图裁剪比例常量
CROP_LEFT_RATIO = 0.25
CROP_RIGHT_RATIO = 0.75
CROP_TOP_RATIO = 0.15
CROP_BOTTOM_RATIO = 0.65

app = Flask(__name__)


# ========== 全局变量 ==========
CHARACTERS_CACHE = None  # 预加载的角色数据缓存
BOX_SESSIONS = {}  # 盲盒会话状态 {user_id: {"boxes": [], "opened": [], "created_at": timestamp}}
MESSAGE_COUNTER = {}  # 消息发送计数器 {target_id: {"count": int, "start_time": timestamp}}
MAX_MESSAGES_PER_MINUTE = 30  # 每分钟最大消息发送量
DAILY_SEND_STATS = {}  # 每日发送消息统计 {日期: 发送次数}
DAILY_USER_SET = set()  # 今日活跃用户ID集合
_RESTORED_DAU = 0       # 重启后从磁盘恢复的日活基准数
GACHA10_COOLDOWN = {}  # 十连冷却时间 {user_id: last_gacha10_timestamp}
BOSS_BATTLE_COOLDOWN = {}  # BOSS战冷却时间 {user_id: last_boss_battle_timestamp}


# ========== 预加载模块 ==========
def preload_characters():
    """预加载角色数据到缓存，避免每次抽卡都重新读取"""
    global CHARACTERS_CACHE
    if CHARACTERS_CACHE is None:
        log_info("开始预加载角色数据...")
        CHARACTERS_CACHE = load_character_data()
        log_info(f"角色数据预加载完成，共 {len(CHARACTERS_CACHE)} 个角色")
    return CHARACTERS_CACHE


def get_characters():
    """获取预加载的角色数据（如果未预加载则自动加载）"""
    return preload_characters()


# 记录配队系统加载状态
if TEAM_SYSTEM_LOADED:
    log_info("配队系统加载成功")
else:
    log_error("配队系统加载失败")


# ========== 抽卡记录存储模块 ==========
def get_pity_file(user_id: str) -> Path:
    """获取用户抽卡记录文件路径"""
    return INFO_DIR / f"pity_{user_id}.json"


def load_pity_data(user_id: str) -> dict:
    """加载用户的抽卡记录数据"""
    pity_file = get_pity_file(user_id)
    if pity_file.exists():
        try:
            with open(pity_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            log_error(f"用户 {user_id} 的抽卡记录文件格式错误: {e}")
            return get_default_pity_data()
        except IOError as e:
            log_error(f"读取用户 {user_id} 的抽卡记录文件失败: {e}")
            return get_default_pity_data()
    return get_default_pity_data()


def get_default_pity_data() -> dict:
    """获取默认的抽卡记录数据"""
    return {
        "total_draws": 0, 
        "pity_count": 0, 
        "fes_pity_count": 0,
        "total_3stars": 0,
        "red_crystal": 0,
        "blue_crystal": 0,
        "recent_3stars": [],  # 最近获得的三星卡记录，按时间排序，存储card_id
        "card_collection": {},  # 卡片收藏 {card_id: {"name": "", "stars": 3, "limit_type": "", "count": 1, "last_time": timestamp}}
        "fes_count": 0,        # フェス限定数量
        "period_count": 0,      # 期間限定数量
        "other_3star_count": 0, # 其他三星数量
        "total_2stars": 0       # 二星数量
    }


def save_pity_data(user_id: str, data: dict):
    """保存用户的抽卡记录数据"""
    pity_file = get_pity_file(user_id)
    with open(pity_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def get_remaining_pity(user_id: str) -> int:
    """获取用户距离フェス限定三星保底还剩多少抽"""
    pity_data = load_pity_data(user_id)
    fes_pity_count = pity_data.get("fes_pity_count", 0)
    return max(0, PITY_LIMIT - fes_pity_count)


def update_pity(user_id: str, got_3star: bool = False, is_fes_3star: bool = False):
    """更新用户的抽卡记录"""
    pity_data = load_pity_data(user_id)
    
    # 如果抽到3星，重置保底计数并增加三星个数
    if got_3star:
        pity_data["pity_count"] = 0
        pity_data["total_3stars"] = pity_data.get("total_3stars", 0) + 1
        # 如果抽到フェス限定三星，重置フェス保底计数
        if is_fes_3star:
            pity_data["fes_pity_count"] = 0
    else:
        pity_data["pity_count"] = pity_data.get("pity_count", 0) + 1
        pity_data["fes_pity_count"] = pity_data.get("fes_pity_count", 0) + 1
    
    pity_data["total_draws"] = pity_data.get("total_draws", 0) + 1
    save_pity_data(user_id, pity_data)
    return pity_data


def get_total_draws(user_id: str) -> int:
    """获取用户累计抽卡次数"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("total_draws", 0)


def get_total_3stars(user_id: str) -> int:
    """获取用户累计获得的三星个数"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("total_3stars", 0)


def get_red_crystal(user_id: str) -> int:
    """获取用户的红色碎片数量"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("red_crystal", 0)


def get_blue_crystal(user_id: str) -> int:
    """获取用户的蓝色碎片数量"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("blue_crystal", 0)


def add_red_crystal(user_id: str, amount: int):
    """增加用户的红色碎片"""
    pity_data = load_pity_data(user_id)
    pity_data["red_crystal"] = pity_data.get("red_crystal", 0) + amount
    save_pity_data(user_id, pity_data)
    return pity_data["red_crystal"]


def add_blue_crystal(user_id: str, amount: int):
    """增加用户的蓝色碎片"""
    pity_data = load_pity_data(user_id)
    pity_data["blue_crystal"] = pity_data.get("blue_crystal", 0) + amount
    save_pity_data(user_id, pity_data)
    return pity_data["blue_crystal"]


def spend_red_crystal(user_id: str, amount: int) -> bool:
    """消耗用户的红色碎片，返回是否成功"""
    pity_data = load_pity_data(user_id)
    current = pity_data.get("red_crystal", 0)
    if current >= amount:
        pity_data["red_crystal"] = current - amount
        save_pity_data(user_id, pity_data)
        return True
    return False


def spend_blue_crystal(user_id: str, amount: int) -> bool:
    """消耗用户的蓝色碎片，返回是否成功"""
    pity_data = load_pity_data(user_id)
    current = pity_data.get("blue_crystal", 0)
    if current >= amount:
        pity_data["blue_crystal"] = current - amount
        save_pity_data(user_id, pity_data)
        return True
    return False


# ========== 三星池子系统 ==========
# 三星池子配置
THREE_STAR_POOL_RED_COST = 1500  # 红色碎片消耗
THREE_STAR_POOL_BLUE_COST = 350  # 蓝色碎片消耗


def select_3star_from_pool(characters: list) -> dict:
    """从三星池子中随机抽取一个角色（只返回三星角色）"""
    three_star_chars = [c for c in characters if c.get("stars") == 3]
    
    if not three_star_chars:
        return None
    
    # 随机选择
    selected = random.choice(three_star_chars)
    return selected


def draw_3star_pool(user_id: str, crystal_type: str = "red") -> dict:
    """
    抽取三星池子
    :param user_id: 用户ID
    :param crystal_type: 碎片类型 "red" 或 "blue"
    :return: 抽卡结果
    """
    characters = get_characters()
    
    # 检查碎片是否足够
    if crystal_type == "red":
        cost = THREE_STAR_POOL_RED_COST
        if not spend_red_crystal(user_id, cost):
            return {
                "success": False,
                "message": f"红色碎片不足！需要{cost}个，当前拥有{get_red_crystal(user_id)}个"
            }
    else:
        cost = THREE_STAR_POOL_BLUE_COST
        if not spend_blue_crystal(user_id, cost):
            return {
                "success": False,
                "message": f"蓝色碎片不足！需要{cost}个，当前拥有{get_blue_crystal(user_id)}个"
            }
    
    # 抽取角色
    selected = select_3star_from_pool(characters)
    
    if not selected:
        return {
            "success": False,
            "message": "三星池子为空，请联系管理员！"
        }
    
    # 处理抽卡结果
    card_id = str(selected.get("card_id", ""))
    chara_id = selected.get("id", "")
    name = selected.get("name", "")
    limit_type = selected.get("limit_type", "")
    stars = selected.get("stars", 3)
    
    # FES统计和提示
    fes_message = ""
    if limit_type == "フェス限定":
        # 增加FES统计
        fes_count = increment_fes_count(card_id, name)
        fes_message = f"✨ 恭喜！这是全服第{fes_count}个「{name}」！"
    
    # 添加到用户卡库
    add_card_collection(user_id, card_id, name, stars, limit_type)
    
    # 添加到最近三星记录
    add_recent_3star(user_id, card_id, chara_id, name, limit_type)
    
    # 更新统计（三星池子抽出的都是三星）
    is_fes_3star = (limit_type == "フェス限定")
    update_pity(user_id, got_3star=True, is_fes_3star=is_fes_3star)
    
    return {
        "success": True,
        "character": selected,
        "message": f"恭喜获得三星角色：{name}",
        "fes_message": fes_message
    }


def add_recent_3star(user_id: str, card_id: str, chara_id: str, name: str, limit_type: str = None):
    """添加最近获得的三星卡记录（最多保存10个）"""
    pity_data = load_pity_data(user_id)
    recent = pity_data.get("recent_3stars", [])
    
    # 创建记录
    record = {
        "card_id": card_id,
        "chara_id": chara_id,
        "name": name,
        "limit_type": limit_type,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    # 添加到开头
    recent.insert(0, record)
    
    # 只保留最近10个
    if len(recent) > 10:
        recent = recent[:10]
    
    pity_data["recent_3stars"] = recent
    save_pity_data(user_id, pity_data)
    return recent


def get_recent_3stars(user_id: str) -> list:
    """获取用户最近获得的三星卡记录"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("recent_3stars", [])


def add_card_collection(user_id: str, card_id: str, name: str, stars: int, limit_type: str = None):
    """添加卡片到收藏并更新数量统计（支持重复计数）"""
    pity_data = load_pity_data(user_id)
    collection = pity_data.get("card_collection", {})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    now_timestamp = datetime.now().timestamp()
    
    # 如果是新卡片，添加到收藏
    if card_id not in collection:
        collection[card_id] = {
            "name": name,
            "stars": stars,
            "limit_type": limit_type,
            "count": 1,
            "first_time": now,
            "last_time": now_timestamp
        }
        
        # 更新数量统计
        if stars == 3:
            if limit_type == "フェス限定":
                pity_data["fes_count"] = pity_data.get("fes_count", 0) + 1
            elif limit_type == "期間限定":
                pity_data["period_count"] = pity_data.get("period_count", 0) + 1
            else:
                pity_data["other_3star_count"] = pity_data.get("other_3star_count", 0) + 1
        elif stars == 2:
            pity_data["total_2stars"] = pity_data.get("total_2stars", 0) + 1
    else:
        # 如果是重复卡片，增加计数并更新时间
        collection[card_id]["count"] = collection[card_id].get("count", 1) + 1
        collection[card_id]["last_time"] = now_timestamp
    
    pity_data["card_collection"] = collection
    save_pity_data(user_id, pity_data)
    return collection


def calculate_power(user_id: str) -> int:
    """计算用户战力：フェス限定数*10 + 期間限定数*8 + 其他三星数*7 + 二星数*3"""
    pity_data = load_pity_data(user_id)
    fes = pity_data.get("fes_count", 0)
    period = pity_data.get("period_count", 0)
    other_3star = pity_data.get("other_3star_count", 0)
    two_star = pity_data.get("total_2stars", 0)
    return fes * 10 + period * 8 + other_3star * 7 + two_star * 3


# ========== FES统计模块 ==========
FES_STATS_FILE = INFO_DIR / "fes_stats.json"

def load_fes_stats() -> dict:
    """加载全服FES角色获取统计"""
    if FES_STATS_FILE.exists():
        try:
            with open(FES_STATS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_fes_stats(stats: dict):
    """保存全服FES角色获取统计"""
    with open(FES_STATS_FILE, "w", encoding="utf-8") as f:
        json.dump(stats, f, indent=2, ensure_ascii=False)

def get_fes_count(card_id: str) -> int:
    """获取某个FES角色的全服获取次数"""
    stats = load_fes_stats()
    return stats.get(card_id, 0)

def increment_fes_count(card_id: str, name: str) -> int:
    """增加FES角色获取计数并返回新的计数"""
    stats = load_fes_stats()
    if card_id not in stats:
        stats[card_id] = {"count": 0, "name": name}
    stats[card_id]["count"] += 1
    save_fes_stats(stats)
    return stats[card_id]["count"]


def get_leaderboard() -> list:
    """获取排行榜（战力前10名）"""
    leaderboard = []
    
    # 遍历所有用户数据文件
    for pity_file in INFO_DIR.glob("pity_*.json"):
        try:
            user_id = pity_file.stem.replace("pity_", "")
            pity_data = load_pity_data(user_id)
            
            # 计算战力
            power = calculate_power(user_id)
            
            # 获取基本信息
            total_draws = pity_data.get("total_draws", 0)
            total_3stars = pity_data.get("total_3stars", 0)
            
            leaderboard.append({
                "user_id": user_id,
                "power": power,
                "total_draws": total_draws,
                "total_3stars": total_3stars,
                "fes_count": pity_data.get("fes_count", 0),
                "period_count": pity_data.get("period_count", 0),
                "other_3star_count": pity_data.get("other_3star_count", 0),
                "total_2stars": pity_data.get("total_2stars", 0)
            })
        except Exception as e:
            log_error(f"读取用户数据失败 {pity_file}: {e}")
    
    # 按战力降序排序
    leaderboard.sort(key=lambda x: x["power"], reverse=True)
    
    # 返回前10名
    return leaderboard[:10]


# ========== 呱太系统 ==========
def get_gacha_file(user_id: str) -> Path:
    """获取用户呱太记录文件路径"""
    return INFO_DIR / f"gacha_{user_id}.json"


def load_gacha_data(user_id: str) -> dict:
    """加载用户的呱太数据"""
    gacha_file = get_gacha_file(user_id)
    if gacha_file.exists():
        try:
            with open(gacha_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {"gacha": 0}
    return {"gacha": 0}


def save_gacha_data(user_id: str, data: dict):
    """保存用户的呱太数据"""
    gacha_file = get_gacha_file(user_id)
    with open(gacha_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def get_gacha_count(user_id: str) -> int:
    """获取用户的呱太数量"""
    gacha_data = load_gacha_data(user_id)
    return gacha_data.get("gacha", 0)


def add_gacha(user_id: str, amount: int) -> int:
    """增加用户的呱太数量"""
    gacha_data = load_gacha_data(user_id)
    gacha_data["gacha"] = gacha_data.get("gacha", 0) + amount
    save_gacha_data(user_id, gacha_data)
    return gacha_data["gacha"]


def spend_gacha(user_id: str, amount: int) -> bool:
    """消耗用户的呱太数量，返回是否成功"""
    gacha_data = load_gacha_data(user_id)
    current = gacha_data.get("gacha", 0)
    if current >= amount:
        gacha_data["gacha"] = current - amount
        save_gacha_data(user_id, gacha_data)
        return True
    return False


# ========== 签到系统 ==========
def get_signin_file(user_id: str) -> Path:
    """获取用户签到记录文件路径"""
    return INFO_DIR / f"signin_{user_id}.json"


def load_signin_data(user_id: str) -> dict:
    """加载用户的签到数据"""
    signin_file = get_signin_file(user_id)
    if signin_file.exists():
        try:
            with open(signin_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {"last_signin": "", "streak": 0}
    return {"last_signin": "", "streak": 0}


def save_signin_data(user_id: str, data: dict):
    """保存用户的签到数据"""
    signin_file = get_signin_file(user_id)
    with open(signin_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def can_signin(user_id: str) -> bool:
    """检查用户今天是否可以签到"""
    signin_data = load_signin_data(user_id)
    last_signin = signin_data.get("last_signin", "")
    today = datetime.now().strftime("%Y-%m-%d")
    return last_signin != today


def do_signin(user_id: str) -> dict:
    """执行签到，返回签到结果"""
    signin_data = load_signin_data(user_id)
    today = datetime.now().strftime("%Y-%m-%d")
    
    if signin_data.get("last_signin") == today:
        return {"success": False, "message": "今天已经签到过了", "streak": signin_data.get("streak", 0)}
    
    # 更新签到记录
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    if signin_data.get("last_signin") == yesterday:
        # 连续签到
        signin_data["streak"] = signin_data.get("streak", 0) + 1
    else:
        # 断签，重置连续天数
        signin_data["streak"] = 1
    
    signin_data["last_signin"] = today
    save_signin_data(user_id, signin_data)
    
    # 添加签到奖励
    add_gacha(user_id, DAILY_REWARD)
    
    return {
        "success": True,
        "message": f"签到成功！获得 {DAILY_REWARD} 呱太",
        "streak": signin_data["streak"],
        "reward": DAILY_REWARD
    }


# ========== 盲盒开箱系统 ==========
def draw_mystery_box(characters: list, user_id: str = None, is_pity: bool = False) -> dict:
    """
    抽取一个盲盒（未开的盒子）
    返回: {"stars": 星级, "is_mystery": 是否黑色盲盒, "character": 角色对象(如果是黑色盲盒则为None)}
    
    参数:
        characters: 角色列表
        user_id: 用户ID（用于判断保底）
        is_pity: 是否触发保底
    """
    # 如果是保底抽卡，跳过黑色盲盒，直接生成3星
    if is_pity:
        # 保底必定出3星，并且是FES限定
        character = select_3star_character(characters, is_fes_pity=True)
        return {
            "stars": 3,
            "is_mystery": False,
            "character": character
        }
    
    # 先判断是否是黑色盲盒
    is_mystery = random.random() < MYSTERY_BOX_CHANCE
    
    if is_mystery:
        # 黑色盲盒概率（从配置读取）
        stars = random.choices([2, 3], weights=[MYSTERY_BOX_2STAR_PROB, MYSTERY_BOX_3STAR_PROB], k=1)[0]
        return {
            "stars": stars,
            "is_mystery": True,
            "character": None  # 黑色盲盒还没开，所以没有角色
        }
    else:
        # 正常盲盒概率（从配置读取）
        stars = random.choices([1, 2, 3], weights=[NORMAL_BOX_1STAR_PROB, NORMAL_BOX_2STAR_PROB, NORMAL_BOX_3STAR_PROB], k=1)[0]
        
        # 如果是3星，根据限定种类概率分配
        if stars == 3:
            character = select_3star_character(characters, is_fes_pity=False)
        else:
            # 1星或2星，按原逻辑抽取
            available = [c for c in characters if c["stars"] == stars]
            if available:
                character = random.choice(available)
            else:
                character = random.choice(characters)
        
        return {
            "stars": stars,
            "is_mystery": False,
            "character": character
        }


def select_3star_character(characters: list, is_fes_pity: bool = False) -> dict:
    """
    根据限定种类概率选择三星角色
    期間限定35%，フェス限定25%，其余三星40%
    如果是フェス保底，则必定返回フェス限定角色
    """
    # 获取所有三星角色
    all_3stars = [c for c in characters if c["stars"] == 3]
    if not all_3stars:
        return random.choice(characters)
    
    # 如果是フェス保底，直接返回フェス限定角色
    if is_fes_pity:
        fes_chars = [c for c in all_3stars if c.get("limit_type") == "フェス限定"]
        if fes_chars:
            return random.choice(fes_chars)
        # 如果没有フェス限定，返回任意三星
        return random.choice(all_3stars)
    
    # 按限定种类分类
    period_limited = [c for c in all_3stars if c.get("limit_type") == "期間限定"]
    fes_limited = [c for c in all_3stars if c.get("limit_type") == "フェス限定"]
    other_3stars = [c for c in all_3stars if c.get("limit_type") not in ["期間限定", "フェス限定"]]
    
    # 概率分配（从配置读取）
    categories = []
    weights = []
    
    if period_limited:
        categories.append(period_limited)
        weights.append(PERIOD_LIMIT_PROB)
    if fes_limited:
        categories.append(fes_limited)
        weights.append(FES_LIMIT_PROB)
    if other_3stars:
        categories.append(other_3stars)
        weights.append(OTHER_3STAR_PROB)
    
    if not categories:
        return random.choice(all_3stars)
    
    # 按权重选择类别，再从该类别中随机选择角色
    selected_category = random.choices(categories, weights=weights, k=1)[0]
    return random.choice(selected_category)


def apply_mutation(box_info: dict, characters: list) -> tuple:
    """
    对盲盒结果应用突变，并在突变后重新抽取对应星级的角色
    返回: (突变后的box_info, 是否发生突变, 突变描述)
    """
    if box_info["is_mystery"]:
        # 黑色盲盒不受突变影响
        return box_info, False, None
    
    original_stars = box_info["stars"]
    roll = random.random()
    new_stars = original_stars
    mutation_occurred = False
    mutation_text = None
    
    if original_stars == 1:
        if roll < MUTATION_1_TO_2:
            new_stars = 2
            mutation_occurred = True
            mutation_text = "1星→2星"
        elif roll > 1 - MUTATION_1_TO_3:
            new_stars = 3
            mutation_occurred = True
            mutation_text = "1星→3星"
    elif original_stars == 2:
        if roll < MUTATION_2_TO_3:
            new_stars = 3
            mutation_occurred = True
            mutation_text = "2星→3星"
    
    if mutation_occurred and new_stars != original_stars:
        box_info["stars"] = new_stars
        # 重新抽取对应新星级的角色
        available = [c for c in characters if c["stars"] == new_stars]
        if available:
            box_info["character"] = random.choice(available)
        else:
            box_info["character"] = random.choice(characters)
    
    return box_info, mutation_occurred, mutation_text


def get_mystery_box_image(stars: int) -> str:
    """获取盲盒图片路径（使用02图层作为盲盒封面）"""
    star_idx = stars - 1
    # 盲盒使用02图层
    filename = f"gacha_tmb_{star_idx:02d}_02.png"
    path = LEVEL_DIR / filename
    if path.exists():
        return str(path)
    
    # 如果02不存在，尝试用01（框）作为后备
    filename = f"gacha_tmb_{star_idx:02d}_01.png"
    path = LEVEL_DIR / filename
    if path.exists():
        return str(path)
    
    log_error(f"找不到图片: gacha_tmb_{star_idx:02d}_02")
    return None


def get_black_box_image() -> str:
    """获取黑色盲盒图片路径"""
    filename = "gacha_tmb_04_02.png"
    path = LEVEL_DIR / filename
    if path.exists():
        return str(path)
    log_error(f"找不到黑色图片: {filename}")
    return None


def create_box_card(box_info: dict, characters: list) -> bytes:
    """
    创建盲盒卡牌图片（未开的样子）
    对于黑色盲盒，使用gacha_tmb_04_02作为中心图
    对于3星卡，如果是フェス限定，使用gacha_tmb_03_02_b.png
    """
    stars = box_info["stars"]
    is_mystery = box_info.get("is_mystery", False)
    character = box_info.get("character", None)
    
    # 获取背景和框
    if is_mystery:
        # 黑色盲盒使用04_00和04_01
        bg_path = LEVEL_DIR / "gacha_tmb_04_00.png"
        frame_path = LEVEL_DIR / "gacha_tmb_04_01.png"
        center_path = LEVEL_DIR / "gacha_tmb_04_02.png"
    else:
        bg_path = get_level_image(stars, "bg")
        frame_path = get_level_image(stars, "frame")
        # 盲盒使用02图层作为中心
        star_idx = stars - 1
        
        # 判断是否是フェス限定3星
        is_fes = False
        if character and stars == 3:
            limit_type = character.get("limit_type", "")
            if limit_type == "フェス限定":
                is_fes = True
        
        # 根据是否是フェス限定选择不同的盲盒图
        if is_fes:
            # フェス限定3星使用特殊盲盒图
            center_path = LEVEL_DIR / "gacha_tmb_03_02_b.png"
            # 如果不存在则使用普通3星盲盒图
            if not center_path.exists():
                center_path = LEVEL_DIR / f"gacha_tmb_{star_idx:02d}_02.png"
        else:
            center_path = LEVEL_DIR / f"gacha_tmb_{star_idx:02d}_02.png"
    
    if not bg_path or not frame_path or not center_path:
        log_error(f"盲盒图片不全: bg={bg_path}, frame={frame_path}, center={center_path}")
        return None
    
    bg_img = Image.open(bg_path).convert('RGBA')
    frame_img = Image.open(frame_path).convert('RGBA')
    center_img = Image.open(center_path).convert('RGBA')
    
    bg_width, bg_height = bg_img.size
    
    # 创建输出画布
    output = Image.new('RGBA', (bg_width, bg_height), (0, 0, 0, 0))
    output.paste(bg_img, (0, 0))
    
    # 调整中心图大小以适应框
    center_width, center_height = center_img.size
    center_ratio = center_width / center_height
    frame_ratio = bg_width / bg_height
    
    if center_ratio > frame_ratio:
        target_width = int(bg_width * 0.85)
        target_height = int(target_width / center_ratio)
    else:
        target_height = int(bg_height * 0.85)
        target_width = int(target_height * center_ratio)
    
    center_resized = center_img.resize((target_width, target_height), Image.Resampling.LANCZOS)
    center_x = (bg_width - target_width) // 2
    center_y = (bg_height - target_height) // 2
    
    output.paste(center_resized, (center_x, center_y), center_resized)
    output.paste(frame_img, (0, 0), frame_img)
    
    # 转换为RGB
    output_rgb = Image.new('RGB', (bg_width, bg_height), (255, 255, 255))
    output_rgb.paste(output, (0, 0), output)
    
    bio = BytesIO()
    output_rgb.save(bio, format='JPEG', optimize=True, quality=60)
    return bio.getvalue()


def open_mystery_box(box_info: dict, characters: list) -> dict:
    """
    开黑色盲盒，获取实际角色
    """
    if not box_info["is_mystery"]:
        # 不是黑色盲盒，直接返回（突变在调用处处理）
        return box_info
    
    # 黑色盲盒：3星40%，2星60%（stars已在创建时确定）
    stars = box_info["stars"]
    available = [c for c in characters if c["stars"] == stars]
    if available:
        character = random.choice(available)
    else:
        character = random.choice(characters)
    
    box_info["character"] = character
    box_info["opened"] = True
    
    return box_info


def create_opened_box_card(box_info: dict) -> bytes:
    """创建已开盲盒的卡牌图片（显示角色）"""
    if box_info["is_mystery"] and not box_info.get("opened"):
        # 黑色盲盒还没开，应该用未开的图片
        return create_box_card(box_info, [])
    
    stars = box_info["stars"]
    character = box_info.get("character")
    
    if not character:
        log_error("已开盲盒没有角色信息")
        return create_box_card(box_info, [])
    
    # 使用正常的卡牌合成逻辑
    return composite_card(character)


def create_box_summary_image(boxes: list, opened_indices: list, characters: list) -> bytes:
    """
    创建盲盒汇总图片（带背景）
    boxes: 所有盲盒列表
    opened_indices: 已开的盲盒索引列表
    """
    if not boxes:
        return None
    
    # 创建单张盲盒图片的尺寸
    single_card_img = None
    for i, box in enumerate(boxes):
        img_bytes = create_box_card(box, characters)
        if img_bytes:
            single_card_img = Image.open(BytesIO(img_bytes))
            break
    
    if not single_card_img:
        return None
    
    card_width, card_height = single_card_img.size
    count = len(boxes)
    
    # 计算行列布局
    if count <= 5:
        cols = count
        rows = 1
    else:
        cols = 5
        rows = (count + 4) // 5
    
    gap = 10
    cards_total_width = card_width * cols + gap * (cols - 1)
    cards_total_height = card_height * rows + gap * (rows - 1)
    
    # 尝试加载抽卡结果背景图片（使用gacha_tmb_bg_11.png）
    bg_path = None
    for bg_name in ["gacha_tmb_bg_11.png", "gacha_tmb_11_bg.png", "gacha_bg_11.png"]:
        test_path = LEVEL_DIR / bg_name
        if test_path.exists():
            bg_path = str(test_path)
            break
    
    # 如果找不到gacha_tmb_bg_11.png，回退到gacha_tmb_bg_10.png
    if not bg_path:
        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
            test_path = LEVEL_DIR / bg_name
            if test_path.exists():
                bg_path = str(test_path)
                break
    
    if bg_path:
        # 使用背景图片
        log_info(f"使用汇总背景图片: {bg_path}")
        bg_img = Image.open(bg_path).convert('RGB')
        
        # 将背景放大到原来的2倍
        bg_w, bg_h = bg_img.size
        final_w = int(bg_w * 2 * 0.264)
        final_h = int(bg_h * 2 *0.264)
        bg_img_resized = bg_img.resize((final_w, final_h), Image.Resampling.LANCZOS)
        
        # 创建最终画布
        output = Image.new('RGB', (final_w, final_h), (50, 50, 50))
        output.paste(bg_img_resized, (0, 0))
        
        # 计算卡牌居中位置
        cards_x = (final_w - cards_total_width) // 2
        cards_y = (final_h - cards_total_height) // 2
    else:
        # 没有背景，使用纯色背景
        output = Image.new('RGB', (cards_total_width, cards_total_height), (50, 50, 50))
        cards_x = 0
        cards_y = 0
    
    for i, box in enumerate(boxes):
        row = i // cols
        col = i % cols
        
        if i in opened_indices:
            # 已开的盲盒显示角色
            img_bytes = create_opened_box_card(box)
        else:
            # 未开的盲盒显示盲盒封面
            img_bytes = create_box_card(box, characters)
        
        if img_bytes:
            card_img = Image.open(BytesIO(img_bytes))
            x = cards_x + col * (card_width + gap)
            y = cards_y + row * (card_height + gap)
            output.paste(card_img, (x, y))
    
    bio = BytesIO()
    output.save(bio, format='JPEG', optimize=True, quality=60)
    return bio.getvalue()


def has_box_session(user_id: str) -> bool:
    """检查用户是否有未完成的盲盒会话"""
    if user_id not in BOX_SESSIONS:
        return False
    
    # 检查是否所有盲盒都已开完
    session = BOX_SESSIONS[user_id]
    boxes = session.get("boxes", [])
    opened = session.get("opened", [])
    
    if len(boxes) == len(opened):
        # 所有盲盒已开完，但保留会话用于详细信息查询
        return False
    
    return True


def get_box_session(user_id: str) -> dict:
    """获取用户的盲盒会话"""
    return BOX_SESSIONS.get(user_id)


def create_box_session(user_id: str, boxes: list):
    """创建盲盒会话"""
    BOX_SESSIONS[user_id] = {
        "boxes": boxes,
        "opened": [],  # 已开的盲盒索引
        "characters": None,  # 角色列表引用
        "created_at": datetime.now()
    }


def clear_box_session(user_id: str):
    """清除盲盒会话"""
    if user_id in BOX_SESSIONS:
        del BOX_SESSIONS[user_id]


def is_valid_box_index(boxes: list, index_str: str, opened_indices: list = None) -> tuple:
    """
    验证盲盒索引是否合法
    返回: (是否有效, 索引列表或错误信息)
    """
    if opened_indices is None:
        opened_indices = []
    opened_set = set(opened_indices)
    
    index_str = index_str.strip()
    
    if index_str == "全部开":
        return True, list(range(len(boxes)))
    
    if index_str == "剩下的全部开":
        return True, [i for i in range(len(boxes)) if i not in opened_set]
    
    # 解析单个或多个索引，支持"选择1"等格式
    import re
    try:
        indices = []
        parts = index_str.replace("，", ",").split(",")
        
        for part in parts:
            part = part.strip()
            # 使用正则提取数字
            num_match = re.search(r'(\d+)', part)
            if num_match:
                num = int(num_match.group(1))
                idx = num - 1  # 用户输入1-10，转为0-9
                if 0 <= idx < len(boxes):
                    indices.append(idx)
                else:
                    return False, f"索引{num}超出范围(1-{len(boxes)})"
            else:
                return False, f"无效的输入: {part}，请输入数字或「选择数字」"
        
        return True, list(set(indices))  # 去重
    except Exception as e:
        return False, f"解析失败: {str(e)}"


# ========== 日活与发送统计模块 ==========
def restore_daily_stats():
    """启动时从 dau_log.json 恢复今日统计数据到内存"""
    global DAILY_SEND_STATS, _RESTORED_DAU
    today = datetime.now().strftime("%Y-%m-%d")
    dau_file = INFO_DIR / "dau_log.json"

    if not dau_file.exists():
        return

    try:
        with open(dau_file, "r", encoding="utf-8") as f:
            raw = json.load(f)
        if today in raw:
            saved = raw[today]
            if isinstance(saved, dict):
                saved_sends = saved.get("send", 0)
                saved_dau = saved.get("dau", 0)
                if DAILY_SEND_STATS.get(today, 0) == 0:
                    DAILY_SEND_STATS[today] = saved_sends
                # 记录重启前的日活基准（内存中无法恢复用户ID，用计数追平）
                _RESTORED_DAU = saved_dau
                log_info(f"从磁盘恢复今日统计: 日活{saved_dau}人 发送{saved_sends}次")
    except Exception as e:
        log_error(f"恢复日活数据失败: {e}")


def save_daily_stats():
    """保存当前内存中的日活和发送统计到磁盘（轻量，不打报告）"""
    today = datetime.now().strftime("%Y-%m-%d")
    dau_file = INFO_DIR / "dau_log.json"

    # 读取历史数据
    data = {}
    if dau_file.exists():
        try:
            with open(dau_file, "r", encoding="utf-8") as f:
                raw = json.load(f)
            for k, v in raw.items():
                if isinstance(v, dict):
                    data[k] = v
                elif isinstance(v, int):
                    data[k] = {"dau": v, "send": 0}
        except Exception:
            pass

    # 更新今天的数据（合并重启前基准 + 本次运行新数据）
    today_dau = max(len(DAILY_USER_SET), _RESTORED_DAU)
    today_sends = max(DAILY_SEND_STATS.get(today, 0), data.get(today, {}).get("send", 0) if isinstance(data.get(today), dict) else 0)
    data[today] = {"dau": today_dau, "send": today_sends}

    # 保存
    with open(dau_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def record_daily_dau():
    """
    保存并输出从6/8至今的统计报告
    日活 = 当日发过消息的独立用户数 (来自 DAILY_USER_SET)
    发送次数 = 当日发送消息总数 (来自 DAILY_SEND_STATS)
    """
    # 先保存最新数据
    save_daily_stats()

    today = datetime.now().strftime("%Y-%m-%d")
    dau_file = INFO_DIR / "dau_log.json"

    # 读取完整数据生成报告
    data = {}
    if dau_file.exists():
        try:
            with open(dau_file, "r", encoding="utf-8") as f:
                raw = json.load(f)
            for k, v in raw.items():
                if isinstance(v, dict):
                    data[k] = v
                elif isinstance(v, int):
                    data[k] = {"dau": v, "send": 0}
        except Exception:
            pass

    today_dau = max(len(DAILY_USER_SET), _RESTORED_DAU)
    today_sends = DAILY_SEND_STATS.get(today, 0)

    # 生成统计报告 (仅统计6/8至今)
    start = "2026-06-08"
    dates = sorted(d for d in data if d >= start)
    if not dates:
        return

    total_dau = sum(data[d]["dau"] for d in dates)
    total_send = sum(data[d]["send"] for d in dates)
    days = len(dates)
    peak_dau = max(dates, key=lambda d: data[d]["dau"])
    peak_send = max(dates, key=lambda d: data[d]["send"])

    report = f"\n===== 统计 ({start} ~ {today}) =====\n"
    for d in dates:
        report += f"  {d}: 日活{data[d]['dau']:>3}人  发送{data[d]['send']:>4}次\n"
    report += f"  ─────────────────────\n"
    report += f"  总天数:{days}  总活跃:{total_dau}人次  总发送:{total_send}次\n"
    report += f"  日均活跃:{total_dau//days}人  日均发送:{total_send//days}次\n"
    report += f"  最高日活:{data[peak_dau]['dau']}人({peak_dau})  最高发送:{data[peak_send]['send']}次({peak_send})\n"
    report += f"  今日日活:{today_dau}人  今日发送:{today_sends}次\n"

    log_info(report)
    print(report)


# ========== 敏感词过滤模块 ==========
SENSITIVE_WORDS = [
    # 政治敏感词
    "习近平", "胡锦涛", "江泽民", "毛泽东", "邓小平",
    "中南海", "人民大会堂", "天安门", "国旗", "国徽",
    # 色情暴力词
    "色情", "色情内容", "黄色", "强奸", "卖淫", "嫖娼",
    "毒品", "赌博", "暴力", "恐怖", "杀人",
    # 广告垃圾词
    "广告", "推广", "刷单", "返利", "红包", "二维码",
    "微信号", "QQ号", "电话号码", "网址", "链接"
]


def filter_sensitive_words(message: str) -> str:
    """
    过滤消息中的敏感词
    :param message: 原始消息
    :return: 过滤后的消息
    """
    if not message:
        return message
    
    filtered_message = message
    for word in SENSITIVE_WORDS:
        filtered_message = filtered_message.replace(word, "*" * len(word))
    
    return filtered_message


# ========== 消息速率限制模块 ==========
def check_message_rate(target_id: str) -> bool:
    """
    检查消息发送速率是否超过限制
    :param target_id: 目标ID（群ID或用户ID）
    :return: True表示可以发送，False表示超过限制
    """
    now = datetime.now().timestamp()
    one_minute = 60  # 60秒
    
    # 获取当前目标的计数器
    if target_id not in MESSAGE_COUNTER:
        MESSAGE_COUNTER[target_id] = {"count": 0, "start_time": now}
    
    counter = MESSAGE_COUNTER[target_id]
    
    # 检查是否需要重置计数器（超过1分钟）
    if now - counter["start_time"] >= one_minute:
        counter["count"] = 0
        counter["start_time"] = now
    
    # 检查是否超过限制
    if counter["count"] >= MAX_MESSAGES_PER_MINUTE:
        log_error(f"消息发送速率超过限制！目标: {target_id}, 当前计数: {counter['count']}")
        return False
    
    # 增加计数
    counter["count"] += 1
    return True


# ========== Napcat消息发送模块 ==========
def send_message(message: str, user_id: str = None, group_id: str = None, image_path: str = None):
    """
    调用Napcat API发送消息（带速率限制）
    :param message: 消息内容
    :param user_id: 用户ID（私聊）
    :param group_id: 群ID（群聊）
    :param image_path: 图片路径（可选）
    """
    # 确定目标ID用于速率限制
    target_id = group_id if group_id else user_id
    if not target_id:
        log_error("发送消息失败：缺少user_id或group_id")
        return False
    
    # 检查速率限制
    if not check_message_rate(target_id):
        log_error(f"消息发送被速率限制阻止，目标: {target_id}")
        return False
    
    # 敏感词过滤
    message = filter_sensitive_words(message)
    
    try:
        if group_id:
            # 发送群消息
            url = f"http://{NAPCAT_HOST}:{NAPCAT_PORT}/send_group_msg"
            data = {
                "group_id": int(group_id),
                "message": message
            }
        elif user_id:
            # 发送私聊消息
            url = f"http://{NAPCAT_HOST}:{NAPCAT_PORT}/send_private_msg"
            data = {
                "user_id": int(user_id),
                "message": message
            }
        else:
            log_error("发送消息失败：缺少user_id或group_id")
            return False

        # NapCatQQ 用 access_token 参数认证
        if NAPCAT_TOKEN:
            url += f"?access_token={NAPCAT_TOKEN}"

        response = requests.post(url, json=data, timeout=10)
        
        if response.status_code != 200:
            log_error(f"发送失败 HTTP{response.status_code}: {response.text[:100]}")
            return False

        try:
            result = response.json()
        except Exception as e:
            log_error(f"响应JSON解析失败: {e}")
            return False

        if result.get("status") == "ok" or result.get("retcode") == 0:
            today = datetime.now().strftime("%Y-%m-%d")
            DAILY_SEND_STATS[today] = DAILY_SEND_STATS.get(today, 0) + 1
            return True
        else:
            log_error(f"Napcat返回错误: {result}")
            return False

    except requests.exceptions.ConnectionError:
        log_error(f"连接Napcat失败，请检查Napcat是否启动以及端口是否正确: {NAPCAT_HOST}:{NAPCAT_PORT}")
        return False
    except Exception as e:
        log_error(f"发送消息异常: {e}")
        return False


def compress_image_bytes(img_path: str, max_kb: int = 300) -> bytes:
    """压缩图片到目标大小以内，返回JPEG字节"""
    try:
        from PIL import Image
        img = Image.open(img_path).convert('RGB')
        quality = 75
        while quality >= 20:
            bio = BytesIO()
            img.save(bio, format='JPEG', optimize=True, quality=quality)
            if bio.tell() <= max_kb * 1024 or quality <= 20:
                return bio.getvalue()
            quality -= 10
        return bio.getvalue()
    except Exception:
        with open(img_path, 'rb') as f:
            return f.read()


def send_image(image_path: str, user_id: str = None, group_id: str = None):
    """
    发送图片消息（自动压缩，发送后删除）
    """
    try:
        if not os.path.exists(image_path):
            log_error(f"图片文件不存在: {image_path}")
            return False

        compressed = compress_image_bytes(image_path)
        image_base64 = base64.b64encode(compressed).decode("utf-8")
        image_message = f"[CQ:image,file=base64://{image_base64}]"
        result = send_message(image_message, user_id, group_id)

        # 发送后删除临时图片
        try:
            os.remove(image_path)
        except Exception:
            pass

        return result

    except Exception as e:
        log_error(f"发送图片失败: {e}")
        return False


# ========== Excel数据读取 ==========
def _parse_direction_offsets(direction_str) -> list:
    """解析攻击方向箭头 → 列偏移列表"""
    if not direction_str:
        return [0]
    s = str(direction_str)
    offsets = set()
    for ch, off in [('↖',-1),('↙',-1),('←',-1),('↑',0),('↓',0),('↗',1),('↘',1),('→',1)]:
        if ch in s:
            offsets.add(off)
    return sorted(offsets) if offsets else [0]


def _derive_limit_type(acquire: str) -> str:
    """从获取方式推导限定种类"""
    if not acquire:
        return ""
    a = str(acquire).lower()
    if 'fes' in a:
        return "フェス限定"
    if '限定' in a or '期間' in a:
        return "期間限定"
    if '活動' in a or '活动' in a:
        return "期間限定"
    return ""


def load_character_data():
    """从卡牌信息.xlsx 加载角色数据（用于抽卡，含1/2/3星）"""
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
        characters = []

        for sheet_name, type_name in [('BattleCard资源', 'battle'), ('AssistCard资源', 'assist')]:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    card_id = str(row[5]) if row[5] is not None else None
                    chara_id = row[6]
                    init_stars = row[1]
                    card_name = row[4]
                    limit_type = row[0] or ""
                    if not (card_id and chara_id and init_stars and card_name):
                        continue
                    stars_val = int(float(str(init_stars)))
                    icon_path = find_character_icon(chara_id, stars_val)
                    characters.append({
                        "card_id": card_id, "chara_id": chara_id,
                        "stars": stars_val, "name": str(card_name),
                        "attribute": str(row[3]) if row[3] else "红",
                        "limit_type": str(limit_type),
                        "icon_path": icon_path, "type": type_name,
                        "hp": 5000, "attack": 3000, "defense": 2000,
                        "dexterity": 1000, "speed": 500,
                        "attack_directions": [0],
                        "attack_type": "物理", "side": "科学",
                        "skill1": {"cd": 0, "description": "", "keywords": "", "condition": ""},
                        "skill2": {"cd": 0, "description": "", "keywords": "", "condition": ""},
                        "skill3": {"cd": 0, "description": "", "keywords": "", "condition": ""},
                        "passive1_text": "", "passive2_text": "",
                    })
                except Exception:
                    continue
            log_info(f"[抽卡] 加载了 {sum(1 for c in characters if c['type']==type_name)} 个{type_name}角色")

        log_info(f"[抽卡] 共加载 {len(characters)} 个角色")
        wb.close()
        return characters
    except Exception as e:
        import traceback
        log_error(f"加载抽卡数据失败: {e}")
        return []


def load_battle_characters():
    """从 cards_completed.xlsx 加载战斗数据（3星卡详细数值）"""
    try:
        wb = openpyxl.load_workbook(BATTLE_XLSX, data_only=True)
        characters = []

        for sheet_name, card_type in [('b卡', 'battle'), ('a卡', 'assist')]:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    card_id = str(int(row[0])) if row[0] is not None else None
                    if not card_id:
                        continue
                    name_raw = str(row[1]) if row[1] else ""
                    name = name_raw.split('|')[0].strip() if '|' in name_raw else name_raw.strip()
                    side = str(row[3]) if row[3] else "科学"
                    direction_raw = str(row[4]) if row[4] else ""
                    attribute = str(row[5]) if row[5] else "红"
                    attack_type_str = str(row[6]) if row[6] else "物理"
                    sk1_desc = str(row[8]) if row[8] else ""
                    sk1_cd = int(float(row[9])) if row[9] is not None else 0
                    sk2_desc = str(row[10]) if row[10] else ""
                    sk2_cd = int(float(row[11])) if row[11] is not None else 0
                    sk3_desc = str(row[12]) if row[12] else ""
                    sk3_cd = int(float(row[13])) if row[13] is not None else 0
                    pas1 = str(row[14]) if row[14] else ""
                    pas2 = str(row[15]) if row[15] else ""
                    phys_atk = int(float(row[16])) if row[16] is not None else 5000
                    magic_atk = int(float(row[17])) if row[17] is not None else 5000
                    phys_def = int(float(row[18])) if row[18] is not None else 3000
                    magic_def = int(float(row[19])) if row[19] is not None else 3000
                    hp = int(float(row[20])) if row[20] is not None else 10000
                    dexterity = int(float(row[21])) if row[21] is not None else 1000
                    is_physical = '物理' in attack_type_str
                    attack = phys_atk if is_physical else magic_atk
                    defense = (phys_def + magic_def) // 2
                    color_map = {'赤': '红', '緑': '绿', '青': '蓝'}
                    attr = attribute.strip()
                    is_super = attr.startswith('超')
                    base = attr[1:] if is_super else attr
                    base = color_map.get(base, base)
                    attr_normalized = ('超' + base) if is_super else base
                    offsets = set()
                    for ch, off in [('↖',-1),('↙',-1),('←',-1),('↑',0),('↓',0),('↗',1),('↘',1),('→',1)]:
                        if ch in str(direction_raw): offsets.add(off)
                    dire_list = sorted(offsets) if offsets else [0]
                    acquire = str(row[2]) if row[2] else ""
                    limit_type = ""
                    if 'fes' in acquire.lower(): limit_type = "フェス限定"
                    elif '限定' in acquire or '期間' in acquire: limit_type = "期間限定"
                    elif '活動' in acquire or '活动' in acquire: limit_type = "期間限定"
                    icon_path = find_character_icon(card_id, 3)
                    characters.append({
                        "card_id": card_id, "chara_id": card_id,
                        "stars": 3, "name": name,
                        "attribute": attr_normalized, "limit_type": limit_type,
                        "icon_path": icon_path, "type": card_type,
                        "hp": hp, "attack": attack, "defense": defense,
                        "dexterity": dexterity, "phys_atk": phys_atk, "magic_atk": magic_atk,
                        "attack_directions": dire_list,
                        "attack_type": attack_type_str, "side": side,
                        "skill1": {"cd": sk1_cd, "description": sk1_desc, "keywords": "", "condition": ""},
                        "skill2": {"cd": sk2_cd, "description": sk2_desc, "keywords": "", "condition": ""},
                        "skill3": {"cd": sk3_cd, "description": sk3_desc, "keywords": "", "condition": ""},
                        "passive1_text": pas1, "passive2_text": pas2,
                    })
                except Exception:
                    continue
            log_info(f"[战斗] 加载了 {sum(1 for c in characters if c['type']==card_type)} 个{card_type}角色")

        log_info(f"[战斗] 共加载 {len(characters)} 个角色")
        wb.close()
        return characters
    except Exception as e:
        import traceback
        log_error(f"加载战斗数据失败: {e}")
        return []


def find_character_icon(chara_id, stars: int) -> str:
    """根据角色ID和星级查找图标文件（支持多种文件名格式）"""
    # 确保 chara_id 是整数
    try:
        # 如果是浮点数（如 100111001.0），先转int再转str去掉小数点
        if isinstance(chara_id, float):
            chara_id_int = int(chara_id)
        elif isinstance(chara_id, str):
            # 尝试解析字符串数字，可能带有小数点
            chara_id_int = int(float(chara_id))
        else:
            chara_id_int = int(chara_id)
    except (ValueError, TypeError):
        log_error(f"无效的chara_id: {chara_id}, 类型: {type(chara_id)}")
        return None
    
    # 尝试多种可能的文件名模式（按优先级排序）
    patterns = [
        # 新格式：card_cutin_前缀（9位数字）
        f"card_cutin_{chara_id_int:09d}.png",
        f"card_cutin_{chara_id_int}.png",
        
        # 新格式：直接用chara_id命名
        f"{chara_id_int:03d}.png",
        f"{chara_id_int}.png",
        f"{chara_id_int:09d}.png",
        
        # chara前缀
        f"chara_{chara_id_int:03d}.png",
        f"chara_{chara_id_int}.png",
        
        # icon前缀
        f"icon_{chara_id_int:03d}.png",
        f"icon_{chara_id_int}.png",
        
        # 带星级的格式
        f"{chara_id_int:03d}_{stars}star.png",
        f"{chara_id_int}_{stars}star.png",
        f"chara_{chara_id_int:03d}_{stars}.png",
        f"chara_{chara_id_int}_{stars}.png",
        f"icon_{chara_id_int:03d}_{stars}.png",
        f"icon_{chara_id_int}_{stars}.png",
        
        # 原始格式（保留兼容）
        f"character_icon_{chara_id_int:03d}{stars:02d}0101.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0201.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0102.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0202.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0103.png",
    ]

    # 先尝试精确匹配
    for pattern in patterns:
        if '*' not in pattern:  # 非通配符模式
            path = ICON_DIR / pattern
            if path.exists():
                return str(path)

    # 如果精确匹配都找不到，尝试通配符匹配
    wildcard_patterns = [
        f"card_cutin_{chara_id_int}*.png",
        f"card_cutin_{chara_id_int:03d}*.png",
        f"card_cutin_{chara_id_int:09d}*.png",
        f"{chara_id_int:03d}*.png",
        f"{chara_id_int:09d}*.png",
        f"*{chara_id_int:03d}*.png",
        f"*{chara_id_int:09d}*.png",
        f"chara_{chara_id_int:03d}*.png",
        f"icon_{chara_id_int:03d}*.png",
        f"character_icon_{chara_id_int:03d}*.png",
    ]
    
    for pattern in wildcard_patterns:
        matches = list(ICON_DIR.glob(pattern))
        if matches:
            # 优先选择包含星级的文件
            star_matches = [m for m in matches if f"{stars}" in m.name or f"{stars:02d}" in m.name]
            if star_matches:
                return str(star_matches[0])
            return str(matches[0])

    log_error(f"未找到角色图标: chara_id={chara_id}, chara_id_int={chara_id_int}, stars={stars}")
    return None


# ========== 抽卡逻辑 ==========
from typing import Optional, List, Dict, Any


def draw_cards(count: int, characters: List[Dict[str, Any]], user_id: Optional[str] = None) -> Dict[str, Any]:
    """
    抽取指定数量的卡牌
    使用权重：从配置读取概率
    保底机制：每150抽必出フェス限定3星
    
    :param count: 抽卡数量（1或10）
    :param characters: 角色列表
    :param user_id: 用户ID（用于保底计数）
    :return: {"results": 抽卡结果列表, "remaining_pity": 剩余保底抽数, "got_3star": 是否抽到3星}
    """
    # 输入验证
    if count not in [1, 10]:
        raise ValueError("抽卡数量只能是1或10")
    
    if not characters:
        raise ValueError("角色列表不能为空")
    
    if user_id and not isinstance(user_id, str):
        raise TypeError("user_id必须是字符串类型")

    results: List[Dict[str, Any]] = []
    got_3star = False
    remaining_pity = 0
    
    for _ in range(count):
        # 检查是否触发保底
        if user_id:
            pity_data = load_pity_data(user_id)
            pity_count = pity_data.get("pity_count", 0)
            remaining_before = max(0, PITY_LIMIT - pity_count)
            
            # 如果是保底抽，直接出3星
            if remaining_before == 1:
                stars = 3
            else:
                # 随机决定星级（使用预定义的权重常量）
                stars = random.choices(
                    population=[1, 2, 3],
                    weights=GACHA_WEIGHTS,
                    k=1
                )[0]
        else:
            # 没有用户ID，正常抽卡（使用预定义的权重常量）
            stars = random.choices(
                population=[1, 2, 3],
                weights=GACHA_WEIGHTS,
                k=1
            )[0]

        # 从对应星级的角色中随机选择一个
        available = [c for c in characters if c["stars"] == stars]
        if available:
            char = random.choice(available)
            results.append(char)
            if stars == 3:
                got_3star = True
        else:
            # 如果没有对应星级的角色，随机选一个
            log_error(f"没有找到 {stars}星级的角色，随机分配")
            char = random.choice(characters)
            results.append(char)
            if char["stars"] == 3:
                got_3star = True
    
    # 更新保底计数
    if user_id:
        for _ in range(count):
            update_pity(user_id, got_3star)
        remaining_pity = get_remaining_pity(user_id)
    
    return {
        "results": results,
        "remaining_pity": remaining_pity,
        "got_3star": got_3star
    }


# ========== 图片合成 ==========
def get_level_image(stars: int, layer_type: str) -> str:
    """
    获取星级框或背景图片
    stars: 1, 2, 3
    layer_type: "bg" (背景) 或 "frame" (框)
    """
    # 星级索引 (0=1星, 1=2星, 2=3星)
    star_idx = stars - 1

    # 背景和框的索引
    layer_idx = 0 if layer_type == "bg" else 1

    # 文件名模式
    filename = f"gacha_tmb_{star_idx:02d}_{layer_idx:02d}"

    # 3星背景特殊处理（有_b后缀）
    if stars == 3 and layer_type == "bg":
        filename += "_b"

    path = LEVEL_DIR / f"{filename}.png"
    if path.exists():
        return str(path)
    else:
        log_error(f"找不到星级图片: {path}")
        return None


def find_attribute_icon(attribute: str) -> str:
    """根据属性名称查找属性图标文件"""
    if not attribute:
        return None
    
    # 属性图标文件名模式（在level文件夹中）
    attr_name = str(attribute).strip()
    
    # 尝试多种文件名模式
    patterns = [
        f"common_tmb_label_element_{attr_name}.png",  # 用户指定的格式
        f"attr_{attr_name}.png",
        f"attribute_{attr_name}.png",
        f"type_{attr_name}.png",
        f"{attr_name}_attr.png",
        f"{attr_name}.png"
    ]
    
    for pattern in patterns:
        path = LEVEL_DIR / pattern
        if path.exists():
            return str(path)
    
    log_error(f"找不到属性图标: attribute={attribute}, 尝试的模式: {patterns}")
    return None


def find_type_icon(card_type: str) -> str:
    """根据卡牌类型查找Battle/Assist图标文件"""
    if not card_type:
        return None
    
    # 类型图标文件名模式（在level文件夹中）
    type_name = str(card_type).strip().lower()
    
    # 尝试多种文件名模式
    patterns = [
        f"battle_{type_name}.png",      # battle_battle.png, battle_assist.png
        f"{type_name}_icon.png",
        f"{type_name}.png",
        f"{type_name}.png"
    ]
    
    for pattern in patterns:
        path = LEVEL_DIR / pattern
        if path.exists():
            return str(path)
    
    log_error(f"找不到类型图标: type={card_type}, 尝试的模式: {patterns}")
    return None


def composite_card(character: dict) -> bytes:
    """
    合成卡牌图片：背景 + 属性图标 + Battle/Assist图标 + 角色图标 + 框
    返回PNG格式的字节数据
    
    角色图裁剪区域：
    - 宽度：25% ~ 75%（中间50%）
    - 高度：15% ~ 65%（中间50%）
    裁剪后压缩到框内显示
    
    图层顺序（从后到前）：
    1. 背景图
    2. 属性图标（框的左下角）
    3. 角色图
    4. Battle/Assist图标（角色图最下方）
    5. 框（最顶层）
    """
    stars = character["stars"]
    icon_path = character.get("icon_path") or find_character_icon(
        character["chara_id"], stars
    )

    if not icon_path or not os.path.exists(icon_path):
        log_error(f"找不到角色图标: icon_path={icon_path}, chara_id={character['chara_id']}, stars={stars}")
        # 返回一个空白图片作为占位
        img = Image.new('RGBA', (200, 300), (100, 100, 100, 255))
        bio = BytesIO()
        img.save(bio, format='JPEG', optimize=True, quality=60)
        return bio.getvalue()

    try:
        # 加载各层图片
        bg_path = get_level_image(stars, "bg")
        frame_path = get_level_image(stars, "frame")

        if not bg_path or not frame_path:
            log_error(f"背景或框图片不存在: bg={bg_path}, frame={frame_path}")
            return None

        bg_img = Image.open(bg_path).convert('RGBA')
        frame_img = Image.open(frame_path).convert('RGBA')
        char_img = Image.open(icon_path).convert('RGBA')
        
        # 加载属性图标（根据角色属性）
        attribute = character.get("attribute")
        attr_icon_path = find_attribute_icon(attribute) if attribute else None
        attr_img = None
        if attr_icon_path and os.path.exists(attr_icon_path):
            attr_img = Image.open(attr_icon_path).convert('RGBA')
        
        # 加载Battle/Assist图标（根据角色类型）
        card_type = character.get("type", "battle")
        type_icon_path = find_type_icon(card_type)
        type_img = None
        if type_icon_path and os.path.exists(type_icon_path):
            type_img = Image.open(type_icon_path).convert('RGBA')

        # 获取各层尺寸
        bg_width, bg_height = bg_img.size
        frame_width, frame_height = frame_img.size
        
        # ========== 裁剪角色图的指定区域 ==========
        # 裁剪区域：宽度25%~75%，高度15%~65%（使用预定义常量）
        char_width, char_height = char_img.size
        crop_left = int(char_width * CROP_LEFT_RATIO)
        crop_right = int(char_width * CROP_RIGHT_RATIO)
        crop_top = int(char_height * CROP_TOP_RATIO)
        crop_bottom = int(char_height * CROP_BOTTOM_RATIO)
        
        # 执行裁剪
        char_img_cropped = char_img.crop((crop_left, crop_top, crop_right, crop_bottom))
        
        # ========== 缩放裁剪后的图片以适应框 ==========
        cropped_width, cropped_height = char_img_cropped.size
        cropped_ratio = cropped_width / cropped_height
        frame_ratio = frame_width / frame_height

        if cropped_ratio > frame_ratio:
            # 裁剪后的图片更宽，以框宽度为基准缩放
            char_target_width = int(frame_width * 0.85)
            char_target_height = int(char_target_width / cropped_ratio)
        else:
            # 裁剪后的图片更高，以框高度为基准缩放
            char_target_height = int(frame_height * 0.85)
            char_target_width = int(char_target_height * cropped_ratio)

        char_img_resized = char_img_cropped.resize(
            (char_target_width, char_target_height),
            Image.Resampling.LANCZOS
        )

        # 创建输出画布（使用背景尺寸）
        output = Image.new('RGBA', (bg_width, bg_height), (0, 0, 0, 0))

        # 逐层合成（从后到前）：
        # 1. 背景
        # 2. 角色图
        # 3. 框
        # 4. 属性图标（框的左下角，在框之上）
        # 5. Battle/Assist图标（卡牌最底部，在框之上）
        
        # 角色图居中放置
        char_x = (bg_width - char_target_width) // 2
        char_y = (bg_height - char_target_height) // 2

        output.paste(bg_img, (0, 0))
        output.paste(char_img_resized, (char_x, char_y), char_img_resized)
        
        # 先粘贴框
        output.paste(frame_img, (0, 0), frame_img)
        
        # 添加属性图标（框的左下角，在框之上）
        if attr_img:
            # 属性图标保持原始大小，不压缩
            attr_width, attr_height = attr_img.size
            # 放在框的左下角（留一点边距）
            attr_x = 5
            attr_y = bg_height - attr_height - 5
            output.paste(attr_img, (attr_x, attr_y), attr_img)
        
        # 添加Battle/Assist图标（卡牌最底部，在框之上）
        if type_img:
            # 不压缩，保持原始大小
            type_width, type_height = type_img.size
            # 放在卡牌最底部居中，距离底部0像素
            type_x = (bg_width - type_width) // 2
            type_y = bg_height - type_height  # 距离底部0像素
            output.paste(type_img, (type_x, type_y), type_img)

        # 转换为RGB（去除alpha通道）用于发送
        output_rgb = Image.new('RGB', (bg_width, bg_height), (255, 255, 255))
        output_rgb.paste(output, (0, 0), output)

        # 保存到BytesIO（使用压缩优化，减小文件大小）
        bio = BytesIO()
        output_rgb.save(bio, format='JPEG', optimize=True, quality=60)
        return bio.getvalue()

    except Exception as e:
        log_error(f"合成卡牌图片失败: {e}")
        return None


def save_card_image(card_data: list, output_idx: int) -> str:
    """保存抽卡结果图片"""
    if not card_data:
        return None

    try:
        if len(card_data) == 1:
            # 单抽：直接保存单张
            img_bytes = composite_card(card_data[0])
            if not img_bytes:
                return None

            filename = f"gacha_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            filepath = OUTPUT_DIR / filename
            with open(filepath, 'wb') as f:
                f.write(img_bytes)
            return str(filepath)
        else:
            # 十连：合成一张大图（2行5列）
            card_images = []
            for char in card_data:
                img_bytes = composite_card(char)
                if img_bytes:
                    card_images.append(Image.open(BytesIO(img_bytes)))

            if not card_images:
                return None

            # 每张卡的尺寸
            card_width, card_height = card_images[0].size

            # 创建2行5列的大图
            gap = 10
            total_width = card_width * 5 + gap * 4
            total_height = card_height * 2 + gap

            # ========== 尝试加载十连背景图片 ==========
            # 查找level目录下的十连背景图片
            bg_path = None
            for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
                test_path = LEVEL_DIR / bg_name
                if test_path.exists():
                    bg_path = str(test_path)
                    break
            
            if bg_path:
                # 使用找到的背景图片
                log_info(f"使用十连背景图片: {bg_path}")
                bg_img = Image.open(bg_path).convert('RGB')
                
                # 将背景图片压缩到原来的50%
                bg_w, bg_h = bg_img.size
                compressed_w = int(bg_w * 0.5 * 0.264)
                compressed_h = int(bg_h * 0.5 * 0.264)
                bg_img_compressed = bg_img.resize((compressed_w, compressed_h), Image.Resampling.LANCZOS)
                
                # 使用压缩后的背景尺寸作为画布尺寸
                output = Image.new('RGB', (compressed_w, compressed_h), (50, 50, 50))
                output.paste(bg_img_compressed, (0, 0))
                
                # 计算卡牌区域在背景中的居中位置
                cards_x = (compressed_w - total_width) // 2
                cards_y = (compressed_h - total_height) // 2
            else:
                # 使用默认背景色
                output = Image.new('RGB', (total_width, total_height), (50, 50, 50))
                cards_x = 0
                cards_y = 0

            # 将卡牌粘贴到背景上（居中放置）
            for idx, img in enumerate(card_images):
                row = idx // 5
                col = idx % 5
                x = cards_x + col * (card_width + gap)
                y = cards_y + row * (card_height + gap)
                output.paste(img, (x, y))

            filename = f"gacha_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            filepath = OUTPUT_DIR / filename
            output.save(filepath, format='JPEG', optimize=True, quality=60)
            return str(filepath)

    except Exception as e:
        log_error(f"保存卡牌图片失败: {e}")
        return None


# ========== Napcat消息处理 ==========
@app.route('/', methods=['POST'])
def handle_message():
    """
    处理Napcat转发的QQ消息
    消息格式为JSON
    """
    try:
        data = request.json

        # 解析消息
        raw_message = data.get('raw_message', '') or data.get('message', '')
        user_id = data.get('user_id', 'unknown')
        group_id = data.get('group_id', None)
        message_id = data.get('message_id', None)

        log_info(f"MSG uid={user_id} gid={group_id}: {raw_message[:60]}")

        # 记录今日活跃用户
        if user_id and user_id != 'unknown':
            DAILY_USER_SET.add(str(user_id))

        # 实时保存每日统计数据到磁盘
        save_daily_stats()

        # 获取Bot自己的QQ号（从数据中提取或配置）
        self_id = data.get('self_id', None)
        
        # 判断是否被艾特
        is_at = False
        if group_id:
            # 群聊：检查是否包含艾特CQ码
            # CQ码格式: [CQ:at,qq=123456789] 或 @昵称
            if f'[CQ:at,qq={self_id}]' in raw_message or f'@' in raw_message:
                is_at = True
        
        # 如果是群聊且没有被艾特，忽略消息
        if group_id and not is_at:
            log_info(f"群聊消息未艾特bot，忽略: {raw_message[:30]}")
            return jsonify({"status": "ignored", "reason": "not_at"})

        # 检查是否有正在进行的盲盒会话
        if has_box_session(user_id):
            # 清理原始消息，移除艾特信息和多余空格，准备判断是否是开箱命令
            cleaned_message = raw_message
            # 移除艾特CQ码
            cleaned_message = cleaned_message.replace(f'[CQ:at,qq={self_id}]', '').strip()
            # 移除@符号（如果有的话）
            cleaned_message = cleaned_message.replace('@', '').strip()
            
            # 处理盲盒开启请求
            # 检查是否是开箱命令（必须是明确的命令才处理）
            open_commands = ['全部开', '剩下的全部开']
            is_open_command = any(cmd in cleaned_message for cmd in open_commands)
            
            # 检查是否是数字输入（支持"选择1"格式）
            import re
            has_valid_input = False
            # 先检查是否是纯数字
            stripped_input = cleaned_message.strip().replace('，', ',').replace(' ', '').replace('\u3000', '').replace(',', '')
            if stripped_input.isdigit():
                has_valid_input = True
            else:
                # 检查是否包含"选择"开头后跟数字
                match = re.search(r'选择[0-9]+', cleaned_message)
                if match:
                    has_valid_input = True
                # 或者直接提取数字
                else:
                    num_match = re.search(r'[0-9]+', cleaned_message)
                    if num_match:
                        has_valid_input = True
            
            if is_open_command or has_valid_input:
                return handle_box_open(user_id, group_id, cleaned_message)
            else:
                # 有盲盒会话但输入无效，提示用户
                session = get_box_session(user_id)
                remaining = len(session["boxes"]) - len(session["opened"])
                reply = f"你还有{remaining}个未开！请输入要开的编号（如1、选择1）或「全部开」"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "pending", "message": "waiting for box open"})

        # ===== 管理员命令 (仅 ADMIN_QQ 可用) =====
        if str(user_id) == str(ADMIN_QQ):
            import re as re_admin
            at_match_admin = re_admin.search(r'\[CQ:at,qq=(\d+)\]', raw_message)

            # 呱太
            amount_gacha = re_admin.search(r'(\d+)\s*呱太', raw_message)
            if at_match_admin and amount_gacha:
                target_qq = at_match_admin.group(1)
                amount = int(amount_gacha.group(1))

                if '扣' in raw_message or '减' in raw_message:
                    success = spend_gacha(target_qq, amount)
                    action = "扣除" if success else "扣除失败(余额不足)"
                else:
                    add_gacha(target_qq, amount)
                    action = "增加"
                    success = True

                new_balance = get_gacha_count(target_qq)
                reply = f"✅ 管理员操作完成: {action} {target_qq} {amount}呱太 (余额: {new_balance})"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success", "message": f"admin: {action}呱太"})

            # 蓝碎片
            amount_blue = re_admin.search(r'(\d+)\s*蓝碎片', raw_message)
            if at_match_admin and amount_blue:
                target_qq = at_match_admin.group(1)
                amount = int(amount_blue.group(1))

                if '扣' in raw_message or '减' in raw_message:
                    success = spend_blue_crystal(target_qq, amount)
                    action = "扣除" if success else "扣除失败(余额不足)"
                else:
                    add_blue_crystal(target_qq, amount)
                    action = "增加"
                    success = True

                new_balance = get_blue_crystal(target_qq)
                reply = f"✅ 管理员操作完成: {action} {target_qq} {amount}蓝碎片 (余额: {new_balance})"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success", "message": f"admin: {action}蓝碎片"})

            # 红碎片
            amount_red = re_admin.search(r'(\d+)\s*红碎片', raw_message)
            if at_match_admin and amount_red:
                target_qq = at_match_admin.group(1)
                amount = int(amount_red.group(1))

                if '扣' in raw_message or '减' in raw_message:
                    success = spend_red_crystal(target_qq, amount)
                    action = "扣除" if success else "扣除失败(余额不足)"
                else:
                    add_red_crystal(target_qq, amount)
                    action = "增加"
                    success = True

                new_balance = get_red_crystal(target_qq)
                reply = f"✅ 管理员操作完成: {action} {target_qq} {amount}红碎片 (余额: {new_balance})"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success", "message": f"admin: {action}红碎片"})

        # 处理私聊和群聊消息（群聊已确保被艾特）
        # 支持"十连"/"单抽"/"帮助"/"获取呱太"/"签到"/"个人记录"/"兑换呱太"
        if '十连' in raw_message or 'gacha10' in raw_message.lower():
            return handle_gacha(10, user_id, group_id, message_id)
        elif '单抽' in raw_message or 'gacha' in raw_message.lower():
            return handle_gacha(1, user_id, group_id, message_id)
        elif '帮助' in raw_message or 'help' in raw_message.lower():
            return handle_help(user_id, group_id)
        elif '获取呱太' in raw_message or 'getgacha' in raw_message.lower():
            return handle_get_gacha(user_id, group_id)
        elif '签到' in raw_message or 'signin' in raw_message.lower():
            return handle_signin(user_id, group_id)
        elif '队伍' in raw_message or '配队' in raw_message.lower():
            return handle_team(user_id, group_id, raw_message)
        elif '下一页' in raw_message:
            return handle_personal_info(user_id, group_id, page_action='next')
        elif '上一页' in raw_message:
            return handle_personal_info(user_id, group_id, page_action='prev')
        elif '个人记录' in raw_message or '记录' in raw_message.lower():
            return handle_personal_info(user_id, group_id)
        elif '兑换呱太' in raw_message or '兑换' in raw_message.lower():
            return handle_exchange_crystal(user_id, group_id)
        elif '排行榜' in raw_message or '排行' in raw_message.lower():
            return show_ranking(user_id, group_id)
        elif '详细信息' in raw_message:
            return handle_show_details(user_id, group_id)
        elif '战斗日志' in raw_message:
            return handle_battle_log(user_id, group_id)
        elif '三王女' in raw_message:
            return handle_sannoujo(user_id, group_id)
        elif '三星池子' in raw_message or '红抽' in raw_message or '蓝抽' in raw_message:
            return handle_3star_pool(user_id, group_id, raw_message)
        elif '挑战' in raw_message:
            # 解析挑战排名
            import re
            rank_match = re.search(r'挑战\s*(\d+)', raw_message)
            if rank_match:
                target_rank = int(rank_match.group(1))
                return challenge_rank(user_id, group_id, target_rank)
            else:
                reply = "请指定挑战的排名！格式：挑战 1~10"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "未指定排名"})
        elif 'BOSS战' in raw_message or 'boss战' in raw_message:
            return handle_boss_battle(user_id, group_id, raw_message)
        elif '战斗' in raw_message or '对战' in raw_message or '决斗' in raw_message:
            return handle_battle(user_id, group_id, raw_message)
        
        return jsonify({"status": "ignored"})

    except Exception as e:
        log_error(f"处理消息失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_gacha(count: int, user_id: str, group_id, message_id, auto_open: bool = False):
    """处理抽卡请求（盲盒模式）"""
    try:
        # 检查呱太数量
        cost = GACHA10_COST if count == 10 else GACHA_COST
        current_gacha = get_gacha_count(user_id)
        
        if current_gacha < cost:
            reply = f"呱太不足！当前呱太: {current_gacha}，需要: {cost}。请艾特我发送「获取呱太」获得10000呱太~"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"抽卡失败 [{user_id}]: 呱太不足 {current_gacha}/{cost}")
            return jsonify({
                "status": "error",
                "message": "呱太不足",
                "current_gacha": current_gacha,
                "required_gacha": cost
            })

        # 十连冷却检查：一分钟内每个用户至多只能抽一次十连
        if count == 10:
            now_ts = datetime.now().timestamp()
            last_gacha10 = GACHA10_COOLDOWN.get(user_id, 0)
            remaining_cooldown = GACHA10_COOLDOWN_SECONDS - (now_ts - last_gacha10)
            if remaining_cooldown > 0:
                reply = f"十连冷却中！请等待 {int(remaining_cooldown)} 秒后再试~"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                log_info(f"十连冷却 [{user_id}]: 还需等待 {int(remaining_cooldown)} 秒")
                return jsonify({
                    "status": "error",
                    "message": "十连冷却中",
                    "remaining_cooldown": int(remaining_cooldown)
                })

        # 消耗呱太
        spend_gacha(user_id, cost)

        # 十连冷却记录：更新最后一次十连时间戳
        if count == 10:
            GACHA10_COOLDOWN[user_id] = datetime.now().timestamp()

        # 使用预加载的角色数据
        characters = get_characters()
        if not characters:
            return jsonify({
                "status": "error",
                "message": "无法加载角色数据"
            })

        # 检查是否触发フェス保底
        pity_data = load_pity_data(user_id)
        fes_pity_count = pity_data.get("fes_pity_count", 0)
        is_fes_pity = fes_pity_count >= PITY_LIMIT
        
        if is_fes_pity:
            log_info(f"用户 {user_id} 触发フェス限定三星保底！")

        # 生成盲盒
        boxes = []
        has_2star = False  # 十连保底二星标记
        
        for i in range(count):
            # 如果是保底的最后一抽，强制触发フェス保底
            is_pity_draw = is_fes_pity and (i == count - 1)
            box = draw_mystery_box(characters, user_id, is_pity_draw)
            
            # 记录是否有二星
            if box["stars"] == 2:
                has_2star = True
            
            boxes.append(box)
        
        # 十连保底：如果没有二星，将最后一个非保底盲盒改为二星
        if count == 10 and not has_2star and not is_fes_pity:
            # 找到最后一个可以修改的盲盒（非黑色盲盒且不是保底）
            for i in range(len(boxes)-1, -1, -1):
                if not boxes[i].get("is_mystery", False):
                    # 改为二星
                    boxes[i]["stars"] = 2
                    # 获取二星角色
                    available = [c for c in characters if c["stars"] == 2]
                    if available:
                        boxes[i]["character"] = random.choice(available)
                    else:
                        boxes[i]["character"] = random.choice(characters)
                    log_info(f"十连保底触发，将第{i+1}个盲盒改为二星")
                    break
        
        # 创建盲盒会话
        create_box_session(user_id, boxes)
        BOX_SESSIONS[user_id]["characters"] = characters
        BOX_SESSIONS[user_id]["is_fes_pity"] = is_fes_pity  # 标记FES保底

        # 生成唯一标识
        output_idx = random.randint(1000, 9999)

        # 生成盲盒图片
        box_images = []
        for box in boxes:
            img_bytes = create_box_card(box, characters)
            if img_bytes:
                box_images.append(Image.open(BytesIO(img_bytes)))

        if not box_images:
            log_error("生成盲盒图片失败")
            clear_box_session(user_id)
            return jsonify({"status": "error", "message": "生成盲盒图片失败"})

        # 合成盲盒大图
        card_width, card_height = box_images[0].size
        if count == 1:
            img_path = OUTPUT_DIR / f"gacha_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            box_images[0].convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)
        else:
            # 十连：使用背景图片
            gap = 10
            cols = 5
            rows = (count + 4) // 5
            cards_total_width = card_width * cols + gap * (cols - 1)
            cards_total_height = card_height * rows + gap * (rows - 1)
            
            # 尝试加载背景图片
            bg_path = None
            for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
                test_path = LEVEL_DIR / bg_name
                if test_path.exists():
                    bg_path = str(test_path)
                    break
            
            if bg_path:
                # 使用背景图片，放大到原来的2倍
                log_info(f"使用十连背景图片: {bg_path}")
                bg_img = Image.open(bg_path).convert('RGB')
                
                # 将背景放大到原来的2倍
                bg_w, bg_h = bg_img.size
                final_w = int(bg_w * 2.0 * 0.264)
                final_h = int(bg_h * 2.0 * 0.264)
                bg_img_resized = bg_img.resize((final_w, final_h), Image.Resampling.LANCZOS)
                
                # 创建最终画布
                output = Image.new('RGB', (final_w, final_h), (50, 50, 50))
                output.paste(bg_img_resized, (0, 0))
                
                # 计算卡牌居中位置
                cards_x = (final_w - cards_total_width) // 2
                cards_y = (final_h - cards_total_height) // 2
            else:
                # 没有背景，使用纯色背景
                output = Image.new('RGB', (cards_total_width, cards_total_height), (50, 50, 50))
                cards_x = 0
                cards_y = 0

            # 粘贴卡牌
            for i, img in enumerate(box_images):
                row = i // cols
                col = i % cols
                x = cards_x + col * (card_width + gap)
                y = cards_y + row * (card_height + gap)
                output.paste(img, (x, y))

            img_path = OUTPUT_DIR / f"gacha_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            output.convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)

        # 获取用户信息
        current_gacha = get_gacha_count(user_id)
        total_draws = get_total_draws(user_id)
        total_3stars = get_total_3stars(user_id)
        remaining_pity = get_remaining_pity(user_id)

        # 构建提示消息，包含呱太和抽数信息
        box_hints = [f"选择{i+1}" for i in range(count)]
        hint_text = " | ".join(box_hints[:5])
        if count > 5:
            hint_text += "\n" + " | ".join(box_hints[5:10])
        
        # 检查是否触发FES保底（is_fes_pity已在前面定义）
        fes_pity_text = ""
        if is_fes_pity:
            fes_pity_text = "🎉 触发FES保底，今日时运为man！\n"
        
        info_text = (
            f"当前呱太: {current_gacha}\n"
            f"累计次数: {total_draws}\n"
            f"累计3星: {total_3stars}\n"
            f"距离FES保底: {remaining_pity}次\n"
        )
        prompt_text = f"{fes_pity_text}{info_text}\n请输入要开的编号：\n{hint_text}\n输入「全部开」一键开启"

        # 发送消息（文字和图片合成一条消息
        if group_id and user_id:
            at_message = f"[CQ:at,qq={user_id}] "
        else:
            at_message = ""

        # 读取图片并合成到文字消息
        with open(img_path, "rb") as f:
            image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
        img_message = f"[CQ:image,file=base64://{image_base64}]"
        # 如果是自动开箱模式，跳过盲盒阶段直接开启所有盲盒
        if auto_open:
            # 直接调用开箱函数，开启所有盲盒
            result = handle_box_open(user_id, group_id, "全部开")
            # 删除本地盲盒图片
            try:
                os.remove(img_path)
            except:
                pass
            return result
        
        full_message = f"{at_message}{prompt_text}\n{img_message}"
        send_message(full_message, user_id, group_id)

        # 删除本地图片
        try:
            os.remove(img_path)
        except:
            pass

        log_info(f"盲盒生成 [{user_id}]: {count}个盲盒")

        return jsonify({
            "status": "success",
            "box_count": count,
            "message": prompt_text
        })

    except Exception as e:
        log_error(f"抽卡处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_sannoujo(user_id: str, group_id):
    """处理三王女命令，输出十个ID为207832001的三星卡"""
    try:
        characters = get_characters()
        if not characters:
            reply = "无法加载角色数据！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "无法加载角色数据"})
        
        # 查找ID为207832001的角色
        target_card = next((c for c in characters if str(c.get("card_id")) == "207832001"), None)
        if not target_card:
            reply = "找不到ID为207832001的角色！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "找不到目标角色"})
        
        # 生成10张相同的卡
        card_imgs = []
        for _ in range(10):
            card_bytes = composite_card(target_card)
            if card_bytes:
                from io import BytesIO
                card_img = Image.open(BytesIO(card_bytes))
                card_imgs.append(card_img)
        
        if not card_imgs:
            reply = "无法生成卡片图片！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "无法生成卡片图片"})
        
        # 创建展示图片（使用十连背景，与抽卡一致）
        bg_path = None
        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
            test_path = LEVEL_DIR / bg_name
            if test_path.exists():
                bg_path = str(test_path)
                break
        
        if bg_path:
            # 使用背景图片，放大到原来的2倍
            bg_img = Image.open(bg_path).convert('RGB')
            bg_w, bg_h = bg_img.size
            final_w = int(bg_w * 1.5 * 0.264)
            final_h = int(bg_h * 1.5 * 0.264)
            bg = bg_img.resize((final_w, final_h), Image.LANCZOS)
            bg_w, bg_h = final_w, final_h
        else:
            bg = Image.new('RGB', (600, 400), (50, 50, 50))
            bg_w, bg_h = bg.size
        
        # 使用第一张卡的实际比例作为基准
        first_card = card_imgs[0]
        orig_w, orig_h = first_card.size
        max_card_width = 90
        max_card_height = 120
        scale = min(max_card_width / orig_w, max_card_height / orig_h)
        card_width = int(orig_w * scale)
        card_height = int(orig_h * scale)
        
        gap = 10
        cols = 5
        rows = 2
        
        total_w = cols * (card_width + gap) - gap
        total_h = rows * (card_height + gap) - gap
        start_x = (bg_w - total_w) // 2
        start_y = (bg_h - total_h) // 2
        
        # 创建最终画布（与抽卡一致）
        output = Image.new('RGB', (bg_w, bg_h), (50, 50, 50))
        output.paste(bg, (0, 0))
        
        for i, img in enumerate(card_imgs):
            col = i % cols
            row = i // cols
            x = start_x + col * (card_width + gap)
            y = start_y + row * (card_height + gap)
            
            # 使用thumbnail保持原比例
            img_copy = img.copy()
            img_copy.thumbnail((card_width, card_height), Image.LANCZOS)
            thumb_w, thumb_h = img_copy.size
            
            # 居中粘贴
            paste_x = x + (card_width - thumb_w) // 2
            paste_y = y + (card_height - thumb_h) // 2
            
            output.paste(img_copy, (paste_x, paste_y))
        
        # 保存图片
        import random
        output_idx = random.randint(1000, 9999)
        img_path = OUTPUT_DIR / f"sannoujo_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
        output.convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)
        
        # 发送消息
        at_message = f"[CQ:at,qq={user_id}] " if group_id and user_id else ""
        with open(img_path, "rb") as f:
            image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
        img_message = f"[CQ:image,file=base64://{image_base64}]"
        full_message = f"{at_message}👑 三王女降临！\n{img_message}"
        send_message(full_message, user_id, group_id)
        
        # 删除临时图片
        try:
            img_path.unlink()
        except:
            pass
        
        log_info(f"三王女命令 [{user_id}]")
        return jsonify({"status": "success", "message": "三王女卡生成成功"})
    
    except Exception as e:
        log_error(f"三王女命令失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_box_open(user_id: str, group_id: str, open_input: str):
    """处理盲盒开启请求"""
    try:
        session = get_box_session(user_id)
        if not session:
            return None  # 没有盲盒会话，不回复

        boxes = session["boxes"]
        characters = session.get("characters") or get_characters()
        opened = session["opened"].copy()

        # 验证输入
        valid, result = is_valid_box_index(boxes, open_input, session["opened"])
        if not valid:
            reply = f"输入无效：{result}\n请输入数字（如1、2、3）或「全部开」"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": result})

        indices = result
        # 过滤掉已经开过的
        indices = [i for i in indices if i not in opened]
        
        if not indices:
            reply = "这些都已经开过了！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "都已开过"})

        # 记录信息
        new_opened = []
        opened_results = []
        mutation_messages = []  # 收集突变信息

        # 记录碎片获得
        red_crystal_gained = 0
        blue_crystal_gained = 0
        
        # 开盲盒
        log_info(f"用户 {user_id} 开启盲盒，共 {len(indices)} 个，索引: {indices}")
        for idx in indices:
            box = boxes[idx]
            log_info(f"处理盲盒 {idx+1}: stars={box.get('stars')}, is_mystery={box.get('is_mystery')}, opened={box.get('opened')}")
            
            # 先开黑色盲盒获取角色
            if box["is_mystery"] and not box.get("opened"):
                box = open_mystery_box(box, characters)
                boxes[idx] = box
            
            # 应用突变（仅非黑色盲盒），突变后重新抽对应星级的卡
            box, mutated, mutation_text = apply_mutation(box, characters)
            boxes[idx] = box
            
            # 记录突变信息
            if mutated and mutation_text:
                mutation_messages.append(f"#{idx+1}发生了{mutation_text}突变！")
            
            # 获取角色（确保有角色）
            character = box.get("character")
            if not character:
                # 正常盲盒需要获取角色
                stars = box["stars"]
                log_info(f"盲盒 {idx+1} 没有角色，stars={stars}，需要随机选择")
                available = [c for c in characters if c["stars"] == stars]
                if available:
                    character = random.choice(available)
                else:
                    character = random.choice(characters)
                box["character"] = character
                boxes[idx] = box
            
            stars = box["stars"]
            log_info(f"盲盒 {idx+1} 最终 stars={stars}，character={character.get('name') if character else 'None'}")
            
            # 更新卡片收藏和数量统计
            card_id = str(character.get("card_id", ""))
            limit_type = character.get("limit_type", "")
            chara_name = character.get("name", "")
            
            # FES统计和提示
            fes_message = ""
            if stars == 3 and limit_type == "フェス限定":
                # 增加FES统计
                fes_count = increment_fes_count(card_id, chara_name)
                fes_message = f"✨ 恭喜！这是全服第{fes_count}个「{chara_name}」！"
            
            # 碎片转化逻辑
            if stars == 1:
                # 1星卡转化为红色碎片
                add_red_crystal(user_id, 1)
                red_crystal_gained += 1
                log_info(f"用户 {user_id} 获得红色碎片 +1（1星卡）")
            elif stars == 2:
                # 2星卡转化为蓝色碎片
                add_blue_crystal(user_id, 1)
                blue_crystal_gained += 1
                log_info(f"用户 {user_id} 获得蓝色碎片 +1（2星卡）")
                # 更新二星数量
                add_card_collection(user_id, card_id, chara_name, stars, limit_type)
            elif stars == 3:
                # 3星卡添加到最近获得记录和收藏
                add_recent_3star(
                    user_id,
                    card_id,
                    str(character.get("chara_id", "")),
                    chara_name,
                    limit_type
                )
                add_card_collection(user_id, card_id, chara_name, stars, limit_type)
            
            # 更新抽卡记录
            got_3star = (stars == 3)
            is_fes_3star = got_3star and (limit_type == "フェス限定")
            update_pity(user_id, got_3star, is_fes_3star)
            
            # 如果是FES角色，添加到结果中
            if fes_message:
                opened_results.append({
                    "index": idx + 1,
                    "stars": stars,
                    "name": chara_name,
                    "fes_message": fes_message
                })
            else:
                opened_results.append({
                    "index": idx + 1,
                    "stars": stars,
                    "name": chara_name
                })
            
            new_opened.append(idx)

        # 更新会话
        session["boxes"] = boxes
        session["opened"].extend(new_opened)

        # 构建结果文本（不显示单个卡片的碎片转化）
        stars_display = {1: "⭐", 2: "⭐⭐", 3: "⭐⭐⭐"}
        result_lines = []
        for r in opened_results:
            result_lines.append(f"{r['index']}. {stars_display.get(r['stars'], '⭐')} {r['name']}")

        result_text = "\n".join(result_lines)
        
        # 添加碎片获得汇总
        crystal_summary = ""
        if red_crystal_gained > 0 or blue_crystal_gained > 0:
            parts = []
            if red_crystal_gained > 0:
                parts.append(f"🔴红色碎片 x{red_crystal_gained}")
            if blue_crystal_gained > 0:
                parts.append(f"🔵蓝色碎片 x{blue_crystal_gained}")
            crystal_summary = f"\n本次获得: {' + '.join(parts)}"

        # 生成汇总图片（已开的显示角色，未开的不变）
        all_opened = session["opened"]
        summary_img_bytes = create_box_summary_image(boxes, all_opened, characters)
        
        img_path = None
        if summary_img_bytes:
            output_idx = random.randint(1000, 9999)
            img_path = OUTPUT_DIR / f"gacha_open_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            with open(img_path, "wb") as f:
                f.write(summary_img_bytes)

        # 发送结果（突变提示、抽卡结果、图片合成一条消息）
        if group_id and user_id:
            at_message = f"[CQ:at,qq={user_id}] "
        else:
            at_message = ""

        # 保存详细结果到会话
        session["last_results"] = {
            "results": opened_results,
            "mutations": mutation_messages,
            "red_crystal": red_crystal_gained,
            "blue_crystal": blue_crystal_gained
        }

        # 检查是否全部开完
        remaining = len(boxes) - len(all_opened)
        if remaining > 0:
            remaining_hint = f"\n还有{remaining}个未开，输入「剩下的全部开」可以一键开启"
            box_hints = [f"选择{i+1}" for i in range(len(boxes)) if i not in all_opened]
            hint = " | ".join(box_hints[:5])
            if len(box_hints) > 5:
                hint += "\n" + " | ".join(box_hints[5:])
            remaining_hint += f"\n{hint}"
        else:
            remaining_hint = "\n所有已开完！"
            # 延迟清除会话，给用户时间查看详细信息
            # 不在此处清除，让详细信息查询可以访问

        # 收集FES消息
        fes_messages = [r.get("fes_message") for r in opened_results if r.get("fes_message")]
        fes_text = "\n".join(fes_messages) if fes_messages else ""
        
        # 合成消息（不显示详细文字信息，只显示图片和简短提示）
        short_text = f"开了{len(new_opened)}个！{fes_text}{crystal_summary}{remaining_hint}\n输入「详细信息」查看详情"

        # 合成一条消息发送
        if img_path:
            with open(img_path, "rb") as f:
                image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
            img_message = f"[CQ:image,file=base64://{image_base64}]"
            full_message = f"{at_message}{short_text}\n{img_message}"
            send_message(full_message, user_id, group_id)

            try:
                os.remove(img_path)
            except:
                pass
        else:
            # 没有图片时，只发送文字
            send_message(f"{at_message}{complete_text}", user_id, group_id)

        log_info(f"盲盒开启 [{user_id}]: 开了{len(new_opened)}个")

        return jsonify({
            "status": "success",
            "opened_count": len(new_opened),
            "results": opened_results,
            "remaining": remaining
        })

    except Exception as e:
        log_error(f"开箱处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_get_gacha(user_id: str, group_id):
    """处理获取呱太请求"""
    try:
        # 添加呱太
        new_gacha = add_gacha(user_id, GET_GACHA_REWARD)
        
        # 构建回复消息
        reply = f"成功获得 {GET_GACHA_REWARD} 呱太！当前呱太: {new_gacha}"
        
        # 发送消息
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        
        log_info(f"获取呱太 [{user_id}]: +{GET_GACHA_REWARD} -> {new_gacha}")
        
        return jsonify({
            "status": "success",
            "user_id": user_id,
            "gacha_added": GET_GACHA_REWARD,
            "current_gacha": new_gacha
        })
    
    except Exception as e:
        log_error(f"获取呱太处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_signin(user_id: str, group_id):
    """处理签到请求"""
    try:
        # 执行签到
        result = do_signin(user_id)
        
        if result["success"]:
            # 签到成功
            streak = result["streak"]
            current_gacha = get_gacha_count(user_id)
            
            if streak > 1:
                reply = f"🎉 签到成功！获得 {result['reward']} 呱太！\n当前呱太: {current_gacha}\n连续签到: {streak} 天"
            else:
                reply = f"🎉 签到成功！获得 {result['reward']} 呱太！\n当前呱太: {current_gacha}"
        else:
            # 今天已经签到过
            current_gacha = get_gacha_count(user_id)
            streak = result["streak"]
            reply = f"今天已经签到过了！当前呱太: {current_gacha}\n连续签到: {streak} 天"
        
        # 发送消息
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        
        log_info(f"签到 [{user_id}]: success={result['success']}, streak={result['streak']}")
        
        return jsonify({
            "status": "success" if result["success"] else "already_signed",
            "user_id": user_id,
            "streak": result["streak"],
            "reward": result.get("reward", 0),
            "current_gacha": get_gacha_count(user_id)
        })
    
    except Exception as e:
        log_error(f"签到处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_personal_info(user_id: str, group_id, page_action: str = None):
    """处理个人记录查询请求（支持分页）"""
    try:
        # 获取用户数据
        total_draws = get_total_draws(user_id)
        total_3stars = get_total_3stars(user_id)
        red_crystal = get_red_crystal(user_id)
        blue_crystal = get_blue_crystal(user_id)
        remaining_pity = get_remaining_pity(user_id)
        current_gacha = get_gacha_count(user_id)
        
        # 获取卡片收藏（所有三星卡，带计数）
        pity_data = load_pity_data(user_id)
        card_collection = pity_data.get("card_collection", {})
        
        # 筛选三星卡并按时间排序（最近到最远）
        three_star_cards = []
        for card_id, info in card_collection.items():
            if info.get("stars") == 3:
                three_star_cards.append({
                    "card_id": card_id,
                    "name": info.get("name", ""),
                    "limit_type": info.get("limit_type", ""),
                    "count": info.get("count", 1),
                    "last_time": info.get("last_time", 0)
                })
        
        # 按时间排序（最近到最远）
        three_star_cards.sort(key=lambda x: x["last_time"], reverse=True)
        
        # 分页处理
        page_size = 10
        current_page = 1
        
        # 从会话获取当前页码
        session = get_box_session(user_id)
        if session and "personal_page" in session:
            current_page = session["personal_page"]
        
        # 根据操作更新页码
        if page_action == 'next':
            current_page += 1
        elif page_action == 'prev':
            current_page = max(1, current_page - 1)
        
        # 计算总页数
        total_pages = max(1, (len(three_star_cards) + page_size - 1) // page_size)
        current_page = max(1, min(current_page, total_pages))
        
        # 更新会话中的页码
        if not session:
            create_box_session(user_id, [])
        BOX_SESSIONS[user_id]["personal_page"] = current_page
        
        # 获取当前页的卡片
        start_idx = (current_page - 1) * page_size
        end_idx = start_idx + page_size
        current_page_cards = three_star_cards[start_idx:end_idx]
        
        # 计算可兑换呱太
        exchange_red = red_crystal * 5
        exchange_blue = blue_crystal * 20
        total_exchange = exchange_red + exchange_blue
        
        # 构建消息
        info_text = (
            f"📊 个人记录\n"
            f"当前呱太: {current_gacha}\n"
            f"累计次数: {total_draws}\n"
            f"累计3星: {total_3stars}\n"
            f"🔴 红色碎片: {red_crystal}\n"
            f"🔵 蓝色碎片: {blue_crystal}\n"
            f"距离保底: {remaining_pity}抽\n\n"
            f"💎 碎片兑换呱太:\n"
            f"红碎片 x{red_crystal} → {exchange_red} 呱太\n"
            f"蓝碎片 x{blue_crystal} → {exchange_blue} 呱太\n"
            f"总计可兑换: {total_exchange} 呱太\n"
            f"输入「兑换呱太」即可兑换"
        )
        
        # 构建三星卡列表（带计数）
        img_path = None
        if three_star_cards:
            info_text += f"\n\n✨ 所有三星卡 (第{current_page}/{total_pages}页):"
            for i, card in enumerate(current_page_cards, 1):
                limit_badge = ""
                if card.get("limit_type") == "期間限定":
                    limit_badge = "🔸"
                elif card.get("limit_type") == "フェス限定":
                    limit_badge = "🔹"
                count = card.get("count", 1)
                count_suffix = f" x{count}" if count > 1 else ""
                info_text += f"\n{i}. {limit_badge}{card['name']}{count_suffix}"
            
            # 创建三星卡展示图片（带计数标记）
            characters = get_characters()
            if characters:
                display_cards = []
                for card in current_page_cards:
                    card_id = card.get("card_id")
                    chara = next((c for c in characters if str(c.get("card_id")) == str(card_id)), None)
                    if chara:
                        display_cards.append({"chara": chara, "count": card.get("count", 1)})
                
                if display_cards:
                    from PIL import Image, ImageDraw, ImageFont
                    from io import BytesIO
                    import random
                    
                    card_imgs = []
                    for item in display_cards:
                        card_bytes = composite_card(item["chara"])
                        if card_bytes:
                            card_img = Image.open(BytesIO(card_bytes))
                            card_imgs.append({"img": card_img, "count": item["count"]})
                    
                    if card_imgs:
                        # 使用十连背景图（与抽卡一致）
                        bg_path = None
                        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
                            test_path = LEVEL_DIR / bg_name
                            if test_path.exists():
                                bg_path = str(test_path)
                                break
                        
                        if bg_path:
                            # 使用背景图片，放大到原来的1.5倍
                            bg_img = Image.open(bg_path).convert('RGB')
                            bg_w, bg_h = bg_img.size
                            final_w = int(bg_w * 1.5 * 0.264)
                            final_h = int(bg_h * 1.5 * 0.264)
                            bg = bg_img.resize((final_w, final_h), Image.LANCZOS)
                            bg_w, bg_h = final_w, final_h
                        else:
                            bg = Image.new('RGB', (600, 400), (50, 50, 50))
                            bg_w, bg_h = bg.size
                        
                        max_card_width = 90
                        max_card_height = 120
                        gap = 10
                        cols = min(len(card_imgs), 5)
                        rows = (len(card_imgs) + cols - 1) // cols
                        
                        first_card = card_imgs[0]["img"]
                        orig_w, orig_h = first_card.size
                        scale = min(max_card_width / orig_w, max_card_height / orig_h)
                        card_width = int(orig_w * scale)
                        card_height = int(orig_h * scale)
                        
                        total_w = cols * (card_width + gap) - gap
                        total_h = rows * (card_height + gap) - gap
                        start_x = (bg_w - total_w) // 2
                        start_y = (bg_h - total_h) // 2
                        
                        # 创建最终画布（与抽卡一致）
                        output = Image.new('RGB', (bg_w, bg_h), (50, 50, 50))
                        output.paste(bg, (0, 0))
                        
                        for i, item in enumerate(card_imgs):
                            img = item["img"]
                            count = item["count"]
                            col = i % cols
                            row = i // cols
                            x = start_x + col * (card_width + gap)
                            y = start_y + row * (card_height + gap)
                            
                            img_copy = img.copy()
                            img_copy.thumbnail((card_width, card_height), Image.LANCZOS)
                            thumb_w, thumb_h = img_copy.size
                            
                            paste_x = x + (card_width - thumb_w) // 2
                            paste_y = y + (card_height - thumb_h) // 2
                            
                            output.paste(img_copy, (paste_x, paste_y))
                            
                            # 如果计数大于1，在右下角添加计数标记
                            if count > 1:
                                draw = ImageDraw.Draw(output)
                                # 创建半透明背景
                                badge_w = 30
                                badge_h = 20
                                badge_x = paste_x + thumb_w - badge_w - 2
                                badge_y = paste_y + thumb_h - badge_h - 2
                                
                                # 绘制圆角矩形背景
                                draw.rounded_rectangle([badge_x, badge_y, badge_x+badge_w, badge_y+badge_h], radius=4, fill=(0, 0, 0, 200))
                                
                                # 绘制数字
                                try:
                                    font = ImageFont.truetype("arial.ttf", 12)
                                except:
                                    font = ImageFont.load_default()
                                text = f"x{count}"
                                text_w, text_h = draw.textsize(text, font=font)
                                text_x = badge_x + (badge_w - text_w) // 2
                                text_y = badge_y + (badge_h - text_h) // 2
                                draw.text((text_x, text_y), text, fill=(255, 255, 255), font=font)
                        
                        output_idx = random.randint(1000, 9999)
                        img_path = OUTPUT_DIR / f"personal_3stars_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
                        output.convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)
        
        # 添加分页提示
        if total_pages > 1:
            page_hints = []
            if current_page > 1:
                page_hints.append("输入「上一页」查看上一页")
            if current_page < total_pages:
                page_hints.append("输入「下一页」查看下一页")
            if page_hints:
                info_text += "\n\n" + " | ".join(page_hints)
        
        # 发送消息（文字+图片）
        if group_id and user_id:
            info_text = f"[CQ:at,qq={user_id}] {info_text}"
        
        if img_path and img_path.exists():
            with open(img_path, "rb") as f:
                image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
            img_message = f"[CQ:image,file=base64://{image_base64}]"
            full_message = f"{info_text}\n{img_message}"
            send_message(full_message, user_id, group_id)
            if img_path.exists():
                img_path.unlink()
        else:
            send_message(info_text, user_id, group_id)
        
        log_info(f"查询个人记录 [{user_id}] 第{current_page}/{total_pages}页")
        
        return jsonify({
            "status": "success",
            "user_id": user_id,
            "total_draws": total_draws,
            "total_3stars": total_3stars,
            "red_crystal": red_crystal,
            "blue_crystal": blue_crystal,
            "remaining_pity": remaining_pity,
            "current_gacha": current_gacha,
            "current_page": current_page,
            "total_pages": total_pages
        })
    
    except Exception as e:
        log_error(f"查询个人记录失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_show_details(user_id: str, group_id):
    """处理详细信息查询请求"""
    try:
        session = get_box_session(user_id)
        if not session or "last_results" not in session:
            reply = "没有可查看的详细信息！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "没有详细信息"})

        last_results = session["last_results"]
        results = last_results.get("results", [])
        mutations = last_results.get("mutations", [])
        red_crystal = last_results.get("red_crystal", 0)
        blue_crystal = last_results.get("blue_crystal", 0)

        # 构建详细结果
        stars_display = {1: "⭐", 2: "⭐⭐", 3: "⭐⭐⭐"}
        result_lines = []
        fes_messages = []
        for r in results:
            result_lines.append(f"{r['index']}. {stars_display.get(r['stars'], '⭐')} {r['name']}")
            # 收集FES提示消息
            if r.get("fes_message"):
                fes_messages.append(r["fes_message"])

        result_text = "\n".join(result_lines)

        # 构建消息
        msg_parts = ["详细信息"]
        if mutations:
            msg_parts.append("\n".join(mutations))
        # 添加FES提示消息
        if fes_messages:
            msg_parts.append("\n".join(fes_messages))
        msg_parts.append(f"\n结果:\n{result_text}")
        
        if red_crystal > 0 or blue_crystal > 0:
            crystal_parts = []
            if red_crystal > 0:
                crystal_parts.append(f"🔴红色碎片 x{red_crystal}")
            if blue_crystal > 0:
                crystal_parts.append(f"🔵蓝色碎片 x{blue_crystal}")
            msg_parts.append(f"\n本次获得: {' + '.join(crystal_parts)}")

        complete_text = "\n".join(msg_parts)

        # 发送消息
        if group_id and user_id:
            complete_text = f"[CQ:at,qq={user_id}] {complete_text}"
        send_message(complete_text, user_id, group_id)

        # 查看详细信息后清除会话
        clear_box_session(user_id)

        log_info(f"查询详细信息 [{user_id}]")

        return jsonify({
            "status": "success",
            "results": results,
            "mutations": mutations,
            "red_crystal": red_crystal,
            "blue_crystal": blue_crystal
        })

    except Exception as e:
        log_error(f"查询详细信息失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_leaderboard(user_id: str, group_id):
    """处理排行榜查询请求"""
    try:
        # 获取排行榜数据
        leaderboard = get_leaderboard()
        
        if not leaderboard:
            reply = "暂无排行榜数据！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "leaderboard": []})
        
        # 构建排行榜消息
        rank_emojis = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]
        
        msg_lines = ["🏆 战力排行榜 🏆"]
        for i, player in enumerate(leaderboard):
            rank = i + 1
            emoji = rank_emojis[i] if i < len(rank_emojis) else f"{rank}."
            
            # 获取用户自己的战力
            if player["user_id"] == user_id:
                power_info = f"⚔️{player['power']} (你)"
            else:
                power_info = f"⚔️{player['power']}"
            
            msg_lines.append(
                f"{emoji} {player['user_id']} {power_info}\n"
                f"   累计: {player['total_draws']} | 累计3星: {player['total_3stars']}"
            )
        
        
        reply = "\n".join(msg_lines)
        
        # 发送消息
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        
        log_info(f"查询排行榜 [{user_id}]")
        
        return jsonify({
            "status": "success",
            "leaderboard": leaderboard
        })
    
    except Exception as e:
        log_error(f"查询排行榜失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_exchange_crystal(user_id: str, group_id):
    """处理碎片兑换呱太请求"""
    try:
        # 获取用户碎片数量
        red_crystal = get_red_crystal(user_id)
        blue_crystal = get_blue_crystal(user_id)
        
        if red_crystal == 0 and blue_crystal == 0:
            reply = "你没有可以兑换的碎片！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "没有碎片可兑换"})
        
        # 计算兑换数量
        red_amount = red_crystal * 5
        blue_amount = blue_crystal * 20
        total_amount = red_amount + blue_amount
        
        # 清空碎片
        pity_data = load_pity_data(user_id)
        pity_data["red_crystal"] = 0
        pity_data["blue_crystal"] = 0
        save_pity_data(user_id, pity_data)
        
        # 添加呱太
        add_gacha(user_id, total_amount)
        
        # 构建回复消息
        parts = []
        if red_crystal > 0:
            parts.append(f"🔴红色碎片 x{red_crystal} → {red_amount} 呱太")
        if blue_crystal > 0:
            parts.append(f"🔵蓝色碎片 x{blue_crystal} → {blue_amount} 呱太")
        
        parts_str = '\n'.join(parts)
        reply = f"💎 兑换成功！\n{parts_str}\n总共获得: {total_amount} 呱太"
        
        # 发送消息
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        
        log_info(f"碎片兑换 [{user_id}]: red={red_crystal}, blue={blue_crystal}, total={total_amount}")
        
        return jsonify({
            "status": "success",
            "user_id": user_id,
            "red_crystal_exchanged": red_crystal,
            "blue_crystal_exchanged": blue_crystal,
            "gacha_added": total_amount,
            "current_gacha": get_gacha_count(user_id)
        })
    
    except Exception as e:
        log_error(f"碎片兑换失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ========== 三星池子命令处理 ==========
def handle_3star_pool(user_id: str, group_id, raw_message: str):
    """
    处理三星池子抽卡请求
    命令格式:
    - 三星池子 - 显示三星池子介绍
    - 红抽 - 使用红色碎片抽卡
    - 蓝抽 - 使用蓝色碎片抽卡
    """
    try:
        # 检测抽卡类型
        if '红抽' in raw_message:
            crystal_type = "red"
            crystal_name = "红色碎片"
            cost = THREE_STAR_POOL_RED_COST
            current = get_red_crystal(user_id)
        elif '蓝抽' in raw_message:
            crystal_type = "blue"
            crystal_name = "蓝色碎片"
            cost = THREE_STAR_POOL_BLUE_COST
            current = get_blue_crystal(user_id)
        else:
            # 显示三星池子介绍
            red_crystal = get_red_crystal(user_id)
            blue_crystal = get_blue_crystal(user_id)
            
            reply = f"""╔══════════════════════════════╗
║     ★★★ 三星池子 ★★★       ║
╠══════════════════════════════╣
║ 消耗说明:                    ║
║ 🔴 红色碎片: {THREE_STAR_POOL_RED_COST}个/抽       ║
║ 🔵 蓝色碎片: {THREE_STAR_POOL_BLUE_COST}个/抽       ║
╠══════════════════════════════╣
║ 你当前拥有:                  ║
║ 🔴 红色碎片: {red_crystal}个            ║
║ 🔵 蓝色碎片: {blue_crystal}个            ║
╠══════════════════════════════╣
║ 可抽取次数:                  ║
║ 🔴 红色碎片: {red_crystal // THREE_STAR_POOL_RED_COST}次          ║
║ 🔵 蓝色碎片: {blue_crystal // THREE_STAR_POOL_BLUE_COST}次          ║
╠══════════════════════════════╣
║ 输入「三星池子红抽」使用红色碎片抽    ║
║ 输入「三星池子蓝抽」使用蓝色碎片抽    ║
╚══════════════════════════════╝
"""
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}]\n{reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "显示三星池子"})
        
        # 执行抽卡
        result = draw_3star_pool(user_id, crystal_type)
        
        if not result["success"]:
            reply = result["message"]
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": result["message"]})
        
        # 抽卡成功，生成结果图片
        selected = result["character"]
        card_id = str(selected.get("card_id", ""))
        fes_message = result.get("fes_message", "")
        
        # 生成单抽结果图片（使用抽卡结果背景）
        img_bytes = composite_card(selected)
        img_path = None
        
        if img_bytes:
            # 保存图片到临时文件
            output_idx = random.randint(1000, 9999)
            img_path = OUTPUT_DIR / f"3star_pool_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            with open(img_path, "wb") as f:
                f.write(img_bytes)
        
        if img_path and os.path.exists(img_path):
            with open(img_path, "rb") as f:
                image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
            img_message = f"[CQ:image,file=base64://{image_base64}]"
            
            # 发送结果（包含FES提示）
            if group_id and user_id:
                at_message = f"[CQ:at,qq={user_id}] "
            else:
                at_message = ""
            
            remaining = current - cost
            # 如果有FES提示，添加到消息中
            fes_text = f"\n{fes_message}" if fes_message else ""
            full_message = f"{at_message}💎 使用{cost}个{crystal_name}！{fes_text}\n{img_message}\n剩余{crystal_name}: {remaining}个"
            send_message(full_message, user_id, group_id)
            
            if os.path.exists(img_path):
                os.remove(img_path)
        else:
            # 如果没有图片，发送文字消息
            fes_text = f"\n{fes_message}" if fes_message else ""
            reply = f"💎 使用{cost}个{crystal_name}！\n{result['message']}{fes_text}\n剩余{crystal_name}: {current - cost}个"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
        
        log_info(f"三星only池 [{user_id}]: type={crystal_type}, card={selected.get('name')}")
        
        return jsonify({
            "status": "success",
            "message": result["message"],
            "character": selected.get("name")
        })
    
    except Exception as e:
        log_error(f"三星池失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def save_rolling_battle_log(user_id: str, result: dict) -> str:
    """保存战斗日志，滚动保留最近3次（单文件 battle_{uid}.json，数组存储）"""
    import time
    log_path = INFO_DIR / f"battle_{user_id}.json"
    logs = []
    if log_path.exists():
        try:
            with open(log_path, "r", encoding="utf-8") as f:
                logs = json.load(f)
            if not isinstance(logs, list):
                logs = []
        except Exception:
            logs = []
    # 添加时间戳
    result["saved_at"] = datetime.now().strftime("%m-%d %H:%M:%S")
    logs.append(result)
    # 只保留最近3次
    if len(logs) > 3:
        logs = logs[-3:]
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)
    return f"battle_{user_id} (共{len(logs)}次)"


def handle_battle(user_id: str, group_id, raw_message: str):
    """
    处理对战请求
    命令格式:
    - 战斗[1-5] - AI对战，可选难度1~5（默认2）
    - 战斗 @玩家QQ - 与指定玩家对战
    - 对战说明 - 显示战斗帮助
    """
    try:
        if not BATTLE_SYSTEM_LOADED or BATTLE_INSTANCE is None:
            reply = "战斗系统未加载，请稍后重试"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "战斗系统未加载"})

        if '说明' in raw_message:
            reply = get_battle_help()
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}]\n{reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "显示战斗帮助"})

        # 解析对手@ 和 AI难度
        import re
        at_match = re.search(r'\[CQ:at,qq=(\d+)\]', raw_message)
        enemy_user_id = None

        # 解析难度: 战斗5, 战斗 3, 对战2 等
        diff_match = re.search(r'(?:战斗|对战|决斗)\s*([1-5])', raw_message)
        ai_difficulty = int(diff_match.group(1)) if diff_match else 2

        if at_match:
            enemy_user_id = at_match.group(1)
            if enemy_user_id == user_id:
                reply = "不能与自己对战！"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "不能与自己对战"})

        # 加载玩家队伍
        player_team = load_team_data(user_id)
        player_battle_cards = player_team.get("battle_cards", [])
        player_has_cards = any(card for card in player_battle_cards)

        # 读取活跃预设信息
        active_slot = 0
        try:
            pdata = load_presets(user_id)
            active_slot = pdata.get("active_slot", 0)
        except Exception:
            pass

        if not player_has_cards:
            reply = "你的队伍还没有配置战斗卡！请先使用「队伍我的卡」命令配置队伍。"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "玩家队伍为空"})

        if enemy_user_id:
            enemy_team = load_team_data(enemy_user_id)
            enemy_battle_cards = enemy_team.get("battle_cards", [])
            enemy_has_cards = any(card for card in enemy_battle_cards)

            if not enemy_has_cards:
                reply = "对手的队伍还没有配置战斗卡！"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "对手队伍为空"})

            enemy_name = f"玩家{enemy_user_id[:4]}****"

            # 发送双方VS配队图
            characters = get_characters()
            vs_img = build_vs_team_image(player_team, enemy_team, characters)
            if vs_img:
                send_image(vs_img, user_id, group_id)
        else:
            # AI对手（带难度）
            difficulty_names = {1: "简单", 2: "普通", 3: "困难", 4: "极难", 5: "地狱"}
            enemy_team = generate_ai_team(ai_difficulty)
            enemy_name = f"AI({difficulty_names[ai_difficulty]})"

            # 发送双方VS配队图
            characters = get_characters()
            at_message = f"[CQ:at,qq={user_id}] " if group_id and user_id else ""
            vs_img = build_vs_team_image(player_team, enemy_team, characters)
            if vs_img:
                send_image(vs_img, user_id, group_id)
            send_message(at_message + f"⚔️ VS {enemy_name}", user_id, group_id)

        # 执行战斗
        log_info(f"战斗开始: {user_id} vs {enemy_user_id or 'AI'}")
        result = BATTLE_INSTANCE.start_battle(player_team, enemy_team, challenger="player")

        # 保存战斗日志（滚动保留最近3次）
        battle_log_key = save_rolling_battle_log(user_id, result)

        # 生成简短结果
        winner = result["winner"]
        rounds = result["rounds"]
        
        if winner == "player":
            result_text = f"🏆 胜利！经过 {rounds} 回合的激战，你击败了 {enemy_name}！"
        else:
            result_text = f"💀 失败... 经过 {rounds} 回合的激战，你被 {enemy_name} 击败了..."
        
        # 统计双方存活情况
        player_alive = sum(1 for u in result["player_units"] if u["alive"] and not u["is_assist"])
        enemy_alive = sum(1 for u in result["enemy_units"] if u["alive"] and not u["is_assist"])
        
        result_text += f"\n📊 最终状态: 我方存活 {player_alive}/6, 敌方存活 {enemy_alive}/6"
        if active_slot > 0:
            result_text += f"\n📋 使用预设: 槽{active_slot}"

        # 发送结果
        at_message = f"[CQ:at,qq={user_id}] " if group_id and user_id else ""
        send_message(at_message + result_text, user_id, group_id)
        
        log_info(f"战斗结束: {user_id} vs {enemy_user_id or 'AI'}, winner={winner}, rounds={rounds}")
        
        return jsonify({
            "status": "success",
            "message": result_text,
            "winner": winner,
            "rounds": rounds,
            "battle_log_key": battle_log_key
        })
    
    except Exception as e:
        log_error(f"战斗失败: {e}")
        reply = f"战斗失败: {str(e)}"
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_boss_battle(user_id: str, group_id, raw_message: str):
    """
    处理BOSS战请求
    命令格式: BOSS战 / boss战
    BOSS由config.py中的BOSS_CARD_ID指定，初始SP=90
    """
    try:
        if not BATTLE_SYSTEM_LOADED or BATTLE_INSTANCE is None:
            reply = "战斗系统未加载，请稍后重试"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "战斗系统未加载"})

        # BOSS战冷却检查
        now_ts = datetime.now().timestamp()
        last_boss = BOSS_BATTLE_COOLDOWN.get(user_id, 0)
        remaining = BOSS_BATTLE_COOLDOWN_SECONDS - (now_ts - last_boss)
        if remaining > 0:
            reply = f"BOSS战冷却中！请等待 {int(remaining)} 秒后再试~"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"BOSS战冷却 [{user_id}]: 还需等待 {int(remaining)} 秒")
            return jsonify({
                "status": "cooldown",
                "message": "BOSS战冷却中",
                "remaining_seconds": int(remaining)
            })

        # 从config获取BOSS卡牌ID
        try:
            from config import BOSS_CARD_ID
        except ImportError:
            BOSS_CARD_ID = "100430006"  # 默认

        # 加载玩家队伍
        player_team = load_team_data(user_id)
        player_battle_cards = player_team.get("battle_cards", [])
        player_has_cards = any(card for card in player_battle_cards)

        # 读取活跃预设信息
        active_slot = 0
        try:
            pdata = load_presets(user_id)
            active_slot = pdata.get("active_slot", 0)
        except Exception:
            pass

        if not player_has_cards:
            reply = "你的队伍还没有配置战斗卡！请先使用「队伍 我的卡」命令配置队伍。"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "玩家队伍为空"})

        # 查找BOSS角色
        boss_char = BATTLE_INSTANCE.get_character(str(BOSS_CARD_ID))
        if not boss_char:
            boss_char = BATTLE_INSTANCE._get_fallback_character(str(BOSS_CARD_ID))
        boss_name = boss_char.name

        # 构建BOSS队伍用于VS图（位置1=中间，单卡+5空位，无A卡）
        boss_team = {
            "battle_cards": [None, str(BOSS_CARD_ID)] + [None] * 4,
            "assist_cards": [None] * 6
        }

        # 发送双方VS配队图
        characters = get_characters()
        at_message = f"[CQ:at,qq={user_id}] " if group_id and user_id else ""
        vs_img = build_vs_team_image(player_team, boss_team, characters)
        if vs_img:
            send_image(vs_img, user_id, group_id)
        send_message(at_message + f"⚔️ BOSS战 VS 【{boss_name}】 (HP: 15,000,000)", user_id, group_id)

        # 执行BOSS战
        log_info(f"BOSS战开始: {user_id} vs BOSS({BOSS_CARD_ID} {boss_name})")
        BOSS_BATTLE_COOLDOWN[user_id] = datetime.now().timestamp()
        result = BATTLE_INSTANCE.start_boss_battle(
            player_team, str(BOSS_CARD_ID), initial_sp=90
        )

        # 保存战斗日志（复用现有）
        save_rolling_battle_log(user_id, result)

        # 格式化并发送结果
        result_text = format_boss_result(result)
        if active_slot > 0:
            result_text += f"\n📋 使用预设: 槽{active_slot}"

        send_message(at_message + result_text, user_id, group_id)

        log_info(f"BOSS战结束: {user_id}, damage={result['damage_dealt']}, pct={result['damage_percent']}%")

        return jsonify({
            "status": "success",
            "message": result_text,
            "boss_name": result["boss_name"],
            "damage_dealt": result["damage_dealt"],
            "damage_percent": result["damage_percent"],
            "rounds": result["rounds"]
        })

    except Exception as e:
        log_error(f"BOSS战失败: {e}")
        reply = f"BOSS战失败: {str(e)}"
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        return jsonify({"status": "error", "message": str(e)}), 500


def generate_ai_team(difficulty: int = 2) -> dict:
    """
    生成AI队伍
    difficulty 1-5:
      1=简单: 随机选卡，不刻意匹配
      2=普通: 随机但偏向同色同攻 (默认)
      3=困难: 同色同攻匹配，FES优先
      4=极难: 完美同色同攻，大量FES，高分卡
      5=地狱: 全完美匹配，全FES，最高评分
    """
    characters = get_characters()

    # 筛选三星及以上卡牌
    all_battle = [c for c in characters if c.get("type") == "battle" and c.get("stars", 0) >= 3]
    all_assist = [c for c in characters if c.get("type") == "assist" and c.get("stars", 0) >= 3]

    if not all_battle:
        log_error("没有可用的战斗卡！")
        return {"battle_cards": [], "assist_cards": []}

    # 给每张卡打分
    def score_card(c):
        s = c.get("stars", 3) * 100
        s += c.get("attack", 5000) * 0.5 + c.get("hp", 10000) * 0.2 + c.get("dexterity", 1000) * 0.2
        lt = c.get("limit_type", "")
        if lt == "フェス限定":
            s *= 2.0
        elif lt == "期間限定":
            s *= 1.5
        return s

    # 按分数排序
    all_battle.sort(key=score_card, reverse=True)
    all_assist.sort(key=score_card, reverse=True)

    # 根据难度决定选择范围
    # 难度越高，可选池越小（只用最强的卡）
    pool_ratios = {1: 1.0, 2: 0.8, 3: 0.5, 4: 0.3, 5: 0.15}
    ratio = pool_ratios.get(difficulty, 0.8)

    b_count = max(6, int(len(all_battle) * ratio))
    a_count = max(6, int(len(all_assist) * ratio))
    b_pool = all_battle[:b_count]
    a_pool = all_assist[:a_count]

    # 匹配概率: 难度越高越要同色同攻
    match_chance = {1: 0.2, 2: 0.5, 3: 0.8, 4: 0.95, 5: 1.0}[difficulty]
    # FES占比目标
    fes_target = {1: 0.1, 2: 0.3, 3: 0.5, 4: 0.7, 5: 0.9}[difficulty]
    # 选全队统一颜色: 按评分加权随机，保证多样性
    BASE_COLORS = ["红", "绿", "蓝", "黄", "紫"]
    color_scores = {c: 0 for c in BASE_COLORS}
    for card in b_pool:
        attr = card.get("attribute", "")
        base = attr[1:] if attr.startswith("超") else attr
        if base in color_scores:
            color_scores[base] += score_card(card)

    # 加权随机选色: 评分高概率大，但不总选同一个
    total = sum(color_scores.values())
    if total > 0:
        weights = [color_scores[c] / total for c in BASE_COLORS]
        team_color = random.choices(BASE_COLORS, weights=weights, k=1)[0]
    else:
        team_color = "红"

    # 全队同色比例: 锦上添花，不压倒单卡强度
    mono_color_ratio = {1: 0.0, 2: 0.15, 3: 0.3, 4: 0.5, 5: 0.7}[difficulty]
    # 同色同攻权重: 单卡强度为主，完美匹配加分
    perfect_pair_weight = {1: 1.0, 2: 2.0, 3: 3.0, 4: 5.0, 5: 8.0}[difficulty]
    # 难度5: B+A必须同色
    strict_same_color = difficulty >= 5

    battle_cards = []
    assist_cards = []
    used_b = set()
    used_a = set()

    for pos in range(6):
        available_b = [c for c in b_pool if str(c.get("card_id")) not in used_b]
        if not available_b:
            break

        # 全队同色倾向
        if random.random() < mono_color_ratio:
            color_filtered = [c for c in available_b
                            if (c.get("attribute", "")[1:] if c.get("attribute", "").startswith("超") else c.get("attribute", "")) == team_color]
            if color_filtered:
                available_b = color_filtered

        # 难度5: B卡只能选有同色A卡可配的
        if strict_same_color:
            has_match = [c for c in available_b if any(
                ac for ac in a_pool
                if str(ac.get("card_id")) not in used_a
                and ac.get("attribute", "") == c.get("attribute", "")
                and ac.get("attack_type", "") == c.get("attack_type", "")
            )]
            if has_match:
                available_b = has_match

        # 给每张候选B卡打分: 自身分 + 完美A匹配加分
        def b_score(bc):
            s = score_card(bc)
            b_attr = bc.get("attribute", "")
            b_type = bc.get("attack_type", "")
            perfect_count = sum(1 for ac in a_pool
                              if str(ac.get("card_id")) not in used_a
                              and ac.get("attribute", "") == b_attr
                              and ac.get("attack_type", "") == b_type)
            if perfect_count > 0:
                s *= perfect_pair_weight
            return s

        available_b.sort(key=b_score, reverse=True)

        top_n = min(3, len(available_b))
        b_card = available_b[random.randint(0, top_n - 1)]

        battle_cards.append(str(b_card.get("card_id", "")))
        used_b.add(str(b_card.get("card_id", "")))

        # 选A卡: 难度5强制同色同攻，不降级
        b_attr = b_card.get("attribute", "")
        b_type = b_card.get("attack_type", "")

        perfect_a = [c for c in a_pool
                     if str(c.get("card_id")) not in used_a
                     and c.get("attribute", "") == b_attr
                     and c.get("attack_type", "") == b_type]
        same_color_a = [c for c in a_pool
                        if str(c.get("card_id")) not in used_a
                        and c.get("attribute", "") == b_attr]
        any_a = [c for c in a_pool if str(c.get("card_id")) not in used_a]

        if strict_same_color:
            # 难度5: 必须同色同攻，否则同色
            a_card = random.choice(perfect_a) if perfect_a else (random.choice(same_color_a) if same_color_a else None)
        elif random.random() < match_chance:
            if perfect_a:
                a_card = random.choice(perfect_a)
            elif same_color_a:
                a_card = random.choice(same_color_a)
            elif any_a:
                a_card = random.choice(any_a)
            else:
                a_card = None
        else:
            a_card = random.choice(any_a) if any_a else None

        assist_cards.append(str(a_card.get("card_id", "")) if a_card else None)
        if a_card:
            used_a.add(str(a_card.get("card_id", "")))

    # 统计实际颜色
    color_count = {}
    for cid in battle_cards:
        for c in characters:
            if str(c.get("card_id")) == cid:
                attr = c.get("attribute", "")
                base = attr[1:] if attr.startswith("超") else attr
                color_count[base] = color_count.get(base, 0) + 1
                break
    dominant = max(color_count, key=color_count.get) if color_count else "?"

    log_info(f"AI队伍(难度{difficulty} 主色{dominant} {color_count.get(dominant,0)}/6 FES目标{fes_target:.0%}): 战斗{battle_cards} 支援{assist_cards}")
    return {"battle_cards": battle_cards, "assist_cards": assist_cards}

def get_user_team(user_id: str) -> dict:
    """获取用户的队伍配置（封装team_system的load_team_data）"""
    try:
        from team_system import load_team_data
        return load_team_data(user_id)
    except Exception as e:
        log_error(f"获取用户队伍失败: {e}")
        return {"battle_cards": [], "assist_cards": []}


# ========== 排行榜系统 ==========
RANKING_FILE = INFO_DIR / "ranking.json"

def init_ranking():
    """初始化排行榜（如果文件不存在）"""
    if not RANKING_FILE.exists():
        # 初始排行榜：10个AI队伍
        ranking = []
        for i in range(10):
            ai_team = generate_ai_team(difficulty=i + 1 if i < 5 else 5)  # 排名越高AI越强
            # 设置一个不满编的队伍（前几个位置为空）
            empty_positions = i // 3  # 越靠前的AI队伍越完整
            for pos in range(empty_positions):
                if pos < len(ai_team["battle_cards"]):
                    ai_team["battle_cards"][pos] = None
                if pos < len(ai_team["assist_cards"]):
                    ai_team["assist_cards"][pos] = None
            
            ranking.append({
                "rank": i + 1,
                "is_ai": True,
                "user_id": f"AI_{i + 1}",
                "nickname": f"AI队伍{i + 1}",
                "team": ai_team,
                "wins": 0,
                "losses": 0
            })
        
        with open(RANKING_FILE, "w", encoding="utf-8") as f:
            json.dump(ranking, f, ensure_ascii=False, indent=2)

def load_ranking():
    """加载排行榜数据"""
    init_ranking()
    with open(RANKING_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_ranking(ranking):
    """保存排行榜数据"""
    with open(RANKING_FILE, "w", encoding="utf-8") as f:
        json.dump(ranking, f, ensure_ascii=False, indent=2)

def get_player_ranking(user_id: str):
    """获取玩家的排名（如果不在排行榜中返回11）"""
    ranking = load_ranking()
    for entry in ranking:
        if not entry["is_ai"] and entry["user_id"] == user_id:
            return entry["rank"]
    return 11

def get_player_entry(user_id: str):
    """获取玩家的排行榜entry（如果不在排行榜中返回None）"""
    ranking = load_ranking()
    for entry in ranking:
        if not entry["is_ai"] and entry["user_id"] == user_id:
            return entry
    return None

def add_player_to_ranking(user_id: str, nickname: str, team: dict, rank: int):
    """将玩家添加到排行榜（替换该位置的AI）"""
    ranking = load_ranking()

    # 检查是否已存在
    for entry in ranking:
        if not entry["is_ai"] and entry["user_id"] == user_id:
            return entry

    # 移除该排名的AI条目，为新玩家腾位置
    ranking = [e for e in ranking if not (e["rank"] == rank and e["is_ai"])]

    # 添加新玩家
    new_entry = {
        "rank": rank,
        "is_ai": False,
        "user_id": user_id,
        "nickname": nickname,
        "team": team,
        "wins": 0,
        "losses": 0
    }
    ranking.append(new_entry)

    # 重新排序
    ranking.sort(key=lambda x: x["rank"])
    for i, entry in enumerate(ranking[:10], 1):
        entry["rank"] = i

    # 确保没有超过10条
    for entry in ranking[10:]:
        if not entry["is_ai"]:
            entry["rank"] = 11

    save_ranking(ranking)
    return new_entry

def update_ranking_after_battle(winner_id: str, loser_id: str, is_winner_ai: bool, is_loser_ai: bool, winner_team: dict = None, winner_nickname: str = None):
    """战斗结束后更新排行榜"""
    ranking = load_ranking()
    
    # 找到获胜者和失败者
    winner_entry = None
    loser_entry = None
    
    for entry in ranking:
        if entry["is_ai"] and is_winner_ai and entry["user_id"] == winner_id:
            winner_entry = entry
        elif not entry["is_ai"] and not is_winner_ai and entry["user_id"] == winner_id:
            winner_entry = entry
        
        if entry["is_ai"] and is_loser_ai and entry["user_id"] == loser_id:
            loser_entry = entry
        elif not entry["is_ai"] and not is_loser_ai and entry["user_id"] == loser_id:
            loser_entry = entry
    
    # 如果玩家首次进入排行榜（winner_entry为None）
    if not is_winner_ai and winner_entry is None and loser_entry is not None:
        # 添加玩家到排行榜（替换该位置AI）
        add_player_to_ranking(winner_id, winner_nickname or f"玩家{winner_id[:4]}****", winner_team, loser_entry["rank"])
        ranking = load_ranking()
        # 重新找到winner_entry（reload后的新对象）
        for entry in ranking:
            if not entry["is_ai"] and entry["user_id"] == winner_id:
                winner_entry = entry
                break
    
    if winner_entry and loser_entry:
        # 增加胜负记录
        winner_entry["wins"] += 1
        loser_entry["losses"] += 1

        # 更新人类玩家的存储队伍为当前队伍
        if not is_winner_ai and winner_team:
            winner_entry["team"] = winner_team
        if not is_loser_ai:
            loser_team = load_team_data(loser_id)
            if loser_team and loser_team.get("battle_cards"):
                loser_entry["team"] = loser_team
        
        # 如果玩家击败了AI或排名更高的玩家，交换位置
        if winner_entry["rank"] > loser_entry["rank"]:
            # 交换排名
            winner_entry["rank"], loser_entry["rank"] = loser_entry["rank"], winner_entry["rank"]
            
            # 重新排序
            ranking.sort(key=lambda x: x["rank"])
            
            # 调整排名序号
            for i, entry in enumerate(ranking[:10], 1):
                entry["rank"] = i
            
            # 将超出前10的设为11
            for entry in ranking[10:]:
                if not entry["is_ai"]:
                    entry["rank"] = 11
    
    save_ranking(ranking)

def show_ranking(user_id: str, group_id):
    """显示排行榜"""
    ranking = load_ranking()
    
    lines = ["🏆 排行榜 TOP 10 🏆"]
    for i, entry in enumerate(ranking[:10], 1):
        if entry["is_ai"]:
            lines.append(f"第{i}名: 🤖 {entry['nickname']} (AI)")
        else:
            lines.append(f"第{i}名: 👤 {entry['nickname']} (玩家)")
    
    player_rank = get_player_ranking(user_id)
    if player_rank > 10:
        lines.append(f"\n你的排名: 第{player_rank}名 (未进入前10)")
        # 显示可挑战的排名范围
        lines.append(f"可挑战排名: 第8-10名")
    else:
        lines.append(f"\n你的排名: 第{player_rank}名")
        # 显示可挑战的排名范围
        min_challenge = max(1, player_rank - 3)
        if min_challenge < player_rank:
            lines.append(f"可挑战排名: 第{min_challenge}-{player_rank-1}名")
        else:
            lines.append("你已是第1名，无法被挑战")
    
    reply = "\n".join(lines)
    if group_id and user_id:
        reply = f"[CQ:at,qq={user_id}]\n{reply}"
    send_message(reply, user_id, group_id)
    
    return jsonify({"status": "success", "message": "显示排行榜"})

def challenge_rank(user_id: str, group_id, target_rank: int):
    """挑战指定排名的玩家/AI"""
    try:
        log_info(f"开始挑战: user_id={user_id}, target_rank={target_rank}")
        
        ranking = load_ranking()
        log_info(f"排行榜加载完成，共{len(ranking)}个条目")
        
        if target_rank < 1 or target_rank > 10:
            reply = "无效的排名！请挑战1-10名之间的对手"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "无效的排名"})
        
        player_rank = get_player_ranking(user_id)
        log_info(f"玩家排名: {player_rank}")
        
        # 检查是否可以挑战
        if player_rank <= 10:
            min_challenge_rank = max(1, player_rank - 3)
            if target_rank < min_challenge_rank or target_rank >= player_rank:
                reply = f"只能挑战排名比你高且不超过3位的对手！\n你的排名: 第{player_rank}名\n可挑战排名: 第{min_challenge_rank}-{player_rank-1}名"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "无法挑战该排名"})
        else:
            if target_rank < 8:
                reply = f"你还未进入排行榜，只能挑战第8-10名！\n你的排名: 第{player_rank}名"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "无法挑战该排名"})
        
        target_entry = None
        for entry in ranking:
            if entry["rank"] == target_rank:
                target_entry = entry
                break
        
        if not target_entry:
            reply = "未找到该排名的对手！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "未找到对手"})
        
        log_info(f"目标对手: {target_entry['nickname']}, is_ai={target_entry['is_ai']}")
        
        # 获取玩家队伍 + 活跃预设
        player_team_data = get_user_team(user_id)
        active_slot = 0
        try:
            pdata = load_presets(user_id)
            active_slot = pdata.get("active_slot", 0)
        except Exception:
            pass
        log_info(f"玩家队伍数据类型: {type(player_team_data)}")
        
        if player_team_data:
            log_info(f"玩家队伍 battle_cards: {player_team_data.get('battle_cards')}")
        
        if not player_team_data or not player_team_data.get("battle_cards"):
            reply = "请先配置队伍！使用【队伍】命令"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "未配置队伍"})
        
        # 获取敌方队伍（人类玩家用当前队伍，AI用存储队伍）
        if target_entry["is_ai"]:
            enemy_team = target_entry["team"]
        else:
            enemy_team = load_team_data(target_entry["user_id"])
        log_info(f"敌方队伍类型: {type(enemy_team)}")
        if enemy_team:
            log_info(f"敌方队伍 battle_cards: {enemy_team.get('battle_cards')}")
        
        # 开始战斗
        enemy_name = target_entry["nickname"]
        
        # 生成并发送双方VS配队图
        try:
            characters = get_characters()
            vs_img = build_vs_team_image(player_team_data, enemy_team, characters)
            if vs_img:
                send_image(vs_img, user_id, group_id)
            else:
                # 如果无法生成图片，显示文字信息
                enemy_battle_cards = enemy_team.get("battle_cards", [])
                enemy_assist_cards = enemy_team.get("assist_cards", [])
                characters = get_characters()
                char_dict = {c["card_id"]: c for c in characters}
                
                enemy_team_display = f"👥 {enemy_name} 的队伍:\n"
                for i in range(6):
                    battle_name = "空"
                    assist_name = "空"
                    
                    if i < len(enemy_battle_cards) and enemy_battle_cards[i]:
                        card_info = char_dict.get(str(enemy_battle_cards[i]))
                        if card_info:
                            battle_name = card_info.get("name", str(enemy_battle_cards[i]))
                        else:
                            battle_name = str(enemy_battle_cards[i])
                    
                    if i < len(enemy_assist_cards) and enemy_assist_cards[i]:
                        card_info = char_dict.get(str(enemy_assist_cards[i]))
                        if card_info:
                            assist_name = card_info.get("name", str(enemy_assist_cards[i]))
                        else:
                            assist_name = str(enemy_assist_cards[i])
                    
                    enemy_team_display += f" 位置{i+1}: {battle_name} + {assist_name}\n"
                
                send_message(enemy_team_display.strip(), user_id, group_id)
        except Exception as e:
            log_error(f"生成敌方配队图片失败: {e}")
        
        reply = f"⚔️ 正在挑战 {enemy_name}（排名第{target_rank}）..."
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        
        # 执行战斗
        log_info("开始执行战斗...")
        result = BATTLE_INSTANCE.start_battle(player_team_data, enemy_team, challenger="player")
        log_info(f"战斗结果: {result}")

        # 保存战斗日志（滚动保留最近3次）
        save_rolling_battle_log(user_id, result)

        winner = result["winner"]
        # 将user_id转换为字符串再切片
        user_id_str = str(user_id)
        player_nickname = f"玩家{user_id_str[:4]}****"
        
        if winner == "player":
            update_ranking_after_battle(user_id, target_entry["user_id"], False, target_entry["is_ai"], player_team_data, player_nickname)
        else:
            update_ranking_after_battle(target_entry["user_id"], user_id, target_entry["is_ai"], False)
        
        rounds = result["rounds"]
        if winner == "player":
            new_rank = get_player_ranking(user_id)
            result_text = f"🏆 胜利！你击败了 {enemy_name}！\n🎉 你的新排名: 第{new_rank}名"
        else:
            result_text = f"💀 失败... 你被 {enemy_name} 击败了..."
        
        player_alive = sum(1 for u in result["player_units"] if u["alive"] and not u["is_assist"])
        enemy_alive = sum(1 for u in result["enemy_units"] if u["alive"] and not u["is_assist"])
        result_text += f"\n📊 我方存活 {player_alive}/6, 敌方存活 {enemy_alive}/6"
        if active_slot > 0:
            result_text += f"\n📋 使用预设: 槽{active_slot}"
        result_text += "\n💡 输入「战斗日志」查看详细记录"

        send_message(result_text, user_id, group_id)
        show_ranking(user_id, group_id)
        
        return jsonify({"status": "success", "message": "挑战完成", "winner": winner})
        
    except Exception as e:
        import traceback
        log_error(f"挑战失败: {e}\n{traceback.format_exc()}")
        reply = f"挑战失败: {str(e)}"
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        return jsonify({"status": "error", "message": f"挑战失败: {str(e)}"})


def handle_battle_log(user_id: str, group_id):
    """显示最近一场战斗的详细日志"""
    try:
        log_path = INFO_DIR / f"battle_{user_id}.json"
        if not log_path.exists():
            reply = "你还没有进行过战斗！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "没有战斗日志"})

        with open(log_path, "r", encoding="utf-8") as f:
            all_logs = json.load(f)
        if not isinstance(all_logs, list) or not all_logs:
            reply = "你还没有进行过战斗！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "没有战斗日志"})

        # 读取最近一次战斗日志
        result = all_logs[-1]

        # 判断是BOSS战还是普通战斗，使用对应的格式化函数
        if "boss_name" in result:
            log_text = format_boss_result(result)
        else:
            log_text = format_battle_result(result)
        # 压缩空行
        import re
        log_text = re.sub(r'\n{3,}', '\n\n', log_text)

        at_message = f"[CQ:at,qq={user_id}] " if group_id and user_id else ""
        # QQ消息限制约5000字符，单条发送
        if len(log_text) > 4000:
            log_text = log_text[:4000] + "\n... (日志过长已截断)"
        send_message(at_message + log_text, user_id, group_id)
        
        return jsonify({"status": "success", "message": "显示战斗日志"})
    
    except Exception as e:
        log_error(f"获取战斗日志失败: {e}")
        reply = f"获取战斗日志失败: {str(e)}"
        if group_id and user_id:
            reply = f"[CQ:at,qq={user_id}] {reply}"
        send_message(reply, user_id, group_id)
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_team(user_id: str, group_id, raw_message: str):
    """处理配队相关命令"""
    try:
        characters = get_characters()
        
        # 解析命令
        # 队伍 - 显示当前队伍（只显示图片）
        # 队伍 我的卡 - 显示三星卡图（50张，可翻页，无文字卡名）
        # 队伍 我的卡 下一页/上一页 - 翻页查看三星卡
        # 队伍 设置 位置 序号(1-50) - 根据当前页序号设置卡牌
        # 队伍 设置 战斗位/支援位 位置 序号 - 手动指定类型
        # 队伍 清除 位置 - 清除该位置的战斗卡和支援卡
        # 队伍 清空 - 清空所有队伍配置
        # 队伍 自动配队 - AI自动配队（B+A同色同攻击类型，FES优先）
        # 队伍 切换 N - 加载预设N (1-6)，设为活跃槽位，后续编辑自动保存到此槽
        # 队伍 预设 - 查看所有预设摘要
        
        # 获取用户当前查看的页码（从session或默认第1页）
        team_session_file = INFO_DIR / f"team_session_{user_id}.json"
        current_page = 1
        if team_session_file.exists():
            try:
                with open(team_session_file, "r", encoding="utf-8") as f:
                    session_data = json.load(f)
                    current_page = session_data.get("cards_page", 1)
            except:
                current_page = 1
        
        # 自动配队命令
        if '自动配队' in raw_message or '自动' in raw_message:
            result = auto_build_team(user_id, characters)

            if not result["success"]:
                reply = result["message"]
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": result["message"]})

            # 只发队伍图片 + 简短提示
            team_data = load_team_data(user_id)
            img_path = build_team_image(team_data, characters)
            if img_path and os.path.exists(img_path):
                with open(img_path, "rb") as f:
                    image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
                at = f"[CQ:at,qq={user_id}] " if group_id and user_id else ""
                msg = f"{at}🤖 自动配队完成\n[CQ:image,file=base64://{image_base64}]"
                os.remove(img_path)
            else:
                at = f"[CQ:at,qq={user_id}] " if group_id and user_id else ""
                msg = f"{at}🤖 自动配队完成"
            send_message(msg, user_id, group_id)
            return jsonify({"status": "success"})
        
        if '我的卡' in raw_message:
            # 先获取总页数（用于验证页码）
            user_cards = get_user_3star_cards(user_id, characters)
            total_pages = max(1, (len(user_cards) + 50 - 1) // 50)
            
            # 处理翻页
            import re
            # 检查是否跳转到指定页码（如"队伍 我的卡 第3页"或"队伍 我的卡 3"）
            page_match = re.search(r'我的卡\s+(第)?(\d+)(页)?', raw_message)
            if page_match:
                target_page = int(page_match.group(2))
                current_page = max(1, min(target_page, total_pages))
            elif '下一页' in raw_message:
                current_page += 1
                if current_page > total_pages:
                    current_page = total_pages
            elif '上一页' in raw_message:
                current_page -= 1
                if current_page < 1:
                    current_page = 1
            
            # 确保页码在有效范围内
            current_page = max(1, min(current_page, total_pages))
            
            # 保存当前页码
            with open(team_session_file, "w", encoding="utf-8") as f:
                json.dump({"cards_page": current_page}, f)
            
            # 显示用户拥有的三星卡（50张一页，只显示图片，无文字卡名）
            img_path, current_cards, total_pages = build_3star_cards_image(user_id, characters, current_page, 50)
            
            if not current_cards:
                reply = "你还没有三星卡~"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success", "message": "没有三星卡"})
            
            # 构建消息（只有图片和页码提示）
            page_info = f"第{current_page}/{total_pages}页"
            if total_pages > 1:
                if current_page < total_pages:
                    page_info += " | 输入「队伍 我的卡 下一页」查看下一页"
                if current_page > 1:
                    page_info += " | 输入「队伍 我的卡 上一页」查看上一页"
                # 添加跳转到指定页码的提示
                page_info += " | 输入「队伍 我的卡 页码」跳转到指定页"
            
            # 使用提示（根据当前页实际卡牌数量）
            current_page_size = len(current_cards)
            usage_hint = (f"设置: 队伍 设置 位置 序号(1-{current_page_size}) | "
                          f"切换预设: 队伍 切换 1~6")
            
            if img_path and os.path.exists(img_path):
                with open(img_path, "rb") as f:
                    image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
                img_message = f"[CQ:image,file=base64://{image_base64}]"
                
                # 发送图片和页码提示
                if group_id and user_id:
                    at_message = f"[CQ:at,qq={user_id}] "
                else:
                    at_message = ""
                
                full_message = f"{at_message}{img_message}\n{page_info}\n{usage_hint}"
                send_message(full_message, user_id, group_id)
                
                if os.path.exists(img_path):
                    os.remove(img_path)
            else:
                reply = f"{page_info}\n{usage_hint}"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
            
            return jsonify({"status": "success", "message": "显示三星卡", "page": current_page, "total_pages": total_pages})
        
        elif '设置' in raw_message:
            # 设置队伍卡牌
            import re
            # 匹配格式1: 队伍 设置 位置 序号（使用当前页的序号1-50）
            match_simple = re.search(r'设置\s+(\d+)\s+(\d+)', raw_message)
            # 匹配格式2: 队伍 设置 战斗位/支援位 位置 序号
            match_full = re.search(r'设置\s+(战斗位|支援位)\s+(\d+)\s+(\d+)', raw_message)
            
            if match_simple and not match_full:
                # 简化格式：使用序号选择卡牌
                position = int(match_simple.group(1))
                card_index = int(match_simple.group(2))  # 序号1-50
                
                if position < 1 or position > 6:
                    reply = "队伍位置必须在1-6之间！"
                else:
                    # 获取当前页的卡牌列表
                    img_path, current_cards, total_pages = build_3star_cards_image(user_id, characters, current_page, 50)
                    
                    if card_index < 1:
                        reply = "序号必须大于0！"
                    elif card_index > len(current_cards):
                        reply = f"当前页只有{len(current_cards)}张卡，序号{card_index}无效！"
                    else:
                        card_info = current_cards[card_index - 1]
                        card_id = card_info.get("card_id")
                        card_type = card_info.get("type", "battle")
                        
                        success = set_team_card(user_id, position, card_id, card_type)
                        type_text = "战斗位" if card_type == "battle" else "支援位"
                        if success:
                            auto_save_preset(user_id)
                            reply = f"成功设置{type_text}{position}！"
                        else:
                            # 检查是否因为重复导致失败
                            team_data = load_team_data(user_id)
                            all_cards = (team_data.get("battle_cards", []) +
                                        team_data.get("assist_cards", []))
                            already_used = any(
                                c and str(c) == str(card_id)
                                for i, c in enumerate(all_cards)
                                if not (i == position - 1 and card_type == "battle") and
                                   not (i == position + 5 and card_type == "assist")
                            )
                            if already_used:
                                reply = "设置失败！该卡牌已在队伍的其他位置使用，不能重复选择同一张卡~"
                            else:
                                reply = "设置失败！"
                            
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success" if '成功' in reply else "error", "message": reply})
            
            elif match_full:
                # 完整格式：手动指定类型
                card_type = "battle" if match_full.group(1) == "战斗位" else "assist"
                position = int(match_full.group(2))
                card_index = int(match_full.group(3))  # 序号1-10
                
                if position < 1 or position > 6:
                    reply = "队伍位置必须在1-6之间！"
                elif card_index < 1 or card_index > 10:
                    reply = "序号必须在1-10之间！"
                else:
                    # 获取当前页的卡牌列表
                    img_path, current_cards, total_pages = build_3star_cards_image(user_id, characters, current_page, 10)
                    
                    if card_index > len(current_cards):
                        reply = f"当前页只有{len(current_cards)}张卡，序号{card_index}无效！"
                    else:
                        card_info = current_cards[card_index - 1]
                        card_id = card_info.get("card_id")
                        
                        success = set_team_card(user_id, position, card_id, card_type)
                        if success:
                            auto_save_preset(user_id)
                            reply = f"成功设置{match_full.group(1)}{position}！"
                        else:
                            # 检查是否因为重复导致失败
                            team_data = load_team_data(user_id)
                            all_cards = (team_data.get("battle_cards", []) +
                                        team_data.get("assist_cards", []))
                            already_used = any(
                                c and str(c) == str(card_id)
                                for i, c in enumerate(all_cards)
                                if not (i == position - 1 and card_type == "battle") and
                                   not (i == position + 5 and card_type == "assist")
                            )
                            if already_used:
                                reply = "设置失败！该卡牌已在队伍的其他位置使用，不能重复选择同一张卡~"
                            else:
                                reply = "设置失败！"
                
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success" if success else "error", "message": reply})
            
            else:
                reply = "格式错误！正确格式：队伍 设置 位置 序号(1-10)"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "格式错误"})
        
        elif '清除' in raw_message:
            # 清除指定位置的卡牌
            import re
            match = re.search(r'清除\s+(\d+)', raw_message)
            
            if match:
                position = int(match.group(1))
                
                if position < 1 or position > 6:
                    reply = "位置必须在1-6之间！"
                else:
                    # 清除战斗位和支援位的对应位置
                    clear_team_card(user_id, position, "battle")
                    clear_team_card(user_id, position, "assist")
                    reply = f"成功清除位置{position}的战斗卡和支援卡！"
                
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success", "message": reply})
            
            else:
                reply = "格式错误！正确格式：队伍 清除 位置"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "格式错误"})
        
        elif '清空' in raw_message:
            # 清空整个队伍
            clear_all_team(user_id)
            reply = "已清空所有队伍配置！"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "队伍已清空"})

        elif '切换' in raw_message:
            # 切换预设: 队伍 切换 N
            import re
            m = re.search(r'切换\s*(\d)', raw_message)
            if m:
                slot = int(m.group(1))
                if load_preset(user_id, slot):
                    reply = f"已切换到预设{slot}！"
                    # 显示新队伍
                    team_data = load_team_data(user_id)
                    img_path = build_team_image(team_data, characters)
                    if img_path and os.path.exists(img_path):
                        with open(img_path, "rb") as f:
                            image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
                        reply += f"\n[CQ:image,file=base64://{image_base64}]"
                        os.remove(img_path)
                else:
                    # 预设为空 → 自动配队并保存到此槽位
                    reply = f"预设{slot}为空，正在自动配队...\n"
                    result = auto_build_team(user_id, characters)
                    if result["success"] and result["team"]:
                        # 保存自动配队结果到当前队伍
                        save_team_data(user_id, result["team"])
                        # 保存到此预设槽位
                        presets_data = load_presets(user_id)
                        presets_data["presets"][slot - 1] = {
                            "battle_cards": list(result["team"].get("battle_cards", [])),
                            "assist_cards": list(result["team"].get("assist_cards", []))
                        }
                        presets_data["active_slot"] = slot
                        save_presets(user_id, presets_data)
                        reply += f"预设{slot}自动配队完成！"
                        # 显示新队伍
                        img_path = build_team_image(result["team"], characters)
                        if img_path and os.path.exists(img_path):
                            with open(img_path, "rb") as f:
                                image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
                            reply += f"\n[CQ:image,file=base64://{image_base64}]"
                            os.remove(img_path)
                    else:
                        reply += result.get("message", "自动配队失败，请先抽卡！")
            else:
                reply = "格式：队伍 切换 1~6"
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}] {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": reply})

        elif '预设' in raw_message:
            # 查看所有预设
            info = list_presets_info(user_id, characters)
            reply = info
            if group_id and user_id:
                reply = f"[CQ:at,qq={user_id}]\n{reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "显示预设列表"})

        else:
            # 显示当前队伍（只显示图片，不显示文字信息）
            team_data = load_team_data(user_id)
            
            # 生成队伍图片
            img_path = build_team_image(team_data, characters)
            
            if img_path and os.path.exists(img_path):
                with open(img_path, "rb") as f:
                    image_base64 = base64.b64encode(compress_image_bytes(f.name)).decode("utf-8")
                img_message = f"[CQ:image,file=base64://{image_base64}]"
                
                # 只发送图片，不发送文字信息
                if group_id and user_id:
                    at_message = f"[CQ:at,qq={user_id}] "
                else:
                    at_message = ""
                
                hints = ("队伍 预设 | 队伍 切换 1~6 | "
                         "队伍 我的卡 | 队伍 自动配队")
                full_message = f"{at_message}{img_message}\n{hints}"
                send_message(full_message, user_id, group_id)
                
                if os.path.exists(img_path):
                    os.remove(img_path)
            else:
                # 如果没有图片，发送提示
                reply = "队伍配置为空！"
                if group_id and user_id:
                    reply = f"[CQ:at,qq={user_id}] {reply}"
                send_message(reply, user_id, group_id)
            
            return jsonify({"status": "success", "message": "显示队伍"})
    
    except Exception as e:
        log_error(f"配队处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


def handle_help(user_id: str, group_id):
    """返回帮助信息"""
    help_text = f"""
╔══════════════════════════════╗
║     ifBot 帮助      ║
╠══════════════════════════════╣
║ 单抽  - ({GACHA_COST}呱太) ║
║ 十连  - ({GACHA10_COST}呱太) ║
║ 获取呱太 - 获得{GET_GACHA_REWARD}呱太   ║
║ 签到  - 每日签到({DAILY_REWARD}呱太)  ║
║ 个人记录 - 查看个人统计    ║
║ 兑换呱太 - 用碎片兑换呱太      ║
║ 排行榜 - 查看战力排行榜        ║
║ 三星池子 - 查看三星only池介绍    ║
║ 队伍  - 查看当前队伍        ║
║ 帮助  - 显示此帮助         ║
╠══════════════════════════════╣                   ║
║ 战斗 @  - 与某个玩家战斗  ║
║ 战斗  - 与随机AI战斗  ║
╠══════════════════════════════╣
║ 配队命令:                      ║
║ 队伍 我的卡 - 用于配队          ║
║ 队伍 设置 位置 序号   ║
║ 队伍 清除 位置             ║
║ 队伍 清空 - 清空队伍       ║
╚══════════════════════════════╝
"""
    # 发送帮助消息
    send_message(help_text.strip(), user_id, group_id)
    
    return jsonify({
        "status": "success",
        "message": help_text.strip()
    })


def handle_cute_reply(user_id: str, group_id):
    """被艾特但不是有效命令时，发送卖萌回复"""
    cute_replies = [
        "喵~ 你叫我吗？😺",
        "来了来了~ 有什么事呀？✨",
        "呜哇！被艾特了好开心~ 🥰",
        "主人~ 你在叫我吗？🐾",
        "嗨~ 找我有事吗？😊",
        "嗷呜~ 我在这里！🐺",
        "你好呀~ 今天也要元气满满哦！☀️",
        "咕噜咕噜~ 你是在叫我吗？🐱",
        "嗯嗯~ 我在听呢！👂",
        "好耶！终于被注意到了~ 🎉",
        "哇~ 有人艾特我！开心~ 😄",
        "喵星人收到信号！📡",
    ]
    
    # 随机选择一个卖萌回复
    reply = random.choice(cute_replies)
    
    # 如果是群聊，艾特用户
    if group_id and user_id:
        reply = f"[CQ:at,qq={user_id}] {reply}"
    
    # 发送回复
    send_message(reply, user_id, group_id)
    
    return jsonify({
        "status": "success",
        "message": reply
    })


# ========== 启动 ==========
if __name__ == '__main__':
    log_info("=" * 50)
    log_info("自动抽卡Bot 启动中...")
    log_info(f"图标目录: {ICON_DIR}")
    log_info(f"星级图片目录: {LEVEL_DIR}")
    log_info(f"Excel文件: {XLSX_FILE}")
    log_info("=" * 50)

    # 预加载抽卡角色数据
    characters = preload_characters()

    # 初始化战斗系统（使用详细战斗数据）
    if BATTLE_SYSTEM_LOADED:
        battle_chars = load_battle_characters()
        BATTLE_INSTANCE = BattleSystem(battle_chars)
        log_info("战斗系统初始化完成")

    # 恢复并记录日活数据
    restore_daily_stats()
    record_daily_dau()

    # 启动Flask服务
    app.run(
        host=FLASK_HOST,
        port=FLASK_PORT,
        debug=False
    )
