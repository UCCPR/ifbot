"""
自动抽卡QQ Bot (WebSocket版) - 基于botpy官方SDK
无需Webhook/Cloudflare/HTTPS, 直接WebSocket连接
用法: python3 qq_bot_ws.py
"""

import os
import sys
import random
import json
import asyncio
import threading
import re as _re
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
import atexit

from PIL import Image
import openpyxl

# botpy WebSocket SDK
import botpy
from botpy import Client, BotAPI, Intents
from botpy.message import Message, DirectMessage, C2CMessage, GroupMessage

# 导入配队系统（延迟到log_info定义后再记录日志）
try:
    from team_system import (
        load_team_data,
        save_team_data,
        get_user_3star_cards,
        build_team_image,
        build_vs_team_image,
        build_3star_cards_image,
        set_team_card,
        clear_team_card,
        clear_all_team,
        get_team_info,
        auto_build_team,
        auto_save_preset,
        load_preset,
        load_presets,
        save_presets,
        list_presets_info,
        get_defense_slot,
        set_defense_slot,
        get_defense_team,
        get_defense_team_info,
        copy_team
    )
    TEAM_SYSTEM_LOADED = True
except ImportError as e:
    raise ImportError(f"team_system.py is required but failed to load: {e}") from e


# 导入战斗系统
try:
    from battle_system import BattleSystem, format_battle_result, format_boss_result, get_battle_help
    BATTLE_SYSTEM_LOADED = True
    BATTLE_INSTANCE = None  # 战斗系统实例
    BATTLE_CHARACTERS = {}  # 战斗角色数据缓存（card_id -> dict）
    _GIF_COOLDOWN = {}  # GIF生成冷却 {user_id: timestamp}
except ImportError as e:
    raise ImportError(f"battle_system.py is required but failed to load: {e}") from e



# ========== 配置 ==========
BASE_DIR = Path(__file__).parent
ICON_DIR = BASE_DIR / "iconimage"
LEVEL_DIR = BASE_DIR / "level"
XLSX_FILE = BASE_DIR / "卡牌信息.xlsx"         # 抽卡卡池（含1/2/3星）
BATTLE_XLSX = BASE_DIR / "cards_completed.xlsx"  # 战斗数值（3星详细数据）
INFO_DIR = BASE_DIR / "info"
OUTPUT_DIR = BASE_DIR / "output"
BACKUP_DIR = BASE_DIR / "backup"

# 确保目录存在
INFO_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

# 强制 stdout 使用 UTF-8，避免 Windows GBK 控制台下 emoji 等字符抛出 UnicodeEncodeError
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ========== 日志模块（必须在其他模块之前定义）==========
def _safe_print(prefix: str, message: str):
    """安全打印，防止 Windows GBK 控制台因 emoji 等字符抛出 UnicodeEncodeError"""
    try:
        print(f"[{prefix}] {message}", flush=True)
    except UnicodeEncodeError:
        # Windows GBK 控制台无法编码 emoji，用 ASCII 安全方式输出
        safe_msg = message.encode("ascii", errors="replace").decode("ascii")
        try:
            print(f"[{prefix}] {safe_msg}", flush=True)
        except Exception:
            pass  # 最后的兜底，放弃控制台输出


def log_info(message: str):
    """记录普通信息"""
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "gacha_info.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")
    _safe_print("INFO", message)


def log_error(message: str):
    """记录错误信息"""
    timestamp = datetime.now().strftime("%m-%d %H:%M:%S")
    log_file = INFO_DIR / "gacha_error.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] ERR {message}\n")
    _safe_print("ERROR", message)


# ========== 抽卡记录备份模块 ==========
BACKUP_RECORD_FILE = INFO_DIR / "last_backup.json"

def get_last_backup_date() -> str:
    """获取最后一次备份的日期"""
    if BACKUP_RECORD_FILE.exists():
        try:
            with open(BACKUP_RECORD_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("last_backup_date", "")
        except:
            return ""
    return ""

def set_last_backup_date(date_str: str):
    """设置最后一次备份的日期"""
    _atomic_json_save(BACKUP_RECORD_FILE, {"last_backup_date": date_str})

def backup_pity_records():
    """备份所有抽卡记录（每天第一次启动时执行）"""
    today = datetime.now().strftime("%Y-%m-%d")
    last_backup = get_last_backup_date()
    
    if last_backup == today:
        log_info(f"今日({today})已备份过抽卡记录，跳过")
        return False
    
    try:
        # 创建今日备份目录
        today_backup_dir = BACKUP_DIR / today
        today_backup_dir.mkdir(exist_ok=True)
        
        # 备份 pity_*.json 文件（抽卡记录）
        pity_files = list(INFO_DIR.glob("pity_*.json"))
        for pity_file in pity_files:
            dest_file = today_backup_dir / pity_file.name
            with open(pity_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 fes_stats.json（FES统计）
        fes_stats = INFO_DIR / "fes_stats.json"
        if fes_stats.exists():
            dest_file = today_backup_dir / fes_stats.name
            with open(fes_stats, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 gacha_*.json 文件（呱太数据）
        gacha_files = list(INFO_DIR.glob("gacha_*.json"))
        for gacha_file in gacha_files:
            dest_file = today_backup_dir / gacha_file.name
            with open(gacha_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 signin_*.json 文件（签到数据）
        signin_files = list(INFO_DIR.glob("signin_*.json"))
        for signin_file in signin_files:
            dest_file = today_backup_dir / signin_file.name
            with open(signin_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 备份 ranking.json（排行榜）
        ranking_file = INFO_DIR / "ranking.json"
        if ranking_file.exists():
            dest_file = today_backup_dir / ranking_file.name
            with open(ranking_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())

        # 备份 dau_log.json（日活记录）
        dau_file = INFO_DIR / "dau_log.json"
        if dau_file.exists():
            dest_file = today_backup_dir / dau_file.name
            with open(dau_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())

        # 备份 team_*.json 文件（队伍配置）
        team_files = list(INFO_DIR.glob("team_*.json"))
        for team_file in team_files:
            dest_file = today_backup_dir / team_file.name
            with open(team_file, "rb") as src:
                with open(dest_file, "wb") as dst:
                    dst.write(src.read())
        
        # 更新备份记录
        set_last_backup_date(today)
        
        total_files = len(pity_files) + len(gacha_files) + len(signin_files) + len(team_files)
        if fes_stats.exists(): total_files += 1
        if ranking_file.exists(): total_files += 1
        if dau_file.exists(): total_files += 1

        log_info(f"数据备份完成！日期: {today}, 文件数: {total_files}")
        return True
        
    except Exception as e:
        log_error(f"抽卡记录备份失败: {e}")
        return False


# 从配置文件导入配置信息
# 请在 config.py 文件中填写实际的配置值
try:
    from config import (
        QQ_BOT_APP_ID,
        QQ_BOT_TOKEN,
        QQ_BOT_SECRET,
        IMAGE_HOST,
        FLASK_HOST,
        FLASK_PORT,
        # 保底配置
        PITY_LIMIT,
        # 呱太配置
        GACHA_COST,
        GACHA10_COST,
        GACHA10_COOLDOWN_SECONDS,
        GET_GACHA_COOLDOWN_SECONDS,
        LIMITED_GACHA_COST,
        LIMITED_GACHA_COOLDOWN_SECONDS,
        GET_GACHA_REWARD,
        DAILY_REWARD,
        # 开箱配置
        MYSTERY_BOX_CHANCE,
        MUTATION_NO_CHANGE,
        MUTATION_1_TO_2,
        MUTATION_1_TO_3,
        MUTATION_2_TO_3,
        BOX_OPEN_TIMEOUT,
        # 三星池子配置
        THREE_STAR_POOL_RED_COST,
        THREE_STAR_POOL_BLUE_COST,
        # 抽卡概率（三星内部分配）
        FES_LIMIT_PROB,
        PERIOD_LIMIT_PROB,
        OTHER_3STAR_PROB,
        # 盲盒星级概率
        MYSTERY_BOX_2STAR_PROB,
        MYSTERY_BOX_3STAR_PROB,
        NORMAL_BOX_1STAR_PROB,
        NORMAL_BOX_2STAR_PROB,
        NORMAL_BOX_3STAR_PROB,
        # 抽卡星级概率（呱太抽卡）
        GACHA_1STAR_PROB,
        GACHA_2STAR_PROB,
        GACHA_3STAR_PROB,
        # 管理员
        ADMIN_QQ,
        # BOSS战
        BOSS_BATTLE_COOLDOWN_SECONDS,
        # CDKEY
        CDKEYS
    )
except ImportError:
    # 如果配置文件不存在，使用默认值
    QQ_BOT_APP_ID = ""
    QQ_BOT_TOKEN = ""
    QQ_BOT_SECRET = ""
    IMAGE_HOST = ""
    FLASK_HOST = "0.0.0.0"
    FLASK_PORT = 5000
    # 保底配置
    PITY_LIMIT = 150  # 150抽必出フェス限定三星
    # 呱太配置
    GACHA_COST = 300        # 单抽价格
    GACHA10_COST = 3000     # 十连价格
    GACHA10_COOLDOWN_SECONDS = 60  # 十连冷却时间（秒）
    GET_GACHA_COOLDOWN_SECONDS = 60  # 获取呱太冷却时间（秒）
    LIMITED_GACHA_COST = 15000  # 限定池价格
    LIMITED_GACHA_COOLDOWN_SECONDS = 600  # 限定池冷却 10分钟
    GET_GACHA_REWARD = 10000  # 获取呱太奖励
    DAILY_REWARD = 30000     # 每日签到奖励
    # 开箱配置
    MYSTERY_BOX_CHANCE = 0.02  # 黑色盲盒概率
    MUTATION_NO_CHANGE = 0.88  # 不突变概率
    MUTATION_1_TO_2 = 0.08    # 1星→2星概率
    MUTATION_1_TO_3 = 0.02    # 1星→3星概率
    MUTATION_2_TO_3 = 0.05    # 2星→3星概率
    BOX_OPEN_TIMEOUT = 300     # 盲盒开启超时时间（秒）
    # 三星池子配置
    THREE_STAR_POOL_RED_COST = 1500   # 红色碎片消耗
    THREE_STAR_POOL_BLUE_COST = 350   # 蓝色碎片消耗
    # 抽卡概率（三星内部分配）
    FES_LIMIT_PROB = 0.25     # フェス限定概率
    PERIOD_LIMIT_PROB = 0.35  # 期間限定概率
    OTHER_3STAR_PROB = 0.40   # 其他三星概率
    # 不跳过的十连和单抽星级概率
    MYSTERY_BOX_2STAR_PROB = 65  # 黑色盲盒2星概率（权重）
    MYSTERY_BOX_3STAR_PROB = 35  # 黑色盲盒3星概率（权重）
    NORMAL_BOX_1STAR_PROB = 72   # 正常盲盒1星概率（权重）
    NORMAL_BOX_2STAR_PROB = 23   # 正常盲盒2星概率（权重）
    NORMAL_BOX_3STAR_PROB = 3    # 正常盲盒3星概率（权重）
    # 十连跳过概率
    GACHA_1STAR_PROB = 72   # 1星概率（权重）
    GACHA_2STAR_PROB = 23   # 2星概率（权重）
    GACHA_3STAR_PROB = 3    # 3星概率（权重）
    # BOSS战
    BOSS_BATTLE_COOLDOWN_SECONDS = 60
    ADMIN_QQ = ""  # 管理员QQ
    # CDKEY
    CDKEYS = {}

# 定义概率权重常量（避免代码重复）
GACHA_WEIGHTS = [GACHA_1STAR_PROB, GACHA_2STAR_PROB, GACHA_3STAR_PROB]
NORMAL_BOX_WEIGHTS = [NORMAL_BOX_1STAR_PROB, NORMAL_BOX_2STAR_PROB, NORMAL_BOX_3STAR_PROB]
MYSTERY_BOX_WEIGHTS = [MYSTERY_BOX_2STAR_PROB, MYSTERY_BOX_3STAR_PROB]

# 定义角色图裁剪比例常量
CROP_LEFT_RATIO = 0.25
CROP_RIGHT_RATIO = 0.75
CROP_TOP_RATIO = 0.15
CROP_BOTTOM_RATIO = 0.65

BOT_API: BotAPI = None  # botpy API句柄，连接后设置
# IMAGE_HOST 从 config.py 导入，此处不覆盖
# 兼容 Flask jsonify（botpy WebSocket 版无条件返回 None，所有消息通过 send_message 发送）
def jsonify(obj=None, **kwargs):
    """[DEPRECATED] 所有消息已通过 send_message 主动发送，返回 None 避免 botpy 重复回复"""
    return None


# ========== 全局变量 ==========
CHARACTERS_CACHE = None  # 预加载的角色数据缓存
BOX_SESSIONS = {}  # 盲盒会话状态 {user_id: {"boxes": [], "opened": [], "created_at": timestamp}}
MESSAGE_COUNTER = {}  # 消息发送计数器 {target_id: {"count": int, "start_time": timestamp}}
MAX_MESSAGES_PER_MINUTE = 30  # 每分钟最大消息发送量
DAILY_SEND_STATS = {}  # 每日发送消息统计 {日期: 发送次数}
DAILY_USER_SET = set()  # 今日活跃用户ID集合
_RESTORED_DAU = 0       # 重启后从磁盘恢复的日活基准数
GACHA10_COOLDOWN = {}  # 十连冷却时间 {user_id: last_gacha10_timestamp}
GET_GACHA_COOLDOWN = {}  # 获取呱太冷却时间 {user_id: last_get_gacha_timestamp}
LIMITED_GACHA_COOLDOWN = {}  # 限定池冷却时间 {user_id: last_limited_gacha_timestamp}
BOSS_BATTLE_COOLDOWN = {}  # BOSS战冷却时间 {user_id: last_boss_battle_timestamp}
USER_DATA = {}  # 用户数据缓存 {user_id: data}


# ========== 预加载模块 ==========
def preload_characters():
    """预加载角色数据到缓存，避免每次抽卡都重新读取"""
    global CHARACTERS_CACHE
    if CHARACTERS_CACHE is None:
        log_info("开始预加载角色数据...")
        CHARACTERS_CACHE = load_character_data()
        log_info(f"角色数据预加载完成，共 {len(CHARACTERS_CACHE)} 个角色")
    return CHARACTERS_CACHE


def get_characters():
    """获取预加载的角色数据（如果未预加载则自动加载）"""
    return preload_characters()


def get_characters_dict():
    """获取角色数据字典（card_id -> data），用于 extra_characters"""
    chars_list = get_characters()
    result = {}
    for c in chars_list:
        cid = c.get("card_id") or c.get("id", "")
        if cid:
            result[str(cid)] = c
    return result


# 记录配队系统加载状态
if TEAM_SYSTEM_LOADED:
    log_info("配队系统加载成功")
else:
    log_error("配队系统加载失败")


# ========== 抽卡记录存储模块 ==========
def get_pity_file(user_id: str) -> Path:
    """获取用户抽卡记录文件路径"""
    return INFO_DIR / f"pity_{user_id}.json"


def load_pity_data(user_id: str) -> dict:
    """加载用户的抽卡记录数据"""
    pity_file = get_pity_file(user_id)
    if pity_file.exists():
        try:
            with open(pity_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            log_error(f"用户 {user_id} 的抽卡记录文件格式错误: {e}")
            return get_default_pity_data()
        except IOError as e:
            log_error(f"读取用户 {user_id} 的抽卡记录文件失败: {e}")
            return get_default_pity_data()
    return get_default_pity_data()


def get_default_pity_data() -> dict:
    """获取默认的抽卡记录数据"""
    return {
        "total_draws": 0, 
        "pity_count": 0, 
        "fes_pity_count": 0,
        "total_3stars": 0,
        "red_crystal": 0,
        "blue_crystal": 0,
        "recent_3stars": [],  # 最近获得的三星卡记录，按时间排序，存储card_id
        "card_collection": {},  # 卡片收藏 {card_id: {"name": "", "stars": 3, "limit_type": "", "count": 1, "last_time": timestamp}}
        "fes_count": 0,        # フェス限定数量
        "period_count": 0,      # 期間限定数量
        "other_3star_count": 0, # 其他三星数量
        "total_2stars": 0       # 二星数量
    }


def save_pity_data(user_id: str, data: dict):
    """保存用户的抽卡记录数据"""
    _atomic_json_save(get_pity_file(user_id), data)


def get_remaining_pity(user_id: str) -> int:
    """获取用户距离フェス限定三星保底还剩多少抽"""
    pity_data = load_pity_data(user_id)
    fes_pity_count = pity_data.get("fes_pity_count", 0)
    return max(0, PITY_LIMIT - fes_pity_count)


def update_pity(user_id: str, got_3star: bool = False, is_fes_3star: bool = False):
    """更新用户的抽卡记录"""
    pity_data = load_pity_data(user_id)
    
    # 如果抽到3星，重置保底计数并增加三星个数
    if got_3star:
        pity_data["pity_count"] = 0
        pity_data["total_3stars"] = pity_data.get("total_3stars", 0) + 1
        # 如果抽到フェス限定三星，重置フェス保底计数
        if is_fes_3star:
            pity_data["fes_pity_count"] = 0
    else:
        pity_data["pity_count"] = pity_data.get("pity_count", 0) + 1
        pity_data["fes_pity_count"] = pity_data.get("fes_pity_count", 0) + 1
    
    pity_data["total_draws"] = pity_data.get("total_draws", 0) + 1
    save_pity_data(user_id, pity_data)
    return pity_data


def get_total_draws(user_id: str) -> int:
    """获取用户累计抽卡次数"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("total_draws", 0)


def get_total_3stars(user_id: str) -> int:
    """获取用户累计获得的三星个数"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("total_3stars", 0)


def get_red_crystal(user_id: str) -> int:
    """获取用户的红色碎片数量"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("red_crystal", 0)


def get_blue_crystal(user_id: str) -> int:
    """获取用户的蓝色碎片数量"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("blue_crystal", 0)


def add_red_crystal(user_id: str, amount: int):
    """增加用户的红色碎片"""
    pity_data = load_pity_data(user_id)
    pity_data["red_crystal"] = pity_data.get("red_crystal", 0) + amount
    save_pity_data(user_id, pity_data)
    return pity_data["red_crystal"]


def add_blue_crystal(user_id: str, amount: int):
    """增加用户的蓝色碎片"""
    pity_data = load_pity_data(user_id)
    pity_data["blue_crystal"] = pity_data.get("blue_crystal", 0) + amount
    save_pity_data(user_id, pity_data)
    return pity_data["blue_crystal"]


def spend_red_crystal(user_id: str, amount: int) -> bool:
    """消耗用户的红色碎片，返回是否成功"""
    pity_data = load_pity_data(user_id)
    current = pity_data.get("red_crystal", 0)
    if current >= amount:
        pity_data["red_crystal"] = current - amount
        save_pity_data(user_id, pity_data)
        return True
    return False


def spend_blue_crystal(user_id: str, amount: int) -> bool:
    """消耗用户的蓝色碎片，返回是否成功"""
    pity_data = load_pity_data(user_id)
    current = pity_data.get("blue_crystal", 0)
    if current >= amount:
        pity_data["blue_crystal"] = current - amount
        save_pity_data(user_id, pity_data)
        return True
    return False


# ========== 三星池子系统 ==========
# 三星池子配置
THREE_STAR_POOL_RED_COST = 1500  # 红色碎片消耗
THREE_STAR_POOL_BLUE_COST = 350  # 蓝色碎片消耗


def select_3star_from_pool(characters: list) -> dict:
    """从三星池子中随机抽取一个角色（只返回三星角色）"""
    three_star_chars = [c for c in characters if c.get("stars") == 3]
    
    if not three_star_chars:
        return None
    
    # 随机选择
    selected = random.choice(three_star_chars)
    return selected


def draw_3star_pool(user_id: str, crystal_type: str = "red") -> dict:
    """
    抽取三星池子
    :param user_id: 用户ID
    :param crystal_type: 碎片类型 "red" 或 "blue"
    :return: 抽卡结果
    """
    characters = get_characters()
    
    # 检查碎片是否足够
    if crystal_type == "red":
        cost = THREE_STAR_POOL_RED_COST
        if not spend_red_crystal(user_id, cost):
            return {
                "success": False,
                "message": f"红色碎片不足！需要{cost}个，当前拥有{get_red_crystal(user_id)}个"
            }
    else:
        cost = THREE_STAR_POOL_BLUE_COST
        if not spend_blue_crystal(user_id, cost):
            return {
                "success": False,
                "message": f"蓝色碎片不足！需要{cost}个，当前拥有{get_blue_crystal(user_id)}个"
            }
    
    # 抽取角色
    selected = select_3star_from_pool(characters)
    
    if not selected:
        return {
            "success": False,
            "message": "三星池子为空，请联系管理员！"
        }
    
    # 处理抽卡结果
    card_id = str(selected.get("card_id", ""))
    chara_id = selected.get("id", "")
    name = selected.get("name", "")
    limit_type = selected.get("limit_type", "")
    stars = selected.get("stars", 3)
    
    # FES统计和提示
    fes_message = ""
    if limit_type == "フェス限定":
        # 增加FES统计
        fes_count = increment_fes_count(card_id, name)
        fes_message = f"✨ 恭喜！这是全服第{fes_count}个「{name}」！"
    
    # 添加到用户卡库
    add_card_collection(user_id, card_id, name, stars, limit_type)
    
    # 添加到最近三星记录
    add_recent_3star(user_id, card_id, chara_id, name, limit_type)
    
    # 更新统计（三星池子抽出的都是三星）
    is_fes_3star = (limit_type == "フェス限定")
    update_pity(user_id, got_3star=True, is_fes_3star=is_fes_3star)
    
    return {
        "success": True,
        "character": selected,
        "message": f"恭喜获得三星角色：{name}",
        "fes_message": fes_message
    }


def add_recent_3star(user_id: str, card_id: str, chara_id: str, name: str, limit_type: str = None):
    """添加最近获得的三星卡记录（最多保存10个）"""
    pity_data = load_pity_data(user_id)
    recent = pity_data.get("recent_3stars", [])
    
    # 创建记录
    record = {
        "card_id": card_id,
        "chara_id": chara_id,
        "name": name,
        "limit_type": limit_type,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    # 添加到开头
    recent.insert(0, record)
    
    # 只保留最近10个
    if len(recent) > 10:
        recent = recent[:10]
    
    pity_data["recent_3stars"] = recent
    save_pity_data(user_id, pity_data)
    return recent


def get_recent_3stars(user_id: str) -> list:
    """获取用户最近获得的三星卡记录"""
    pity_data = load_pity_data(user_id)
    return pity_data.get("recent_3stars", [])


def add_card_collection(user_id: str, card_id: str, name: str, stars: int, limit_type: str = None):
    """添加卡片到收藏并更新数量统计（支持重复计数）"""
    pity_data = load_pity_data(user_id)
    collection = pity_data.get("card_collection", {})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    now_timestamp = datetime.now().timestamp()
    
    # 如果是新卡片，添加到收藏
    if card_id not in collection:
        collection[card_id] = {
            "name": name,
            "stars": stars,
            "limit_type": limit_type,
            "count": 1,
            "first_time": now,
            "last_time": now_timestamp
        }
        
        # 更新数量统计
        if stars == 3:
            if limit_type == "フェス限定":
                pity_data["fes_count"] = pity_data.get("fes_count", 0) + 1
            elif limit_type == "期間限定":
                pity_data["period_count"] = pity_data.get("period_count", 0) + 1
            else:
                pity_data["other_3star_count"] = pity_data.get("other_3star_count", 0) + 1
        elif stars == 2:
            pity_data["total_2stars"] = pity_data.get("total_2stars", 0) + 1
    else:
        # 如果是重复卡片，增加计数并更新时间
        collection[card_id]["count"] = collection[card_id].get("count", 1) + 1
        collection[card_id]["last_time"] = now_timestamp
    
    pity_data["card_collection"] = collection
    save_pity_data(user_id, pity_data)
    return collection


def calculate_power(user_id: str) -> int:
    """计算用户战力：フェス限定数*10 + 期間限定数*8 + 其他三星数*7 + 二星数*3"""
    pity_data = load_pity_data(user_id)
    fes = pity_data.get("fes_count", 0)
    period = pity_data.get("period_count", 0)
    other_3star = pity_data.get("other_3star_count", 0)
    two_star = pity_data.get("total_2stars", 0)
    return fes * 10 + period * 8 + other_3star * 7 + two_star * 3


# ========== FES统计模块 ==========
FES_STATS_FILE = INFO_DIR / "fes_stats.json"

def load_fes_stats() -> dict:
    """加载全服FES角色获取统计"""
    if FES_STATS_FILE.exists():
        try:
            with open(FES_STATS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_fes_stats(stats: dict):
    """保存全服FES角色获取统计"""
    _atomic_json_save(FES_STATS_FILE, stats)

def get_fes_count(card_id: str) -> int:
    """获取某个FES角色的全服获取次数"""
    stats = load_fes_stats()
    return stats.get(card_id, 0)

def increment_fes_count(card_id: str, name: str) -> int:
    """增加FES角色获取计数并返回新的计数"""
    stats = load_fes_stats()
    if card_id not in stats:
        stats[card_id] = {"count": 0, "name": name}
    stats[card_id]["count"] += 1
    save_fes_stats(stats)
    return stats[card_id]["count"]


def get_leaderboard() -> list:
    """获取排行榜（战力前10名）"""
    leaderboard = []
    
    # 遍历所有用户数据文件
    for pity_file in INFO_DIR.glob("pity_*.json"):
        try:
            user_id = pity_file.stem.replace("pity_", "")
            pity_data = load_pity_data(user_id)
            
            # 计算战力
            power = calculate_power(user_id)
            
            # 获取基本信息
            total_draws = pity_data.get("total_draws", 0)
            total_3stars = pity_data.get("total_3stars", 0)
            
            leaderboard.append({
                "user_id": user_id,
                "power": power,
                "total_draws": total_draws,
                "total_3stars": total_3stars,
                "fes_count": pity_data.get("fes_count", 0),
                "period_count": pity_data.get("period_count", 0),
                "other_3star_count": pity_data.get("other_3star_count", 0),
                "total_2stars": pity_data.get("total_2stars", 0)
            })
        except Exception as e:
            log_error(f"读取用户数据失败 {pity_file}: {e}")
    
    # 按战力降序排序
    leaderboard.sort(key=lambda x: x["power"], reverse=True)
    
    # 返回前10名
    return leaderboard[:10]


def get_gacha_leaderboard() -> list:
    """获取抽卡榜单（三星个数前10名）"""
    leaderboard = []

    for pity_file in INFO_DIR.glob("pity_*.json"):
        try:
            user_id = pity_file.stem.replace("pity_", "")
            pity_data = load_pity_data(user_id)

            total_draws = pity_data.get("total_draws", 0)
            total_3stars = pity_data.get("total_3stars", 0)

            # 计算三星率
            rate = round(total_3stars / total_draws * 100, 2) if total_draws > 0 else 0.0

            leaderboard.append({
                "user_id": user_id,
                "total_draws": total_draws,
                "total_3stars": total_3stars,
                "rate": rate,
                "fes_count": pity_data.get("fes_count", 0),
                "period_count": pity_data.get("period_count", 0),
                "other_3star_count": pity_data.get("other_3star_count", 0),
                "total_2stars": pity_data.get("total_2stars", 0)
            })
        except Exception as e:
            log_error(f"读取用户数据失败 {pity_file}: {e}")

    # 按三星个数降序排序，相同则按总抽卡数升序
    leaderboard.sort(key=lambda x: (-x["total_3stars"], x["total_draws"]))

    return leaderboard[:10]


# ========== 呱太系统 ==========
def get_gacha_file(user_id: str) -> Path:
    """获取用户呱太记录文件路径"""
    return INFO_DIR / f"gacha_{user_id}.json"


def load_gacha_data(user_id: str) -> dict:
    """加载用户的呱太数据"""
    gacha_file = get_gacha_file(user_id)
    if gacha_file.exists():
        try:
            with open(gacha_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {"gacha": 0}
    return {"gacha": 0}


def _atomic_json_save(file_path: Path, data: dict):
    """原子写入 JSON：先写临时文件再 rename，防止崩溃损坏数据"""
    import tempfile
    file_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        fd, tmp_path = tempfile.mkstemp(
            suffix='.json', prefix='tmp_', dir=str(file_path.parent))
        try:
            with os.fdopen(fd, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            os.replace(tmp_path, file_path)
        except:
            os.unlink(tmp_path)
            raise
    except OSError:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

def save_gacha_data(user_id: str, data: dict):
    """保存用户的呱太数据"""
    _atomic_json_save(get_gacha_file(user_id), data)


def get_gacha_count(user_id: str) -> int:
    """获取用户的呱太数量"""
    gacha_data = load_gacha_data(user_id)
    return gacha_data.get("gacha", 0)


def add_gacha(user_id: str, amount: int) -> int:
    """增加用户的呱太数量"""
    gacha_data = load_gacha_data(user_id)
    gacha_data["gacha"] = gacha_data.get("gacha", 0) + amount
    save_gacha_data(user_id, gacha_data)
    return gacha_data["gacha"]


def spend_gacha(user_id: str, amount: int) -> bool:
    """消耗用户的呱太数量，返回是否成功"""
    gacha_data = load_gacha_data(user_id)
    current = gacha_data.get("gacha", 0)
    if current >= amount:
        gacha_data["gacha"] = current - amount
        save_gacha_data(user_id, gacha_data)
        return True
    return False


# ========== 签到系统 ==========
def get_signin_file(user_id: str) -> Path:
    """获取用户签到记录文件路径"""
    return INFO_DIR / f"signin_{user_id}.json"


def load_signin_data(user_id: str) -> dict:
    """加载用户的签到数据"""
    signin_file = get_signin_file(user_id)
    if signin_file.exists():
        try:
            with open(signin_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {"last_signin": "", "streak": 0}
    return {"last_signin": "", "streak": 0}


def save_signin_data(user_id: str, data: dict):
    """保存用户的签到数据"""
    _atomic_json_save(get_signin_file(user_id), data)


def can_signin(user_id: str) -> bool:
    """检查用户今天是否可以签到"""
    signin_data = load_signin_data(user_id)
    last_signin = signin_data.get("last_signin", "")
    today = datetime.now().strftime("%Y-%m-%d")
    return last_signin != today


def do_signin(user_id: str) -> dict:
    """执行签到，返回签到结果"""
    signin_data = load_signin_data(user_id)
    today = datetime.now().strftime("%Y-%m-%d")
    
    if signin_data.get("last_signin") == today:
        return {"success": False, "message": "今天已经签到过了", "streak": signin_data.get("streak", 0)}
    
    # 更新签到记录
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    if signin_data.get("last_signin") == yesterday:
        # 连续签到
        signin_data["streak"] = signin_data.get("streak", 0) + 1
    else:
        # 断签，重置连续天数
        signin_data["streak"] = 1
    
    signin_data["last_signin"] = today
    save_signin_data(user_id, signin_data)
    
    # 添加签到奖励
    add_gacha(user_id, DAILY_REWARD)
    
    return {
        "success": True,
        "message": f"签到成功！获得 {DAILY_REWARD} 呱太",
        "streak": signin_data["streak"],
        "reward": DAILY_REWARD
    }


# ========== 盲盒开箱系统 (box_system.py) ==========
from box_system import (draw_mystery_box, select_3star_character, apply_mutation,
    get_mystery_box_image, get_black_box_image, create_box_card, open_mystery_box,
    create_opened_box_card, create_box_summary_image, has_box_session, get_box_session,
    create_box_session, clear_box_session, is_valid_box_index, _pil_open)
# ========== 日活与发送统计模块 ==========
def restore_daily_stats():
    """启动时从 dau_log.json 恢复今日统计数据到内存"""
    global DAILY_SEND_STATS, _RESTORED_DAU
    today = datetime.now().strftime("%Y-%m-%d")
    dau_file = INFO_DIR / "dau_log.json"

    if not dau_file.exists():
        return

    try:
        with open(dau_file, "r", encoding="utf-8") as f:
            raw = json.load(f)
        if today in raw:
            saved = raw[today]
            if isinstance(saved, dict):
                saved_sends = saved.get("send", 0)
                saved_dau = saved.get("dau", 0)
                if DAILY_SEND_STATS.get(today, 0) == 0:
                    DAILY_SEND_STATS[today] = saved_sends
                # 记录重启前的日活基准（内存中无法恢复用户ID，用计数追平）
                _RESTORED_DAU = saved_dau
                log_info(f"从磁盘恢复今日统计: 日活{saved_dau}人 发送{saved_sends}次")
    except Exception as e:
        log_error(f"恢复日活数据失败: {e}")


def save_daily_stats(force=False):
    """保存当前内存中的日活和发送统计到磁盘（轻量，不打报告）"""
    global _DAILY_SAVE_COUNTER
    _DAILY_SAVE_COUNTER += 1
    if not force and _DAILY_SAVE_COUNTER % 50 != 0:
        return  # 每50条消息才执行一次文件IO
    today = datetime.now().strftime("%Y-%m-%d")
    dau_file = INFO_DIR / "dau_log.json"

    # 读取历史数据
    data = {}
    if dau_file.exists():
        try:
            with open(dau_file, "r", encoding="utf-8") as f:
                raw = json.load(f)
            for k, v in raw.items():
                if isinstance(v, dict):
                    data[k] = v
                elif isinstance(v, int):
                    data[k] = {"dau": v, "send": 0}
        except Exception:
            pass

    # 更新今天的数据（合并重启前基准 + 本次运行新数据）
    today_dau = max(len(DAILY_USER_SET), _RESTORED_DAU)
    today_sends = max(DAILY_SEND_STATS.get(today, 0), data.get(today, {}).get("send", 0) if isinstance(data.get(today), dict) else 0)
    data[today] = {"dau": today_dau, "send": today_sends}

    # 保存（写前备份，保留最近 7 天）
    if dau_file.exists():
        try:
            import shutil
            backup_dir = BACKUP_DIR / "dau"
            backup_dir.mkdir(parents=True, exist_ok=True)
            backup_path = backup_dir / f"dau_log_{today}.json"
            if not backup_path.exists():
                shutil.copy2(dau_file, backup_path)
            # 清理 7 天前的旧备份
            cutoff = datetime.now().timestamp() - 7 * 86400
            for old in backup_dir.glob("dau_log_*.json"):
                if old.stat().st_mtime < cutoff:
                    old.unlink()
        except Exception:
            pass
    _atomic_json_save(dau_file, data)


def record_daily_dau():
    """
    保存并输出从6/8至今的统计报告
    日活 = 当日发过消息的独立用户数 (来自 DAILY_USER_SET)
    发送次数 = 当日发送消息总数 (来自 DAILY_SEND_STATS)
    """
    # 先保存最新数据
    save_daily_stats(force=True)

    today = datetime.now().strftime("%Y-%m-%d")
    dau_file = INFO_DIR / "dau_log.json"

    # 读取完整数据生成报告
    data = {}
    if dau_file.exists():
        try:
            with open(dau_file, "r", encoding="utf-8") as f:
                raw = json.load(f)
            for k, v in raw.items():
                if isinstance(v, dict):
                    data[k] = v
                elif isinstance(v, int):
                    data[k] = {"dau": v, "send": 0}
        except Exception:
            pass

    today_dau = max(len(DAILY_USER_SET), _RESTORED_DAU)
    today_sends = DAILY_SEND_STATS.get(today, 0)

    # 生成统计报告 (仅统计6/8至今)
    start = "2026-06-08"
    dates = sorted(d for d in data if d >= start)
    if not dates:
        return

    total_dau = sum(data[d]["dau"] for d in dates)
    total_send = sum(data[d]["send"] for d in dates)
    days = len(dates)
    peak_dau = max(dates, key=lambda d: data[d]["dau"])
    peak_send = max(dates, key=lambda d: data[d]["send"])

    report = f"\n===== 统计 ({start} ~ {today}) =====\n"
    for d in dates:
        report += f"  {d}: 日活{data[d]['dau']:>3}人  发送{data[d]['send']:>4}次\n"
    report += f"  ─────────────────────\n"
    report += f"  总天数:{days}  总活跃:{total_dau}人次  总发送:{total_send}次\n"
    report += f"  日均活跃:{total_dau//days}人  日均发送:{total_send//days}次\n"
    report += f"  最高日活:{data[peak_dau]['dau']}人({peak_dau})  最高发送:{data[peak_send]['send']}次({peak_send})\n"
    report += f"  今日日活:{today_dau}人  今日发送:{today_sends}次\n"

    log_info(report)
    print(report)


# ========== 敏感词过滤模块 ==========
SENSITIVE_WORDS = [
    # 政治敏感词
    "习近平", "胡锦涛", "江泽民", "毛泽东", "邓小平",
    "中南海", "人民大会堂", "天安门", "国旗", "国徽",
    # 色情暴力词
    "色情", "色情内容", "黄色", "强奸", "卖淫", "嫖娼",
    "毒品", "赌博", "暴力", "恐怖", "杀人",
    # 广告垃圾词
    "广告", "推广", "刷单", "返利", "红包", "二维码",
    "微信号", "QQ号", "电话号码", "网址", "链接"
]


def filter_sensitive_words(message: str) -> str:
    """
    过滤消息中的敏感词
    :param message: 原始消息
    :return: 过滤后的消息
    """
    if not message:
        return message
    
    filtered_message = message
    for word in SENSITIVE_WORDS:
        filtered_message = filtered_message.replace(word, "*" * len(word))
    
    return filtered_message


# ========== 消息速率限制模块 ==========
_RATE_CLEANUP_COUNTER = 0  # 定期清理计数器
_DAILY_SAVE_COUNTER = 0  # save_daily_stats 节流计数器
_LAST_SETTLEMENT_CHECK = ""  # settle_ranking_rewards 每日去重标志（日期字符串）
_IMAGE_HOST_CACHE = {"value": None, "mtime": 0}  # config.py IMAGE_HOST 缓存

def _cleanup_rate_limits():
    """清理超过2分钟的过期速率计数器，防止内存泄漏"""
    now = datetime.now().timestamp()
    stale = [tid for tid, c in MESSAGE_COUNTER.items() if now - c["start_time"] >= 120]
    for tid in stale:
        del MESSAGE_COUNTER[tid]
    if stale:
        log_info(f"速率限制清理: 移除 {len(stale)} 个过期条目")

def check_message_rate(target_id: str) -> bool:
    """
    检查消息发送速率是否超过限制
    :param target_id: 目标ID（群ID或用户ID）
    :return: True表示可以发送，False表示超过限制
    """
    now = datetime.now().timestamp()
    one_minute = 60  # 60秒
    
    # 获取当前目标的计数器
    if target_id not in MESSAGE_COUNTER:
        MESSAGE_COUNTER[target_id] = {"count": 0, "start_time": now}
    
    counter = MESSAGE_COUNTER[target_id]
    
    # 检查是否需要重置计数器（超过1分钟）
    if now - counter["start_time"] >= one_minute:
        counter["count"] = 0
        counter["start_time"] = now
    
    # 检查是否超过限制
    if counter["count"] >= MAX_MESSAGES_PER_MINUTE:
        log_error(f"消息发送速率超过限制！目标: {target_id}, 当前计数: {counter['count']}")
        return False
    
    # 增加计数
    counter["count"] += 1
    return True


# ========== botpy 消息发送模块 ==========

def _cleanup_temp_images():
    """清理 static_images/ 中的过期临时文件（启动和退出时调用）"""
    import time as _time
    img_dir = BASE_DIR / "static_images"
    if not img_dir.exists():
        return
    now = _time.time()
    count = 0
    for f in img_dir.glob("bot_*.png"):
        try:
            if now - f.stat().st_mtime > 600:
                f.unlink()
                count += 1
        except OSError:
            pass
    if count:
        log_info(f"清理过期临时文件: {count} 个")

def _start_image_server():
    """启动内置HTTP文件服务（后台线程，端口18080，多线程）"""
    import threading, http.server, socketserver
    img_dir = BASE_DIR / "static_images"
    img_dir.mkdir(exist_ok=True)
    def _serve():
        os.chdir(str(img_dir))
        import socket
        class ThreadingHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
            allow_reuse_address = True
            daemon_threads = True
            def server_bind(self):
                self.socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                super().server_bind()
        try:
            with ThreadingHTTPServer(("0.0.0.0", 18080), http.server.SimpleHTTPRequestHandler) as h:
                log_info("图片HTTP: 0.0.0.0:18080 \u2713 (多线程)")
                h.serve_forever()
        except OSError as e:
            log_error(f"图片HTTP启动失败(端口18080被占用): {e}")
        except Exception as e:
            log_error(f"图片HTTP服务异常: {e}")
    threading.Thread(target=_serve, daemon=True).start()

    # 定期清理过期临时图片（每30分钟）
    def _periodic_cleanup():
        import time as _time
        while True:
            _time.sleep(1800)
            try:
                _cleanup_temp_images()
            except Exception:
                pass
    threading.Thread(target=_periodic_cleanup, daemon=True).start()


async def _upload_and_send_image(target_id: str, file_bytes: bytes, content: str = "", is_group: bool = True, msg_id: str = ""):
    """图片发送: 本地文件 -> Cloudflare Tunnel -> QQ下载 -> 发送"""
    global BOT_API, IMAGE_HOST
    import time as _t

    fname = f"bot_{_t.time():.0f}_{random.randint(1000,9999)}.png"
    fpath = BASE_DIR / "static_images" / fname
    fpath.parent.mkdir(exist_ok=True)
    with open(fpath, "wb") as f:
        f.write(file_bytes)

    # 动态读取 IMAGE_HOST，支持运行时更新无需重启（带 mtime 缓存）
    img_host = ""
    try:
        _config_path = BASE_DIR / "config.py"
        _config_mtime = _config_path.stat().st_mtime if _config_path.exists() else 0
        if _IMAGE_HOST_CACHE["mtime"] != _config_mtime:
            with open(_config_path, "r", encoding="utf-8") as _f:
                for _line in _f:
                    _m = _re.search(r'IMAGE_HOST\s*=\s*"([^"]*)"', _line)
                    if _m:
                        _IMAGE_HOST_CACHE["value"] = _m.group(1)
                        break
            _IMAGE_HOST_CACHE["mtime"] = _config_mtime
        img_host = _IMAGE_HOST_CACHE["value"] or ""
    except:
        pass
    img_host = img_host or IMAGE_HOST or "localhost"
    host = img_host.replace("https://", "").replace("http://", "").rstrip("/")
    image_url = f"https://{host}/{fname}"
    log_info(f"图片URL: {image_url}, 文件大小: {os.path.getsize(fpath)} bytes")

    try:
        from botpy.types.message import Media

        if is_group:
            media_resp = await BOT_API.post_group_file(group_openid=target_id, file_type=1, url=image_url)
        else:
            media_resp = await BOT_API.post_c2c_file(openid=target_id, file_type=1, url=image_url)
        if not media_resp:
            log_error("QQ上传返回空")
            return
        log_info(f"QQ上传成功: uuid={media_resp.get('file_uuid','?')[:20]}...")

        media = Media(file_uuid=media_resp.get("file_uuid",""), file_info=media_resp.get("file_info",""), ttl=media_resp.get("ttl",0))
        content_clean = filter_sensitive_words(_re.sub(r'<@[A-F0-9]+>', '', content or '').strip())
        for attempt in range(3):
            import time as _time2
            try:
                if is_group:
                    result = await BOT_API.post_group_message(
                        group_openid=target_id, content=content_clean,
                        msg_type=7, msg_id=msg_id, media=media
                    )
                else:
                    result = await _c2c_send_raw(
                        openid=target_id, content=content_clean,
                        msg_type=7, msg_id=msg_id, media=media
                    )
                log_info(f"消息发送成功: id={getattr(result, 'id', '?')}")
                break
            except Exception as send_err:
                err_msg = str(send_err)
                if 'msgseq' in err_msg.lower() or '去重' in err_msg:
                    if attempt < 2:
                        await asyncio.sleep(0.5 + attempt * 0.3)
                        continue
                raise send_err
    except Exception as e:
        log_error(f"图片发送失败: {e}")
    finally:
        # QQ异步下载需要时间，延迟5秒再删
        await asyncio.sleep(5)
        try: os.remove(fpath)
        except: pass
async def _c2c_send_raw(openid: str, content: str = "", msg_type: int = 0, msg_id: str = None, media=None):
    """绕过 botpy post_c2c_message 的 msg_seq bug，直接构建干净请求"""
    from botpy.http import Route
    payload = {"openid": openid, "msg_type": msg_type}
    if content:
        payload["content"] = content
    if msg_id:
        payload["msg_id"] = msg_id
    if media is not None:
        payload["media"] = dict(media)  # Media 是 TypedDict，直接转 dict
    route = Route("POST", "/v2/users/{openid}/messages", openid=openid)
    return await BOT_API._http.request(route, json=payload)


def _bot_send(target_id: str, content: str = "", file_image: bytes = None, is_group: bool = True) -> bool:
    """通过 botpy API 发送被动回复"""
    global BOT_API, _CURRENT_MSG_ID
    if not BOT_API or not target_id:
        return False
    try:
        content = _re.sub(r'<@[A-F0-9]+>', '', content or '').strip()
        content = filter_sensitive_words(content)

        # 有图片：保存本地 → 公网URL → 上传QQ → media发送（群聊/C2C均支持）
        if file_image:
            coro = _upload_and_send_image(target_id, file_image, content=content,
                                          is_group=is_group, msg_id=_CURRENT_MSG_ID)
        elif is_group:
            coro = BOT_API.post_group_message(
                group_openid=target_id, content=content, msg_type=0,
                msg_id=_CURRENT_MSG_ID
            )
        else:
            # 绕过 botpy post_c2c_message 的 msg_seq bug，直接发干净请求
            coro = _c2c_send_raw(openid=target_id, content=content, msg_type=0,
                                 msg_id=_CURRENT_MSG_ID)

        loop = _get_main_loop()
        if loop is not None and loop.is_running():
            if threading.current_thread() is not threading.main_thread():
                # 从to_thread线程池调用：线程安全调度协程
                loop.call_soon_threadsafe(loop.create_task, coro)
            else:
                asyncio.ensure_future(coro)
        else:
            asyncio.run(coro)
        today = datetime.now().strftime("%Y-%m-%d")
        DAILY_SEND_STATS[today] = DAILY_SEND_STATS.get(today, 0) + 1
        return True
    except Exception as e:
        log_error(f"发送消息失败: {e}")
        return False


_MSG_IS_GROUP = True    # 当前消息是否为群聊（由 _route 设置）
_CURRENT_MSG_ID = ""     # 当前消息ID，用于被动回复
_CURRENT_MESSAGE = None  # 当前消息对象
_MAIN_LOOP = None        # 主线程事件循环引用（由on_ready设置，供to_thread中使用）

def _get_main_loop():
    """获取主线程事件循环（线程安全）"""
    return _MAIN_LOOP

def send_message(message: str, user_id: str = None, group_id: str = None, image_path: str = None):
    """统一消息发送入口（兼容旧接口）"""
    global _MSG_IS_GROUP
    target = group_id if group_id else user_id
    if not target:
        return False
    content = message if message else ""
    if image_path and os.path.exists(image_path):
        with open(image_path, "rb") as f:
            return _bot_send(target, content=content, file_image=f.read(), is_group=_MSG_IS_GROUP)
    return _bot_send(target, content=content, is_group=_MSG_IS_GROUP)


def send_image_from_path(channel_id: str, img_path: str, content: str = "") -> bool:
    """从文件路径发送图片"""
    global _MSG_IS_GROUP
    if not os.path.exists(img_path): return False
    with open(img_path, "rb") as f:
        img_bytes = f.read()
    try: os.remove(img_path)
    except: pass
    return _bot_send(channel_id, content=content, file_image=img_bytes, is_group=_MSG_IS_GROUP)


def send_message_with_image(channel_id: str, text: str, img_path: str) -> bool:
    """发送文字+图片"""
    return send_image_from_path(channel_id, img_path, content=text)


def send_image(image_path: str, user_id: str = None, group_id: str = None):
    """发送图片（兼容旧接口）"""
    if not os.path.exists(image_path): return False
    return send_image_from_path(group_id or user_id, image_path)


def send_qq_gif(channel_id: str, gif_buffer: BytesIO, content: str = "") -> bool:
    """发送 GIF"""
    global _MSG_IS_GROUP
    return _bot_send(channel_id, content=content, file_image=gif_buffer.getvalue(), is_group=_MSG_IS_GROUP)


def send_qq_image(channel_id: str, image: Image.Image, content: str = "") -> bool:
    """发送 PIL Image"""
    global _MSG_IS_GROUP
    buf = BytesIO()
    image.save(buf, format="PNG")
    return _bot_send(channel_id, content=content, file_image=buf.getvalue(), is_group=_MSG_IS_GROUP)


# ========== Excel数据读取 ==========
def _parse_direction_offsets(direction_str) -> list:
    """解析攻击方向箭头 → 列偏移列表"""
    if not direction_str:
        return [0]
    s = str(direction_str)
    offsets = set()
    for ch, off in [('↖',-1),('↙',-1),('←',-1),('↑',0),('↓',0),('↗',1),('↘',1),('→',1)]:
        if ch in s:
            offsets.add(off)
    return sorted(offsets) if offsets else [0]


def _derive_limit_type(acquire: str) -> str:
    """从获取方式推导限定种类"""
    if not acquire:
        return ""
    a = str(acquire).lower()
    if 'fes' in a:
        return "フェス限定"
    if '限定' in a or '期間' in a:
        return "期間限定"
    if '活動' in a or '活动' in a:
        return "期間限定"
    return ""


def load_character_data():
    """从卡牌信息.xlsx 加载角色数据（用于抽卡，含1/2/3星）"""
    try:
        wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
        characters = []

        for sheet_name, type_name in [('BattleCard资源', 'battle'), ('AssistCard资源', 'assist')]:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    card_id = str(row[5]) if row[5] is not None else None
                    chara_id = row[6]
                    init_stars = row[1]
                    card_name = row[4]
                    limit_type = row[0] or ""
                    if not (card_id and chara_id and init_stars and card_name):
                        continue
                    stars_val = int(float(str(init_stars)))
                    icon_path = find_character_icon(chara_id, stars_val)
                    characters.append({
                        "card_id": card_id, "chara_id": chara_id,
                        "stars": stars_val, "name": str(card_name),
                        "attribute": str(row[3]) if row[3] else "红",
                        "limit_type": str(limit_type),
                        "icon_path": icon_path, "type": type_name,
                        "hp": 5000, "attack": 3000, "defense": 2000,
                        "dexterity": 1000, "speed": 500,
                        "attack_directions": [0],
                        "attack_type": "物理", "side": "科学",
                        "skill1": {"cd": 0, "description": "", "keywords": "", "condition": ""},
                        "skill2": {"cd": 0, "description": "", "keywords": "", "condition": ""},
                        "skill3": {"cd": 0, "description": "", "keywords": "", "condition": ""},
                        "passive1_text": "", "passive2_text": "",
                    })
                except Exception:
                    continue
            log_info(f"[抽卡] 加载了 {sum(1 for c in characters if c['type']==type_name)} 个{type_name}角色")

        log_info(f"[抽卡] 共加载 {len(characters)} 个角色")
        wb.close()
        return characters
    except Exception as e:
        import traceback
        log_error(f"加载抽卡数据失败: {e}")
        return []


def load_battle_characters():
    """从 cards_completed.xlsx 加载战斗数据（3星卡详细数值）"""
    global BATTLE_CHARACTERS
    try:
        wb = openpyxl.load_workbook(BATTLE_XLSX, data_only=True)
        characters = []
        BATTLE_CHARACTERS = {}

        for sheet_name, card_type in [('b卡', 'battle'), ('a卡', 'assist')]:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    card_id = str(int(row[0])) if row[0] is not None else None
                    if not card_id:
                        continue
                    name_raw = str(row[1]) if row[1] else ""
                    name = name_raw.split('|')[0].strip() if '|' in name_raw else name_raw.strip()
                    side = str(row[3]) if row[3] else "科学"
                    direction_raw = str(row[4]) if row[4] else ""
                    attribute = str(row[5]) if row[5] else "红"
                    attack_type_str = str(row[6]) if row[6] else "物理"
                    sk1_desc = str(row[8]) if row[8] else ""
                    sk1_cd = int(float(row[9])) if row[9] is not None else 0
                    sk2_desc = str(row[10]) if row[10] else ""
                    sk2_cd = int(float(row[11])) if row[11] is not None else 0
                    sk3_desc = str(row[12]) if row[12] else ""
                    sk3_cd = int(float(row[13])) if row[13] is not None else 0
                    pas1 = str(row[14]) if row[14] else ""
                    pas2 = str(row[15]) if row[15] else ""
                    phys_atk = int(float(row[16])) if row[16] is not None else 5000
                    magic_atk = int(float(row[17])) if row[17] is not None else 5000
                    phys_def = int(float(row[18])) if row[18] is not None else 3000
                    magic_def = int(float(row[19])) if row[19] is not None else 3000
                    hp = int(float(row[20])) if row[20] is not None else 10000
                    dexterity = int(float(row[21])) if row[21] is not None else 1000
                    is_physical = '物理' in attack_type_str
                    attack = phys_atk if is_physical else magic_atk
                    defense = (phys_def + magic_def) // 2
                    color_map = {'赤': '红', '緑': '绿', '青': '蓝'}
                    attr = attribute.strip()
                    is_super = attr.startswith('超')
                    base = attr[1:] if is_super else attr
                    base = color_map.get(base, base)
                    attr_normalized = ('超' + base) if is_super else base
                    offsets = set()
                    for ch, off in [('↖',-1),('↙',-1),('←',-1),('↑',0),('↓',0),('↗',1),('↘',1),('→',1)]:
                        if ch in str(direction_raw): offsets.add(off)
                    dire_list = sorted(offsets) if offsets else [0]
                    acquire = str(row[2]) if row[2] else ""
                    limit_type = ""
                    if 'fes' in acquire.lower(): limit_type = "フェス限定"
                    elif '限定' in acquire or '期間' in acquire: limit_type = "期間限定"
                    elif '活動' in acquire or '活动' in acquire: limit_type = "期間限定"
                    icon_path = find_character_icon(card_id, 3)
                    char_data = {
                        "card_id": card_id, "chara_id": card_id,
                        "stars": 3, "name": name,
                        "attribute": attr_normalized, "limit_type": limit_type,
                        "icon_path": icon_path, "type": card_type,
                        "hp": hp, "attack": attack, "defense": defense,
                        "dexterity": dexterity, "phys_atk": phys_atk, "magic_atk": magic_atk,
                        "attack_directions": dire_list,
                        "attack_type": attack_type_str, "side": side,
                        "skill1": {"cd": sk1_cd, "description": sk1_desc, "keywords": "", "condition": ""},
                        "skill2": {"cd": sk2_cd, "description": sk2_desc, "keywords": "", "condition": ""},
                        "skill3": {"cd": sk3_cd, "description": sk3_desc, "keywords": "", "condition": ""},
                        "passive1_text": pas1, "passive2_text": pas2,
                    }
                    characters.append(char_data)
                    BATTLE_CHARACTERS[card_id] = char_data
                except Exception:
                    continue
            log_info(f"[战斗] 加载了 {sum(1 for c in characters if c['type']==card_type)} 个{card_type}角色")

        log_info(f"[战斗] 共加载 {len(characters)} 个角色")
        wb.close()
        return characters
    except Exception as e:
        import traceback
        log_error(f"加载战斗数据失败: {e}")
        return []


def find_character_icon(chara_id, stars: int) -> str:
    """根据角色ID和星级查找图标文件（支持多种文件名格式）"""
    # 确保 chara_id 是整数
    try:
        # 如果是浮点数（如 100111001.0），先转int再转str去掉小数点
        if isinstance(chara_id, float):
            chara_id_int = int(chara_id)
        elif isinstance(chara_id, str):
            # 尝试解析字符串数字，可能带有小数点
            chara_id_int = int(float(chara_id))
        else:
            chara_id_int = int(chara_id)
    except (ValueError, TypeError):
        log_error(f"无效的chara_id: {chara_id}, 类型: {type(chara_id)}")
        return None
    
    # 尝试多种可能的文件名模式（按优先级排序）
    patterns = [
        # 新格式：card_cutin_前缀（9位数字）
        f"card_cutin_{chara_id_int:09d}.png",
        f"card_cutin_{chara_id_int}.png",
        
        # 新格式：直接用chara_id命名
        f"{chara_id_int:03d}.png",
        f"{chara_id_int}.png",
        f"{chara_id_int:09d}.png",
        
        # chara前缀
        f"chara_{chara_id_int:03d}.png",
        f"chara_{chara_id_int}.png",
        
        # icon前缀
        f"icon_{chara_id_int:03d}.png",
        f"icon_{chara_id_int}.png",
        
        # 带星级的格式
        f"{chara_id_int:03d}_{stars}star.png",
        f"{chara_id_int}_{stars}star.png",
        f"chara_{chara_id_int:03d}_{stars}.png",
        f"chara_{chara_id_int}_{stars}.png",
        f"icon_{chara_id_int:03d}_{stars}.png",
        f"icon_{chara_id_int}_{stars}.png",
        
        # 原始格式（保留兼容）
        f"character_icon_{chara_id_int:03d}{stars:02d}0101.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0201.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0102.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0202.png",
        f"character_icon_{chara_id_int:03d}{stars:02d}0103.png",
    ]

    # 先尝试精确匹配
    for pattern in patterns:
        if '*' not in pattern:  # 非通配符模式
            path = ICON_DIR / pattern
            if path.exists():
                return str(path)

    # 如果精确匹配都找不到，尝试通配符匹配
    wildcard_patterns = [
        f"card_cutin_{chara_id_int}*.png",
        f"card_cutin_{chara_id_int:03d}*.png",
        f"card_cutin_{chara_id_int:09d}*.png",
        f"{chara_id_int:03d}*.png",
        f"{chara_id_int:09d}*.png",
        f"*{chara_id_int:03d}*.png",
        f"*{chara_id_int:09d}*.png",
        f"chara_{chara_id_int:03d}*.png",
        f"icon_{chara_id_int:03d}*.png",
        f"character_icon_{chara_id_int:03d}*.png",
    ]
    
    for pattern in wildcard_patterns:
        matches = list(ICON_DIR.glob(pattern))
        if matches:
            # 优先选择包含星级的文件
            star_matches = [m for m in matches if f"{stars}" in m.name or f"{stars:02d}" in m.name]
            if star_matches:
                return str(star_matches[0])
            return str(matches[0])

    log_error(f"未找到角色图标: chara_id={chara_id}, chara_id_int={chara_id_int}, stars={stars}")
    return None


# ========== 抽卡逻辑 ==========
from typing import Optional, List, Dict, Any


def draw_cards(count: int, characters: List[Dict[str, Any]], user_id: Optional[str] = None) -> Dict[str, Any]:
    """
    抽取指定数量的卡牌
    使用权重：从配置读取概率
    保底机制：每150抽必出フェス限定3星
    
    :param count: 抽卡数量（1或10）
    :param characters: 角色列表
    :param user_id: 用户ID（用于保底计数）
    :return: {"results": 抽卡结果列表, "remaining_pity": 剩余保底抽数, "got_3star": 是否抽到3星}
    """
    # 输入验证
    if count not in [1, 10]:
        raise ValueError("抽卡数量只能是1或10")
    
    if not characters:
        raise ValueError("角色列表不能为空")
    
    if user_id and not isinstance(user_id, str):
        raise TypeError("user_id必须是字符串类型")

    results: List[Dict[str, Any]] = []
    got_3star = False
    remaining_pity = 0
    
    for _ in range(count):
        # 检查是否触发保底
        if user_id:
            pity_data = load_pity_data(user_id)
            pity_count = pity_data.get("pity_count", 0)
            remaining_before = max(0, PITY_LIMIT - pity_count)
            
            # 如果是保底抽，直接出3星
            if remaining_before == 1:
                stars = 3
            else:
                # 随机决定星级（使用预定义的权重常量）
                stars = random.choices(
                    population=[1, 2, 3],
                    weights=GACHA_WEIGHTS,
                    k=1
                )[0]
        else:
            # 没有用户ID，正常抽卡（使用预定义的权重常量）
            stars = random.choices(
                population=[1, 2, 3],
                weights=GACHA_WEIGHTS,
                k=1
            )[0]

        # 从对应星级的角色中随机选择一个
        available = [c for c in characters if c["stars"] == stars]
        if available:
            char = random.choice(available)
            results.append(char)
            if stars == 3:
                got_3star = True
        else:
            # 如果没有对应星级的角色，随机选一个
            log_error(f"没有找到 {stars}星级的角色，随机分配")
            char = random.choice(characters)
            results.append(char)
            if char["stars"] == 3:
                got_3star = True
    
    # 更新保底计数
    if user_id:
        for _ in range(count):
            update_pity(user_id, got_3star)
        remaining_pity = get_remaining_pity(user_id)
    
    return {
        "results": results,
        "remaining_pity": remaining_pity,
        "got_3star": got_3star
    }


# ========== 图片合成 (已迁移至 card_image.py) ==========
from card_image import (get_level_image, find_attribute_icon, find_type_icon,
    composite_card, save_card_image)
# ========== 数据迁移 ==========

def _at_user(user_id: str, group_id: str) -> str:
    return f"<@{user_id}> " if (group_id and user_id) else ""


def _reply(text: str, user_id: str, group_id: str):
    send_message(f"{_at_user(user_id, group_id)}{text}", user_id, group_id)



def _handle_bind(user_id: str, group_id: str, raw_message: str):
    """绑定旧QQ号，迁移数据: /绑定 QQ号"""
    m = _re.search(r'绑定\s*(\d{5,15})', raw_message)
    if not m:
        send_message("格式: /绑定 QQ号\n例如: /绑定 1463421422", user_id, group_id)
        return

    old_qq = m.group(1)
    old_file = INFO_DIR / f"gacha_{old_qq}.json"
    old_pity = INFO_DIR / f"pity_{old_qq}.json"
    old_team = INFO_DIR / f"team_{old_qq}.json"
    new_file = INFO_DIR / f"gacha_{user_id}.json"
    new_pity = INFO_DIR / f"pity_{user_id}.json"
    new_team = INFO_DIR / f"team_{user_id}.json"

    overridden = []
    for old_path, new_path, name in [
        (old_file, new_file, "呱太数据"),
        (old_pity, new_pity, "抽卡记录"),
        (old_team, new_team, "队伍配置"),
    ]:
        if old_path.exists():
            import shutil
            was_overwrite = new_path.exists()
            shutil.copy(str(old_path), str(new_path))
            overridden.append(f"{name}{'(覆盖)' if was_overwrite else ''}")

    if overridden:
        if user_id in USER_DATA:
            del USER_DATA[user_id]
        send_message(f"数据迁移成功！已继承旧QQ号 {old_qq} 的数据: {', '.join(overridden)}", user_id, group_id)
        log_info(f"数据迁移: {old_qq} -> {user_id} ({', '.join(overridden)})")
        # 排行榜：把旧QQ号替换为新openid
        try:
            ranking = load_ranking()
            updated = False
            for entry in ranking:
                if entry.get("user_id") == old_qq:
                    entry["user_id"] = user_id
                    entry["nickname"] = get_nickname(user_id)
                    log_info(f"排行榜更新: {old_qq} -> {user_id}")
                    updated = True
            if updated:
                save_ranking(ranking)
        except Exception as e:
            log_error(f"排行榜同步失败: {e}")
    else:
        send_message(f"未找到旧QQ号 {old_qq} 的数据。", user_id, group_id)


# ========== 昵称系统 ==========

def _load_nicknames() -> dict:
    fp = INFO_DIR / "nicknames.json"
    if fp.exists():
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                return json.load(f)
        except: pass
    return {}

def _save_nicknames(data: dict):
    with open(INFO_DIR / "nicknames.json", 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_nickname(uid: str) -> str:
    uid = str(uid)  # 兼容整数型 user_id
    nick = _load_nicknames().get(uid, "")
    return nick if nick else uid[:8]  # 未设置昵称显示ID前8位

def _handle_nickname(user_id: str, group_id: str, raw_message: str):
    """设置/查看昵称: 昵称 新名字"""
    m = _re.search(r'昵称\s+(.+)', raw_message)
    if m:
        nick = m.group(1).strip()[:16]  # 最多16字
        data = _load_nicknames()
        data[str(user_id)] = nick
        _save_nicknames(data)
        send_message(f"昵称已设置为「{nick}」", user_id, group_id)
    else:
        nick = get_nickname(user_id)
        send_message(f"当前昵称: {nick}\n修改: 昵称 新名字", user_id, group_id)


# ========== QQ Bot 消息处理 (botpy WebSocket) ==========

class QQBotClient(Client):
    """QQ机器人 - 基于botpy WebSocket"""

    async def on_ready(self):
        global BOT_API, _MAIN_LOOP
        BOT_API = self.api
        _MAIN_LOOP = asyncio.get_running_loop()
        log_info(f"Bot已连接! AppID={QQ_BOT_APP_ID}")

        # 启动时检查是否有未结算的排行榜奖励
        result = settle_ranking_rewards()
        if result and result.get("rewards"):
            names = ", ".join(f"#{r['rank']} {r['nickname']}" for r in result["rewards"])
            log_info(f"启动补结算: {result['date']} → {names}")

    async def on_group_at_message_create(self, message: GroupMessage):
        """群聊@消息"""
        await self._route(message)

    async def on_c2c_message_create(self, message: C2CMessage):
        """私聊消息"""
        await self._route(message)

    async def _route(self, message):
        """统一消息路由（复用原有所有命令处理逻辑）"""
        try:
            await self._route_impl(message)
        except Exception as e:
            import traceback
            log_error(f"消息路由异常: {e}\n{traceback.format_exc()}")
            try:
                raw = (getattr(message, 'content', '') or '').strip()[:80]
                log_error(f"崩溃消息预览: {raw}")
            except:
                pass

    async def _route_impl(self, message):
        """消息路由实现（由 _route 的 try/except 保护）"""
        global BOT_API, _LAST_SETTLEMENT_CHECK
        BOT_API = self.api

        raw_message = (getattr(message, 'content', '') or '').strip()

        # 提取 @mentions（管理命令需要）
        raw_mentions = _re.findall(r'<@([A-F0-9]+)>', raw_message)

        # 去掉 @mention 用于命令匹配
        cleaned_message = _re.sub(r'<@[A-F0-9]+>', '', raw_message).strip()

        if not cleaned_message and not raw_mentions:
            return

        # 定期清理过期的速率限制条目（每100条消息触发一次）
        global _RATE_CLEANUP_COUNTER
        _RATE_CLEANUP_COUNTER += 1
        if _RATE_CLEANUP_COUNTER % 100 == 0:
            _cleanup_rate_limits()

        # 判断消息类型，设置全局标志（供 send_message 使用）
        global _MSG_IS_GROUP, _CURRENT_MSG_ID, _CURRENT_MESSAGE
        _MSG_IS_GROUP = hasattr(message, 'group_openid')
        _CURRENT_MSG_ID = getattr(message, 'id', '')
        _CURRENT_MESSAGE = message

        # 提取用户ID（群聊 author 有 member_openid，私聊 author 有 user_openid）
        author = getattr(message, 'author', None)
        if author is None:
            user_id = 'unknown'
        elif isinstance(author, dict):
            user_id = str(author.get('member_openid') or author.get('user_openid') or author.get('id', 'unknown'))
        else:
            user_id = str(getattr(author, 'member_openid', None) or getattr(author, 'user_openid', None) or getattr(author, 'id', 'unknown'))

        # 目标ID: 群聊用 group_openid，私聊用 user_id
        target_id = str(getattr(message, 'group_openid', '') or user_id or '')
        message_id = getattr(message, 'id', '')
        # 兼容旧代码
        group_id = target_id
        channel_id = target_id

        log_info(f"MSG type={type(message).__name__} uid={user_id} target={target_id} msgid={_CURRENT_MSG_ID}: {raw_message[:60]}")

        if user_id and user_id != 'unknown':
            DAILY_USER_SET.add(str(user_id))
        save_daily_stats()

        # 检查排行榜每日结算（每天最多一次，12:00触发）
        _today = datetime.now().strftime("%Y-%m-%d")
        if _LAST_SETTLEMENT_CHECK != _today:
            settlement = settle_ranking_rewards()
            # RAID每日重置（发券/解锁/次数重置）
            try:
                from rescue_event import load_event_data, save_event_data, daily_raid_reset
                raid_data = load_event_data()
                if raid_data.get("active", False):
                    reset_msg = daily_raid_reset(raid_data)
                    log_info(f"RAID每日重置: {reset_msg}")
            except Exception as e:
                log_error(f"RAID每日重置失败: {e}")
            _LAST_SETTLEMENT_CHECK = _today
            if settlement and settlement.get("rewards"):
                lines = [f"🏆 排行榜每日结算 ({settlement['date']})"]
                for r in settlement["rewards"]:
                    lines.append(f"  {['','🥇','🥈','🥉'][r['rank']]} 第{r['rank']}名: {r['nickname']} +{r['amount']}呱太")
                send_message("\n".join(lines), user_id, group_id)

        self_id = str(QQ_BOT_APP_ID)

        # 盲盒会话
        if has_box_session(user_id):
            open_commands = ['全部开', '剩下的全部开']
            is_open = any(cmd in cleaned_message for cmd in open_commands)
            has_input = False
            s = cleaned_message.strip().replace('，', ',').replace(' ', '').replace('　', '').replace(',', '')
            if s.isdigit(): has_input = True
            elif _re.search(r'选择[0-9]+', cleaned_message): has_input = True
            elif _re.search(r'[0-9]+', cleaned_message): has_input = True
            # 允许放弃旧盲盒：取消 / 新抽卡覆盖
            abandon_commands = ['取消', '不要了', '放弃']
            is_abandon = any(cmd in cleaned_message for cmd in abandon_commands)
            is_new_gacha = cleaned_message.lstrip('/').startswith('十连') or cleaned_message.lstrip('/').startswith('单抽') or cleaned_message.lstrip('/').startswith('限定十连')
            if is_abandon or is_new_gacha:
                if is_abandon:
                    session = get_box_session(user_id)
                    if session and len(session["opened"]) > 0:
                        send_message("已开过盲盒，无法取消退回呱太！", user_id, group_id)
                        remaining = max(0, len(session["boxes"]) - len(session["opened"]))
                        reply = f"你还有{remaining}个未开！请输入要开的编号（如1、选择1）或「全部开」"
                        send_message(reply, user_id, group_id)
                        return
                    send_message("盲盒已取消，呱太不退还~", user_id, group_id)
                clear_box_session(user_id)
                if is_abandon:
                    return
            elif is_open or has_input:
                return handle_box_open(user_id, group_id, cleaned_message)
            else:
                session = get_box_session(user_id)
                remaining = max(0, len(session["boxes"]) - len(session["opened"]))
                reply = f"你还有{remaining}个未开！请输入要开的编号（如1、选择1）或「全部开」\n输入「取消」放弃本次盲盒"
                send_message(reply, user_id, group_id)
                return

        # 管理员命令
        if str(user_id) == str(ADMIN_QQ):
            # 用 raw_mentions 获取被 @ 的目标（第一个非bot用户）
            target_user = raw_mentions[1] if len(raw_mentions) > 1 else (raw_mentions[0] if len(raw_mentions) > 0 and raw_mentions[0] != str(QQ_BOT_APP_ID) else None)
            amt_gacha = _re.search(r'(\d+)\s*呱太', cleaned_message)
            if target_user and amt_gacha:
                target = target_user; amt = int(amt_gacha.group(1))
                if '扣' in cleaned_message or '减' in cleaned_message:
                    ok = spend_gacha(target, amt); act = "扣除" if ok else "扣除失败(余额不足)"
                else:
                    add_gacha(target, amt); act = "增加"
                send_message(f"管理员操作完成: {act} {amt}呱太 (余额: {get_gacha_count(target)})", user_id, group_id)
                return
            amt_blue = _re.search(r'(\d+)\s*蓝碎片', cleaned_message)
            if target_user and amt_blue:
                target = target_user; amt = int(amt_blue.group(1))
                if '扣' in cleaned_message or '减' in cleaned_message:
                    ok = spend_blue_crystal(target, amt); act = "扣除" if ok else "扣除失败"
                else:
                    add_blue_crystal(target, amt); act = "增加"
                send_message(f"管理员操作完成: {act} {amt}蓝碎片", user_id, group_id)
                return
            amt_red = _re.search(r'(\d+)\s*红碎片', cleaned_message)
            if target_user and amt_red:
                target = target_user; amt = int(amt_red.group(1))
                if '扣' in cleaned_message or '减' in cleaned_message:
                    ok = spend_red_crystal(target, amt); act = "扣除" if ok else "扣除失败"
                else:
                    add_red_crystal(target, amt); act = "增加"
                send_message(f"管理员操作完成: {act} {amt}红碎片", user_id, group_id)
                return
            # 全榜替: 排行榜中把旧QQ号替换为新openid
            m_replace = _re.search(r'全榜替\s+([A-F0-9]+)\s+(\d+)', cleaned_message)
            if m_replace:
                new_id, old_qq = m_replace.group(1), m_replace.group(2)
                ranking = load_ranking()
                count = 0
                for entry in ranking:
                    if entry.get("user_id") == old_qq:
                        entry["user_id"] = new_id
                        count += 1
                if count > 0:
                    save_ranking(ranking)
                    send_message(f"排行榜已更新: {old_qq} -> {new_id} ({count}条)", user_id, group_id)
                else:
                    send_message(f"排行榜中未找到 {old_qq}", user_id, group_id)
                return

        # ===== 命令路由（全部复用原有handler）=====
        # 去掉开头的 /（兼容有/无/两种情况）
        cmd = cleaned_message.lstrip('/')

        if cmd.startswith('签到'):
            return handle_signin(user_id, group_id)
        if cmd.startswith('获取呱太'):
            return handle_get_gacha(user_id, group_id)
        if cmd.startswith('十连') or cmd.startswith('单抽'):
            return await asyncio.to_thread(handle_gacha, 10 if '十连' in cmd else 1, user_id, group_id, message_id)
        elif '战斗GIF' in cmd:
            return await asyncio.to_thread(handle_battle_log, user_id, group_id, True)
        elif 'gif列表' in cmd.lower() or 'GIF列表' in cmd:
            return _list_cached_gifs(user_id, group_id)
        elif '战斗日志' in cmd:
            return await asyncio.to_thread(handle_battle_log, user_id, group_id, False)
        if cmd.startswith('战斗') or cmd.startswith('对战'):
            return await asyncio.to_thread(handle_battle, user_id, group_id, cmd, raw_mentions)
        if cmd.startswith('挑战'):
            target_mention = None
            for mid in raw_mentions:
                if mid != str(QQ_BOT_APP_ID):
                    target_mention = mid
                    break
            if target_mention:
                return await asyncio.to_thread(challenge_player, user_id, group_id, target_mention)
            m = _re.search(r'挑战\s*(\d+)', cmd)
            target_rank = int(m.group(1)) if m else None
            return await asyncio.to_thread(challenge_rank, user_id, group_id, target_rank)
        if cmd.startswith('BOSS战') or cmd.startswith('boss战'):
            return await asyncio.to_thread(handle_boss_battle, user_id, group_id, cmd)
        if cmd.startswith('防守队'):
            return handle_defense_team(user_id, group_id, cmd)
        if cmd.startswith('队伍'):
            return handle_team(user_id, group_id, cmd)
        if cmd.startswith('昵称'):
            return _handle_nickname(user_id, group_id, cmd)
        if cmd.startswith('我的呱太') or cmd.startswith('余额'):
            send_message(f"当前呱太: {get_gacha_count(user_id)}", user_id, group_id)
        if cmd.startswith('限定十连'):
            return await asyncio.to_thread(handle_limited_gacha, user_id, group_id)
        if cmd.startswith('兑换'):
            # 兑换 ZMDBOT2026 → CDKEY 兑换（大写字母+数字串）
            import re
            key_match = re.match(r'兑换\s*([A-Z0-9]{4,})\s*$', cmd)
            if key_match:
                return handle_cdkey_redeem(user_id, group_id, key_match.group(1))
            # 兑换红碎片 / 兑换红 → 仅兑红色碎片
            if '红' in cmd:
                return handle_exchange_crystal(user_id, group_id, crystal_type="red")
            # 兑换蓝碎片 / 兑换蓝 → 仅兑蓝色碎片
            if '蓝' in cmd:
                return handle_exchange_crystal(user_id, group_id, crystal_type="blue")
            # 兑换 / 兑换呱太 → 兑全部碎片
            return handle_exchange_crystal(user_id, group_id)
        elif '抽卡排行' in cmd or cmd.startswith('呱太排行'):
            return handle_gacha_leaderboard(user_id, group_id)
        elif '排行榜' in cmd or '排行' in cmd:
            return show_ranking(user_id, group_id)
        elif '详细信息' in cmd or '详情' in cmd or cmd.startswith('查看'):
            return handle_show_details(user_id, group_id)
        elif '三王女' in cmd:
            return handle_sannoujo(user_id, group_id)
        elif '红抽' in cmd or '蓝抽' in cmd or '三星池' in cmd:
            return await asyncio.to_thread(handle_3star_pool, user_id, group_id, cmd)
        if cmd.startswith('help') or cmd.startswith('帮助'):
            return handle_help(user_id, group_id, cmd)
        if cmd.startswith('绑定'):
            return _handle_bind(user_id, group_id, cmd)
        if cmd.startswith('个人记录') or cmd.startswith('记录') or cmd.startswith('我的记录'):
            return handle_personal_info(user_id, group_id)
        if cmd.startswith('可爱') or cmd.startswith('卖萌'):
            return handle_cute_reply(user_id, group_id)

        # ===== RAID救援活动（独立模块，不影响现有功能）=====
        if cmd.startswith('raid'):
            try:
                from rescue_event import handle_admin_command, handle_player_command
                if user_id == ADMIN_QQ:
                    admin_result = handle_admin_command(None, user_id, cmd)
                    if admin_result is not None:
                        send_message(admin_result, user_id, group_id)
                        return jsonify({"status": "success"})
                _bs = BATTLE_INSTANCE if BATTLE_SYSTEM_LOADED else None
                player_result = handle_player_command(None, user_id, cmd, _bs)
                if player_result is not None:
                    # 检查是否有VS图需要发送
                    if "[RAID_VS_IMAGE]" in player_result:
                        parts = player_result.split("[RAID_VS_IMAGE]", 1)
                        text_part = parts[0]
                        img_path_part = parts[1].strip() if len(parts) > 1 else ""
                        send_message_with_image(group_id or user_id, text_part, img_path_part)
                        try:
                            if img_path_part and os.path.exists(img_path_part):
                                os.unlink(img_path_part)
                        except Exception:
                            pass
                    else:
                        send_message(player_result, user_id, group_id)
                    return jsonify({"status": "success"})
            except Exception as e:
                log_error(f"RAID处理失败: {e}")
                send_message(f"RAID系统出错: {e}", user_id, group_id)


def handle_gacha(count: int, user_id: str, group_id, message_id, auto_open: bool = False):
    """处理抽卡请求（盲盒模式）"""
    gacha_before = None  # 用于异常退还
    try:
        # 检查呱太数量
        cost = GACHA10_COST if count == 10 else GACHA_COST
        current_gacha = get_gacha_count(user_id)
        
        if current_gacha < cost:
            reply = f"呱太不足！当前呱太: {current_gacha}，需要: {cost}。请艾特我发送「获取呱太」获得10000呱太~"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"抽卡失败 [{user_id}]: 呱太不足 {current_gacha}/{cost}")
            return jsonify({
                "status": "error",
                "message": "呱太不足",
                "current_gacha": current_gacha,
                "required_gacha": cost
            })

        # 十连冷却检查：一分钟内每个用户至多只能抽一次十连
        if count == 10:
            now_ts = datetime.now().timestamp()
            last_gacha10 = GACHA10_COOLDOWN.get(user_id, 0)
            remaining_cooldown = GACHA10_COOLDOWN_SECONDS - (now_ts - last_gacha10)
            if remaining_cooldown > 0:
                reply = f"十连冷却中！请等待 {int(remaining_cooldown)} 秒后再试~"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                log_info(f"十连冷却 [{user_id}]: 还需等待 {int(remaining_cooldown)} 秒")
                return jsonify({
                    "status": "error",
                    "message": "十连冷却中",
                    "remaining_cooldown": int(remaining_cooldown)
                })

        # 记录扣款前呱太数量（用于异常退还）
        gacha_before = get_gacha_count(user_id)

        # 消耗呱太
        spend_gacha(user_id, cost)

        # 十连冷却记录：更新最后一次十连时间戳
        if count == 10:
            GACHA10_COOLDOWN[user_id] = datetime.now().timestamp()

        # 使用预加载的角色数据
        characters = get_characters()
        if not characters:
            return jsonify({
                "status": "error",
                "message": "无法加载角色数据"
            })

        # 检查是否触发フェス保底
        pity_data = load_pity_data(user_id)
        fes_pity_count = pity_data.get("fes_pity_count", 0)
        is_fes_pity = fes_pity_count >= PITY_LIMIT
        
        if is_fes_pity:
            log_info(f"用户 {user_id} 触发フェス限定三星保底！")

        # 生成盲盒
        boxes = []
        has_2star = False  # 十连保底二星标记
        
        for i in range(count):
            # 如果是保底的最后一抽，强制触发フェス保底
            is_pity_draw = is_fes_pity and (i == count - 1)
            box = draw_mystery_box(characters, user_id, is_pity_draw)
            
            # 记录是否有二星
            if box["stars"] == 2:
                has_2star = True
            
            boxes.append(box)
        
        # 十连保底：如果没有二星，将最后一个非保底盲盒改为二星
        if count == 10 and not has_2star and not is_fes_pity:
            # 找到最后一个可以修改的盲盒（非黑色盲盒且不是保底）
            for i in range(len(boxes)-1, -1, -1):
                if not boxes[i].get("is_mystery", False):
                    # 改为二星
                    boxes[i]["stars"] = 2
                    # 获取二星角色
                    available = [c for c in characters if c["stars"] == 2]
                    if available:
                        boxes[i]["character"] = random.choice(available)
                    else:
                        boxes[i]["character"] = random.choice(characters)
                    log_info(f"十连保底触发，将第{i+1}个盲盒改为二星")
                    break
        
        # 创建盲盒会话（通过box_system统一管理，避免__main__ vs qq_bot_ws模块分裂）
        create_box_session(user_id, boxes)
        get_box_session(user_id)["characters"] = characters
        get_box_session(user_id)["is_fes_pity"] = is_fes_pity  # 标记FES保底

        # 生成唯一标识
        output_idx = random.randint(1000, 9999)

        # 生成盲盒图片
        box_images = []
        for box in boxes:
            img_bytes = create_box_card(box, characters)
            if img_bytes:
                box_images.append(Image.open(BytesIO(img_bytes)))

        if not box_images:
            log_error("生成盲盒图片失败")
            clear_box_session(user_id)
            return jsonify({"status": "error", "message": "生成盲盒图片失败"})

        # 合成盲盒大图
        card_width, card_height = box_images[0].size
        if count == 1:
            img_path = OUTPUT_DIR / f"gacha_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            box_images[0].convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)
        else:
            # 十连：使用背景图片
            gap = 18
            cols = 5
            rows = (count + 4) // 5
            cards_total_width = card_width * cols + gap * (cols - 1)
            cards_total_height = card_height * rows + gap * (rows - 1)
            
            # 尝试加载背景图片
            bg_path = None
            for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
                test_path = LEVEL_DIR / bg_name
                if test_path.exists():
                    bg_path = str(test_path)
                    break
            
            if bg_path:
                # 使用背景图片，放大到原来的2倍
                log_info(f"使用十连背景图片: {bg_path}")
                bg_img = _pil_open(bg_path).convert('RGB')
                
                # 将背景放大到原来的2倍
                bg_w, bg_h = bg_img.size
                final_w = int(bg_w * 2.0 * 0.264)
                final_h = int(bg_h * 2.0 * 0.264)
                bg_img_resized = bg_img.resize((final_w, final_h), Image.Resampling.LANCZOS)
                
                # 创建最终画布
                output = Image.new('RGB', (final_w, final_h), (50, 50, 50))
                output.paste(bg_img_resized, (0, 0))
                
                # 计算卡牌居中位置
                cards_x = (final_w - cards_total_width) // 2
                cards_y = (final_h - cards_total_height) // 2
            else:
                # 没有背景，使用纯色背景
                output = Image.new('RGB', (cards_total_width, cards_total_height), (50, 50, 50))
                cards_x = 0
                cards_y = 0

            # 粘贴卡牌
            for i, img in enumerate(box_images):
                row = i // cols
                col = i % cols
                x = cards_x + col * (card_width + gap)
                y = cards_y + row * (card_height + gap)
                output.paste(img, (x, y))

            img_path = OUTPUT_DIR / f"gacha_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            output.convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)

        # 获取用户信息
        current_gacha = get_gacha_count(user_id)
        total_draws = get_total_draws(user_id)
        total_3stars = get_total_3stars(user_id)
        remaining_pity = get_remaining_pity(user_id)

        # 构建提示消息，包含呱太和抽数信息
        box_hints = [f"选择{i+1}" for i in range(count)]
        hint_text = " | ".join(box_hints[:5])
        if count > 5:
            hint_text += "\n" + " | ".join(box_hints[5:10])
        
        # 检查是否触发FES保底（is_fes_pity已在前面定义）
        fes_pity_text = ""
        if is_fes_pity:
            fes_pity_text = "🎉 触发FES保底，今日时运为man！\n"
        
        info_text = (
            f"当前呱太: {current_gacha}\n"
            f"累计次数: {total_draws}\n"
            f"累计3星: {total_3stars}\n"
            f"距离FES保底: {remaining_pity}次\n"
        )
        prompt_text = f"{fes_pity_text}{info_text}\n请输入要开的编号：\n{hint_text}\n输入「全部开」一键开启"

        # 发送消息（文字和图片合成一条消息
        if group_id and user_id:
            at_message = f"<@{user_id}> "
        else:
            at_message = ""

        # 通过QQ API发送图片+文字
        # 如果是自动开箱模式，跳过盲盒阶段直接开启所有盲盒
        if auto_open:
            result = handle_box_open(user_id, group_id, "全部开")
            try:
                os.remove(img_path)
            except:
                pass
            return result

        send_message_with_image(group_id or user_id, f"{at_message}{prompt_text}", img_path)

        # 删除本地图片
        try:
            os.remove(img_path)
        except:
            pass

        log_info(f"盲盒生成 [{user_id}]: {count}个盲盒")

        return jsonify({
            "status": "success",
            "box_count": count,
            "message": prompt_text
        })

    except Exception as e:
        import traceback as _tb
        log_error(f"抽卡处理失败: {e}\n{_tb.format_exc()}")
        # 异常退还呱太：如果扣款后操作失败，退还消耗的呱太
        try:
            if gacha_before is not None and gacha_before >= cost:
                add_gacha(user_id, cost)
                log_info(f"异常退还 [{user_id}]: +{cost} 呱太")
        except Exception as refund_err:
            log_error(f"退还呱太失败: {refund_err}")
        try:
            clear_box_session(user_id)
        except:
            pass
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


# ========== 配队系统 ==========

def handle_team(user_id: str, group_id, cmd: str) -> dict:
    """处理配队相关命令
    队伍 - 显示当前队伍（只显示图片）
    队伍 我的卡 - 显示三星卡图（50张，可翻页，无文字卡名）
    队伍 我的卡 下一页/上一页 - 翻页查看三星卡
    队伍 设置 位置 序号(1-50) - 根据当前页序号设置卡牌
    队伍 设置 战斗位/支援位 位置 序号 - 手动指定类型
    队伍 清除 位置 - 清除该位置的战斗卡和支援卡
    队伍 清空 - 清空所有队伍配置
    队伍 自动配队 - AI自动配队
    队伍 切换 N - 加载预设N (1-11)
    队伍 预设 - 查看所有预设摘要
    """
    try:
        characters = get_characters()
        target_id = group_id or user_id

        # Raid锁定检查辅助函数
        def _check_raid_locked():
            """检查当前活跃槽位是否为已锁定的raid队，返回错误消息或None"""
            try:
                presets_data = load_presets(user_id)
                active_slot = presets_data.get("active_slot", 0)
                if 7 <= active_slot <= 11:
                    try:
                        from rescue_event import load_event_data
                        raid_data = load_event_data()
                        if raid_data.get("active", False):
                            slot_key = f"slot_{active_slot}"
                            player = raid_data.get("players", {}).get(str(user_id), {})
                            if player.get("daily_locked_teams", {}).get(slot_key):
                                return f"该Raid队伍(槽{active_slot})今日已锁定，无法修改。请明日再战或切换其他队伍。"
                    except Exception:
                        pass
            except Exception:
                pass
            return None

        # 获取用户当前查看的页码和筛选状态（从session恢复）
        team_session_file = INFO_DIR / f"team_session_{user_id}.json"
        current_page = 1
        saved_filter_color = None
        saved_filter_type = None
        if team_session_file.exists():
            try:
                with open(team_session_file, "r", encoding="utf-8") as f:
                    session_data = json.load(f)
                    current_page = session_data.get("cards_page", 1)
                    saved_filter_color = session_data.get("filter_color")
                    saved_filter_type = session_data.get("filter_type")
            except Exception:
                current_page = 1

        # 自动配队命令
        if '自动配队' in cmd or '自动' in cmd:
            raid_lock = _check_raid_locked()
            if raid_lock:
                reply = f"<@{user_id}> {raid_lock}" if group_id and user_id else raid_lock
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": raid_lock})

            # RAID槽位自动配队时排除已用角色
            presets = load_presets(user_id)
            active_slot = presets.get("active_slot", 0)
            check_raid = 7 <= active_slot <= 11
            result = auto_build_team(user_id, characters, check_raid_duplicates=check_raid)

            if not result["success"]:
                reply = result["message"]
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": result["message"]})

            # 发队伍图片 + 简短提示
            team_data = load_team_data(user_id)
            img_path = build_team_image(team_data, characters)
            text = f"🤖 自动配队完成"
            if group_id and user_id:
                text = f"<@{user_id}> {text}"
            if img_path:
                send_message_with_image(target_id, text, img_path)
                # 发送后删除临时图片
                try:
                    if img_path and os.path.exists(img_path):
                        os.unlink(img_path)
                except Exception:
                    pass
            else:
                send_message(text, user_id, group_id)
            return jsonify({"status": "success"})

        if '我的卡' in cmd:
            # 解析颜色/类型筛选参数
            # 支持格式: 队伍 我的卡 红/蓝/绿/黄/紫  |  队伍 我的卡 战斗/支援
            # 也可以组合: 队伍 我的卡 红 战斗
            filter_color = None
            filter_type = None
            color_keywords = {'红': '红', '赤': '红', '蓝': '蓝', '青': '蓝',
                              '绿': '绿', '緑': '绿', '黄': '黄', '紫': '紫',
                              '金': '黄', '橙': '橙'}  # 金→黄(用户友好别名)
            # 支持缩写格式: 蓝B=蓝色战斗, 超黄A=超黄支援
            type_abbr_match = _re.search(r'([超]?(?:红|赤|蓝|青|绿|緑|黄|紫|金|橙))\s*([AB])\b', cmd)
            if type_abbr_match:
                color_kw = type_abbr_match.group(1)
                # 解析颜色（处理超X前缀）
                if color_kw.startswith('超'):
                    base_kw = color_kw[1:]
                    base_val = color_keywords.get(base_kw, base_kw)
                    filter_color = '超' + base_val
                else:
                    filter_color = color_keywords.get(color_kw, color_kw)
                # A=支援, B=战斗
                filter_type = "assist" if type_abbr_match.group(2) == 'A' else "battle"
            else:
                for kw, color_val in color_keywords.items():
                    if kw in cmd:
                        if '超' + kw in cmd:
                            filter_color = '超' + color_val
                        else:
                            filter_color = color_val
                        break
                if '战斗' in cmd and '支援' not in cmd:
                    filter_type = "battle"
                elif '支援' in cmd and '战斗' not in cmd:
                    filter_type = "assist"
            # 翻页时保留上次筛选状态（命令中未指定时用session值）
            if filter_color is None and saved_filter_color:
                filter_color = saved_filter_color
            if filter_type is None and saved_filter_type:
                filter_type = saved_filter_type

            # 先获取总页数（用于验证页码，注意筛选后的总数）
            user_cards = get_user_3star_cards(user_id, characters,
                                              filter_color=filter_color,
                                              filter_type=filter_type)
            total_pages = max(1, (len(user_cards) + 50 - 1) // 50)

            # 处理翻页
            page_match = _re.search(r'我的卡\s+(第)?(\d+)(页)?', cmd)
            if page_match:
                target_page = int(page_match.group(2))
                current_page = max(1, min(target_page, total_pages))
            elif '下一页' in cmd:
                current_page += 1
                if current_page > total_pages:
                    current_page = total_pages
            elif '上一页' in cmd:
                current_page -= 1
                if current_page < 1:
                    current_page = 1

            # 确保页码在有效范围内
            current_page = max(1, min(current_page, total_pages))

            # 保存当前页码和筛选状态
            session_data = {"cards_page": current_page}
            if filter_color:
                session_data["filter_color"] = filter_color
            if filter_type:
                session_data["filter_type"] = filter_type
            with open(team_session_file, "w", encoding="utf-8") as f:
                json.dump(session_data, f)

            # 显示用户拥有的三星卡（50张一页，支持颜色/类型筛选）
            img_path, current_cards, total_pages = build_3star_cards_image(
                user_id, characters, current_page, 50,
                filter_color=filter_color, filter_type=filter_type)

            if not current_cards:
                reply = "你还没有三星卡~"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success", "message": "没有三星卡"})

            # 构建消息（只有图片和页码提示）
            page_info = f"第{current_page}/{total_pages}页"
            if filter_color or filter_type:
                filter_desc = ""
                if filter_color:
                    # 输出时将"黄"显示为"金"（避免敏感词）
                    display_color = filter_color.replace("黄", "金")
                    filter_desc += f"属性:{display_color}"
                if filter_type:
                    filter_desc += f" 类型:{'战斗' if filter_type == 'battle' else '支援'}"
                page_info += f" [{filter_desc}]"
            if total_pages > 1:
                if current_page < total_pages:
                    page_info += " | 下一页"
                if current_page > 1:
                    page_info += " | 上一页"
                page_info += " | 队伍 我的卡 页码"

            # 使用提示（根据当前页实际卡牌数量）
            current_page_size = len(current_cards)
            usage_hint = (f"设置: 队伍 设置 位置 序号(1-{current_page_size}) | "
                          f"切换预设: 队伍 切换 1~11 | "
                          f"筛选: 队伍 我的卡 红/蓝/绿/金/紫/超X/B(战斗)/A(支援)")

            if img_path:
                text = f"{page_info}\n{usage_hint}"
                if group_id and user_id:
                    text = f"<@{user_id}> {text}"
                send_message_with_image(target_id, text, img_path)
                # 发送后删除临时图片
                try:
                    if img_path and os.path.exists(img_path):
                        os.unlink(img_path)
                except Exception:
                    pass
            else:
                reply = f"{page_info}\n{usage_hint}"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)

            return jsonify({"status": "success", "message": "显示三星卡", "page": current_page, "total_pages": total_pages})

        elif '设置' in cmd:
            # Raid锁定检查
            raid_lock = _check_raid_locked()
            if raid_lock:
                reply = f"<@{user_id}> {raid_lock}" if group_id and user_id else raid_lock
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": raid_lock})

            # 设置队伍卡牌
            # 匹配格式1: 队伍 设置 位置 序号（使用当前页的序号1-50）
            match_simple = _re.search(r'设置\s+(\d+)\s+(\d+)', cmd)
            # 匹配格式2: 队伍 设置 战斗位/支援位 位置 序号
            match_full = _re.search(r'设置\s+(战斗位|支援位)\s+(\d+)\s+(\d+)', cmd)

            if match_simple and not match_full:
                # 简化格式：使用序号选择卡牌
                position = int(match_simple.group(1))
                card_index = int(match_simple.group(2))  # 序号1-50

                if position < 1 or position > 6:
                    reply = "队伍位置必须在1-6之间！"
                else:
                    # 获取当前页的卡牌列表（使用筛选状态）
                    _, current_cards, _ = build_3star_cards_image(
                        user_id, characters, current_page, 50,
                        filter_color=saved_filter_color,
                        filter_type=saved_filter_type
                    )

                    if card_index < 1:
                        reply = "序号必须大于0！"
                    elif card_index > len(current_cards):
                        reply = f"当前页只有{len(current_cards)}张卡，序号{card_index}无效！"
                    else:
                        card_info = current_cards[card_index - 1]
                        card_id = card_info.get("card_id")
                        card_type = card_info.get("type", "battle")

                        success = set_team_card(user_id, position, card_id, card_type)
                        type_text = "战斗位" if card_type == "battle" else "支援位"
                        if success:
                            auto_save_preset(user_id)
                            reply = f"成功设置{type_text}{position}！"
                        elif success == -1:
                            reply = "设置失败！RAID队伍(7-11)不能使用重复角色~"
                        else:
                            # 检查是否因为重复导致失败
                            team_data = load_team_data(user_id)
                            all_cards = (team_data.get("battle_cards", []) +
                                        team_data.get("assist_cards", []))
                            already_used = any(
                                c and str(c) == str(card_id)
                                for i, c in enumerate(all_cards)
                                if not (i == position - 1 and card_type == "battle") and
                                   not (i == position + 5 and card_type == "assist")
                            )
                            if already_used:
                                reply = "设置失败！该卡牌已在队伍的其他位置使用，不能重复选择同一张卡~"
                            else:
                                reply = "设置失败！"

                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success" if '成功' in reply else "error", "message": reply})

            elif match_full:
                # 完整格式：手动指定类型
                card_type = "battle" if match_full.group(1) == "战斗位" else "assist"
                position = int(match_full.group(2))
                card_index = int(match_full.group(3))  # 序号1-10

                success = False
                if position < 1 or position > 6:
                    reply = "队伍位置必须在1-6之间！"
                elif card_index < 1 or card_index > 10:
                    reply = "序号必须在1-10之间！"
                else:
                    # 获取当前页的卡牌列表
                    _, current_cards, _ = build_3star_cards_image(user_id, characters, current_page, 10)

                    if card_index > len(current_cards):
                        reply = f"当前页只有{len(current_cards)}张卡，序号{card_index}无效！"
                    else:
                        card_info = current_cards[card_index - 1]
                        card_id = card_info.get("card_id")

                        success = set_team_card(user_id, position, card_id, card_type)
                        if success:
                            auto_save_preset(user_id)
                            reply = f"成功设置{match_full.group(1)}{position}！"
                        elif success == -1:
                            reply = "设置失败！RAID队伍(7-11)不能使用重复角色~"
                        else:
                            # 检查是否因为重复导致失败
                            team_data = load_team_data(user_id)
                            all_cards = (team_data.get("battle_cards", []) +
                                        team_data.get("assist_cards", []))
                            already_used = any(
                                c and str(c) == str(card_id)
                                for i, c in enumerate(all_cards)
                                if not (i == position - 1 and card_type == "battle") and
                                   not (i == position + 5 and card_type == "assist")
                            )
                            if already_used:
                                reply = "设置失败！该卡牌已在队伍的其他位置使用，不能重复选择同一张卡~"
                            else:
                                reply = "设置失败！"

                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success" if success else "error", "message": reply})

            else:
                reply = "格式错误！正确格式：队伍 设置 位置 序号(1-10)"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "格式错误"})

        elif '清除' in cmd:
            # Raid锁定检查
            raid_lock = _check_raid_locked()
            if raid_lock:
                reply = f"<@{user_id}> {raid_lock}" if group_id and user_id else raid_lock
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": raid_lock})

            # 清除指定位置的卡牌
            match = _re.search(r'清除\s+(\d+)', cmd)

            if match:
                position = int(match.group(1))

                if position < 1 or position > 6:
                    reply = "位置必须在1-6之间！"
                else:
                    # 清除战斗位和支援位的对应位置
                    clear_team_card(user_id, position, "battle")
                    clear_team_card(user_id, position, "assist")
                    reply = f"成功清除位置{position}的战斗卡和支援卡！"

                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success", "message": reply})

            else:
                reply = "格式错误！正确格式：队伍 清除 位置"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "格式错误"})

        elif '清空' in cmd:
            # Raid锁定检查
            raid_lock = _check_raid_locked()
            if raid_lock:
                reply = f"<@{user_id}> {raid_lock}" if group_id and user_id else raid_lock
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": raid_lock})

            # 清空整个队伍
            clear_all_team(user_id)
            reply = "已清空所有队伍配置！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "队伍已清空"})

        elif '切换' in cmd:
            # 切换预设: 队伍 切换 N
            m = _re.search(r'切换\s*(\d+)', cmd)
            if m:
                slot = int(m.group(1))
                if load_preset(user_id, slot):
                    reply = f"已切换到预设{slot}！"
                    if 7 <= slot <= 11:
                        reply += "\n[RAID队伍槽位 - 此队伍可用于raid挑战]"
                    # 显示新队伍
                    team_data = load_team_data(user_id)
                    img_path = build_team_image(team_data, characters)
                    if img_path:
                        if group_id and user_id:
                            text = f"<@{user_id}> {reply}"
                        else:
                            text = reply
                        send_message_with_image(target_id, text, img_path)
                        # 发送后删除临时图片
                        try:
                            if img_path and os.path.exists(img_path):
                                os.unlink(img_path)
                        except Exception:
                            pass
                    else:
                        if group_id and user_id:
                            reply = f"<@{user_id}> {reply}"
                        send_message(reply, user_id, group_id)
                else:
                    # 预设为空 → 自动配队并保存到此槽位
                    reply = f"预设{slot}为空，正在自动配队...\n"
                    result = auto_build_team(user_id, characters, check_raid_duplicates=(7 <= slot <= 11))
                    if result["success"] and result["team"]:
                        save_team_data(user_id, result["team"])
                        # 保存到此预设槽位
                        presets_data = load_presets(user_id)
                        presets_data["presets"][slot - 1] = {
                            "battle_cards": list(result["team"].get("battle_cards", [])),
                            "assist_cards": list(result["team"].get("assist_cards", []))
                        }
                        presets_data["active_slot"] = slot
                        save_presets(user_id, presets_data)
                        reply += f"预设{slot}自动配队完成！"
                        # 显示新队伍
                        img_path = build_team_image(result["team"], characters)
                        if img_path:
                            if group_id and user_id:
                                text = f"<@{user_id}> {reply}"
                            else:
                                text = reply
                            send_message_with_image(target_id, text, img_path)
                            # 发送后删除临时图片
                            try:
                                if img_path and os.path.exists(img_path):
                                    os.unlink(img_path)
                            except Exception:
                                pass
                        else:
                            if group_id and user_id:
                                reply = f"<@{user_id}> {reply}"
                            send_message(reply, user_id, group_id)
                    else:
                        reply += result.get("message", "自动配队失败，请先抽卡！")
                        if group_id and user_id:
                            reply = f"<@{user_id}> {reply}"
                        send_message(reply, user_id, group_id)
            else:
                reply = "格式：队伍 切换 1~11"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": reply})

        elif '复制' in cmd:
            # 队伍复制 x 到 y
            copy_match = _re.search(r'复制\s*(\d+)\s*到\s*(\d+)', cmd)
            if not copy_match:
                copy_match = _re.search(r'复制\s*(\d+)\s*(\d+)', cmd)
            if copy_match:
                from_slot = int(copy_match.group(1))
                to_slot = int(copy_match.group(2))
                ok = copy_team(user_id, from_slot, to_slot)
                if ok:
                    reply = f"已将队伍 {from_slot} 复制到队伍 {to_slot}"
                else:
                    reply = f"复制失败：槽位无效或源队伍为空"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "success"})
            else:
                reply = "格式：队伍 复制 源槽位 到 目标槽位（1-11）"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "invalid copy format"})

        elif '预设' in cmd:
            # 查看所有预设
            info = list_presets_info(user_id, characters)
            reply = info
            if group_id and user_id:
                reply = f"<@{user_id}>\n{reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "显示预设列表"})

        else:
            # 显示当前队伍（只显示图片，不显示文字信息）
            team_data = load_team_data(user_id)

            # 如果当前是raid槽位(7-11)，读取保存的血量数据用于绘制血条
            raid_hp_data = None
            try:
                active_slot_info = load_presets(user_id)
                active_slot = active_slot_info.get("active_slot", 0)
                if 7 <= active_slot <= 11:
                    try:
                        from rescue_event import load_event_data
                        raid_data = load_event_data()
                        if raid_data.get("active", False):
                            slot_key = f"slot_{active_slot}"
                            raid_player = raid_data.get("players", {}).get(str(user_id), {})
                            raid_hp_data = raid_player.get("team_hp_saved", {}).get(slot_key)
                    except Exception:
                        pass
            except Exception:
                pass

            # 生成队伍图片（raid队时传入血量数据）
            img_path = build_team_image(team_data, characters, hp_data=raid_hp_data)

            if img_path:
                # 显示当前槽位信息
                try:
                    slot_info = load_presets(user_id)
                    current_slot = slot_info.get("active_slot", 0)
                    slot_label = f"第{current_slot}队" if current_slot > 0 else "未选择队伍"
                    if 7 <= current_slot <= 11:
                        slot_label += " (RAID)"
                except Exception:
                    slot_label = ""
                hints = f"当前: {slot_label}\n队伍 预设 | 队伍 切换 1~11 | 队伍 我的卡 | 队伍 自动配队"
                if group_id and user_id:
                    text = f"<@{user_id}> {hints}"
                else:
                    text = hints
                send_message_with_image(target_id, text, img_path)
                # 发送后删除临时图片
                try:
                    if img_path and os.path.exists(img_path):
                        os.unlink(img_path)
                except Exception:
                    pass
            else:
                # 如果没有图片，发送提示
                reply = "队伍配置为空！"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)

            return jsonify({"status": "success", "message": "显示队伍"})

    except Exception as e:
        log_error(f"配队处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)})


def handle_defense_team(user_id: str, group_id, cmd: str) -> dict:
    """处理防守队命令
    防守队         - 查看当前防守队（配队图片）
    防守队 设置 N  - 设置防守队为预设槽位N（1-6）
    """
    try:
        characters = get_characters()
        target_id = group_id or user_id

        # 设置防守队槽位: 防守队 设置 N 或 防守队 N
        set_match = _re.search(r'设置\s*(\d)', cmd)
        if not set_match:
            set_match = _re.search(r'防守队\s+(\d)', cmd)

        if set_match:
            slot = int(set_match.group(1))
            if slot < 1 or slot > 6:
                reply = "防守队槽位必须在1-6之间！"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "无效的槽位"})

            # 检查预设槽位是否有队伍
            presets_data = load_presets(user_id)
            preset = presets_data["presets"][slot - 1]
            if preset is None or not any(preset.get("battle_cards", [])):
                reply = f"预设槽位{slot}为空！请先使用「队伍 切换 {slot}」配置该预设的队伍。"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "预设为空"})

            set_defense_slot(user_id, slot)
            reply = f"🛡️ 防守队已设置为「预设槽位{slot}」！被挑战时将使用该队伍迎战。"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"防守队 [{user_id}]: 设置为预设槽位{slot}")
            return jsonify({"status": "success", "message": f"防守队设置为槽位{slot}"})

        # 查看当前防守队 — 生成配队图片
        defense_team = get_defense_team(user_id)
        defense_slot = get_defense_slot(user_id)

        # 生成防守队配队图片
        img_path = build_team_image(defense_team, characters)
        if img_path:
            text = f"🛡️ 防守队 (预设槽{defense_slot})"
            if group_id and user_id:
                text = f"<@{user_id}> {text}"
            send_message_with_image(target_id, text, img_path)
            # 发送后删除临时图片
            try:
                if img_path and os.path.exists(img_path):
                    os.unlink(img_path)
            except Exception:
                pass
        else:
            # 图片生成失败，回退到文字
            info = get_defense_team_info(user_id, characters)
            if group_id and user_id:
                info = f"<@{user_id}> {info}"
            send_message(info, user_id, group_id)

        return jsonify({"status": "success", "message": "显示防守队"})

    except Exception as e:
        log_error(f"防守队处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)})


def handle_limited_gacha(user_id: str, group_id):
    """
    限定池十连：15000呱太，至少必出一个FES限定或期間限定
    每用户8分钟CD
    """
    gacha_before = None  # 用于异常退还
    try:
        cost = LIMITED_GACHA_COST
        count = 10

        # 检查呱太
        current_gacha = get_gacha_count(user_id)
        if current_gacha < cost:
            reply = f"呱太不足！当前呱太: {current_gacha}，需要: {cost}。请艾特我发送「获取呱太」获得{GET_GACHA_REWARD}呱太~"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"限定池失败 [{user_id}]: 呱太不足 {current_gacha}/{cost}")
            return jsonify({"status": "error", "message": "呱太不足", "current_gacha": current_gacha, "required_gacha": cost})

        # 冷却检查（8分钟）
        now_ts = datetime.now().timestamp()
        last_limited = LIMITED_GACHA_COOLDOWN.get(user_id, 0)
        remaining = LIMITED_GACHA_COOLDOWN_SECONDS - (now_ts - last_limited)
        if remaining > 0:
            mins = int(remaining // 60)
            secs = int(remaining % 60)
            reply = f"限定池冷却中！请等待 {mins}分{secs}秒后再试~"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"限定池冷却 [{user_id}]: 还需 {mins}分{secs}秒")
            return jsonify({"status": "cooldown", "message": "限定池冷却中", "remaining_seconds": int(remaining)})

        # 记录扣款前呱太数量（用于异常退还）
        gacha_before = get_gacha_count(user_id)

        # 消耗呱太 + 记录冷却
        spend_gacha(user_id, cost)
        LIMITED_GACHA_COOLDOWN[user_id] = now_ts

        # 加载角色
        characters = get_characters()
        if not characters:
            return jsonify({"status": "error", "message": "无法加载角色数据"})

        # 检查保底
        pity_data = load_pity_data(user_id)
        fes_pity_count = pity_data.get("fes_pity_count", 0)
        is_fes_pity = fes_pity_count >= PITY_LIMIT

        # 抽取10个盲盒
        boxes = []
        has_limited = False
        for i in range(count):
            is_pity_draw = is_fes_pity and (i == count - 1)
            box = draw_mystery_box(characters, user_id, is_pity_draw)
            # 检查是否已经是FES/期间限定
            if not box.get("is_mystery"):
                char = box.get("character", {})
                if char.get("stars") == 3 and char.get("limit_type", "") in ("フェス限定", "期間限定"):
                    has_limited = True
            else:
                # 黑盒可能开出3星限定
                pass
            boxes.append(box)

        # 限定保底：如果没有FES/期间限定，替换最后一个非保底盲盒
        if not has_limited and not is_fes_pity:
            limited_chars = [c for c in characters if c["stars"] == 3 and c.get("limit_type", "") in ("フェス限定", "期間限定")]
            if limited_chars:
                # 从后往前找可替换的盲盒（优先替换1星）
                replace_idx = -1
                for i in range(len(boxes) - 1, -1, -1):
                    if boxes[i].get("stars") == 1 and not boxes[i].get("is_mystery"):
                        replace_idx = i
                        break
                if replace_idx < 0:
                    # 没有1星，找2星
                    for i in range(len(boxes) - 1, -1, -1):
                        if not boxes[i].get("is_mystery") and boxes[i].get("stars") != 3:
                            replace_idx = i
                            break
                if replace_idx < 0:
                    replace_idx = count - 1  # fallback: 最后一个

                boxes[replace_idx] = {
                    "stars": 3,
                    "is_mystery": False,
                    "character": random.choice(limited_chars)
                }
                log_info(f"限定池保底触发 [{user_id}]: 替换第{replace_idx+1}个盲盒为限定3星")

        # 创建盲盒会话
        create_box_session(user_id, boxes)
        BOX_SESSIONS[user_id]["characters"] = characters
        BOX_SESSIONS[user_id]["is_fes_pity"] = is_fes_pity

        # 生成盲盒图片
        box_images = []
        for box in boxes:
            img_bytes = create_box_card(box, characters)
            if img_bytes:
                box_images.append(Image.open(BytesIO(img_bytes)))

        if not box_images:
            clear_box_session(user_id)
            return jsonify({"status": "error", "message": "生成盲盒图片失败"})

        card_width, card_height = box_images[0].size
        gap = 18
        cols = 5
        rows = 2
        cards_total_width = card_width * cols + gap * (cols - 1)
        cards_total_height = card_height * rows + gap * (rows - 1)

        bg_path = None
        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
            test_path = LEVEL_DIR / bg_name
            if test_path.exists():
                bg_path = str(test_path)
                break

        output = None
        if bg_path:
            bg_img = Image.open(bg_path).convert('RGB')
            bg_w, bg_h = bg_img.size
            final_w = int(bg_w * 2.0 * 0.264)
            final_h = int(bg_h * 2.0 * 0.264)
            bg_img_resized = bg_img.resize((final_w, final_h), Image.Resampling.LANCZOS)
            output = Image.new('RGB', (final_w, final_h), (50, 50, 50))
            output.paste(bg_img_resized, (0, 0))
            cards_x = (final_w - cards_total_width) // 2
            cards_y = (final_h - cards_total_height) // 2
        else:
            pad = 20
            final_w = cards_total_width + pad * 2
            final_h = cards_total_height + pad * 2
            output = Image.new('RGB', (final_w, final_h), (40, 35, 50))
            cards_x = pad
            cards_y = pad

        for i, img in enumerate(box_images):
            col = i % cols
            row = i // cols
            x = cards_x + col * (card_width + gap)
            y = cards_y + row * (card_height + gap)
            output.paste(img.resize((card_width, card_height), Image.Resampling.LANCZOS), (x, y))

        output_idx = random.randint(1000, 9999)
        img_path = OUTPUT_DIR / f"limited_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.jpg"
        output.convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)

        current_gacha = get_gacha_count(user_id)
        total_draws = get_total_draws(user_id)
        prompt_text = f"🎯 限定池十连 (消耗{cost}呱太) | 余额: {current_gacha}呱太 | 累计: {total_draws}抽"

        at_message = _at_user(user_id, group_id)
        send_message_with_image(group_id or user_id, f"{at_message}{prompt_text}", img_path)

        try:
            os.remove(img_path)
        except:
            pass

        log_info(f"限定池 [{user_id}]: 10盲盒已生成")
        return jsonify({"status": "success", "box_count": count, "message": prompt_text})

    except Exception as e:
        import traceback as _tb2
        log_error(f"限定池处理失败: {e}\n{_tb2.format_exc()}")
        # 异常退还呱太：如果扣款后操作失败，退还消耗的呱太
        try:
            if gacha_before is not None and gacha_before >= cost:
                add_gacha(user_id, cost)
                log_info(f"限定池异常退还 [{user_id}]: +{cost} 呱太")
        except Exception as refund_err:
            log_error(f"限定池退还呱太失败: {refund_err}")
        try:
            clear_box_session(user_id)
        except:
            pass
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_sannoujo(user_id: str, group_id):
    """处理三王女命令，输出十个ID为207832001的三星卡"""
    try:
        characters = get_characters()
        if not characters:
            reply = "无法加载角色数据！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "无法加载角色数据"})
        
        # 查找ID为207832001的角色
        target_card = next((c for c in characters if str(c.get("card_id")) == "207832001"), None)
        if not target_card:
            reply = "找不到ID为207832001的角色！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "找不到目标角色"})
        
        # 生成10张相同的卡
        card_imgs = []
        for _ in range(10):
            card_bytes = composite_card(target_card)
            if card_bytes:
                from io import BytesIO
                card_img = Image.open(BytesIO(card_bytes))
                card_imgs.append(card_img)
        
        if not card_imgs:
            reply = "无法生成卡片图片！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "无法生成卡片图片"})
        
        # 创建展示图片（使用十连背景，与抽卡一致）
        bg_path = None
        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
            test_path = LEVEL_DIR / bg_name
            if test_path.exists():
                bg_path = str(test_path)
                break
        
        if bg_path:
            # 使用背景图片，放大到原来的2倍
            bg_img = Image.open(bg_path).convert('RGB')
            bg_w, bg_h = bg_img.size
            final_w = int(bg_w * 1.5 * 0.264)
            final_h = int(bg_h * 1.5 * 0.264)
            bg = bg_img.resize((final_w, final_h), Image.LANCZOS)
            bg_w, bg_h = final_w, final_h
        else:
            bg = Image.new('RGB', (600, 400), (50, 50, 50))
            bg_w, bg_h = bg.size
        
        # 使用第一张卡的实际比例作为基准
        first_card = card_imgs[0]
        orig_w, orig_h = first_card.size
        max_card_width = 90
        max_card_height = 120
        scale = min(max_card_width / orig_w, max_card_height / orig_h)
        card_width = int(orig_w * scale)
        card_height = int(orig_h * scale)

        gap = 18
        cols = 5
        rows = 2
        
        total_w = cols * (card_width + gap) - gap
        total_h = rows * (card_height + gap) - gap
        start_x = (bg_w - total_w) // 2
        start_y = (bg_h - total_h) // 2
        
        # 创建最终画布（与抽卡一致）
        output = Image.new('RGB', (bg_w, bg_h), (50, 50, 50))
        output.paste(bg, (0, 0))
        
        for i, img in enumerate(card_imgs):
            col = i % cols
            row = i // cols
            x = start_x + col * (card_width + gap)
            y = start_y + row * (card_height + gap)
            
            # 使用thumbnail保持原比例
            img_copy = img.copy()
            img_copy.thumbnail((card_width, card_height), Image.LANCZOS)
            thumb_w, thumb_h = img_copy.size
            
            # 居中粘贴
            paste_x = x + (card_width - thumb_w) // 2
            paste_y = y + (card_height - thumb_h) // 2
            
            output.paste(img_copy, (paste_x, paste_y))
        
        # 保存图片
        import random
        output_idx = random.randint(1000, 9999)
        img_path = OUTPUT_DIR / f"sannoujo_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
        output.convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)
        
        # 发送消息
        at_message = _at_user(user_id, group_id)
        send_message_with_image(group_id or user_id, f"{at_message}👑 三王女降临！", img_path)
        
        # 删除临时图片
        try:
            img_path.unlink()
        except:
            pass
        
        log_info(f"三王女命令 [{user_id}]")
        return jsonify({"status": "success", "message": "三王女卡生成成功"})
    
    except Exception as e:
        log_error(f"三王女命令失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_box_open(user_id: str, group_id: str, open_input: str):
    """处理盲盒开启请求"""
    try:
        session = get_box_session(user_id)
        if not session:
            return None  # 没有盲盒会话，不回复

        boxes = session["boxes"]
        characters = session.get("characters") or get_characters()
        opened = session["opened"].copy()

        # 验证输入
        valid, result = is_valid_box_index(boxes, open_input, session["opened"])
        if not valid:
            reply = f"输入无效：{result}\n请输入数字（如1、2、3）或「全部开」"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": result})

        indices = result
        # 过滤掉已经开过的
        indices = [i for i in indices if i not in opened]
        
        if not indices:
            reply = "这些都已经开过了！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "都已开过"})

        # 记录信息
        new_opened = []
        opened_results = []
        mutation_messages = []  # 收集突变信息

        # 记录碎片获得
        red_crystal_gained = 0
        blue_crystal_gained = 0
        auto_conversion_messages = []  # 3星重复自动转化提示

        # 开盲盒
        log_info(f"用户 {user_id} 开启盲盒，共 {len(indices)} 个，索引: {indices}")
        for idx in indices:
            box = boxes[idx]
            log_info(f"处理盲盒 {idx+1}: stars={box.get('stars')}, is_mystery={box.get('is_mystery')}, opened={box.get('opened')}")
            
            # 先开黑色盲盒获取角色
            if box["is_mystery"] and not box.get("opened"):
                box = open_mystery_box(box, characters)
                boxes[idx] = box
            
            # 应用突变（仅非黑色盲盒），突变后重新抽对应星级的卡
            box, mutated, mutation_text = apply_mutation(box, characters)
            boxes[idx] = box
            
            # 记录突变信息
            if mutated and mutation_text:
                mutation_messages.append(f"#{idx+1}发生了{mutation_text}突变！")
            
            # 获取角色（确保有角色）
            character = box.get("character")
            if not character:
                # 正常盲盒需要获取角色
                stars = box["stars"]
                log_info(f"盲盒 {idx+1} 没有角色，stars={stars}，需要随机选择")
                available = [c for c in characters if c["stars"] == stars]
                if available:
                    character = random.choice(available)
                else:
                    character = random.choice(characters)
                box["character"] = character
                boxes[idx] = box
            
            stars = box["stars"]
            log_info(f"盲盒 {idx+1} 最终 stars={stars}，character={character.get('name') if character else 'None'}")
            
            # 更新卡片收藏和数量统计
            card_id = str(character.get("card_id", ""))
            limit_type = character.get("limit_type", "")
            chara_name = character.get("name", "")
            
            # FES统计和提示
            fes_message = ""
            if stars == 3 and limit_type == "フェス限定":
                # 增加FES统计
                fes_count = increment_fes_count(card_id, chara_name)
                fes_message = f"✨ 恭喜！这是全服第{fes_count}个「{chara_name}」！"
            
            # 碎片转化逻辑
            if stars == 1:
                # 1星卡转化为红色碎片
                add_red_crystal(user_id, 1)
                red_crystal_gained += 1
                log_info(f"用户 {user_id} 获得红色碎片 +1（1星卡）")
            elif stars == 2:
                # 2星卡转化为蓝色碎片
                add_blue_crystal(user_id, 1)
                blue_crystal_gained += 1
                log_info(f"用户 {user_id} 获得蓝色碎片 +1（2星卡）")
                # 更新二星数量
                add_card_collection(user_id, card_id, chara_name, stars, limit_type)
            elif stars == 3:
                # 3星卡添加到最近获得记录和收藏
                add_recent_3star(
                    user_id,
                    card_id,
                    str(character.get("chara_id", "")),
                    chara_name,
                    limit_type
                )
                collection = add_card_collection(user_id, card_id, chara_name, stars, limit_type)
                # 自动转化：已有≥6张重复，再抽到同一3星卡→100蓝碎片
                dup_count = collection.get(card_id, {}).get("count", 0)
                if dup_count >= 7:
                    add_blue_crystal(user_id, 100)
                    blue_crystal_gained += 100
                    auto_conversion_messages.append(f"🔄 {chara_name} (x{dup_count}) → 100蓝碎片")
                    log_info(f"用户 {user_id} 3星重复自动转化: {chara_name} x{dup_count} → +100蓝碎片")
            
            # 更新抽卡记录
            got_3star = (stars == 3)
            is_fes_3star = got_3star and (limit_type == "フェス限定")
            update_pity(user_id, got_3star, is_fes_3star)
            
            # 如果是FES角色，添加到结果中
            if fes_message:
                opened_results.append({
                    "index": idx + 1,
                    "stars": stars,
                    "name": chara_name,
                    "fes_message": fes_message
                })
            else:
                opened_results.append({
                    "index": idx + 1,
                    "stars": stars,
                    "name": chara_name
                })
            
            new_opened.append(idx)

        # 更新会话
        session["boxes"] = boxes
        session["opened"].extend(new_opened)

        # 构建结果文本（不显示单个卡片的碎片转化）
        stars_display = {1: "⭐", 2: "⭐⭐", 3: "⭐⭐⭐"}
        result_lines = []
        for r in opened_results:
            result_lines.append(f"{r['index']}. {stars_display.get(r['stars'], '⭐')} {r['name']}")

        result_text = "\n".join(result_lines)
        
        # 添加碎片获得汇总
        crystal_summary = ""
        if red_crystal_gained > 0 or blue_crystal_gained > 0:
            parts = []
            if red_crystal_gained > 0:
                parts.append(f"🔴红色碎片 x{red_crystal_gained}")
            if blue_crystal_gained > 0:
                parts.append(f"🔵蓝色碎片 x{blue_crystal_gained}")
            crystal_summary = f"\n本次获得: {' + '.join(parts)}"

        # 生成汇总图片（已开的显示角色，未开的不变）
        all_opened = session["opened"]
        summary_img_bytes = create_box_summary_image(boxes, all_opened, characters)
        
        img_path = None
        if summary_img_bytes:
            output_idx = random.randint(1000, 9999)
            img_path = OUTPUT_DIR / f"gacha_open_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            with open(img_path, "wb") as f:
                f.write(summary_img_bytes)

        # 发送结果（突变提示、抽卡结果、图片合成一条消息）
        if group_id and user_id:
            at_message = f"<@{user_id}> "
        else:
            at_message = ""

        # 保存详细结果到会话
        session["last_results"] = {
            "results": opened_results,
            "mutations": mutation_messages,
            "red_crystal": red_crystal_gained,
            "blue_crystal": blue_crystal_gained,
            "auto_conversions": auto_conversion_messages
        }

        # 检查是否全部开完
        remaining = len(boxes) - len(all_opened)
        if remaining > 0:
            remaining_hint = f"\n还有{remaining}个未开，输入「剩下的全部开」可以一键开启"
            box_hints = [f"选择{i+1}" for i in range(len(boxes)) if i not in all_opened]
            hint = " | ".join(box_hints[:5])
            if len(box_hints) > 5:
                hint += "\n" + " | ".join(box_hints[5:])
            remaining_hint += f"\n{hint}"
        else:
            remaining_hint = "\n所有已开完！"
            # 延迟清除会话，给用户时间查看详细信息
            # 不在此处清除，让详细信息查询可以访问

        # 收集FES消息
        fes_messages = [r.get("fes_message") for r in opened_results if r.get("fes_message")]
        fes_text = "\n".join(fes_messages) if fes_messages else ""

        # 自动转化文本
        auto_conv_text = "\n".join(auto_conversion_messages) if auto_conversion_messages else ""
        if auto_conv_text:
            auto_conv_text = "\n" + auto_conv_text

        # 合成消息（不显示详细文字信息，只显示图片和简短提示）
        short_text = f"开了{len(new_opened)}个！{fes_text}{auto_conv_text}{crystal_summary}{remaining_hint}\n输入「详细信息」查看详情"

        # 合成一条消息发送
        if img_path:
            send_message_with_image(group_id or user_id, f"{at_message}{short_text}", img_path)

            try:
                os.remove(img_path)
            except:
                pass
        else:
            # 没有图片时，只发送文字
            send_message(f"{at_message}{complete_text}", user_id, group_id)

        log_info(f"盲盒开启 [{user_id}]: 开了{len(new_opened)}个")

        return jsonify({
            "status": "success",
            "opened_count": len(new_opened),
            "results": opened_results,
            "remaining": remaining
        })

    except Exception as e:
        log_error(f"开箱处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_get_gacha(user_id: str, group_id):
    """处理获取呱太请求（每分钟限1次）"""
    try:
        # 冷却检查
        now_ts = datetime.now().timestamp()
        last_get = GET_GACHA_COOLDOWN.get(user_id, 0)
        remaining = GET_GACHA_COOLDOWN_SECONDS - (now_ts - last_get)
        if remaining > 0:
            reply = f"获取呱太冷却中！请等待 {int(remaining)} 秒后再试~"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"获取呱太冷却 [{user_id}]: 还需等待 {int(remaining)} 秒")
            return jsonify({
                "status": "cooldown",
                "message": "获取呱太冷却中",
                "remaining_seconds": int(remaining)
            })

        # 记录冷却时间戳
        GET_GACHA_COOLDOWN[user_id] = now_ts

        # 添加呱太
        new_gacha = add_gacha(user_id, GET_GACHA_REWARD)
        
        # 构建回复消息
        reply = f"成功获得 {GET_GACHA_REWARD} 呱太！当前呱太: {new_gacha}"
        
        # 发送消息
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)
        
        log_info(f"获取呱太 [{user_id}]: +{GET_GACHA_REWARD} -> {new_gacha}")
        
        return jsonify({
            "status": "success",
            "user_id": user_id,
            "gacha_added": GET_GACHA_REWARD,
            "current_gacha": new_gacha
        })
    
    except Exception as e:
        log_error(f"获取呱太处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_signin(user_id: str, group_id):
    """处理签到请求"""
    try:
        # 执行签到
        result = do_signin(user_id)
        
        if result["success"]:
            # 签到成功
            streak = result["streak"]
            current_gacha = get_gacha_count(user_id)
            
            if streak > 1:
                reply = f"🎉 签到成功！获得 {result['reward']} 呱太！\n当前呱太: {current_gacha}\n连续签到: {streak} 天"
            else:
                reply = f"🎉 签到成功！获得 {result['reward']} 呱太！\n当前呱太: {current_gacha}"
        else:
            # 今天已经签到过
            current_gacha = get_gacha_count(user_id)
            streak = result["streak"]
            reply = f"今天已经签到过了！当前呱太: {current_gacha}\n连续签到: {streak} 天"
        
        # 发送消息
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)
        
        log_info(f"签到 [{user_id}]: success={result['success']}, streak={result['streak']}")
        
        return jsonify({
            "status": "success" if result["success"] else "already_signed",
            "user_id": user_id,
            "streak": result["streak"],
            "reward": result.get("reward", 0),
            "current_gacha": get_gacha_count(user_id)
        })
    
    except Exception as e:
        log_error(f"签到处理失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_personal_info(user_id: str, group_id, page_action: str = None):
    """处理个人记录查询请求（支持分页）"""
    try:
        # 获取用户数据
        total_draws = get_total_draws(user_id)
        total_3stars = get_total_3stars(user_id)
        red_crystal = get_red_crystal(user_id)
        blue_crystal = get_blue_crystal(user_id)
        remaining_pity = get_remaining_pity(user_id)
        current_gacha = get_gacha_count(user_id)
        
        # 获取卡片收藏（所有三星卡，带计数）
        pity_data = load_pity_data(user_id)
        card_collection = pity_data.get("card_collection", {})
        
        # 筛选三星卡并按时间排序（最近到最远）
        three_star_cards = []
        for card_id, info in card_collection.items():
            if info.get("stars") == 3:
                three_star_cards.append({
                    "card_id": card_id,
                    "name": info.get("name", ""),
                    "limit_type": info.get("limit_type", ""),
                    "count": info.get("count", 1),
                    "last_time": info.get("last_time", 0)
                })
        
        # 按时间排序（最近到最远）
        three_star_cards.sort(key=lambda x: x["last_time"], reverse=True)
        
        # 分页处理
        page_size = 10
        current_page = 1
        
        # 从会话获取当前页码
        session = get_box_session(user_id)
        if session and "personal_page" in session:
            current_page = session["personal_page"]
        
        # 根据操作更新页码
        if page_action == 'next':
            current_page += 1
        elif page_action == 'prev':
            current_page = max(1, current_page - 1)
        
        # 计算总页数
        total_pages = max(1, (len(three_star_cards) + page_size - 1) // page_size)
        current_page = max(1, min(current_page, total_pages))
        
        # 更新会话中的页码
        if not session:
            create_box_session(user_id, [])
        BOX_SESSIONS[user_id]["personal_page"] = current_page
        
        # 获取当前页的卡片
        start_idx = (current_page - 1) * page_size
        end_idx = start_idx + page_size
        current_page_cards = three_star_cards[start_idx:end_idx]
        
        # 计算可兑换呱太
        exchange_red = red_crystal * 5
        exchange_blue = blue_crystal * 20
        total_exchange = exchange_red + exchange_blue
        
        # 构建消息
        info_text = (
            f"📊 个人记录\n"
            f"当前呱太: {current_gacha}\n"
            f"累计次数: {total_draws}\n"
            f"累计3星: {total_3stars}\n"
            f"🔴 红色碎片: {red_crystal}\n"
            f"🔵 蓝色碎片: {blue_crystal}\n"
            f"距离保底: {remaining_pity}抽\n\n"
            f"💎 碎片兑换呱太:\n"
            f"红碎片 x{red_crystal} → {exchange_red} 呱太\n"
            f"蓝碎片 x{blue_crystal} → {exchange_blue} 呱太\n"
            f"总计可兑换: {total_exchange} 呱太\n"
            f"输入「兑换」全部兑换 |「兑换红碎片」仅兑红 |「兑换蓝碎片」仅兑蓝"
        )
        
        # 排行榜奖励历史
        rewards_data = load_ranking_rewards()
        player_rewards = rewards_data.get("players", {}).get(user_id)
        if player_rewards and any(player_rewards.values()):
            info_text += f"\n\n🏆 排行榜奖励历史:"
            if player_rewards.get("first", 0) > 0:
                info_text += f"\n🥇 第1名: {player_rewards['first']}次"
            if player_rewards.get("second", 0) > 0:
                info_text += f"\n🥈 第2名: {player_rewards['second']}次"
            if player_rewards.get("third", 0) > 0:
                info_text += f"\n🥉 第3名: {player_rewards['third']}次"

        # 构建三星卡列表（带计数）
        img_path = None
        if three_star_cards:
            info_text += f"\n\n✨ 所有三星卡 (第{current_page}/{total_pages}页):"
            for i, card in enumerate(current_page_cards, 1):
                limit_badge = ""
                if card.get("limit_type") == "期間限定":
                    limit_badge = "🔸"
                elif card.get("limit_type") == "フェス限定":
                    limit_badge = "🔹"
                count = card.get("count", 1)
                count_suffix = f" x{count}" if count > 1 else ""
                info_text += f"\n{i}. {limit_badge}{card['name']}{count_suffix}"
            
            # 创建三星卡展示图片（带计数标记）
            characters = get_characters()
            if characters:
                display_cards = []
                for card in current_page_cards:
                    card_id = card.get("card_id")
                    chara = next((c for c in characters if str(c.get("card_id")) == str(card_id)), None)
                    if chara:
                        display_cards.append({"chara": chara, "count": card.get("count", 1)})
                
                if display_cards:
                    from PIL import Image, ImageDraw, ImageFont
                    from io import BytesIO
                    import random
                    
                    card_imgs = []
                    for item in display_cards:
                        card_bytes = composite_card(item["chara"])
                        if card_bytes:
                            card_img = Image.open(BytesIO(card_bytes))
                            card_imgs.append({"img": card_img, "count": item["count"]})
                    
                    if card_imgs:
                        # 使用十连背景图（与抽卡一致）
                        bg_path = None
                        for bg_name in ["gacha_tmb_bg_10.png", "gacha_tmb_10_bg.png", "gacha_bg_10.png"]:
                            test_path = LEVEL_DIR / bg_name
                            if test_path.exists():
                                bg_path = str(test_path)
                                break
                        
                        if bg_path:
                            # 使用背景图片，放大到原来的1.5倍
                            bg_img = Image.open(bg_path).convert('RGB')
                            bg_w, bg_h = bg_img.size
                            final_w = int(bg_w * 1.5 * 0.264)
                            final_h = int(bg_h * 1.5 * 0.264)
                            bg = bg_img.resize((final_w, final_h), Image.LANCZOS)
                            bg_w, bg_h = final_w, final_h
                        else:
                            bg = Image.new('RGB', (600, 400), (50, 50, 50))
                            bg_w, bg_h = bg.size
                        
                        max_card_width = 90
                        max_card_height = 120
                        gap = 18
                        cols = min(len(card_imgs), 5)
                        rows = (len(card_imgs) + cols - 1) // cols
                        
                        first_card = card_imgs[0]["img"]
                        orig_w, orig_h = first_card.size
                        scale = min(max_card_width / orig_w, max_card_height / orig_h)
                        card_width = int(orig_w * scale)
                        card_height = int(orig_h * scale)
                        
                        total_w = cols * (card_width + gap) - gap
                        total_h = rows * (card_height + gap) - gap
                        start_x = (bg_w - total_w) // 2
                        start_y = (bg_h - total_h) // 2
                        
                        # 创建最终画布（与抽卡一致）
                        output = Image.new('RGB', (bg_w, bg_h), (50, 50, 50))
                        output.paste(bg, (0, 0))
                        
                        for i, item in enumerate(card_imgs):
                            img = item["img"]
                            count = item["count"]
                            col = i % cols
                            row = i // cols
                            x = start_x + col * (card_width + gap)
                            y = start_y + row * (card_height + gap)
                            
                            img_copy = img.copy()
                            img_copy.thumbnail((card_width, card_height), Image.LANCZOS)
                            thumb_w, thumb_h = img_copy.size
                            
                            paste_x = x + (card_width - thumb_w) // 2
                            paste_y = y + (card_height - thumb_h) // 2
                            
                            output.paste(img_copy, (paste_x, paste_y))
                            
                            # 如果计数大于1，在右下角添加计数标记
                            if count > 1:
                                draw = ImageDraw.Draw(output)
                                # 创建半透明背景
                                badge_w = 30
                                badge_h = 20
                                badge_x = paste_x + thumb_w - badge_w - 2
                                badge_y = paste_y + thumb_h - badge_h - 2
                                
                                # 绘制圆角矩形背景
                                draw.rounded_rectangle([badge_x, badge_y, badge_x+badge_w, badge_y+badge_h], radius=4, fill=(0, 0, 0, 200))
                                
                                # 绘制数字
                                try:
                                    font = ImageFont.truetype("arial.ttf", 12)
                                except:
                                    font = ImageFont.load_default()
                                text = f"x{count}"
                                # Pillow >=10.0.0 移除了 textsize，用 textbbox 替代
                                bbox = draw.textbbox((0, 0), text, font=font)
                                text_w = bbox[2] - bbox[0]
                                text_h = bbox[3] - bbox[1]
                                text_x = badge_x + (badge_w - text_w) // 2
                                text_y = badge_y + (badge_h - text_h) // 2
                                draw.text((text_x, text_y), text, fill=(255, 255, 255), font=font)
                        
                        output_idx = random.randint(1000, 9999)
                        img_path = OUTPUT_DIR / f"personal_3stars_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
                        output.convert('RGB').save(img_path, format='JPEG', optimize=True, quality=55)
        
        # 添加分页提示
        if total_pages > 1:
            page_hints = []
            if current_page > 1:
                page_hints.append("输入「上一页」查看上一页")
            if current_page < total_pages:
                page_hints.append("输入「下一页」查看下一页")
            if page_hints:
                info_text += "\n\n" + " | ".join(page_hints)
        
        # 发送消息（文字+图片）
        if group_id and user_id:
            info_text = f"<@{user_id}> {info_text}"
        
        if img_path and img_path.exists():
            send_message_with_image(group_id or user_id, info_text, str(img_path))
            if img_path.exists():
                img_path.unlink()
        else:
            send_message(info_text, user_id, group_id)
        
        log_info(f"查询个人记录 [{user_id}] 第{current_page}/{total_pages}页")
        
        return jsonify({
            "status": "success",
            "user_id": user_id,
            "total_draws": total_draws,
            "total_3stars": total_3stars,
            "red_crystal": red_crystal,
            "blue_crystal": blue_crystal,
            "remaining_pity": remaining_pity,
            "current_gacha": current_gacha,
            "current_page": current_page,
            "total_pages": total_pages
        })
    
    except Exception as e:
        log_error(f"查询个人记录失败: {e}")
        return None  # 消息已通过 send_message 发送，返回 None 避免 botpy 重复回复


def handle_show_details(user_id: str, group_id):
    """处理详细信息查询请求"""
    try:
        session = get_box_session(user_id)
        if not session or "last_results" not in session:
            reply = "没有可查看的详细信息！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "没有详细信息"})

        last_results = session["last_results"]
        results = last_results.get("results", [])
        mutations = last_results.get("mutations", [])
        red_crystal = last_results.get("red_crystal", 0)
        blue_crystal = last_results.get("blue_crystal", 0)
        auto_conversions = last_results.get("auto_conversions", [])

        # 构建详细结果
        stars_display = {1: "⭐", 2: "⭐⭐", 3: "⭐⭐⭐"}
        result_lines = []
        fes_messages = []
        for r in results:
            result_lines.append(f"{r['index']}. {stars_display.get(r['stars'], '⭐')} {r['name']}")
            # 收集FES提示消息
            if r.get("fes_message"):
                fes_messages.append(r["fes_message"])

        result_text = "\n".join(result_lines)

        # 构建消息
        msg_parts = ["详细信息"]
        if mutations:
            msg_parts.append("\n".join(mutations))
        # 添加FES提示消息
        if fes_messages:
            msg_parts.append("\n".join(fes_messages))
        msg_parts.append(f"\n结果:\n{result_text}")
        
        if red_crystal > 0 or blue_crystal > 0:
            crystal_parts = []
            if red_crystal > 0:
                crystal_parts.append(f"🔴红色碎片 x{red_crystal}")
            if blue_crystal > 0:
                crystal_parts.append(f"🔵蓝色碎片 x{blue_crystal}")
            msg_parts.append(f"\n本次获得: {' + '.join(crystal_parts)}")

        if auto_conversions:
            msg_parts.append(f"\n🔄 重复转化:\n" + "\n".join(auto_conversions))

        complete_text = "\n".join(msg_parts)

        # 发送消息
        if group_id and user_id:
            complete_text = f"<@{user_id}> {complete_text}"
        send_message(complete_text, user_id, group_id)

        # 查看详细信息后清除会话
        clear_box_session(user_id)

        log_info(f"查询详细信息 [{user_id}]")

        return jsonify({
            "status": "success",
            "results": results,
            "mutations": mutations,
            "red_crystal": red_crystal,
            "blue_crystal": blue_crystal
        })

    except Exception as e:
        log_error(f"查询详细信息失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_leaderboard(user_id: str, group_id):
    """处理排行榜查询请求"""
    try:
        # 获取排行榜数据
        leaderboard = get_leaderboard()
        
        if not leaderboard:
            reply = "暂无排行榜数据！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "leaderboard": []})
        
        # 构建排行榜消息
        rank_emojis = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]
        
        msg_lines = ["🏆 战力排行榜 🏆"]
        for i, player in enumerate(leaderboard):
            rank = i + 1
            emoji = rank_emojis[i] if i < len(rank_emojis) else f"{rank}."
            
            # 获取用户自己的战力
            if player["user_id"] == user_id:
                power_info = f"⚔️{player['power']} (你)"
            else:
                power_info = f"⚔️{player['power']}"
            
            nick = get_nickname(player['user_id'])
            msg_lines.append(
                f"{emoji} {nick} {power_info}\n"
                f"   累计: {player['total_draws']} | 累计3星: {player['total_3stars']}"
            )
        
        
        reply = "\n".join(msg_lines)
        
        # 发送消息
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)
        
        log_info(f"查询排行榜 [{user_id}]")
        
        return jsonify({
            "status": "success",
            "leaderboard": leaderboard
        })
    
    except Exception as e:
        log_error(f"查询排行榜失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_gacha_leaderboard(user_id: str, group_id):
    """处理抽卡榜单查询请求（按三星个数排行）"""
    try:
        leaderboard = get_gacha_leaderboard()

        if not leaderboard:
            reply = "暂无抽卡数据！"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "leaderboard": []})

        rank_emojis = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣", "6️⃣", "7️⃣", "8️⃣", "9️⃣", "🔟"]

        msg_lines = ["🎰 抽卡榜单 TOP 10 🎰"]
        msg_lines.append("（按三星个数排行）\n")

        for i, player in enumerate(leaderboard):
            rank = i + 1
            emoji = rank_emojis[i] if i < len(rank_emojis) else f"{rank}."

            # 标记查询者
            is_self = player["user_id"] == user_id
            self_mark = " 👈(你)" if is_self else ""

            # 格式化三星率
            rate_str = f"{player['rate']:.1f}%" if player['total_draws'] > 0 else "0.0%"

            nick = get_nickname(player['user_id'])
            msg_lines.append(
                f"{emoji} {nick}{self_mark}\n"
                f"   ⭐三星: {player['total_3stars']} | 🎫总抽: {player['total_draws']} | 📊三星率: {rate_str}"
            )

        reply = "\n".join(msg_lines)

        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)

        log_info(f"查询抽卡榜单 [{user_id}]")

        return jsonify({
            "status": "success",
            "leaderboard": leaderboard
        })

    except Exception as e:
        log_error(f"查询抽卡榜单失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_exchange_crystal(user_id: str, group_id, crystal_type: str = None):
    """处理碎片兑换呱太请求
    :param crystal_type: None=全部, "red"=仅红碎片, "blue"=仅蓝碎片
    """
    try:
        # 获取用户碎片数量
        red_crystal = get_red_crystal(user_id)
        blue_crystal = get_blue_crystal(user_id)

        # 根据类型决定兑哪些碎片
        if crystal_type == "red":
            if red_crystal == 0:
                reply = "你没有红色碎片可以兑换！"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "没有红碎片"})
            exchange_red = red_crystal
            exchange_blue = 0
        elif crystal_type == "blue":
            if blue_crystal == 0:
                reply = "你没有蓝色碎片可以兑换！"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "没有蓝碎片"})
            exchange_red = 0
            exchange_blue = blue_crystal
        else:
            # 全部兑换（现有逻辑）
            if red_crystal == 0 and blue_crystal == 0:
                reply = "你没有可以兑换的碎片！"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "没有碎片可兑换"})
            exchange_red = red_crystal
            exchange_blue = blue_crystal

        # 计算兑换数量（红碎片 1:5呱太，蓝碎片 1:20呱太）
        red_amount = exchange_red * 5
        blue_amount = exchange_blue * 20
        total_amount = red_amount + blue_amount

        # 增量扣除碎片（load → 修改特定字段 → atomic save）
        pity_data = load_pity_data(user_id)
        if exchange_red > 0:
            pity_data["red_crystal"] = max(0, pity_data.get("red_crystal", 0) - exchange_red)
        if exchange_blue > 0:
            pity_data["blue_crystal"] = max(0, pity_data.get("blue_crystal", 0) - exchange_blue)
        save_pity_data(user_id, pity_data)

        # 增量添加呱太
        add_gacha(user_id, total_amount)

        # 构建回复消息
        parts = []
        if exchange_red > 0:
            parts.append(f"🔴红色碎片 x{exchange_red} → {red_amount} 呱太")
        if exchange_blue > 0:
            parts.append(f"🔵蓝色碎片 x{exchange_blue} → {blue_amount} 呱太")

        parts_str = '\n'.join(parts)
        reply = f"💎 兑换成功！\n{parts_str}\n总共获得: {total_amount} 呱太"

        # 发送消息
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)

        log_info(f"碎片兑换 [{user_id}]: type={crystal_type or 'all'}, "
                 f"red={exchange_red}, blue={exchange_blue}, total={total_amount}")

        return jsonify({
            "status": "success",
            "user_id": user_id,
            "red_crystal_exchanged": exchange_red,
            "blue_crystal_exchanged": exchange_blue,
            "gacha_added": total_amount,
            "current_gacha": get_gacha_count(user_id)
        })
    
    except Exception as e:
        log_error(f"碎片兑换失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


# ========== CDKEY 兑换系统 ==========
def get_cdkey_file(user_id: str) -> Path:
    """获取用户 CDKEY 兑换记录文件路径"""
    return INFO_DIR / f"cdkey_{user_id}.json"


def load_cdkey_data(user_id: str) -> dict:
    """加载用户 CDKEY 兑换记录（文件损坏自动回退到空记录）"""
    f = get_cdkey_file(user_id)
    if f.exists():
        try:
            with open(f, "r", encoding="utf-8") as fp:
                data = json.load(fp)
            # 校验结构
            if not isinstance(data, dict) or not isinstance(data.get("used_keys", []), list):
                log_error(f"用户 {user_id} CDKEY 记录格式异常，重置为空")
                return {"used_keys": []}
            return data
        except (json.JSONDecodeError, IOError) as e:
            log_error(f"用户 {user_id} CDKEY 记录读取失败: {e}，回退空记录")
            return {"used_keys": []}
    return {"used_keys": []}


def save_cdkey_data(user_id: str, data: dict):
    """保存用户 CDKEY 兑换记录（原子写入，崩溃安全）"""
    _atomic_json_save(get_cdkey_file(user_id), data)


def handle_cdkey_redeem(user_id: str, group_id, key: str):
    """处理 CDKEY 兑换请求
    - 每个用户每个 CDKEY 只能兑换一次
    - 资源增量追加，不做全量覆盖
    """
    try:
        # 1. 校验 CDKEY 是否存在
        if not CDKEYS or key not in CDKEYS:
            reply = f"「{key}」不是有效的兑换码~"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "无效兑换码"})

        # 2. 加载用户 CDKEY 记录
        cdkey_data = load_cdkey_data(user_id)
        used_keys = cdkey_data.get("used_keys", [])

        # 3. 检查是否已兑换
        if key in used_keys:
            reply = f"你已经兑换过「{key}」啦~ 每个兑换码只能使用一次哦"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "已兑换"})

        # 4. 获取奖励配置
        reward = CDKEYS[key]
        reward_gacha = int(reward.get("gacha", 0) or 0)
        reward_red = int(reward.get("red_crystal", 0) or 0)
        reward_blue = int(reward.get("blue_crystal", 0) or 0)
        desc = str(reward.get("desc", ""))
        if desc:
            desc = f" ({desc})"

        # 5. 增量发放奖励（load → modify specific fields → atomic save）
        parts = []
        total_gacha = 0

        if reward_gacha > 0:
            gacha_data = load_gacha_data(user_id)
            gacha_data["gacha"] = gacha_data.get("gacha", 0) + reward_gacha
            save_gacha_data(user_id, gacha_data)
            total_gacha += reward_gacha
            parts.append(f"🪙 呱太 +{reward_gacha}")

        if reward_red > 0:
            pity_data = load_pity_data(user_id)
            pity_data["red_crystal"] = pity_data.get("red_crystal", 0) + reward_red
            save_pity_data(user_id, pity_data)
            parts.append(f"🔴 红碎片 +{reward_red}")

        if reward_blue > 0:
            pity_data = load_pity_data(user_id)
            pity_data["blue_crystal"] = pity_data.get("blue_crystal", 0) + reward_blue
            save_pity_data(user_id, pity_data)
            parts.append(f"🔵 蓝碎片 +{reward_blue}")

        if not parts:
            reply = f"兑换码「{key}」没有配置奖励，请联系管理员~"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "无奖励"})

        # 6. 记录兑换（追加 used_keys，不做全量覆盖）
        used_keys.append(key)
        cdkey_data["used_keys"] = used_keys
        save_cdkey_data(user_id, cdkey_data)

        # 7. 回复
        parts_str = "\n".join(parts)
        reply = f"🎁 兑换成功！{desc}\n{parts_str}"
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)

        log_info(f"CDKEY兑换 [{user_id}]: key={key}, gacha={reward_gacha}, "
                 f"red={reward_red}, blue={reward_blue}")

        return jsonify({
            "status": "success",
            "key": key,
            "gacha_added": reward_gacha,
            "red_crystal_added": reward_red,
            "blue_crystal_added": reward_blue,
            "current_gacha": get_gacha_count(user_id),
        })

    except Exception as e:
        log_error(f"CDKEY兑换失败 [{user_id}] key={key}: {e}")
        reply = f"兑换失败，请稍后再试~"
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)
        return jsonify({"status": "error", "message": str(e)})


# ========== 三星池子命令处理 ==========
def handle_3star_pool(user_id: str, group_id, raw_message: str):
    """
    处理三星池子抽卡请求
    命令格式:
    - 三星池子 - 显示三星池子介绍
    - 红抽 - 使用红色碎片抽卡
    - 蓝抽 - 使用蓝色碎片抽卡
    """
    try:
        # 检测抽卡类型
        if '红抽' in raw_message:
            crystal_type = "red"
            crystal_name = "红色碎片"
            cost = THREE_STAR_POOL_RED_COST
            current = get_red_crystal(user_id)
        elif '蓝抽' in raw_message:
            crystal_type = "blue"
            crystal_name = "蓝色碎片"
            cost = THREE_STAR_POOL_BLUE_COST
            current = get_blue_crystal(user_id)
        else:
            # 显示三星池子介绍
            red_crystal = get_red_crystal(user_id)
            blue_crystal = get_blue_crystal(user_id)
            
            reply = f"""╔══════════════════════════════╗
║     ★★★ 三星池子 ★★★       ║
╠══════════════════════════════╣
║ 消耗说明:                    ║
║ 🔴 红色碎片: {THREE_STAR_POOL_RED_COST}个/抽       ║
║ 🔵 蓝色碎片: {THREE_STAR_POOL_BLUE_COST}个/抽       ║
╠══════════════════════════════╣
║ 你当前拥有:                  ║
║ 🔴 红色碎片: {red_crystal}个            ║
║ 🔵 蓝色碎片: {blue_crystal}个            ║
╠══════════════════════════════╣
║ 可抽取次数:                  ║
║ 🔴 红色碎片: {red_crystal // THREE_STAR_POOL_RED_COST}次          ║
║ 🔵 蓝色碎片: {blue_crystal // THREE_STAR_POOL_BLUE_COST}次          ║
╠══════════════════════════════╣
║ 输入「三星池子红抽」使用红色碎片抽    ║
║ 输入「三星池子蓝抽」使用蓝色碎片抽    ║
╚══════════════════════════════╝
"""
            if group_id and user_id:
                reply = f"<@{user_id}>\n{reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "显示三星池子"})
        
        # 执行抽卡
        result = draw_3star_pool(user_id, crystal_type)
        
        if not result["success"]:
            reply = result["message"]
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": result["message"]})
        
        # 抽卡成功，生成结果图片
        selected = result["character"]
        card_id = str(selected.get("card_id", ""))
        fes_message = result.get("fes_message", "")
        
        # 生成单抽结果图片（使用抽卡结果背景）
        img_bytes = composite_card(selected)
        img_path = None
        
        if img_bytes:
            # 保存图片到临时文件
            output_idx = random.randint(1000, 9999)
            img_path = OUTPUT_DIR / f"3star_pool_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{output_idx}.png"
            with open(img_path, "wb") as f:
                f.write(img_bytes)
        
        # 如果有FES提示，添加到消息中
        at_message = _at_user(user_id, group_id)
        fes_text = f"\n{fes_message}" if fes_message else ""
        remaining = current - cost
        text_msg = f"💎 使用{cost}个{crystal_name}！{fes_text}\n剩余{crystal_name}: {remaining}个"

        if img_path and os.path.exists(img_path):
            send_message_with_image(group_id or user_id, f"{at_message}{text_msg}", img_path)
            if os.path.exists(img_path):
                os.remove(img_path)
        else:
            reply = f"💎 使用{cost}个{crystal_name}！\n{result['message']}{fes_text}\n剩余{crystal_name}: {current - cost}个"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, group_id or user_id)
        
        log_info(f"三星only池 [{user_id}]: type={crystal_type}, card={selected.get('name')}")
        
        return jsonify({
            "status": "success",
            "message": result["message"],
            "character": selected.get("name")
        })
    
    except Exception as e:
        log_error(f"三星池失败: {e}")
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def save_rolling_battle_log(user_id: str, result: dict) -> str:
    """保存战斗日志，只保留最近1次（单文件 battle_{uid}.json，数组存储）"""
    log_path = INFO_DIR / f"battle_{user_id}.json"
    # 添加时间戳
    result["saved_at"] = datetime.now().strftime("%m-%d %H:%M:%S")
    # 只保留最近一次战斗
    logs = [result]
    _atomic_json_save(log_path, logs)
    return f"battle_{user_id} (共1次)"


def handle_battle(user_id: str, group_id, raw_message: str, raw_mentions: list = None):
    """
    处理对战请求
    命令格式:
    - 战斗[1-5] - AI对战，可选难度1~5（默认2）
    - 战斗 @玩家 - 与指定玩家对战
    - 对战说明 - 显示战斗帮助
    """
    try:
        if not BATTLE_SYSTEM_LOADED or BATTLE_INSTANCE is None:
            reply = "战斗系统未加载，请稍后重试"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "战斗系统未加载"})

        if '说明' in raw_message:
            reply = get_battle_help()
            if group_id and user_id:
                reply = f"<@{user_id}>\n{reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "success", "message": "显示战斗帮助"})

        # 解析对手@（使用主循环已解析的 mentions，与挑战命令一致）
        if raw_mentions is None:
            raw_mentions = []
        enemy_user_id = None
        for mid in raw_mentions:
            if mid != str(QQ_BOT_APP_ID) and mid != str(user_id):
                enemy_user_id = mid
                break

        # 解析难度: 战斗5, 战斗 3, 对战2 等
        import re
        diff_match = re.search(r'(?:战斗|对战|决斗)\s*([1-5])', raw_message)
        ai_difficulty = int(diff_match.group(1)) if diff_match else 2

        if enemy_user_id:
            if enemy_user_id == user_id:
                reply = "不能与自己对战！"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "不能与自己对战"})

        # 加载玩家队伍
        player_team = load_team_data(user_id)
        player_battle_cards = player_team.get("battle_cards", [])
        player_has_cards = any(card for card in player_battle_cards)

        # 读取活跃预设信息
        active_slot = 0
        try:
            pdata = load_presets(user_id)
            active_slot = pdata.get("active_slot", 0)
        except Exception:
            pass

        if not player_has_cards:
            reply = "你的队伍还没有配置战斗卡！请先使用「队伍我的卡」命令配置队伍。"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "玩家队伍为空"})

        if enemy_user_id:
            # 对手使用防守队迎战
            enemy_team = get_defense_team(enemy_user_id)
            enemy_battle_cards = enemy_team.get("battle_cards", [])
            enemy_has_cards = any(card for card in enemy_battle_cards)

            if not enemy_has_cards:
                reply = "对手还没有配置防守队！"
                if group_id and user_id:
                    reply = f"<@{user_id}> {reply}"
                send_message(reply, user_id, group_id)
                return jsonify({"status": "error", "message": "对手队伍为空"})

            enemy_name = get_nickname(enemy_user_id)
            enemy_defense_slot = get_defense_slot(enemy_user_id)

            # 生成双方VS配队图
            characters = get_characters_dict()
            vs_img = build_vs_team_image(player_team, enemy_team, characters)
            at_message = _at_user(user_id, group_id)
            # vs_img saved for result merge
        else:
            difficulty_names = {1: "简单", 2: "普通", 3: "困难", 4: "极难", 5: "地狱"}
            enemy_team = generate_ai_team(ai_difficulty)
            enemy_name = f"AI({difficulty_names[ai_difficulty]})"
            characters = get_characters_dict()
            at_message = _at_user(user_id, group_id)
            vs_img = build_vs_team_image(player_team, enemy_team, characters)
            # vs_img saved for result merge

        # 执行战斗
        log_info(f"战斗开始: {user_id} vs {enemy_user_id or 'AI'}")
        result = BATTLE_INSTANCE.start_battle(player_team, enemy_team, challenger="player",
                                              extra_characters={**get_characters_dict(), **BATTLE_CHARACTERS})

        # 保存战斗日志（滚动保留最近3次）
        battle_log_key = save_rolling_battle_log(user_id, result)

        # 生成简短结果
        winner = result.get("winner", "unknown")
        rounds = result.get("rounds", 0)
        
        if winner == "player":
            result_text = f"🏆 胜利！经过 {rounds} 回合的激战，你击败了 {enemy_name}！"
        else:
            result_text = f"💀 失败... 经过 {rounds} 回合的激战，你被 {enemy_name} 击败了..."
        
        # 统计双方存活情况
        player_alive = sum(1 for u in result.get("player_units", []) if u.get("alive") and not u.get("is_assist"))
        enemy_alive = sum(1 for u in result.get("enemy_units", []) if u.get("alive") and not u.get("is_assist"))
        
        result_text += f"\n📊 最终状态: 我方存活 {player_alive}/6, 敌方存活 {enemy_alive}/6"
        if active_slot > 0:
            result_text += f"\n📋 使用预设: 槽{active_slot}"

        log_info(f"BATTLE-SEND: vs_img={vs_img}, exists={os.path.exists(vs_img) if vs_img else 'N/A'}")
        at_message = _at_user(user_id, group_id)
        full_text = f"⚔️ VS {enemy_name}\n{result_text}"
        if vs_img and os.path.exists(vs_img):
            send_message_with_image(group_id or user_id, at_message + full_text, vs_img)
            try:
                os.remove(vs_img)
            except FileNotFoundError:
                pass
        else:
            send_message(at_message + full_text, user_id, group_id)
        
        log_info(f"战斗结束: {user_id} vs {enemy_user_id or 'AI'}, winner={winner}, rounds={rounds}")
        
        return jsonify({
            "status": "success",
            "message": result_text,
            "winner": winner,
            "rounds": rounds,
            "battle_log_key": battle_log_key
        })
    
    except Exception as e:
        log_error(f"战斗失败: {e}")
        reply = f"战斗失败: {str(e)}"
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def handle_boss_battle(user_id: str, group_id, raw_message: str):
    """
    处理BOSS战请求
    命令格式: BOSS战 / boss战
    BOSS由config.py中的BOSS_CARD_ID指定，初始SP=90
    """
    try:
        if not BATTLE_SYSTEM_LOADED or BATTLE_INSTANCE is None:
            reply = "战斗系统未加载，请稍后重试"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "战斗系统未加载"})

        # BOSS战冷却检查
        now_ts = datetime.now().timestamp()
        last_boss = BOSS_BATTLE_COOLDOWN.get(user_id, 0)
        remaining = BOSS_BATTLE_COOLDOWN_SECONDS - (now_ts - last_boss)
        if remaining > 0:
            reply = f"BOSS战冷却中！请等待 {int(remaining)} 秒后再试~"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            log_info(f"BOSS战冷却 [{user_id}]: 还需等待 {int(remaining)} 秒")
            return jsonify({
                "status": "cooldown",
                "message": "BOSS战冷却中",
                "remaining_seconds": int(remaining)
            })

        # 从config获取BOSS卡牌ID（支持热切换，每次reload）
        try:
            import importlib
            import config
            importlib.reload(config)
            BOSS_CARD_ID = getattr(config, 'BOSS_CARD_ID', '100430006')
        except Exception:
            BOSS_CARD_ID = "100430006"  # 默认

        # 加载玩家队伍
        player_team = load_team_data(user_id)
        player_battle_cards = player_team.get("battle_cards", [])
        player_has_cards = any(card for card in player_battle_cards)

        # 读取活跃预设信息
        active_slot = 0
        try:
            pdata = load_presets(user_id)
            active_slot = pdata.get("active_slot", 0)
        except Exception:
            pass

        if not player_has_cards:
            reply = "你的队伍还没有配置战斗卡！请先使用「队伍 我的卡」命令配置队伍。"
            if group_id and user_id:
                reply = f"<@{user_id}> {reply}"
            send_message(reply, user_id, group_id)
            return jsonify({"status": "error", "message": "玩家队伍为空"})

        # 查找BOSS角色
        boss_char = BATTLE_INSTANCE.get_character(str(BOSS_CARD_ID))
        if not boss_char:
            boss_char = BATTLE_INSTANCE._get_fallback_character(str(BOSS_CARD_ID))
        boss_name = boss_char.name if boss_char else "未知BOSS"

        # 构建BOSS队伍用于VS图（位置1=中间，单卡+5空位，无A卡）
        boss_team = {
            "battle_cards": [None, str(BOSS_CARD_ID)] + [None] * 4,
            "assist_cards": [None] * 6
        }

        # 构建BOSS队伍用于VS图
        boss_team = {"battle_cards": [None, str(BOSS_CARD_ID)] + [None] * 4, "assist_cards": [None] * 6}
        characters = get_characters_dict()
        vs_img = build_vs_team_image(player_team, boss_team, characters)

        # 执行BOSS战
        at_message = _at_user(user_id, group_id)
        log_info(f"BOSS战开始: {user_id} vs BOSS({BOSS_CARD_ID} {boss_name})")
        # 清理超过24小时的过期冷却条目，防止内存泄漏
        _now = datetime.now().timestamp()
        _expire = 24 * 3600  # 24小时
        expired_keys = [uid for uid, ts in BOSS_BATTLE_COOLDOWN.items() if _now - ts >= _expire]
        for uid in expired_keys:
            BOSS_BATTLE_COOLDOWN.pop(uid, None)
        BOSS_BATTLE_COOLDOWN[user_id] = _now
        result = BATTLE_INSTANCE.start_boss_battle(
            player_team, str(BOSS_CARD_ID), initial_sp=300,
            extra_characters={**get_characters_dict(), **BATTLE_CHARACTERS}
        )

        # 保存战斗日志（复用现有）
        save_rolling_battle_log(user_id, result)

        # 格式化并发送结果
        result_text = format_boss_result(result)
        if active_slot > 0:
            result_text += f"\n📋 使用预设: 槽{active_slot}"

        full_text = at_message + f"⚔️ BOSS战 VS {boss_name}\n{result_text}"
        if vs_img and os.path.exists(vs_img):
            send_message_with_image(group_id or user_id, full_text, vs_img)
            try:
                os.remove(vs_img)
            except FileNotFoundError:
                pass
        else:
            send_message(full_text, user_id, group_id)

        log_info(f"BOSS战结束: {user_id}, damage={result.get('damage_dealt')}, pct={result.get('damage_percent')}%")

        return jsonify({
            "status": "success",
            "message": result_text,
            "boss_name": result.get("boss_name", boss_name),
            "damage_dealt": result.get("damage_dealt", 0),
            "damage_percent": result.get("damage_percent", 0),
            "rounds": result.get("rounds", 0)
        })

    except Exception as e:
        log_error(f"BOSS战失败: {e}")
        reply = f"BOSS战失败: {str(e)}"
        if group_id and user_id:
            reply = f"<@{user_id}> {reply}"
        send_message(reply, user_id, group_id)
        return jsonify({"status": "error", "message": str(e)})  # jsonify 返回 None，避免 tuple


def generate_ai_team(difficulty: int = 2) -> dict:
    """
    生成AI队伍
    difficulty 1-5:
      1=简单: 随机选卡，不刻意匹配
      2=普通: 随机但偏向同色同攻 (默认)
      3=困难: 同色同攻匹配，FES优先
      4=极难: 完美同色同攻，大量FES，高分卡
      5=地狱: 全完美匹配，全FES，最高评分
    """
    characters = get_characters()

    # 筛选三星及以上卡牌
    all_battle = [c for c in characters if c.get("type") == "battle" and c.get("stars", 0) >= 3]
    all_assist = [c for c in characters if c.get("type") == "assist" and c.get("stars", 0) >= 3]

    if not all_battle:
        log_error("没有可用的战斗卡！")
        return {"battle_cards": [], "assist_cards": []}

    # 给每张卡打分
    def score_card(c):
        s = c.get("stars", 3) * 100
        s += c.get("attack", 5000) * 0.5 + c.get("hp", 10000) * 0.2 + c.get("dexterity", 1000) * 0.2
        lt = c.get("limit_type", "")
        if lt == "フェス限定":
            s *= 2.0
        elif lt == "期間限定":
            s *= 1.5
        return s

    # 按分数排序
    all_battle.sort(key=score_card, reverse=True)
    all_assist.sort(key=score_card, reverse=True)

    # 根据难度决定选择范围
    # 难度越高，可选池越小（只用最强的卡）
    pool_ratios = {1: 1.0, 2: 0.8, 3: 0.5, 4: 0.3, 5: 0.15}
    ratio = pool_ratios.get(difficulty, 0.8)

    b_count = max(6, int(len(all_battle) * ratio))
    a_count = max(6, int(len(all_assist) * ratio))
    b_pool = all_battle[:b_count]
    a_pool = all_assist[:a_count]

    # 匹配概率: 难度越高越要同色同攻
    match_chance = {1: 0.2, 2: 0.5, 3: 0.8, 4: 0.95, 5: 1.0}[difficulty]
    # FES占比目标
    fes_target = {1: 0.1, 2: 0.3, 3: 0.5, 4: 0.7, 5: 0.9}[difficulty]
    # 选全队统一颜色: 按评分加权随机，保证多样性
    BASE_COLORS = ["红", "绿", "蓝", "黄", "紫"]
    color_scores = {c: 0 for c in BASE_COLORS}
    for card in b_pool:
        attr = card.get("attribute", "")
        base = attr[1:] if attr.startswith("超") else attr
        if base in color_scores:
            color_scores[base] += score_card(card)

    # 加权随机选色: 评分高概率大，但不总选同一个
    total = sum(color_scores.values())
    if total > 0:
        weights = [color_scores[c] / total for c in BASE_COLORS]
        team_color = random.choices(BASE_COLORS, weights=weights, k=1)[0]
    else:
        team_color = "红"

    # 全队同色比例: 锦上添花，不压倒单卡强度
    mono_color_ratio = {1: 0.0, 2: 0.15, 3: 0.3, 4: 0.5, 5: 0.7}[difficulty]
    # 同色同攻权重: 单卡强度为主，完美匹配加分
    perfect_pair_weight = {1: 1.0, 2: 2.0, 3: 3.0, 4: 5.0, 5: 8.0}[difficulty]
    # 难度5: B+A必须同色
    strict_same_color = difficulty >= 5

    battle_cards = []
    assist_cards = []
    used_b = set()
    used_a = set()

    for pos in range(6):
        available_b = [c for c in b_pool if str(c.get("card_id")) not in used_b]
        if not available_b:
            break

        # 全队同色倾向
        if random.random() < mono_color_ratio:
            color_filtered = [c for c in available_b
                            if (c.get("attribute", "")[1:] if c.get("attribute", "").startswith("超") else c.get("attribute", "")) == team_color]
            if color_filtered:
                available_b = color_filtered

        # 难度5: B卡只能选有同色A卡可配的
        if strict_same_color:
            has_match = [c for c in available_b if any(
                ac for ac in a_pool
                if str(ac.get("card_id")) not in used_a
                and ac.get("attribute", "") == c.get("attribute", "")
                and ac.get("attack_type", "") == c.get("attack_type", "")
            )]
            if has_match:
                available_b = has_match

        # 给每张候选B卡打分: 自身分 + 完美A匹配加分
        def b_score(bc):
            s = score_card(bc)
            b_attr = bc.get("attribute", "")
            b_type = bc.get("attack_type", "")
            perfect_count = sum(1 for ac in a_pool
                              if str(ac.get("card_id")) not in used_a
                              and ac.get("attribute", "") == b_attr
                              and ac.get("attack_type", "") == b_type)
            if perfect_count > 0:
                s *= perfect_pair_weight
            return s

        available_b.sort(key=b_score, reverse=True)

        top_n = min(3, len(available_b))
        b_card = available_b[random.randint(0, top_n - 1)]

        battle_cards.append(str(b_card.get("card_id", "")))
        used_b.add(str(b_card.get("card_id", "")))

        # 选A卡: 难度5强制同色同攻，不降级
        b_attr = b_card.get("attribute", "")
        b_type = b_card.get("attack_type", "")

        perfect_a = [c for c in a_pool
                     if str(c.get("card_id")) not in used_a
                     and c.get("attribute", "") == b_attr
                     and c.get("attack_type", "") == b_type]
        same_color_a = [c for c in a_pool
                        if str(c.get("card_id")) not in used_a
                        and c.get("attribute", "") == b_attr]
        any_a = [c for c in a_pool if str(c.get("card_id")) not in used_a]

        if strict_same_color:
            # 难度5: 必须同色同攻，否则同色
            a_card = random.choice(perfect_a) if perfect_a else (random.choice(same_color_a) if same_color_a else None)
        elif random.random() < match_chance:
            if perfect_a:
                a_card = random.choice(perfect_a)
            elif same_color_a:
                a_card = random.choice(same_color_a)
            elif any_a:
                a_card = random.choice(any_a)
            else:
                a_card = None
        else:
            a_card = random.choice(any_a) if any_a else None

        assist_cards.append(str(a_card.get("card_id", "")) if a_card else None)
        if a_card:
            used_a.add(str(a_card.get("card_id", "")))

    # 统计实际颜色
    color_count = {}
    for cid in battle_cards:
        for c in characters:
            if str(c.get("card_id")) == cid:
                attr = c.get("attribute", "")
                base = attr[1:] if attr.startswith("超") else attr
                color_count[base] = color_count.get(base, 0) + 1
                break
    dominant = max(color_count, key=color_count.get) if color_count else "?"

    log_info(f"AI队伍(难度{difficulty} 主色{dominant} {color_count.get(dominant,0)}/6 FES目标{fes_target:.0%}): 战斗{battle_cards} 支援{assist_cards}")
    return {"battle_cards": battle_cards, "assist_cards": assist_cards}

def get_user_team(user_id: str) -> dict:
    """获取用户的队伍配置（封装team_system的load_team_data）"""
    try:
        from team_system import load_team_data
        return load_team_data(user_id)
    except Exception as e:
        log_error(f"获取用户队伍失败: {e}")
        return {"battle_cards": [], "assist_cards": []}


# ========== 排行榜系统 (ranking.py) ==========
from ranking import (init_ranking, load_ranking, save_ranking, load_ranking_rewards,
    save_ranking_rewards, get_player_ranking, get_player_entry, settle_ranking_rewards,
    add_player_to_ranking, update_ranking_after_battle, _get_ranking_text,
    show_ranking, challenge_player, challenge_rank,
    handle_battle_log, _list_cached_gifs)
# ========== 帮助章节 ==========

def _help_overview() -> str:
    """帮助总览 — 列出所有章节供用户选择"""
    return f"""
╔══════════════════════════════╗
║     🃏 小千 帮助总览         ║
╠══════════════════════════════╣
║                              ║
║  📖 请选择章节：              ║
║                              ║
║  🎴 帮助 抽卡 — 抽卡相关     ║
║  ⚔️ 帮助 战斗 — 对战/BOSS/排行║
║  📋 帮助 队伍 — 配队/防守队  ║
║  💰 帮助 经济 — 货币/签到    ║
║  📦 帮助 其他 — 个人记录等   ║
║                              ║
║  例：艾特我发送「帮助 抽卡」  ║
║                              ║
║  常用命令速查：               ║
║  签到  |  单抽  |  十连       ║
║  战斗  | BOSS战 | 挑战 N      ║
║  队伍  | 记录  | 排行榜       ║
║  昵称  | 绑定  | 帮助         ║
╚══════════════════════════════╝
"""


def _help_gacha() -> str:
    """抽卡帮助"""
    return f"""
╔══════════════════════════════╗
║     🎴 抽卡系统              ║
╠══════════════════════════════╣
║                              ║
║  ▸ 单抽                      ║
║    消耗 {GACHA_COST} 呱太，抽取 1 个盲盒  ║
║                              ║
║  ▸ 十连                      ║
║    消耗 {GACHA10_COST} 呱太，抽取10个盲盒 ║
║    保底至少 1 张二星          ║
║    冷却 {GACHA10_COOLDOWN_SECONDS} 秒                ║
║                              ║
║  ▸ 限定十连 / 限定池          ║
║    消耗 {LIMITED_GACHA_COST} 呱太，必出FES/期间限定 ║
║    冷却 {LIMITED_GACHA_COOLDOWN_SECONDS // 60} 分钟               ║
║                              ║
║  ▸ 三星池子 / 红抽 / 蓝抽     ║
║    消耗碎片必出3星            ║
║    红抽: 红色碎片×{THREE_STAR_POOL_RED_COST}            ║
║    蓝抽: 蓝色碎片×{THREE_STAR_POOL_BLUE_COST}            ║
║                              ║
║  ▸ 保底机制                   ║
║    每 {PITY_LIMIT} 抽必出FES限定3星     ║
║    保底进度可在个人记录中查看   ║
║                              ║
║  💡 盲盒需开箱查看结果         ║
╚══════════════════════════════╝
"""


def _help_battle() -> str:
    """战斗帮助"""
    return f"""
╔══════════════════════════════╗
║     ⚔️ 战斗系统              ║
╠══════════════════════════════╣
║                              ║
║  ▸ 战斗 / 对战               ║
║    与AI对战（默认难度2）      ║
║    战斗5 = 地狱难度           ║
║                              ║
║  ▸ 战斗 @玩家                 ║
║    与指定玩家对战             ║
║    对手使用防守队迎战         ║
║                              ║
║  ▸ BOSS战                    ║
║    挑战1500万HP的BOSS         ║
║    限12回合，自动战斗         ║
║    冷却 {BOSS_BATTLE_COOLDOWN_SECONDS} 秒                 ║
║                              ║
║  ▸ 挑战 排名 (1~10)          ║
║    挑战排行榜上的玩家/AI      ║
║    只能挑战比你高≤3位的       ║
║    对手使用防守队迎战         ║
║                              ║
║  ▸ 排行榜 / 排行              ║
║    查看TOP10战力排行榜        ║
║                              ║
║  ▸ 战斗日志 / 战斗GIF         ║
║    查看详细记录 / 生成GIF动画  ║
║                              ║
║  💡 战斗力= FES×10 + 期间×8  ║
║     + 其他3星×7 + 2星×3      ║
╚══════════════════════════════╝
"""


def _help_team() -> str:
    """配队帮助"""
    return f"""
╔══════════════════════════════╗
║     📋 配队系统              ║
╠══════════════════════════════╣
║                              ║
║  ▸ 队伍                      ║
║    查看当前队伍（图片）       ║
║                              ║
║  ▸ 队伍 我的卡               ║
║    查看拥有的3星卡（配队用）  ║
║    筛选: 红/绿/蓝/黄/紫       ║
║    超红/超绿... B/A 支持翻页  ║
║                              ║
║  ▸ 队伍 设置 位置 序号       ║
║    把卡放入队伍指定位置(1~6)  ║
║    例：队伍 设置 1 3         ║
║                              ║
║  ▸ 队伍 清除 位置             ║
║    清除指定位置的卡(1~6)     ║
║                              ║
║  ▸ 队伍 清空                  ║
║    清空整个队伍               ║
║                              ║
║  ▸ 队伍 自动配队              ║
║    AI自动组建最优队伍          ║
║                              ║
║  ▸ 队伍 切换 N (1~11)         ║
║    切换到预设槽位的队伍       ║
║                              ║
║  ▸ 队伍 预设                  ║
║    查看所有6个预设槽位        ║
║                              ║
║  ▸ 防守队                     ║
║    查看当前防守队             ║
║    防守队 设置 N — 设为槽位N  ║
║    (被挑战时使用防守队迎战)   ║
║                              ║
║  💡 前3位=上场位，后3位=后补 ║
╚══════════════════════════════╝
"""


def _help_economy() -> str:
    """经济帮助"""
    return f"""
╔══════════════════════════════╗
║     💰 经济系统              ║
╠══════════════════════════════╣
║                              ║
║  ▸ 获取呱太                   ║
║    获得 {GET_GACHA_REWARD} 呱太             ║
║    冷却 {GET_GACHA_COOLDOWN_SECONDS} 秒                ║
║                              ║
║  ▸ 签到                      ║
║    每日签到获得 {DAILY_REWARD} 呱太       ║
║                              ║
║  ▸ 兑换 / 兑换红碎片/蓝碎片   ║
║    红碎片1:5呱太 蓝碎片1:20    ║
║  ▸ 兑换 ABCDEFG               ║
║    CDKEY兑换码（限一次）       ║
║                              ║
║  📊 消费价格表：              ║
║  ├ 单抽: {GACHA_COST} 呱太           ║
║  ├ 十连: {GACHA10_COST} 呱太          ║
║  ├ 限定十连: {LIMITED_GACHA_COST} 呱太       ║
║  ├ 红抽: 红色碎片×{THREE_STAR_POOL_RED_COST}       ║
║  └ 蓝抽: 蓝色碎片×{THREE_STAR_POOL_BLUE_COST}        ║
║                              ║
║  💡 呱太通过获取/签到/兑换    ║
║     三种渠道获得              ║
╚══════════════════════════════╝
"""


def _help_other() -> str:
    """其他功能帮助"""
    return """
╔══════════════════════════════╗
║     📦 其他功能              ║
╠══════════════════════════════╣
║                              ║
║  ▸ 昵称 XX                   ║
║    设置你的昵称（排行榜显示） ║
║    昵称 (不带参数) 查看当前   ║
║                              ║
║  ▸ 绑定 QQ号                 ║
║    从旧Bot迁移数据到新账号    ║
║    例: 绑定 3590876913       ║
║                              ║
║  ▸ 个人记录 / 记录            ║
║    查看个人统计:              ║
║    总抽数、3星数、保底进度    ║
║    卡牌收藏、战力图、最近3星  ║
║    支持翻页：下一页/上一页    ║
║                              ║
║  ▸ 抽卡排行                  ║
║    查看三星数量TOP10排行      ║
║                              ║
║  ▸ 三王女 / 详细信息          ║
║    隐藏彩蛋 & 卡牌详情        ║
║                              ║
║  💡 邮件反馈问题和建议        ║
╚══════════════════════════════╝
"""


def handle_help(user_id: str, group_id, cmd: str = ""):
    """分章节帮助系统"""
    chapter = ""
    for kw in ["抽卡", "gacha", "战斗", "对战", "battle", "队伍", "配队",
               "team", "经济", "economy", "货币", "其他", "other"]:
        if kw in cmd:
            chapter = kw
            break

    if chapter in ("抽卡", "gacha"):
        help_text = _help_gacha()
    elif chapter in ("战斗", "对战", "battle"):
        help_text = _help_battle()
    elif chapter in ("队伍", "配队", "team"):
        help_text = _help_team()
    elif chapter in ("经济", "economy", "货币"):
        help_text = _help_economy()
    elif chapter in ("其他", "other"):
        help_text = _help_other()
    else:
        help_text = _help_overview()

    at = f"<@{user_id}> " if group_id and user_id else ""
    send_message(at + help_text, user_id, group_id)
    return jsonify({"status": "success", "chapter": chapter or "overview"})


def handle_cute_reply(user_id: str, group_id):
    """被艾特但不是有效命令时，发送卖萌回复"""
    cute_replies = [
        "喵~ 你叫我吗？😺",
        "来了来了~ 有什么事呀？✨",
        "呜哇！被艾特了好开心~ 🥰",
        "主人~ 你在叫我吗？🐾",
        "嗨~ 找我有事吗？😊",
        "嗷呜~ 我在这里！🐺",
        "你好呀~ 今天也要元气满满哦！☀️",
        "咕噜咕噜~ 你是在叫我吗？🐱",
        "嗯嗯~ 我在听呢！👂",
        "好耶！终于被注意到了~ 🎉",
        "哇~ 有人艾特我！开心~ 😄",
        "喵星人收到信号！📡",
    ]
    
    # 随机选择一个卖萌回复
    reply = random.choice(cute_replies)
    
    # 如果是群聊，艾特用户
    if group_id and user_id:
        reply = f"<@{user_id}> {reply}"
    
    # 发送回复
    send_message(reply, user_id, group_id)
    
    return jsonify({
        "status": "success",
        "message": reply
    })


# ========== 启动 ==========
if __name__ == '__main__':
    # PID 文件锁：防止重复启动
    import platform
    PID_FILE = Path("/tmp/qq_bot_ws.pid")
    if platform.system() != "Windows":
        import fcntl
        try:
            _pid_fd = open(PID_FILE, "w")
            fcntl.flock(_pid_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            _pid_fd.write(str(os.getpid()))
            _pid_fd.flush()
        except (IOError, OSError):
            print(f"[FATAL] Bot 已在运行中 (PID锁: {PID_FILE})")
            try:
                with open(PID_FILE) as pf:
                    print(f"  运行中的 PID: {pf.read().strip()}")
            except:
                pass
            print(f"  如需强制重启: kill $(cat {PID_FILE}) && rm {PID_FILE}")
            sys.exit(1)
        atexit.register(lambda: os.unlink(PID_FILE) if PID_FILE.exists() else None)
    else:
        import ctypes
        SYNCHRONIZE = 0x00100000
        PROCESS_QUERY_LIMITED_INFO = 0x1000
        if PID_FILE.exists():
            try:
                old_pid = int(PID_FILE.read_text().strip())
                handle = ctypes.windll.kernel32.OpenProcess(PROCESS_QUERY_LIMITED_INFO, False, old_pid)
                if handle:
                    ctypes.windll.kernel32.CloseHandle(handle)
                    print(f"[FATAL] Bot 已在运行中 (PID: {old_pid})")
                    sys.exit(1)
            except (ValueError, OSError):
                pass
        PID_FILE.write_text(str(os.getpid()))
        atexit.register(lambda: os.unlink(PID_FILE) if PID_FILE.exists() else None)

    log_info("=" * 50)
    log_info("自动抽卡Bot 启动中...")
    log_info(f"PID: {os.getpid()}")
    log_info(f"图标目录: {ICON_DIR}")
    log_info(f"星级图片目录: {LEVEL_DIR}")
    log_info(f"Excel文件: {XLSX_FILE}")
    log_info("=" * 50)

    # 清理上次崩溃残留的临时图片，注册退出时清理
    _cleanup_temp_images()
    atexit.register(_cleanup_temp_images)

    # 每天第一次启动时备份抽卡记录
    backup_pity_records()

    # 预加载抽卡角色数据
    characters = preload_characters()

    # 初始化战斗系统（使用详细战斗数据）
    if BATTLE_SYSTEM_LOADED:
        battle_chars = load_battle_characters()
        BATTLE_INSTANCE = BattleSystem(battle_chars)
        log_info("战斗系统初始化完成")

    # 恢复并记录日活数据
    restore_daily_stats()
    record_daily_dau()

    _start_image_server()

    # 启动 botpy WebSocket Client
    log_info("启动 botpy WebSocket 客户端...")
    intents = Intents(public_messages=True)  # 群聊@消息
    bot = QQBotClient(intents=intents)
    bot.run(appid=QQ_BOT_APP_ID, secret=QQ_BOT_SECRET)
